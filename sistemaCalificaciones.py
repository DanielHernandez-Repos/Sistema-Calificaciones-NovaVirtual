import unicodedata
import streamlit as st
import io
import pandas as pd
import sqlite3
import plotly.express as px  # Para gráficos del reporte
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

DB_FILE = "sistema_calificaciones.db"

# ============================================================================
# 2. ACTUALIZAR LA LISTA DE COLUMNAS_DB (línea aproximada 10-20)
# ============================================================================

COLUMNAS_DB = [
    "matricula", "nombre", "grupo", "materia", "fecha_ingreso_materia", "fecha_ingreso_original", "fecha_recursamiento",
    "cuatrimestre", "carrera", "email_personal", "usuario_email_nova", "contraseña",
    "tipo_asignacion", "n_recursamientos", "calificacion", "fecha_calificacion",
    "profesor", "estatus", "origen_asignacion", "cuatrimestre_historico"  # ← AGREGAR ESTA COLUMNA
]

# ================================================================================================
# SISTEMA DE TUTORÍAS - INTEGRACIÓN PARA modificable44_CorteFuncionaTodo.py
# ================================================================================================
# Este código se debe integrar en el archivo existente modificable44_CorteFuncionaTodo.py

# ================================================================================================
# 1. AGREGAR ESTAS FUNCIONES DESPUÉS DE inicializar_base()
# ================================================================================================

def inicializar_tablas_tutoria():
    """Inicializa las tablas necesarias para el sistema de tutorías"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Tabla de tutores
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS tutores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre_tutor TEXT NOT NULL UNIQUE,
            email_tutor TEXT,
            telefono TEXT,
            activo BOOLEAN DEFAULT 1,
            fecha_registro DATE DEFAULT CURRENT_DATE,
            observaciones TEXT
        )
    """)
    
    # Tabla de asignaciones tutor-grupo
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS asignaciones_tutoria (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tutor_id INTEGER,
            grupo INTEGER,
            carrera TEXT,
            fecha_asignacion DATE DEFAULT CURRENT_DATE,
            activo BOOLEAN DEFAULT 1,
            FOREIGN KEY (tutor_id) REFERENCES tutores(id)
        )
    """)
    
    # Tabla de asignaciones manuales (alumnos sin grupo)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS asignaciones_manuales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            matricula TEXT,
            tutor_id INTEGER,
            fecha_asignacion DATE DEFAULT CURRENT_DATE,
            activo BOOLEAN DEFAULT 1,
            observaciones TEXT,
            FOREIGN KEY (tutor_id) REFERENCES tutores(id)
        )
    """)
    
    conn.commit()
    conn.close()

def obtener_tutores_activos():
    """Obtiene la lista de tutores activos"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT id, nombre_tutor FROM tutores WHERE activo = 1 ORDER BY nombre_tutor")
    tutores = cursor.fetchall()
    conn.close()
    return tutores

def obtener_grupos_disponibles():
    """Obtiene los grupos disponibles de la tabla calificaciones"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT grupo, carrera FROM calificaciones WHERE grupo IS NOT NULL ORDER BY carrera, grupo")
    grupos = cursor.fetchall()
    conn.close()
    return grupos

def obtener_alumnos_sin_grupo():
    """Obtiene alumnos que no tienen grupo asignado"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT matricula, nombre, carrera 
        FROM calificaciones 
        WHERE grupo IS NULL OR grupo = '' 
        ORDER BY nombre
    """)
    alumnos = cursor.fetchall()
    conn.close()
    return alumnos

def obtener_alumnos_por_tutor(tutor_id):
    """Obtiene todos los alumnos asignados a un tutor específico"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Alumnos por grupo
    cursor.execute("""
        SELECT DISTINCT c.matricula, c.nombre, c.grupo, c.carrera, 'Grupo' as tipo_asignacion
        FROM calificaciones c
        INNER JOIN asignaciones_tutoria at ON c.grupo = at.grupo AND c.carrera = at.carrera
        WHERE at.tutor_id = ? AND at.activo = 1 AND c.grupo IS NOT NULL
        
        UNION
        
        SELECT DISTINCT c.matricula, c.nombre, c.grupo, c.carrera, 'Manual' as tipo_asignacion
        FROM calificaciones c
        INNER JOIN asignaciones_manuales am ON c.matricula = am.matricula
        WHERE am.tutor_id = ? AND am.activo = 1
        
        ORDER BY nombre
    """, (tutor_id, tutor_id))
    
    alumnos = cursor.fetchall()
    conn.close()
    return alumnos

def verificar_conflicto_asignacion(grupo, carrera, tutor_id_actual=None):
    """Verifica si ya existe un tutor asignado al grupo/carrera"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    query = "SELECT tutor_id FROM asignaciones_tutoria WHERE grupo = ? AND carrera = ? AND activo = 1"
    params = [grupo, carrera]
    
    if tutor_id_actual:
        query += " AND tutor_id != ?"
        params.append(tutor_id_actual)
    
    cursor.execute(query, params)
    resultado = cursor.fetchone()
    conn.close()
    
    return resultado is not None

# ================================================================================================
# 2. NUEVA FUNCIÓN PARA EXPORTAR A EXCEL (agregar después de verificar_conflicto_asignacion)
# ================================================================================================

def exportar_reporte_excel(reporte_df, filtros_aplicados):
    """Exporta el reporte de tutorías a Excel con formato profesional"""
    try:
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Tutorías"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Título principal
        ws.merge_cells('A1:G1')
        ws['A1'] = "REPORTE DE ASIGNACIONES DE TUTORÍAS"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = center_alignment
        ws['A1'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        # Información de filtros
        row_filtros = 3
        ws[f'A{row_filtros}'] = f"Filtros aplicados: {filtros_aplicados}"
        ws[f'A{row_filtros}'].font = Font(italic=True)
        
        # Fecha y hora del reporte
        ws[f'A{row_filtros + 1}'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws[f'A{row_filtros + 1}'].font = Font(italic=True)
        
        # Estadísticas
        row_stats = row_filtros + 3
        total_alumnos = len(reporte_df['matricula'].unique())
        total_tutores = len(reporte_df['nombre_tutor'].unique())
        total_carreras = len(reporte_df['carrera'].unique())
        
        ws[f'A{row_stats}'] = f"Total Alumnos: {total_alumnos}"
        ws[f'C{row_stats}'] = f"Total Tutores: {total_tutores}"
        ws[f'E{row_stats}'] = f"Total Carreras: {total_carreras}"
        
        # Encabezados de la tabla
        row_headers = row_stats + 2
        headers = ['Tutor', 'Tipo Asignación', 'Carrera', 'Grupo', 'Matrícula', 'Alumno', 'Fecha Asignación']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_headers, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Datos
        for row_idx, (_, row_data) in enumerate(reporte_df.iterrows(), row_headers + 1):
            for col_idx, value in enumerate(row_data[['nombre_tutor', 'tipo_asignacion', 'carrera', 'grupo', 'matricula', 'nombre_alumno', 'fecha_asignacion']], 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if col_idx == 4:  # Columna grupo
                    cell.alignment = center_alignment
        
        # Ajustar anchos de columna
        column_widths = [25, 15, 20, 10, 15, 30, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Crear buffer en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Error al generar Excel: {e}")
        return None

# ========= FUNCIÓN AUXILIAR PARA EJECUTAR EL MOVIMIENTO =========
# ========= FUNCIÓN AUXILIAR PARA EJECUTAR EL MOVIMIENTO =========
# ========= FUNCIÓN AUXILIAR PARA EJECUTAR EL MOVIMIENTO =========
def ejecutar_movimiento_materias(materias_data, cuatrimestre_destino, tipo_operacion, columna_cuatrimestre_seleccionada=None):
    """
    Ejecuta el movimiento de materias al cuatrimestre destino
    
    Args:
        materias_data: Lista de materias (individual) o DataFrame (grupo)
        cuatrimestre_destino: Cuatrimestre destino (string)
        tipo_operacion: "individual" o "grupo"
        columna_cuatrimestre_seleccionada: Columna específica a usar (opcional)
    """
    try:
        # Detectar columna de cuatrimestre
        df_temp = cargar_datos_db()
        columnas_cuatrimestre_disponibles = [col for col in ['cuatrimestre', 'cuatrimestre_historico'] if col in df_temp.columns]
        
        if not columnas_cuatrimestre_disponibles:
            st.error("❌ No se encontraron columnas de cuatrimestre en la base de datos")
            return
        
        # Usar la columna especificada o la primera disponible
        if columna_cuatrimestre_seleccionada and columna_cuatrimestre_seleccionada in columnas_cuatrimestre_disponibles:
            columna_cuatrimestre = columna_cuatrimestre_seleccionada
        elif tipo_operacion == "individual" and len(materias_data) > 0:
            if 'columna_cuatrimestre' in materias_data[0]:
                columna_cuatrimestre = materias_data[0]['columna_cuatrimestre']
            else:
                columna_cuatrimestre = columnas_cuatrimestre_disponibles[0]
        else:
            columna_cuatrimestre = columnas_cuatrimestre_disponibles[0]
        
        st.info(f"🔄 Actualizando columna: **{columna_cuatrimestre}**")
        
        # Crear respaldo
        if 'respaldo_movimiento' not in st.session_state:
            st.session_state["respaldo_movimiento"] = cargar_datos_db()
        
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        registros_actualizados = 0
        errores = []
        
        if tipo_operacion == "individual":
            # Procesar lista de materias seleccionadas
            for materia_info in materias_data:
                try:
                    idx = materia_info['idx']
                    # Obtener el registro original del DataFrame
                    df = cargar_datos_db()
                    registro = df.loc[idx]
                    
                    matricula = str(registro.get('matricula', ''))
                    fecha_ingreso = str(registro.get('fecha_ingreso_materia', ''))
                    materia_nombre = str(registro.get('materia', ''))
                    
                    cursor.execute(f"""
                        UPDATE calificaciones 
                        SET {columna_cuatrimestre} = ?
                        WHERE matricula = ? AND fecha_ingreso_materia = ? AND materia = ?
                    """, (cuatrimestre_destino, matricula, fecha_ingreso, materia_nombre))
                    
                    if cursor.rowcount > 0:
                        registros_actualizados += 1
                    else:
                        errores.append(f"No se pudo actualizar: {materia_nombre}")
                        
                except Exception as e:
                    errores.append(f"Error procesando {materia_info.get('materia', 'materia desconocida')}: {str(e)}")
        
        else:  # tipo_operacion == "grupo"
            # Procesar DataFrame de grupo
            for idx, row in materias_data.iterrows():
                try:
                    matricula = str(row.get('matricula', ''))
                    fecha_ingreso = str(row.get('fecha_ingreso_materia', ''))
                    materia_nombre = str(row.get('materia', ''))
                    
                    cursor.execute(f"""
                        UPDATE calificaciones 
                        SET {columna_cuatrimestre} = ?
                        WHERE matricula = ? AND fecha_ingreso_materia = ? AND materia = ?
                    """, (cuatrimestre_destino, matricula, fecha_ingreso, materia_nombre))
                    
                    if cursor.rowcount > 0:
                        registros_actualizados += 1
                    else:
                        errores.append(f"No se pudo actualizar: {matricula} - {materia_nombre}")
                        
                except Exception as e:
                    errores.append(f"Error procesando registro {idx}: {str(e)}")
        
        # Confirmar cambios
        if registros_actualizados > 0:
            conn.commit()
            st.success(f"✅ Se movieron exitosamente {registros_actualizados} registros al cuatrimestre {cuatrimestre_destino}")
            st.balloons()
            
            if errores:
                st.warning(f"⚠️ Se encontraron {len(errores)} errores:")
                with st.expander("Ver errores"):
                    for error in errores:
                        st.write(f"• {error}")
        else:
            st.error("❌ No se pudo actualizar ningún registro")
            if errores:
                with st.expander("Ver errores"):
                    for error in errores:
                        st.write(f"• {error}")
        
        conn.close()
        
    except Exception as e:
        st.error(f"❌ Error crítico durante la operación: {str(e)}")


# ========= FUNCIÓN DE RESPALDO Y RESTAURACIÓN =========
def mostrar_opciones_respaldo_movimiento():
    """Muestra opciones para respaldar y restaurar después de movimientos"""
    st.markdown("### 🛡️ Gestión de respaldos")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("💾 Crear respaldo manual", key="respaldo_manual_movimiento"):
            try:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                respaldo_file = f"respaldo_movimiento_{timestamp}.csv"
                
                df = cargar_datos_db()
                df.to_csv(respaldo_file, index=False)
                
                st.success(f"✅ Respaldo creado: {respaldo_file}")
            except Exception as e:
                st.error(f"Error creando respaldo: {str(e)}")
    
    with col2:
        if 'respaldo_movimiento' in st.session_state:
            if st.button("⏮️ Restaurar último respaldo", key="restaurar_movimiento"):
                try:
                    # Confirmar restauración
                    if st.checkbox("Confirmo que quiero restaurar el respaldo (perderé cambios recientes)", key="confirmar_restaurar"):
                        conn = sqlite3.connect(DB_FILE)
                        
                        # Limpiar tabla actual
                        cursor = conn.cursor()
                        cursor.execute("DELETE FROM calificaciones")
                        
                        # Restaurar desde respaldo
                        respaldo_df = st.session_state["respaldo_movimiento"]
                        respaldo_df.to_sql('calificaciones', conn, if_exists='append', index=False)
                        
                        conn.commit()
                        conn.close()
                        
                        st.success("✅ Base de datos restaurada exitosamente")
                        st.experimental_rerun()
                        
                except Exception as e:
                    st.error(f"Error restaurando respaldo: {str(e)}")

def exportar_detalle_por_cuatrimestre_excel(matricula, nombre_alumno, conn):
    """
    Exporta el detalle por cuatrimestre EXACTAMENTE como se ve en la interfaz:
    - Una hoja por cada cuatrimestre
    - Estadísticas de aprobadas/reprobadas/promedio
    - Tabla con materias, calificaciones, estatus, fecha y profesor
    """
    try:
        # Obtener información básica del alumno
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DISTINCT nombre, matricula, carrera 
            FROM calificaciones 
            WHERE matricula = ?
            LIMIT 1
        """, (matricula,))
        
        info_alumno = cursor.fetchone()
        if not info_alumno:
            st.error("No se encontró información del alumno")
            return None
            
        nombre_db, matricula_db, carrera = info_alumno
        
        # Consulta para obtener los datos agrupados por cuatrimestre
        cursor.execute("""
            SELECT 
                COALESCE(CAST(cuatrimestre_historico AS TEXT), CAST(cuatrimestre AS TEXT), '1') as cuatrimestre_efectivo,
                materia,
                calificacion,
                estatus,
                fecha_ingreso_materia,
                profesor
            FROM calificaciones 
            WHERE matricula = ?
            AND materia IS NOT NULL 
            AND materia != ''
            AND calificacion IS NOT NULL
            ORDER BY 
                CAST(COALESCE(cuatrimestre_historico, cuatrimestre, 1) AS INTEGER) ASC,
                materia ASC
        """, (matricula,))
        
        datos_raw = cursor.fetchall()
        
        if not datos_raw:
            st.warning("No hay datos para exportar")
            return None
        
        # Agrupar datos por cuatrimestre
        datos_por_cuatrimestre = {}
        for row in datos_raw:
            cuatrimestre, materia, calificacion, estatus, fecha, profesor = row
            
            if cuatrimestre not in datos_por_cuatrimestre:
                datos_por_cuatrimestre[cuatrimestre] = []
            
            # Validar calificación
            try:
                calif_float = float(calificacion) if calificacion is not None else 0.0
            except (ValueError, TypeError):
                calif_float = 0.0
            
            datos_por_cuatrimestre[cuatrimestre].append({
                'materia': str(materia) if materia else 'Sin Materia',
                'calificacion': calif_float,
                'estatus': str(estatus) if estatus else 'Sin Estatus',
                'fecha_ingreso_materia': str(fecha) if fecha else 'Sin Fecha',
                'profesor': str(profesor) if profesor else 'Sin Profesor'
            })
        
        # Crear archivo Excel
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            
            # ===== HOJA RESUMEN GENERAL =====
            resumen_data = []
            for cuatrimestre in sorted(datos_por_cuatrimestre.keys(), key=lambda x: int(x) if x.isdigit() else 999):
                materias = datos_por_cuatrimestre[cuatrimestre]
                total_materias = len(materias)
                aprobadas = len([m for m in materias if m['calificacion'] >= 6.0])
                reprobadas = total_materias - aprobadas
                promedio = sum([m['calificacion'] for m in materias]) / total_materias if total_materias > 0 else 0.0
                
                resumen_data.append({
                    'Cuatrimestre': f"Cuatrimestre {cuatrimestre}",
                    'Total Materias': total_materias,
                    'Aprobadas': aprobadas,
                    'Reprobadas': reprobadas,
                    'Promedio': round(promedio, 2)
                })
            
            df_resumen = pd.DataFrame(resumen_data)
            df_resumen.to_excel(writer, sheet_name='Resumen General', index=False)
            
            # ===== HOJA POR CADA CUATRIMESTRE =====
            for cuatrimestre in sorted(datos_por_cuatrimestre.keys(), key=lambda x: int(x) if x.isdigit() else 999):
                materias = datos_por_cuatrimestre[cuatrimestre]
                
                # Calcular estadísticas del cuatrimestre
                total_materias = len(materias)
                aprobadas = len([m for m in materias if m['calificacion'] >= 6.0])
                reprobadas = total_materias - aprobadas
                promedio = sum([m['calificacion'] for m in materias]) / total_materias if total_materias > 0 else 0.0
                
                # Crear DataFrame con la información del cuatrimestre
                cuatrimestre_data = []
                
                # Agregar fila de estadísticas
                cuatrimestre_data.append({
                    'INFORMACIÓN': f'Cuatrimestre {cuatrimestre} ({total_materias} materias)',
                    'DETALLE': '',
                    'VALOR': '',
                    'COMPLEMENTO': '',
                    'ADICIONAL': ''
                })
                
                cuatrimestre_data.append({
                    'INFORMACIÓN': '✅ Aprobadas',
                    'DETALLE': str(aprobadas),
                    'VALOR': '',
                    'COMPLEMENTO': '',
                    'ADICIONAL': ''
                })
                
                cuatrimestre_data.append({
                    'INFORMACIÓN': '❌ Reprobadas', 
                    'DETALLE': str(reprobadas),
                    'VALOR': '',
                    'COMPLEMENTO': '',
                    'ADICIONAL': ''
                })
                
                cuatrimestre_data.append({
                    'INFORMACIÓN': '📊 Promedio',
                    'DETALLE': str(round(promedio, 2)),
                    'VALOR': '',
                    'COMPLEMENTO': '',
                    'ADICIONAL': ''
                })
                
                # Fila separadora
                cuatrimestre_data.append({
                    'INFORMACIÓN': '',
                    'DETALLE': '',
                    'VALOR': '',
                    'COMPLEMENTO': '',
                    'ADICIONAL': ''
                })
                
                # Encabezados de la tabla de materias
                cuatrimestre_data.append({
                    'INFORMACIÓN': 'MATERIA',
                    'DETALLE': 'CALIFICACIÓN',
                    'VALOR': 'ESTATUS',
                    'COMPLEMENTO': 'FECHA INGRESO',
                    'ADICIONAL': 'PROFESOR'
                })
                
                # Datos de cada materia
                for materia in materias:
                    cuatrimestre_data.append({
                        'INFORMACIÓN': materia['materia'],
                        'DETALLE': materia['calificacion'],
                        'VALOR': materia['estatus'],
                        'COMPLEMENTO': materia['fecha_ingreso_materia'],
                        'ADICIONAL': materia['profesor']
                    })
                
                # Crear DataFrame y exportar
                df_cuatrimestre = pd.DataFrame(cuatrimestre_data)
                sheet_name = f'Cuatrimestre {cuatrimestre}'[:31]  # Excel limita a 31 caracteres
                df_cuatrimestre.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Formatear la hoja (opcional - requiere openpyxl)
                worksheet = writer.sheets[sheet_name]
                
                # Ajustar ancho de columnas
                worksheet.column_dimensions['A'].width = 30  # INFORMACIÓN
                worksheet.column_dimensions['B'].width = 15  # DETALLE
                worksheet.column_dimensions['C'].width = 15  # VALOR
                worksheet.column_dimensions['D'].width = 20  # COMPLEMENTO
                worksheet.column_dimensions['E'].width = 25  # ADICIONAL
            
            # ===== HOJA DE INFORMACIÓN DEL ESTUDIANTE =====
            info_estudiante = {
                'Campo': [
                    'Nombre Completo',
                    'Matrícula', 
                    'Carrera',
                    'Total Cuatrimestres',
                    'Total Materias Cursadas',
                    'Fecha de Exportación'
                ],
                'Valor': [
                    str(nombre_db),
                    str(matricula_db),
                    str(carrera) if carrera else 'No especificada',
                    len(datos_por_cuatrimestre),
                    len(datos_raw),
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ]
            }
            
            df_info = pd.DataFrame(info_estudiante)
            df_info.to_excel(writer, sheet_name='Información Estudiante', index=False)
        
        output.seek(0)
        
        # Nombre del archivo
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M')
        nombre_limpio = str(nombre_db).replace(' ', '_').replace('/', '_')[:30]
        nombre_archivo = f"Detalle_Cuatrimestres_{nombre_limpio}_{matricula}_{fecha_actual}.xlsx"
        
        return {
            'buffer': output.getvalue(),
            'filename': nombre_archivo,
            'total_cuatrimestres': len(datos_por_cuatrimestre),
            'total_materias': len(datos_raw)
        }
        
    except Exception as e:
        st.error(f"❌ Error al exportar: {e}")
        import traceback
        st.error(f"Detalles: {traceback.format_exc()}")
        return None

def exportar_kardex_formato_oficial(matricula, nombre_alumno, conn):
    """
    Genera un kardex con formato oficial IDÉNTICO al archivo original de NovaUniversitas,
    replicando exactamente la estructura, layout y disposición.
    """
    try:
        # Obtener información básica del alumno
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DISTINCT nombre, matricula, carrera, email_personal, usuario_email_nova,
                   fecha_ingreso_original, cuatrimestre
            FROM calificaciones 
            WHERE matricula = ?
            LIMIT 1
        """, (matricula,))
        
        info_alumno = cursor.fetchone()
        if not info_alumno:
            st.error("No se encontró información del alumno en la base de datos")
            return None
            
        nombre_db, matricula_db, carrera, email_personal, email_nova, fecha_ingreso, cuatrimestre_actual = info_alumno
        
        # Obtener todas las materias con calificaciones organizadas por cuatrimestre
        cursor.execute("""
            SELECT 
                materia,
                calificacion,
                estatus,
                fecha_ingreso_materia,
                fecha_calificacion,
                profesor,
                cuatrimestre,
                cuatrimestre_historico,
                tipo_asignacion,
                n_recursamientos
            FROM calificaciones 
            WHERE matricula = ?
            AND materia IS NOT NULL 
            AND materia != ''
            AND calificacion IS NOT NULL
            ORDER BY 
                CAST(COALESCE(cuatrimestre_historico, cuatrimestre, 1) AS INTEGER) ASC,
                fecha_ingreso_materia ASC
        """, (matricula,))
        
        materias_data = cursor.fetchall()
        
        if not materias_data:
            st.warning("No hay materias con calificaciones para generar el kardex")
            return None
        
        # Organizar materias por cuatrimestre histórico
        materias_por_cuatrimestre = {}
        for materia_info in materias_data:
            (materia, calificacion, estatus, fecha_materia, fecha_calif, 
             profesor, cuatr_actual, cuatr_historico, tipo_asig, n_recurs) = materia_info
            
            cuatrimestre_key = str(cuatr_historico) if cuatr_historico and cuatr_historico != "None" else str(cuatr_actual)
            
            if cuatrimestre_key not in materias_por_cuatrimestre:
                materias_por_cuatrimestre[cuatrimestre_key] = []
            
            materias_por_cuatrimestre[cuatrimestre_key].append(materia_info)
        
        # Crear el archivo Excel con formato oficial EXACTO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from datetime import datetime
        import io
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "KARDEX"
        
        # ==================== CONFIGURACIÓN DE FORMATO EXACTO ====================
        
        # Fuentes exactas del formato original
        font_titulo_principal = Font(name='Arial', size=14, bold=True)
        font_subtitulos = Font(name='Arial', size=11, bold=True)
        font_normal = Font(name='Arial', size=9)
        font_small = Font(name='Arial', size=8)
        font_header_table = Font(name='Arial', size=8, bold=True)
        
        # Alineaciones
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        # Bordes exactos del formato
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        
        # ==================== ENCABEZADO PRINCIPAL EXACTO ====================
        
        current_row = 1
        
        # Logo - intentar cargar desde carpeta, sino usar placeholder
        logo_agregado = False
        rutas_logotipo = [
            "assets/logo.png",
            "assets/logo.jpg", 
            "assets/logotipo.png",
            "assets/logotipo.jpg",
            "logo.png",
            "logo.jpg",
            "logotipo.png",
            "logotipo.jpg"
        ]
        
        try:
            # Importar de forma más específica para evitar conflictos
            from openpyxl.drawing.image import Image
            from PIL import Image as PILImage
            import os
            
            st.success("✅ PIL y openpyxl.drawing importados correctamente")
            
            # Intentar encontrar el logotipo en las rutas definidas
            for ruta in rutas_logotipo:
                if os.path.exists(ruta):
                    try:
                        st.info(f"📷 Intentando cargar imagen: {ruta}")
                        
                        # Validar que la imagen se puede abrir con PIL primero
                        test_img = PILImage.open(ruta)
                        test_img.close()
                        
                        # Crear imagen para openpyxl
                        img = Image(ruta)
                        
                        # Configurar tamaño apropiado para el kardex
                        img.width = 80
                        img.height = 100
                        
                        # Posicionar en A1
                        img.anchor = 'A1'
                        
                        # Agregar imagen a la hoja
                        ws.add_image(img)
                        
                        # Fusionar celdas para el espacio del logo pero sin texto
                        ws.merge_cells('A1:A6')
                        ws['A1'] = ""  # Sin texto ya que hay imagen
                        ws['A1'].border = thick_border
                        
                        logo_agregado = True
                        st.success(f"🎉 Logotipo cargado exitosamente desde: {ruta}")
                        break
                        
                    except Exception as e:
                        st.error(f"❌ Error cargando {ruta}: {str(e)}")
                        continue  # Intentar con la siguiente ruta
            
        except ImportError as e:
            st.error(f"❌ Error importando módulos de imagen: {str(e)}")
            st.info("💡 Intenta actualizar openpyxl: pip install --upgrade openpyxl")
        
        # Si no se pudo cargar ningún logotipo, usar placeholder de texto
        if not logo_agregado:
            st.warning("⚠️ No se pudo cargar logotipo, usando placeholder de texto")
            ws.merge_cells('A1:A6')
            ws['A1'] = "NOVA\nUNIVERSITAS"
            ws['A1'].font = Font(name='Arial', size=10, bold=True)
            ws['A1'].alignment = center_alignment
            ws['A1'].border = thick_border
        
        # Título principal - NovaUniversitas
        ws.merge_cells('B1:H1')
        ws['B1'] = "NovaUniversitas"
        ws['B1'].font = font_titulo_principal
        ws['B1'].alignment = center_alignment
        
        # Subtítulo - Organismo Público Descentralizado
        ws.merge_cells('B2:H2')
        ws['B2'] = "ORGANISMO PÚBLICO DESCENTRALIZADO"
        ws['B2'].font = font_normal
        ws['B2'].alignment = center_alignment
        
        # Departamento
        ws.merge_cells('B3:H3')
        ws['B3'] = "DEPARTAMENTO DE SERVICIOS ESCOLARES"
        ws['B3'].font = font_normal
        ws['B3'].alignment = center_alignment
        
        # KARDEX title
        ws.merge_cells('B4:H4')
        ws['B4'] = "KARDEX"
        ws['B4'].font = Font(name='Arial', size=16, bold=True)
        ws['B4'].alignment = center_alignment
        
        # FOTO placeholder (columna I, filas 1-6)
        ws.merge_cells('I1:I6')
        ws['I1'] = "FOTO"
        ws['I1'].font = font_normal
        ws['I1'].alignment = center_alignment
        ws['I1'].border = thick_border
        
        current_row = 7
        
        # ==================== INFORMACIÓN DEL ESTUDIANTE EXACTA ====================
        
        # Línea de separación
        for col in range(1, 10):  # A-I
            ws.cell(row=current_row, column=col).border = Border(top=Side(style='thick'))
        current_row += 1
        
        # CARRERA
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = "CARRERA"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        
        ws[f'I{current_row}'] = "MATRÍCULA"
        ws[f'I{current_row}'].font = font_small
        ws[f'I{current_row}'].alignment = center_alignment
        current_row += 1
        
        # Valor de carrera y matrícula
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = f"No Escolarizada (virtual)"  # Formato exacto del original
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].border = thin_border
        
        ws[f'I{current_row}'] = str(matricula_db)
        ws[f'I{current_row}'].font = font_normal
        ws[f'I{current_row}'].alignment = center_alignment
        ws[f'I{current_row}'].border = thin_border
        current_row += 1
        
        # MODALIDAD
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = "MODALIDAD"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        # Campos de nombre - exacto como el original
        ws[f'A{current_row}'] = "APELLIDO PATERNO"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].border = thin_border
        
        ws[f'D{current_row}'] = "APELLIDO MATERNO"
        ws[f'D{current_row}'].font = font_small
        ws[f'D{current_row}'].alignment = center_alignment
        ws[f'D{current_row}'].border = thin_border
        
        ws.merge_cells(f'G{current_row}:I{current_row}')
        ws[f'G{current_row}'] = "NOMBRE(S)"
        ws[f'G{current_row}'].font = font_small
        ws[f'G{current_row}'].alignment = center_alignment
        ws[f'G{current_row}'].border = thin_border
        current_row += 1
        
        # Valores del nombre (dividir el nombre completo)
        nombre_parts = str(nombre_db).split()
        apellido_paterno = nombre_parts[0] if len(nombre_parts) > 0 else ""
        apellido_materno = nombre_parts[1] if len(nombre_parts) > 1 else ""
        nombres = " ".join(nombre_parts[2:]) if len(nombre_parts) > 2 else ""
        
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = apellido_paterno.upper()
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].border = thin_border
        
        ws.merge_cells(f'D{current_row}:F{current_row}')
        ws[f'D{current_row}'] = apellido_materno.upper()
        ws[f'D{current_row}'].font = font_normal
        ws[f'D{current_row}'].alignment = center_alignment
        ws[f'D{current_row}'].border = thin_border
        
        ws.merge_cells(f'G{current_row}:I{current_row}')
        ws[f'G{current_row}'] = nombres.upper()
        ws[f'G{current_row}'].font = font_normal
        ws[f'G{current_row}'].alignment = center_alignment
        ws[f'G{current_row}'].border = thin_border
        current_row += 1
        
        # SEXO y CURP
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = "SEXO"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        
        ws.merge_cells(f'E{current_row}:I{current_row}')
        ws[f'E{current_row}'] = "CURP"
        ws[f'E{current_row}'].font = font_small
        ws[f'E{current_row}'].alignment = center_alignment
        current_row += 1
        
        # Valores vacíos (no disponibles)
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = ""
        ws[f'A{current_row}'].border = thin_border
        
        ws.merge_cells(f'E{current_row}:I{current_row}')
        ws[f'E{current_row}'] = ""
        ws[f'E{current_row}'].border = thin_border
        current_row += 2
        
        # ==================== GENERAR CUATRIMESTRES EXACTOS ====================
        
        # Nombres de cuatrimestres exactos como en el original
        nombres_cuatrimestres = {
            '1': 'PRIMER CUATRIMESTRE',
            '2': 'SEGUNDO CUATRIMESTRE', 
            '3': 'TERCER CUATRIMESTRE',
            '4': 'CUARTO CUATRIMESTRE',
            '5': 'QUINTO CUATRIMESTRE',
            '6': 'SEXTO CUATRIMESTRE',
            '7': 'SÉPTIMO CUATRIMESTRE',
            '8': 'OCTAVO CUATRIMESTRE',
            '9': 'NOVENO CUATRIMESTRE',
            '10': 'DÉCIMO CUATRIMESTRE'
        }
        
        # Generar cada cuatrimestre con formato exacto
        cuatrimestres_ordenados = sorted([int(k) for k in materias_por_cuatrimestre.keys()])
        
        for cuatr_num in cuatrimestres_ordenados:
            cuatr_str = str(cuatr_num)
            if cuatr_str in nombres_cuatrimestres:
                materias_cuatr = materias_por_cuatrimestre[cuatr_str]
                
                # Período Escolar
                ws[f'A{current_row}'] = "Periodo Escolar:"
                ws[f'A{current_row}'].font = font_normal
                current_row += 1
                
                # Encabezados de tabla exactos
                headers = ['Clave', 'Seriación', 'Créditos', 'MATERIAS', 'C.F.', 'Tipo de\nexamen', 'Calif.', 'Fecha', 'Observaciones']
                col_widths = [8, 8, 8, 35, 8, 10, 8, 12, 20]
                
                # Headers de tabla - PRIMERO establecer todos los headers individuales
                header_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
                for i, (header, col_letter) in enumerate(zip(headers, header_cols)):
                    ws[f'{col_letter}{current_row}'] = header
                    ws[f'{col_letter}{current_row}'].font = font_header_table
                    ws[f'{col_letter}{current_row}'].alignment = center_alignment
                    ws[f'{col_letter}{current_row}'].border = thin_border
                
                # DESPUÉS fusionar y establecer el título del cuatrimestre
                ws.merge_cells(f'D{current_row}:E{current_row}')
                ws[f'D{current_row}'] = nombres_cuatrimestres[cuatr_str]
                ws[f'D{current_row}'].font = font_header_table
                ws[f'D{current_row}'].alignment = center_alignment
                
                current_row += 1
                
                # Filas de materias (exacto como el original - 6 filas por cuatrimestre)
                for i in range(6):
                    if i < len(materias_cuatr):
                        materia_info = materias_cuatr[i]
                        (materia, calificacion, estatus, fecha_materia, fecha_calif, 
                         profesor, cuatr_actual, cuatr_historico, tipo_asig, n_recurs) = materia_info
                        
                        # Datos de la materia
                        ws[f'A{current_row}'] = ""  # Clave (vacío)
                        ws[f'B{current_row}'] = ""  # Seriación (vacío)
                        ws[f'C{current_row}'] = ""  # Créditos (vacío)
                        ws[f'D{current_row}'] = str(materia).upper()
                        ws[f'E{current_row}'] = ""  # C.F. (vacío)
                        ws[f'F{current_row}'] = ""  # Tipo de examen (vacío)
                        
                        # Calificación
                        try:
                            calif_num = float(calificacion)
                            ws[f'G{current_row}'] = f"{calif_num:.1f}"
                        except:
                            ws[f'G{current_row}'] = str(calificacion)
                        
                        # Fecha
                        fecha_mostrar = str(fecha_calif) if fecha_calif and fecha_calif != "None" else str(fecha_materia)
                        if fecha_mostrar and fecha_mostrar not in ["None", "N/A", ""]:
                            try:
                                fecha_dt = pd.to_datetime(fecha_mostrar)
                                ws[f'H{current_row}'] = fecha_dt.strftime('%d/%m/%Y')
                            except:
                                ws[f'H{current_row}'] = fecha_mostrar
                        else:
                            ws[f'H{current_row}'] = ""
                        
                        # Observaciones
                        obs = "RECURSAMIENTO" if tipo_asig == "recursamiento" else ""
                        ws[f'I{current_row}'] = obs
                        
                    else:
                        # Filas vacías
                        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                            ws[f'{col}{current_row}'] = ""
                    
                    # Bordes en todas las celdas de la fila
                    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                        ws[f'{col}{current_row}'].border = thin_border
                        ws[f'{col}{current_row}'].font = font_normal
                    
                    current_row += 1
                
                # Elaboró y Cotejó exacto como el original
                ws[f'A{current_row}'] = "Elaboró:"
                ws[f'A{current_row}'].font = font_small
                
                ws[f'B{current_row}'] = "COORD."
                ws[f'B{current_row}'].font = font_small
                
                ws[f'H{current_row}'] = "Cotejó:"
                ws[f'H{current_row}'].font = font_small
                
                ws[f'I{current_row}'] = "DSE"
                ws[f'I{current_row}'].font = font_small
                
                current_row += 2  # Espacio entre cuatrimestres
        
        # ==================== PIE DE PÁGINA EXACTO ====================
        
        # Calcular promedio general
        calificaciones_num = []
        for materia_info in materias_data:
            try:
                if materia_info[1] is not None and str(materia_info[1]) != "None":
                    cal_float = float(materia_info[1])
                    calificaciones_num.append(cal_float)
            except (ValueError, TypeError):
                continue
        
        promedio_general = sum(calificaciones_num) / len(calificaciones_num) if calificaciones_num else 0.0
        
        current_row += 2
        
        # PROMEDIO GENERAL exacto
        ws.merge_cells(f'A{current_row}:G{current_row}')
        ws[f'A{current_row}'] = "PROMEDIO GENERAL:"
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = right_alignment
        
        ws.merge_cells(f'H{current_row}:I{current_row}')
        ws[f'H{current_row}'] = f"{promedio_general:.1f}"
        ws[f'H{current_row}'].font = font_normal
        ws[f'H{current_row}'].alignment = center_alignment
        ws[f'H{current_row}'].border = thick_border
        
        current_row += 1
        
        # Escala de evaluación exacta
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = "ESCALA DE EVALUACIÓN DEL 0.0 AL 10.0 MÍNIMA APROBATORIA 6.0"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        
        current_row += 3
        
        # Firmas exactas
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = "Vice-Rector Académico"
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = center_alignment
        
        ws.merge_cells(f'F{current_row}:I{current_row}')
        ws[f'F{current_row}'] = "Jefa del Departamento de Servicios Escolares"
        ws[f'F{current_row}'].font = font_normal
        ws[f'F{current_row}'].alignment = center_alignment
        
        # ==================== AJUSTAR DIMENSIONES EXACTAS ====================
        
        # Anchos de columna exactos del formato original
        column_widths = [12, 12, 12, 40, 8, 12, 8, 15, 25]
        column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        
        for i, width in enumerate(column_widths):
            if i < len(column_letters):
                ws.column_dimensions[column_letters[i]].width = width
        
        # ==================== GUARDAR ARCHIVO ====================
        
        # Crear buffer en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M')
        nombre_limpio = str(apellido_paterno + apellido_materno).replace(' ', '_')[:20]
        nombre_archivo = f"KARDEX_OFICIAL_{nombre_limpio}_{matricula}_{fecha_actual}.xlsx"
        
        return {
            'buffer': output.getvalue(),
            'filename': nombre_archivo,
            'estadisticas': {
                'materias_cursadas': len(materias_data),
                'aprobadas': len([c for c in calificaciones_num if c >= 6.0]),
                'reprobadas': len(materias_data) - len([c for c in calificaciones_num if c >= 6.0]),
                'promedio_general': round(promedio_general, 2)
            }
        }
        
    except Exception as e:
        st.error(f"Error al generar kardex oficial: {e}")
        import traceback
        st.error(f"Detalles técnicos: {traceback.format_exc()}")
        return None
    """
    Genera un kardex con formato oficial IDÉNTICO al archivo original de NovaUniversitas,
    replicando exactamente la estructura, layout y disposición.
    """
    try:
        # Obtener información básica del alumno
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DISTINCT nombre, matricula, carrera, email_personal, usuario_email_nova,
                   fecha_ingreso_original, cuatrimestre
            FROM calificaciones 
            WHERE matricula = ?
            LIMIT 1
        """, (matricula,))
        
        info_alumno = cursor.fetchone()
        if not info_alumno:
            st.error("No se encontró información del alumno en la base de datos")
            return None
            
        nombre_db, matricula_db, carrera, email_personal, email_nova, fecha_ingreso, cuatrimestre_actual = info_alumno
        
        # Obtener todas las materias con calificaciones organizadas por cuatrimestre
        cursor.execute("""
            SELECT 
                materia,
                calificacion,
                estatus,
                fecha_ingreso_materia,
                fecha_calificacion,
                profesor,
                cuatrimestre,
                cuatrimestre_historico,
                tipo_asignacion,
                n_recursamientos
            FROM calificaciones 
            WHERE matricula = ?
            AND materia IS NOT NULL 
            AND materia != ''
            AND calificacion IS NOT NULL
            ORDER BY 
                CAST(COALESCE(cuatrimestre_historico, cuatrimestre, 1) AS INTEGER) ASC,
                fecha_ingreso_materia ASC
        """, (matricula,))
        
        materias_data = cursor.fetchall()
        
        if not materias_data:
            st.warning("No hay materias con calificaciones para generar el kardex")
            return None
        
        # Organizar materias por cuatrimestre histórico
        materias_por_cuatrimestre = {}
        for materia_info in materias_data:
            (materia, calificacion, estatus, fecha_materia, fecha_calif, 
             profesor, cuatr_actual, cuatr_historico, tipo_asig, n_recurs) = materia_info
            
            cuatrimestre_key = str(cuatr_historico) if cuatr_historico and cuatr_historico != "None" else str(cuatr_actual)
            
            if cuatrimestre_key not in materias_por_cuatrimestre:
                materias_por_cuatrimestre[cuatrimestre_key] = []
            
            materias_por_cuatrimestre[cuatrimestre_key].append(materia_info)
        
        # Crear el archivo Excel con formato oficial EXACTO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from datetime import datetime
        import io
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "KARDEX"
        
        # ==================== CONFIGURACIÓN DE FORMATO EXACTO ====================
        
        # Fuentes exactas del formato original
        font_titulo_principal = Font(name='Arial', size=14, bold=True)
        font_subtitulos = Font(name='Arial', size=11, bold=True)
        font_normal = Font(name='Arial', size=9)
        font_small = Font(name='Arial', size=8)
        font_header_table = Font(name='Arial', size=8, bold=True)
        
        # Alineaciones
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        # Bordes exactos del formato
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        
        # ==================== ENCABEZADO PRINCIPAL EXACTO ====================
        
        current_row = 1
        
        # Logo - intentar cargar desde carpeta, sino usar placeholder
        logo_agregado = False
        rutas_logotipo = [
            "assets/logo.png",
            "assets/logo.jpg", 
            "assets/logotipo.png",
            "assets/logotipo.jpg",
            "logo.png",
            "logo.jpg",
            "logotipo.png",
            "logotipo.jpg"
        ]
        
        # DIAGNÓSTICO - mostrar información de depuración
        import os
        st.info(f"🔍 Directorio actual: {os.getcwd()}")
        st.info("📁 Buscando logotipo en estas rutas:")
        for ruta in rutas_logotipo:
            existe = os.path.exists(ruta)
            st.write(f"  • {ruta} -> {'✅ EXISTE' if existe else '❌ NO EXISTE'}")
        
        try:
            from PIL import Image as PILImage
            from openpyxl.drawing import Image
            st.success("✅ PIL y openpyxl.drawing importados correctamente")
            
            # Intentar encontrar el logotipo en las rutas definidas
            for ruta in rutas_logotipo:
                if os.path.exists(ruta):
                    try:
                        st.info(f"📷 Intentando cargar imagen: {ruta}")
                        
                        # Crear imagen
                        img = Image(ruta)
                        
                        # Configurar tamaño apropiado para el kardex
                        img.width = 80
                        img.height = 100
                        
                        # Posicionar en A1
                        img.anchor = 'A1'
                        
                        # Agregar imagen a la hoja
                        ws.add_image(img)
                        
                        # Fusionar celdas para el espacio del logo pero sin texto
                        ws.merge_cells('A1:A6')
                        ws['A1'] = ""  # Sin texto ya que hay imagen
                        ws['A1'].border = thick_border
                        
                        logo_agregado = True
                        st.success(f"🎉 Logotipo cargado exitosamente desde: {ruta}")
                        break
                        
                    except Exception as e:
                        st.error(f"❌ Error cargando {ruta}: {str(e)}")
                        continue  # Intentar con la siguiente ruta
            
        except ImportError as e:
            st.error(f"❌ Error importando PIL: {str(e)}")
        
        # Si no se pudo cargar ningún logotipo, usar placeholder de texto
        if not logo_agregado:
            st.warning("⚠️ No se pudo cargar logotipo, usando placeholder de texto")
            ws.merge_cells('A1:A6')
            ws['A1'] = "NOVA\nUNIVERSITAS"
            ws['A1'].font = Font(name='Arial', size=10, bold=True)
            ws['A1'].alignment = center_alignment
            ws['A1'].border = thick_border
        
        # Título principal - NovaUniversitas
        ws.merge_cells('B1:H1')
        ws['B1'] = "NovaUniversitas"
        ws['B1'].font = font_titulo_principal
        ws['B1'].alignment = center_alignment
        
        # Subtítulo - Organismo Público Descentralizado
        ws.merge_cells('B2:H2')
        ws['B2'] = "ORGANISMO PÚBLICO DESCENTRALIZADO"
        ws['B2'].font = font_normal
        ws['B2'].alignment = center_alignment
        
        # Departamento
        ws.merge_cells('B3:H3')
        ws['B3'] = "DEPARTAMENTO DE SERVICIOS ESCOLARES"
        ws['B3'].font = font_normal
        ws['B3'].alignment = center_alignment
        
        # KARDEX title
        ws.merge_cells('B4:H4')
        ws['B4'] = "KARDEX"
        ws['B4'].font = Font(name='Arial', size=16, bold=True)
        ws['B4'].alignment = center_alignment
        
        # FOTO placeholder (columna I, filas 1-6)
        ws.merge_cells('I1:I6')
        ws['I1'] = "FOTO"
        ws['I1'].font = font_normal
        ws['I1'].alignment = center_alignment
        ws['I1'].border = thick_border
        
        current_row = 7
        
        # ==================== INFORMACIÓN DEL ESTUDIANTE EXACTA ====================
        
        # Línea de separación
        for col in range(1, 10):  # A-I
            ws.cell(row=current_row, column=col).border = Border(top=Side(style='thick'))
        current_row += 1
        
        # CARRERA
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = "CARRERA"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        
        ws[f'I{current_row}'] = "MATRÍCULA"
        ws[f'I{current_row}'].font = font_small
        ws[f'I{current_row}'].alignment = center_alignment
        current_row += 1
        
        # Valor de carrera y matrícula
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = f"No Escolarizada (virtual)"  # Formato exacto del original
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].border = thin_border
        
        ws[f'I{current_row}'] = str(matricula_db)
        ws[f'I{current_row}'].font = font_normal
        ws[f'I{current_row}'].alignment = center_alignment
        ws[f'I{current_row}'].border = thin_border
        current_row += 1
        
        # MODALIDAD
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = "MODALIDAD"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        # Campos de nombre - exacto como el original
        ws[f'A{current_row}'] = "APELLIDO PATERNO"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].border = thin_border
        
        ws[f'D{current_row}'] = "APELLIDO MATERNO"
        ws[f'D{current_row}'].font = font_small
        ws[f'D{current_row}'].alignment = center_alignment
        ws[f'D{current_row}'].border = thin_border
        
        ws.merge_cells(f'G{current_row}:I{current_row}')
        ws[f'G{current_row}'] = "NOMBRE(S)"
        ws[f'G{current_row}'].font = font_small
        ws[f'G{current_row}'].alignment = center_alignment
        ws[f'G{current_row}'].border = thin_border
        current_row += 1
        
        # Valores del nombre (dividir el nombre completo)
        nombre_parts = str(nombre_db).split()
        apellido_paterno = nombre_parts[0] if len(nombre_parts) > 0 else ""
        apellido_materno = nombre_parts[1] if len(nombre_parts) > 1 else ""
        nombres = " ".join(nombre_parts[2:]) if len(nombre_parts) > 2 else ""
        
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = apellido_paterno.upper()
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].border = thin_border
        
        ws.merge_cells(f'D{current_row}:F{current_row}')
        ws[f'D{current_row}'] = apellido_materno.upper()
        ws[f'D{current_row}'].font = font_normal
        ws[f'D{current_row}'].alignment = center_alignment
        ws[f'D{current_row}'].border = thin_border
        
        ws.merge_cells(f'G{current_row}:I{current_row}')
        ws[f'G{current_row}'] = nombres.upper()
        ws[f'G{current_row}'].font = font_normal
        ws[f'G{current_row}'].alignment = center_alignment
        ws[f'G{current_row}'].border = thin_border
        current_row += 1
        
        # SEXO y CURP
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = "SEXO"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        
        ws.merge_cells(f'E{current_row}:I{current_row}')
        ws[f'E{current_row}'] = "CURP"
        ws[f'E{current_row}'].font = font_small
        ws[f'E{current_row}'].alignment = center_alignment
        current_row += 1
        
        # Valores vacíos (no disponibles)
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = ""
        ws[f'A{current_row}'].border = thin_border
        
        ws.merge_cells(f'E{current_row}:I{current_row}')
        ws[f'E{current_row}'] = ""
        ws[f'E{current_row}'].border = thin_border
        current_row += 2
        
        # ==================== GENERAR CUATRIMESTRES EXACTOS ====================
        
        # Nombres de cuatrimestres exactos como en el original
        nombres_cuatrimestres = {
            '1': 'PRIMER CUATRIMESTRE',
            '2': 'SEGUNDO CUATRIMESTRE', 
            '3': 'TERCER CUATRIMESTRE',
            '4': 'CUARTO CUATRIMESTRE',
            '5': 'QUINTO CUATRIMESTRE',
            '6': 'SEXTO CUATRIMESTRE',
            '7': 'SÉPTIMO CUATRIMESTRE',
            '8': 'OCTAVO CUATRIMESTRE',
            '9': 'NOVENO CUATRIMESTRE',
            '10': 'DÉCIMO CUATRIMESTRE'
        }
        
        # Generar cada cuatrimestre con formato exacto
        cuatrimestres_ordenados = sorted([int(k) for k in materias_por_cuatrimestre.keys()])
        
        for cuatr_num in cuatrimestres_ordenados:
            cuatr_str = str(cuatr_num)
            if cuatr_str in nombres_cuatrimestres:
                materias_cuatr = materias_por_cuatrimestre[cuatr_str]
                
                # Período Escolar
                ws[f'A{current_row}'] = "Periodo Escolar:"
                ws[f'A{current_row}'].font = font_normal
                current_row += 1
                
                # Encabezados de tabla exactos
                headers = ['Clave', 'Seriación', 'Créditos', 'MATERIAS', 'C.F.', 'Tipo de\nexamen', 'Calif.', 'Fecha', 'Observaciones']
                col_widths = [8, 8, 8, 35, 8, 10, 8, 12, 20]
                
                # Headers de tabla - PRIMERO establecer todos los headers individuales
                header_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
                for i, (header, col_letter) in enumerate(zip(headers, header_cols)):
                    ws[f'{col_letter}{current_row}'] = header
                    ws[f'{col_letter}{current_row}'].font = font_header_table
                    ws[f'{col_letter}{current_row}'].alignment = center_alignment
                    ws[f'{col_letter}{current_row}'].border = thin_border
                
                # DESPUÉS fusionar y establecer el título del cuatrimestre
                ws.merge_cells(f'D{current_row}:E{current_row}')
                ws[f'D{current_row}'] = nombres_cuatrimestres[cuatr_str]
                ws[f'D{current_row}'].font = font_header_table
                ws[f'D{current_row}'].alignment = center_alignment
                
                current_row += 1
                
                # Filas de materias (exacto como el original - 6 filas por cuatrimestre)
                for i in range(6):
                    if i < len(materias_cuatr):
                        materia_info = materias_cuatr[i]
                        (materia, calificacion, estatus, fecha_materia, fecha_calif, 
                         profesor, cuatr_actual, cuatr_historico, tipo_asig, n_recurs) = materia_info
                        
                        # Datos de la materia
                        ws[f'A{current_row}'] = ""  # Clave (vacío)
                        ws[f'B{current_row}'] = ""  # Seriación (vacío)
                        ws[f'C{current_row}'] = ""  # Créditos (vacío)
                        ws[f'D{current_row}'] = str(materia).upper()
                        ws[f'E{current_row}'] = ""  # C.F. (vacío)
                        ws[f'F{current_row}'] = ""  # Tipo de examen (vacío)
                        
                        # Calificación
                        try:
                            calif_num = float(calificacion)
                            ws[f'G{current_row}'] = f"{calif_num:.1f}"
                        except:
                            ws[f'G{current_row}'] = str(calificacion)
                        
                        # Fecha
                        fecha_mostrar = str(fecha_calif) if fecha_calif and fecha_calif != "None" else str(fecha_materia)
                        if fecha_mostrar and fecha_mostrar not in ["None", "N/A", ""]:
                            try:
                                fecha_dt = pd.to_datetime(fecha_mostrar)
                                ws[f'H{current_row}'] = fecha_dt.strftime('%d/%m/%Y')
                            except:
                                ws[f'H{current_row}'] = fecha_mostrar
                        else:
                            ws[f'H{current_row}'] = ""
                        
                        # Observaciones
                        obs = "RECURSAMIENTO" if tipo_asig == "recursamiento" else ""
                        ws[f'I{current_row}'] = obs
                        
                    else:
                        # Filas vacías
                        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                            ws[f'{col}{current_row}'] = ""
                    
                    # Bordes en todas las celdas de la fila
                    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                        ws[f'{col}{current_row}'].border = thin_border
                        ws[f'{col}{current_row}'].font = font_normal
                    
                    current_row += 1
                
                # Elaboró y Cotejó exacto como el original
                ws[f'A{current_row}'] = "Elaboró:"
                ws[f'A{current_row}'].font = font_small
                
                ws[f'B{current_row}'] = "COORD."
                ws[f'B{current_row}'].font = font_small
                
                ws[f'H{current_row}'] = "Cotejó:"
                ws[f'H{current_row}'].font = font_small
                
                ws[f'I{current_row}'] = "DSE"
                ws[f'I{current_row}'].font = font_small
                
                current_row += 2  # Espacio entre cuatrimestres
        
        # ==================== PIE DE PÁGINA EXACTO ====================
        
        # Calcular promedio general
        calificaciones_num = []
        for materia_info in materias_data:
            try:
                if materia_info[1] is not None and str(materia_info[1]) != "None":
                    cal_float = float(materia_info[1])
                    calificaciones_num.append(cal_float)
            except (ValueError, TypeError):
                continue
        
        promedio_general = sum(calificaciones_num) / len(calificaciones_num) if calificaciones_num else 0.0
        
        current_row += 2
        
        # PROMEDIO GENERAL exacto
        ws.merge_cells(f'A{current_row}:G{current_row}')
        ws[f'A{current_row}'] = "PROMEDIO GENERAL:"
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = right_alignment
        
        ws.merge_cells(f'H{current_row}:I{current_row}')
        ws[f'H{current_row}'] = f"{promedio_general:.1f}"
        ws[f'H{current_row}'].font = font_normal
        ws[f'H{current_row}'].alignment = center_alignment
        ws[f'H{current_row}'].border = thick_border
        
        current_row += 1
        
        # Escala de evaluación exacta
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = "ESCALA DE EVALUACIÓN DEL 0.0 AL 10.0 MÍNIMA APROBATORIA 6.0"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        
        current_row += 3
        
        # Firmas exactas
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = "Vice-Rector Académico"
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = center_alignment
        
        ws.merge_cells(f'F{current_row}:I{current_row}')
        ws[f'F{current_row}'] = "Jefa del Departamento de Servicios Escolares"
        ws[f'F{current_row}'].font = font_normal
        ws[f'F{current_row}'].alignment = center_alignment
        
        # ==================== AJUSTAR DIMENSIONES EXACTAS ====================
        
        # Anchos de columna exactos del formato original
        column_widths = [12, 12, 12, 40, 8, 12, 8, 15, 25]
        column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        
        for i, width in enumerate(column_widths):
            if i < len(column_letters):
                ws.column_dimensions[column_letters[i]].width = width
        
        # ==================== GUARDAR ARCHIVO ====================
        
        # Crear buffer en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M')
        nombre_limpio = str(apellido_paterno + apellido_materno).replace(' ', '_')[:20]
        nombre_archivo = f"KARDEX_OFICIAL_{nombre_limpio}_{matricula}_{fecha_actual}.xlsx"
        
        return {
            'buffer': output.getvalue(),
            'filename': nombre_archivo,
            'estadisticas': {
                'materias_cursadas': len(materias_data),
                'aprobadas': len([c for c in calificaciones_num if c >= 6.0]),
                'reprobadas': len(materias_data) - len([c for c in calificaciones_num if c >= 6.0]),
                'promedio_general': round(promedio_general, 2)
            }
        }
        
    except Exception as e:
        st.error(f"Error al generar kardex oficial: {e}")
        import traceback
        st.error(f"Detalles técnicos: {traceback.format_exc()}")
        return None
    """
    Genera un kardex con formato oficial IDÉNTICO al archivo original de NovaUniversitas,
    replicando exactamente la estructura, layout y disposición.
    """
    try:
        # Obtener información básica del alumno
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DISTINCT nombre, matricula, carrera, email_personal, usuario_email_nova,
                   fecha_ingreso_original, cuatrimestre
            FROM calificaciones 
            WHERE matricula = ?
            LIMIT 1
        """, (matricula,))
        
        info_alumno = cursor.fetchone()
        if not info_alumno:
            st.error("No se encontró información del alumno en la base de datos")
            return None
            
        nombre_db, matricula_db, carrera, email_personal, email_nova, fecha_ingreso, cuatrimestre_actual = info_alumno
        
        # Obtener todas las materias con calificaciones organizadas por cuatrimestre
        cursor.execute("""
            SELECT 
                materia,
                calificacion,
                estatus,
                fecha_ingreso_materia,
                fecha_calificacion,
                profesor,
                cuatrimestre,
                cuatrimestre_historico,
                tipo_asignacion,
                n_recursamientos
            FROM calificaciones 
            WHERE matricula = ?
            AND materia IS NOT NULL 
            AND materia != ''
            AND calificacion IS NOT NULL
            ORDER BY 
                CAST(COALESCE(cuatrimestre_historico, cuatrimestre, 1) AS INTEGER) ASC,
                fecha_ingreso_materia ASC
        """, (matricula,))
        
        materias_data = cursor.fetchall()
        
        if not materias_data:
            st.warning("No hay materias con calificaciones para generar el kardex")
            return None
        
        # Organizar materias por cuatrimestre histórico
        materias_por_cuatrimestre = {}
        for materia_info in materias_data:
            (materia, calificacion, estatus, fecha_materia, fecha_calif, 
             profesor, cuatr_actual, cuatr_historico, tipo_asig, n_recurs) = materia_info
            
            cuatrimestre_key = str(cuatr_historico) if cuatr_historico and cuatr_historico != "None" else str(cuatr_actual)
            
            if cuatrimestre_key not in materias_por_cuatrimestre:
                materias_por_cuatrimestre[cuatrimestre_key] = []
            
            materias_por_cuatrimestre[cuatrimestre_key].append(materia_info)
        
        # Crear el archivo Excel con formato oficial EXACTO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from datetime import datetime
        import io
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "KARDEX"
        
        # ==================== CONFIGURACIÓN DE FORMATO EXACTO ====================
        
        # Fuentes exactas del formato original
        font_titulo_principal = Font(name='Arial', size=14, bold=True)
        font_subtitulos = Font(name='Arial', size=11, bold=True)
        font_normal = Font(name='Arial', size=9)
        font_small = Font(name='Arial', size=8)
        font_header_table = Font(name='Arial', size=8, bold=True)
        
        # Alineaciones
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        # Bordes exactos del formato
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        
        # ==================== ENCABEZADO PRINCIPAL EXACTO ====================
        
        current_row = 1
        
        # Logo - intentar cargar desde carpeta, sino usar placeholder
        logo_agregado = False
        rutas_logotipo = [
            "assets/logo.png",
            "assets/logo.jpg", 
            "assets/logotipo.png",
            "assets/logotipo.jpg",
            "logo.png",
            "logo.jpg",
            "logotipo.png",
            "logotipo.jpg"
        ]
        
        try:
            from PIL import Image as PILImage
            from openpyxl.drawing import Image
            import os
            
            # Intentar encontrar el logotipo en las rutas definidas
            for ruta in rutas_logotipo:
                if os.path.exists(ruta):
                    try:
                        # Crear imagen
                        img = Image(ruta)
                        
                        # Configurar tamaño apropiado para el kardex
                        img.width = 80
                        img.height = 100
                        
                        # Posicionar en A1
                        img.anchor = 'A1'
                        
                        # Agregar imagen a la hoja
                        ws.add_image(img)
                        
                        # Fusionar celdas para el espacio del logo pero sin texto
                        ws.merge_cells('A1:A6')
                        ws['A1'] = ""  # Sin texto ya que hay imagen
                        ws['A1'].border = thick_border
                        
                        logo_agregado = True
                        break
                        
                    except Exception as e:
                        continue  # Intentar con la siguiente ruta
            
        except ImportError:
            pass  # PIL no disponible, usar placeholder
        
        # Si no se pudo cargar ningún logotipo, usar placeholder de texto
        if not logo_agregado:
            ws.merge_cells('A1:A6')
            ws['A1'] = "NOVA\nUNIVERSITAS"
            ws['A1'].font = Font(name='Arial', size=10, bold=True)
            ws['A1'].alignment = center_alignment
            ws['A1'].border = thick_border
        
        # Título principal - NovaUniversitas
        ws.merge_cells('B1:H1')
        ws['B1'] = "NovaUniversitas"
        ws['B1'].font = font_titulo_principal
        ws['B1'].alignment = center_alignment
        
        # Subtítulo - Organismo Público Descentralizado
        ws.merge_cells('B2:H2')
        ws['B2'] = "ORGANISMO PÚBLICO DESCENTRALIZADO"
        ws['B2'].font = font_normal
        ws['B2'].alignment = center_alignment
        
        # Departamento
        ws.merge_cells('B3:H3')
        ws['B3'] = "DEPARTAMENTO DE SERVICIOS ESCOLARES"
        ws['B3'].font = font_normal
        ws['B3'].alignment = center_alignment
        
        # KARDEX title
        ws.merge_cells('B4:H4')
        ws['B4'] = "KARDEX"
        ws['B4'].font = Font(name='Arial', size=16, bold=True)
        ws['B4'].alignment = center_alignment
        
        # FOTO placeholder (columna I, filas 1-6)
        ws.merge_cells('I1:I6')
        ws['I1'] = "FOTO"
        ws['I1'].font = font_normal
        ws['I1'].alignment = center_alignment
        ws['I1'].border = thick_border
        
        current_row = 7
        
        # ==================== INFORMACIÓN DEL ESTUDIANTE EXACTA ====================
        
        # Línea de separación
        for col in range(1, 10):  # A-I
            ws.cell(row=current_row, column=col).border = Border(top=Side(style='thick'))
        current_row += 1
        
        # CARRERA
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = "CARRERA"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        
        ws[f'I{current_row}'] = "MATRÍCULA"
        ws[f'I{current_row}'].font = font_small
        ws[f'I{current_row}'].alignment = center_alignment
        current_row += 1
        
        # Valor de carrera y matrícula
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = f"No Escolarizada (virtual)"  # Formato exacto del original
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].border = thin_border
        
        ws[f'I{current_row}'] = str(matricula_db)
        ws[f'I{current_row}'].font = font_normal
        ws[f'I{current_row}'].alignment = center_alignment
        ws[f'I{current_row}'].border = thin_border
        current_row += 1
        
        # MODALIDAD
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = "MODALIDAD"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        # Campos de nombre - exacto como el original
        ws[f'A{current_row}'] = "APELLIDO PATERNO"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].border = thin_border
        
        ws[f'D{current_row}'] = "APELLIDO MATERNO"
        ws[f'D{current_row}'].font = font_small
        ws[f'D{current_row}'].alignment = center_alignment
        ws[f'D{current_row}'].border = thin_border
        
        ws.merge_cells(f'G{current_row}:I{current_row}')
        ws[f'G{current_row}'] = "NOMBRE(S)"
        ws[f'G{current_row}'].font = font_small
        ws[f'G{current_row}'].alignment = center_alignment
        ws[f'G{current_row}'].border = thin_border
        current_row += 1
        
        # Valores del nombre (dividir el nombre completo)
        nombre_parts = str(nombre_db).split()
        apellido_paterno = nombre_parts[0] if len(nombre_parts) > 0 else ""
        apellido_materno = nombre_parts[1] if len(nombre_parts) > 1 else ""
        nombres = " ".join(nombre_parts[2:]) if len(nombre_parts) > 2 else ""
        
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = apellido_paterno.upper()
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].border = thin_border
        
        ws.merge_cells(f'D{current_row}:F{current_row}')
        ws[f'D{current_row}'] = apellido_materno.upper()
        ws[f'D{current_row}'].font = font_normal
        ws[f'D{current_row}'].alignment = center_alignment
        ws[f'D{current_row}'].border = thin_border
        
        ws.merge_cells(f'G{current_row}:I{current_row}')
        ws[f'G{current_row}'] = nombres.upper()
        ws[f'G{current_row}'].font = font_normal
        ws[f'G{current_row}'].alignment = center_alignment
        ws[f'G{current_row}'].border = thin_border
        current_row += 1
        
        # SEXO y CURP
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = "SEXO"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        
        ws.merge_cells(f'E{current_row}:I{current_row}')
        ws[f'E{current_row}'] = "CURP"
        ws[f'E{current_row}'].font = font_small
        ws[f'E{current_row}'].alignment = center_alignment
        current_row += 1
        
        # Valores vacíos (no disponibles)
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = ""
        ws[f'A{current_row}'].border = thin_border
        
        ws.merge_cells(f'E{current_row}:I{current_row}')
        ws[f'E{current_row}'] = ""
        ws[f'E{current_row}'].border = thin_border
        current_row += 2
        
        # ==================== GENERAR CUATRIMESTRES EXACTOS ====================
        
        # Nombres de cuatrimestres exactos como en el original
        nombres_cuatrimestres = {
            '1': 'PRIMER CUATRIMESTRE',
            '2': 'SEGUNDO CUATRIMESTRE', 
            '3': 'TERCER CUATRIMESTRE',
            '4': 'CUARTO CUATRIMESTRE',
            '5': 'QUINTO CUATRIMESTRE',
            '6': 'SEXTO CUATRIMESTRE',
            '7': 'SÉPTIMO CUATRIMESTRE',
            '8': 'OCTAVO CUATRIMESTRE',
            '9': 'NOVENO CUATRIMESTRE',
            '10': 'DÉCIMO CUATRIMESTRE'
        }
        
        # Generar cada cuatrimestre con formato exacto
        cuatrimestres_ordenados = sorted([int(k) for k in materias_por_cuatrimestre.keys()])
        
        for cuatr_num in cuatrimestres_ordenados:
            cuatr_str = str(cuatr_num)
            if cuatr_str in nombres_cuatrimestres:
                materias_cuatr = materias_por_cuatrimestre[cuatr_str]
                
                # Período Escolar
                ws[f'A{current_row}'] = "Periodo Escolar:"
                ws[f'A{current_row}'].font = font_normal
                current_row += 1
                
                # Encabezados de tabla exactos
                headers = ['Clave', 'Seriación', 'Créditos', 'MATERIAS', 'C.F.', 'Tipo de\nexamen', 'Calif.', 'Fecha', 'Observaciones']
                col_widths = [8, 8, 8, 35, 8, 10, 8, 12, 20]
                
                # Headers de tabla - PRIMERO establecer todos los headers individuales
                header_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
                for i, (header, col_letter) in enumerate(zip(headers, header_cols)):
                    ws[f'{col_letter}{current_row}'] = header
                    ws[f'{col_letter}{current_row}'].font = font_header_table
                    ws[f'{col_letter}{current_row}'].alignment = center_alignment
                    ws[f'{col_letter}{current_row}'].border = thin_border
                
                # DESPUÉS fusionar y establecer el título del cuatrimestre
                ws.merge_cells(f'D{current_row}:E{current_row}')
                ws[f'D{current_row}'] = nombres_cuatrimestres[cuatr_str]
                ws[f'D{current_row}'].font = font_header_table
                ws[f'D{current_row}'].alignment = center_alignment
                
                current_row += 1
                
                # Filas de materias (exacto como el original - 6 filas por cuatrimestre)
                for i in range(6):
                    if i < len(materias_cuatr):
                        materia_info = materias_cuatr[i]
                        (materia, calificacion, estatus, fecha_materia, fecha_calif, 
                         profesor, cuatr_actual, cuatr_historico, tipo_asig, n_recurs) = materia_info
                        
                        # Datos de la materia
                        ws[f'A{current_row}'] = ""  # Clave (vacío)
                        ws[f'B{current_row}'] = ""  # Seriación (vacío)
                        ws[f'C{current_row}'] = ""  # Créditos (vacío)
                        ws[f'D{current_row}'] = str(materia).upper()
                        ws[f'E{current_row}'] = ""  # C.F. (vacío)
                        ws[f'F{current_row}'] = ""  # Tipo de examen (vacío)
                        
                        # Calificación
                        try:
                            calif_num = float(calificacion)
                            ws[f'G{current_row}'] = f"{calif_num:.1f}"
                        except:
                            ws[f'G{current_row}'] = str(calificacion)
                        
                        # Fecha
                        fecha_mostrar = str(fecha_calif) if fecha_calif and fecha_calif != "None" else str(fecha_materia)
                        if fecha_mostrar and fecha_mostrar not in ["None", "N/A", ""]:
                            try:
                                fecha_dt = pd.to_datetime(fecha_mostrar)
                                ws[f'H{current_row}'] = fecha_dt.strftime('%d/%m/%Y')
                            except:
                                ws[f'H{current_row}'] = fecha_mostrar
                        else:
                            ws[f'H{current_row}'] = ""
                        
                        # Observaciones
                        obs = "RECURSAMIENTO" if tipo_asig == "recursamiento" else ""
                        ws[f'I{current_row}'] = obs
                        
                    else:
                        # Filas vacías
                        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                            ws[f'{col}{current_row}'] = ""
                    
                    # Bordes en todas las celdas de la fila
                    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                        ws[f'{col}{current_row}'].border = thin_border
                        ws[f'{col}{current_row}'].font = font_normal
                    
                    current_row += 1
                
                # Elaboró y Cotejó exacto como el original
                ws[f'A{current_row}'] = "Elaboró:"
                ws[f'A{current_row}'].font = font_small
                
                ws[f'B{current_row}'] = "COORD."
                ws[f'B{current_row}'].font = font_small
                
                ws[f'H{current_row}'] = "Cotejó:"
                ws[f'H{current_row}'].font = font_small
                
                ws[f'I{current_row}'] = "DSE"
                ws[f'I{current_row}'].font = font_small
                
                current_row += 2  # Espacio entre cuatrimestres
        
        # ==================== PIE DE PÁGINA EXACTO ====================
        
        # Calcular promedio general
        calificaciones_num = []
        for materia_info in materias_data:
            try:
                if materia_info[1] is not None and str(materia_info[1]) != "None":
                    cal_float = float(materia_info[1])
                    calificaciones_num.append(cal_float)
            except (ValueError, TypeError):
                continue
        
        promedio_general = sum(calificaciones_num) / len(calificaciones_num) if calificaciones_num else 0.0
        
        current_row += 2
        
        # PROMEDIO GENERAL exacto
        ws.merge_cells(f'A{current_row}:G{current_row}')
        ws[f'A{current_row}'] = "PROMEDIO GENERAL:"
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = right_alignment
        
        ws.merge_cells(f'H{current_row}:I{current_row}')
        ws[f'H{current_row}'] = f"{promedio_general:.1f}"
        ws[f'H{current_row}'].font = font_normal
        ws[f'H{current_row}'].alignment = center_alignment
        ws[f'H{current_row}'].border = thick_border
        
        current_row += 1
        
        # Escala de evaluación exacta
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = "ESCALA DE EVALUACIÓN DEL 0.0 AL 10.0 MÍNIMA APROBATORIA 6.0"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        
        current_row += 3
        
        # Firmas exactas
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = "Vice-Rector Académico"
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = center_alignment
        
        ws.merge_cells(f'F{current_row}:I{current_row}')
        ws[f'F{current_row}'] = "Jefa del Departamento de Servicios Escolares"
        ws[f'F{current_row}'].font = font_normal
        ws[f'F{current_row}'].alignment = center_alignment
        
        # ==================== AJUSTAR DIMENSIONES EXACTAS ====================
        
        # Anchos de columna exactos del formato original
        column_widths = [12, 12, 12, 40, 8, 12, 8, 15, 25]
        column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        
        for i, width in enumerate(column_widths):
            if i < len(column_letters):
                ws.column_dimensions[column_letters[i]].width = width
        
        # ==================== GUARDAR ARCHIVO ====================
        
        # Crear buffer en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M')
        nombre_limpio = str(apellido_paterno + apellido_materno).replace(' ', '_')[:20]
        nombre_archivo = f"KARDEX_OFICIAL_{nombre_limpio}_{matricula}_{fecha_actual}.xlsx"
        
        return {
            'buffer': output.getvalue(),
            'filename': nombre_archivo,
            'estadisticas': {
                'materias_cursadas': len(materias_data),
                'aprobadas': len([c for c in calificaciones_num if c >= 6.0]),
                'reprobadas': len(materias_data) - len([c for c in calificaciones_num if c >= 6.0]),
                'promedio_general': round(promedio_general, 2)
            }
        }
        
    except Exception as e:
        st.error(f"Error al generar kardex oficial: {e}")
        import traceback
        st.error(f"Detalles técnicos: {traceback.format_exc()}")
        return None
    """
    Genera un kardex con formato oficial IDÉNTICO al archivo original de NovaUniversitas,
    replicando exactamente la estructura, layout y disposición.
    """
    try:
        # Obtener información básica del alumno
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DISTINCT nombre, matricula, carrera, email_personal, usuario_email_nova,
                   fecha_ingreso_original, cuatrimestre
            FROM calificaciones 
            WHERE matricula = ?
            LIMIT 1
        """, (matricula,))
        
        info_alumno = cursor.fetchone()
        if not info_alumno:
            st.error("No se encontró información del alumno en la base de datos")
            return None
            
        nombre_db, matricula_db, carrera, email_personal, email_nova, fecha_ingreso, cuatrimestre_actual = info_alumno
        
        # Obtener todas las materias con calificaciones organizadas por cuatrimestre
        cursor.execute("""
            SELECT 
                materia,
                calificacion,
                estatus,
                fecha_ingreso_materia,
                fecha_calificacion,
                profesor,
                cuatrimestre,
                cuatrimestre_historico,
                tipo_asignacion,
                n_recursamientos
            FROM calificaciones 
            WHERE matricula = ?
            AND materia IS NOT NULL 
            AND materia != ''
            AND calificacion IS NOT NULL
            ORDER BY 
                CAST(COALESCE(cuatrimestre_historico, cuatrimestre, 1) AS INTEGER) ASC,
                fecha_ingreso_materia ASC
        """, (matricula,))
        
        materias_data = cursor.fetchall()
        
        if not materias_data:
            st.warning("No hay materias con calificaciones para generar el kardex")
            return None
        
        # Organizar materias por cuatrimestre histórico
        materias_por_cuatrimestre = {}
        for materia_info in materias_data:
            (materia, calificacion, estatus, fecha_materia, fecha_calif, 
             profesor, cuatr_actual, cuatr_historico, tipo_asig, n_recurs) = materia_info
            
            cuatrimestre_key = str(cuatr_historico) if cuatr_historico and cuatr_historico != "None" else str(cuatr_actual)
            
            if cuatrimestre_key not in materias_por_cuatrimestre:
                materias_por_cuatrimestre[cuatrimestre_key] = []
            
            materias_por_cuatrimestre[cuatrimestre_key].append(materia_info)
        
        # Crear el archivo Excel con formato oficial EXACTO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from datetime import datetime
        import io
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "KARDEX"
        
        # ==================== CONFIGURACIÓN DE FORMATO EXACTO ====================
        
        # Fuentes exactas del formato original
        font_titulo_principal = Font(name='Arial', size=14, bold=True)
        font_subtitulos = Font(name='Arial', size=11, bold=True)
        font_normal = Font(name='Arial', size=9)
        font_small = Font(name='Arial', size=8)
        font_header_table = Font(name='Arial', size=8, bold=True)
        
        # Alineaciones
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        # Bordes exactos del formato
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        
        # ==================== ENCABEZADO PRINCIPAL EXACTO ====================
        
        current_row = 1
        
        # Logo placeholder (columna A, filas 1-6)
        ws.merge_cells('A1:A6')
        ws['A1'] = "NOVA\nUNIVERSITAS"
        ws['A1'].font = Font(name='Arial', size=10, bold=True)
        ws['A1'].alignment = center_alignment
        ws['A1'].border = thick_border
        
        # Título principal - NovaUniversitas
        ws.merge_cells('B1:H1')
        ws['B1'] = "NovaUniversitas"
        ws['B1'].font = font_titulo_principal
        ws['B1'].alignment = center_alignment
        
        # Subtítulo - Organismo Público Descentralizado
        ws.merge_cells('B2:H2')
        ws['B2'] = "ORGANISMO PÚBLICO DESCENTRALIZADO"
        ws['B2'].font = font_normal
        ws['B2'].alignment = center_alignment
        
        # Departamento
        ws.merge_cells('B3:H3')
        ws['B3'] = "DEPARTAMENTO DE SERVICIOS ESCOLARES"
        ws['B3'].font = font_normal
        ws['B3'].alignment = center_alignment
        
        # KARDEX title
        ws.merge_cells('B4:H4')
        ws['B4'] = "KARDEX"
        ws['B4'].font = Font(name='Arial', size=16, bold=True)
        ws['B4'].alignment = center_alignment
        
        # FOTO placeholder (columna I, filas 1-6)
        ws.merge_cells('I1:I6')
        ws['I1'] = "FOTO"
        ws['I1'].font = font_normal
        ws['I1'].alignment = center_alignment
        ws['I1'].border = thick_border
        
        current_row = 7
        
        # ==================== INFORMACIÓN DEL ESTUDIANTE EXACTA ====================
        
        # Línea de separación
        for col in range(1, 10):  # A-I
            ws.cell(row=current_row, column=col).border = Border(top=Side(style='thick'))
        current_row += 1
        
        # CARRERA
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = "CARRERA"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        
        ws[f'I{current_row}'] = "MATRÍCULA"
        ws[f'I{current_row}'].font = font_small
        ws[f'I{current_row}'].alignment = center_alignment
        current_row += 1
        
        # Valor de carrera y matrícula
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = f"No Escolarizada (virtual)"  # Formato exacto del original
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].border = thin_border
        
        ws[f'I{current_row}'] = str(matricula_db)
        ws[f'I{current_row}'].font = font_normal
        ws[f'I{current_row}'].alignment = center_alignment
        ws[f'I{current_row}'].border = thin_border
        current_row += 1
        
        # MODALIDAD
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = "MODALIDAD"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        # Campos de nombre - exacto como el original
        ws[f'A{current_row}'] = "APELLIDO PATERNO"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].border = thin_border
        
        ws[f'D{current_row}'] = "APELLIDO MATERNO"
        ws[f'D{current_row}'].font = font_small
        ws[f'D{current_row}'].alignment = center_alignment
        ws[f'D{current_row}'].border = thin_border
        
        ws.merge_cells(f'G{current_row}:I{current_row}')
        ws[f'G{current_row}'] = "NOMBRE(S)"
        ws[f'G{current_row}'].font = font_small
        ws[f'G{current_row}'].alignment = center_alignment
        ws[f'G{current_row}'].border = thin_border
        current_row += 1
        
        # Valores del nombre (dividir el nombre completo)
        nombre_parts = str(nombre_db).split()
        apellido_paterno = nombre_parts[0] if len(nombre_parts) > 0 else ""
        apellido_materno = nombre_parts[1] if len(nombre_parts) > 1 else ""
        nombres = " ".join(nombre_parts[2:]) if len(nombre_parts) > 2 else ""
        
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = apellido_paterno.upper()
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].border = thin_border
        
        ws.merge_cells(f'D{current_row}:F{current_row}')
        ws[f'D{current_row}'] = apellido_materno.upper()
        ws[f'D{current_row}'].font = font_normal
        ws[f'D{current_row}'].alignment = center_alignment
        ws[f'D{current_row}'].border = thin_border
        
        ws.merge_cells(f'G{current_row}:I{current_row}')
        ws[f'G{current_row}'] = nombres.upper()
        ws[f'G{current_row}'].font = font_normal
        ws[f'G{current_row}'].alignment = center_alignment
        ws[f'G{current_row}'].border = thin_border
        current_row += 1
        
        # SEXO y CURP
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = "SEXO"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        
        ws.merge_cells(f'E{current_row}:I{current_row}')
        ws[f'E{current_row}'] = "CURP"
        ws[f'E{current_row}'].font = font_small
        ws[f'E{current_row}'].alignment = center_alignment
        current_row += 1
        
        # Valores vacíos (no disponibles)
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = ""
        ws[f'A{current_row}'].border = thin_border
        
        ws.merge_cells(f'E{current_row}:I{current_row}')
        ws[f'E{current_row}'] = ""
        ws[f'E{current_row}'].border = thin_border
        current_row += 2
        
        # ==================== GENERAR CUATRIMESTRES EXACTOS ====================
        
        # Nombres de cuatrimestres exactos como en el original
        nombres_cuatrimestres = {
            '1': 'PRIMER CUATRIMESTRE',
            '2': 'SEGUNDO CUATRIMESTRE', 
            '3': 'TERCER CUATRIMESTRE',
            '4': 'CUARTO CUATRIMESTRE',
            '5': 'QUINTO CUATRIMESTRE',
            '6': 'SEXTO CUATRIMESTRE',
            '7': 'SÉPTIMO CUATRIMESTRE',
            '8': 'OCTAVO CUATRIMESTRE',
            '9': 'NOVENO CUATRIMESTRE',
            '10': 'DÉCIMO CUATRIMESTRE'
        }
        
        # Generar cada cuatrimestre con formato exacto
        cuatrimestres_ordenados = sorted([int(k) for k in materias_por_cuatrimestre.keys()])
        
        for cuatr_num in cuatrimestres_ordenados:
            cuatr_str = str(cuatr_num)
            if cuatr_str in nombres_cuatrimestres:
                materias_cuatr = materias_por_cuatrimestre[cuatr_str]
                
                # Período Escolar
                ws[f'A{current_row}'] = "Periodo Escolar:"
                ws[f'A{current_row}'].font = font_normal
                current_row += 1
                
                # Encabezados de tabla exactos
                headers = ['Clave', 'Seriación', 'Créditos', 'MATERIAS', 'C.F.', 'Tipo de\nexamen', 'Calif.', 'Fecha', 'Observaciones']
                col_widths = [8, 8, 8, 35, 8, 10, 8, 12, 20]
                
                # Headers de tabla - PRIMERO establecer todos los headers individuales
                header_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
                for i, (header, col_letter) in enumerate(zip(headers, header_cols)):
                    ws[f'{col_letter}{current_row}'] = header
                    ws[f'{col_letter}{current_row}'].font = font_header_table
                    ws[f'{col_letter}{current_row}'].alignment = center_alignment
                    ws[f'{col_letter}{current_row}'].border = thin_border
                
                # DESPUÉS fusionar y establecer el título del cuatrimestre
                ws.merge_cells(f'D{current_row}:E{current_row}')
                ws[f'D{current_row}'] = nombres_cuatrimestres[cuatr_str]
                ws[f'D{current_row}'].font = font_header_table
                ws[f'D{current_row}'].alignment = center_alignment
                
                current_row += 1
                
                # Filas de materias (exacto como el original - 6 filas por cuatrimestre)
                for i in range(6):
                    if i < len(materias_cuatr):
                        materia_info = materias_cuatr[i]
                        (materia, calificacion, estatus, fecha_materia, fecha_calif, 
                         profesor, cuatr_actual, cuatr_historico, tipo_asig, n_recurs) = materia_info
                        
                        # Datos de la materia
                        ws[f'A{current_row}'] = ""  # Clave (vacío)
                        ws[f'B{current_row}'] = ""  # Seriación (vacío)
                        ws[f'C{current_row}'] = ""  # Créditos (vacío)
                        ws[f'D{current_row}'] = str(materia).upper()
                        ws[f'E{current_row}'] = ""  # C.F. (vacío)
                        ws[f'F{current_row}'] = ""  # Tipo de examen (vacío)
                        
                        # Calificación
                        try:
                            calif_num = float(calificacion)
                            ws[f'G{current_row}'] = f"{calif_num:.1f}"
                        except:
                            ws[f'G{current_row}'] = str(calificacion)
                        
                        # Fecha
                        fecha_mostrar = str(fecha_calif) if fecha_calif and fecha_calif != "None" else str(fecha_materia)
                        if fecha_mostrar and fecha_mostrar not in ["None", "N/A", ""]:
                            try:
                                fecha_dt = pd.to_datetime(fecha_mostrar)
                                ws[f'H{current_row}'] = fecha_dt.strftime('%d/%m/%Y')
                            except:
                                ws[f'H{current_row}'] = fecha_mostrar
                        else:
                            ws[f'H{current_row}'] = ""
                        
                        # Observaciones
                        obs = "RECURSAMIENTO" if tipo_asig == "recursamiento" else ""
                        ws[f'I{current_row}'] = obs
                        
                    else:
                        # Filas vacías
                        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                            ws[f'{col}{current_row}'] = ""
                    
                    # Bordes en todas las celdas de la fila
                    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                        ws[f'{col}{current_row}'].border = thin_border
                        ws[f'{col}{current_row}'].font = font_normal
                    
                    current_row += 1
                
                # Elaboró y Cotejó exacto como el original
                ws[f'A{current_row}'] = "Elaboró:"
                ws[f'A{current_row}'].font = font_small
                
                ws[f'B{current_row}'] = "COORD."
                ws[f'B{current_row}'].font = font_small
                
                ws[f'H{current_row}'] = "Cotejó:"
                ws[f'H{current_row}'].font = font_small
                
                ws[f'I{current_row}'] = "DSE"
                ws[f'I{current_row}'].font = font_small
                
                current_row += 2  # Espacio entre cuatrimestres
        
        # ==================== PIE DE PÁGINA EXACTO ====================
        
        # Calcular promedio general
        calificaciones_num = []
        for materia_info in materias_data:
            try:
                if materia_info[1] is not None and str(materia_info[1]) != "None":
                    cal_float = float(materia_info[1])
                    calificaciones_num.append(cal_float)
            except (ValueError, TypeError):
                continue
        
        promedio_general = sum(calificaciones_num) / len(calificaciones_num) if calificaciones_num else 0.0
        
        current_row += 2
        
        # PROMEDIO GENERAL exacto
        ws.merge_cells(f'A{current_row}:G{current_row}')
        ws[f'A{current_row}'] = "PROMEDIO GENERAL:"
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = right_alignment
        
        ws.merge_cells(f'H{current_row}:I{current_row}')
        ws[f'H{current_row}'] = f"{promedio_general:.1f}"
        ws[f'H{current_row}'].font = font_normal
        ws[f'H{current_row}'].alignment = center_alignment
        ws[f'H{current_row}'].border = thick_border
        
        current_row += 1
        
        # Escala de evaluación exacta
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = "ESCALA DE EVALUACIÓN DEL 0.0 AL 10.0 MÍNIMA APROBATORIA 6.0"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        
        current_row += 3
        
        # Firmas exactas
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = "Vice-Rector Académico"
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = center_alignment
        
        ws.merge_cells(f'F{current_row}:I{current_row}')
        ws[f'F{current_row}'] = "Jefa del Departamento de Servicios Escolares"
        ws[f'F{current_row}'].font = font_normal
        ws[f'F{current_row}'].alignment = center_alignment
        
        # ==================== AJUSTAR DIMENSIONES EXACTAS ====================
        
        # Anchos de columna exactos del formato original
        column_widths = [12, 12, 12, 40, 8, 12, 8, 15, 25]
        column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        
        for i, width in enumerate(column_widths):
            if i < len(column_letters):
                ws.column_dimensions[column_letters[i]].width = width
        
        # ==================== GUARDAR ARCHIVO ====================
        
        # Crear buffer en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M')
        nombre_limpio = str(apellido_paterno + apellido_materno).replace(' ', '_')[:20]
        nombre_archivo = f"KARDEX_OFICIAL_{nombre_limpio}_{matricula}_{fecha_actual}.xlsx"
        
        return {
            'buffer': output.getvalue(),
            'filename': nombre_archivo,
            'estadisticas': {
                'materias_cursadas': len(materias_data),
                'aprobadas': len([c for c in calificaciones_num if c >= 6.0]),
                'reprobadas': len(materias_data) - len([c for c in calificaciones_num if c >= 6.0]),
                'promedio_general': round(promedio_general, 2)
            }
        }
        
    except Exception as e:
        st.error(f"Error al generar kardex oficial: {e}")
        import traceback
        st.error(f"Detalles técnicos: {traceback.format_exc()}")
        return None
    """
    Genera un kardex con formato oficial IDÉNTICO al archivo original de NovaUniversitas,
    replicando exactamente la estructura, layout y disposición.
    """
    try:
        # Obtener información básica del alumno
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DISTINCT nombre, matricula, carrera, email_personal, usuario_email_nova,
                   fecha_ingreso_original, cuatrimestre
            FROM calificaciones 
            WHERE matricula = ?
            LIMIT 1
        """, (matricula,))
        
        info_alumno = cursor.fetchone()
        if not info_alumno:
            st.error("No se encontró información del alumno en la base de datos")
            return None
            
        nombre_db, matricula_db, carrera, email_personal, email_nova, fecha_ingreso, cuatrimestre_actual = info_alumno
        
        # Obtener todas las materias con calificaciones organizadas por cuatrimestre
        cursor.execute("""
            SELECT 
                materia,
                calificacion,
                estatus,
                fecha_ingreso_materia,
                fecha_calificacion,
                profesor,
                cuatrimestre,
                cuatrimestre_historico,
                tipo_asignacion,
                n_recursamientos
            FROM calificaciones 
            WHERE matricula = ?
            AND materia IS NOT NULL 
            AND materia != ''
            AND calificacion IS NOT NULL
            ORDER BY 
                CAST(COALESCE(cuatrimestre_historico, cuatrimestre, 1) AS INTEGER) ASC,
                fecha_ingreso_materia ASC
        """, (matricula,))
        
        materias_data = cursor.fetchall()
        
        if not materias_data:
            st.warning("No hay materias con calificaciones para generar el kardex")
            return None
        
        # Organizar materias por cuatrimestre histórico
        materias_por_cuatrimestre = {}
        for materia_info in materias_data:
            (materia, calificacion, estatus, fecha_materia, fecha_calif, 
             profesor, cuatr_actual, cuatr_historico, tipo_asig, n_recurs) = materia_info
            
            cuatrimestre_key = str(cuatr_historico) if cuatr_historico and cuatr_historico != "None" else str(cuatr_actual)
            
            if cuatrimestre_key not in materias_por_cuatrimestre:
                materias_por_cuatrimestre[cuatrimestre_key] = []
            
            materias_por_cuatrimestre[cuatrimestre_key].append(materia_info)
        
        # Crear el archivo Excel con formato oficial EXACTO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from datetime import datetime
        import io
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "KARDEX"
        
        # ==================== CONFIGURACIÓN DE FORMATO EXACTO ====================
        
        # Fuentes exactas del formato original
        font_titulo_principal = Font(name='Arial', size=14, bold=True)
        font_subtitulos = Font(name='Arial', size=11, bold=True)
        font_normal = Font(name='Arial', size=9)
        font_small = Font(name='Arial', size=8)
        font_header_table = Font(name='Arial', size=8, bold=True)
        
        # Alineaciones
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        # Bordes exactos del formato
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        
        # ==================== ENCABEZADO PRINCIPAL EXACTO ====================
        
        current_row = 1
        
        # Logo placeholder (columna A, filas 1-6)
        ws.merge_cells('A1:A6')
        ws['A1'] = "NOVA\nUNIVERSITAS"
        ws['A1'].font = Font(name='Arial', size=10, bold=True)
        ws['A1'].alignment = center_alignment
        ws['A1'].border = thick_border
        
        # Título principal - NovaUniversitas
        ws.merge_cells('B1:H1')
        ws['B1'] = "NovaUniversitas"
        ws['B1'].font = font_titulo_principal
        ws['B1'].alignment = center_alignment
        
        # Subtítulo - Organismo Público Descentralizado
        ws.merge_cells('B2:H2')
        ws['B2'] = "ORGANISMO PÚBLICO DESCENTRALIZADO"
        ws['B2'].font = font_normal
        ws['B2'].alignment = center_alignment
        
        # Departamento
        ws.merge_cells('B3:H3')
        ws['B3'] = "DEPARTAMENTO DE SERVICIOS ESCOLARES"
        ws['B3'].font = font_normal
        ws['B3'].alignment = center_alignment
        
        # KARDEX title
        ws.merge_cells('B4:H4')
        ws['B4'] = "KARDEX"
        ws['B4'].font = Font(name='Arial', size=16, bold=True)
        ws['B4'].alignment = center_alignment
        
        # FOTO placeholder (columna I, filas 1-6)
        ws.merge_cells('I1:I6')
        ws['I1'] = "FOTO"
        ws['I1'].font = font_normal
        ws['I1'].alignment = center_alignment
        ws['I1'].border = thick_border
        
        current_row = 7
        
        # ==================== INFORMACIÓN DEL ESTUDIANTE EXACTA ====================
        
        # Línea de separación
        for col in range(1, 10):  # A-I
            ws.cell(row=current_row, column=col).border = Border(top=Side(style='thick'))
        current_row += 1
        
        # CARRERA
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = "CARRERA"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        
        ws[f'I{current_row}'] = "MATRÍCULA"
        ws[f'I{current_row}'].font = font_small
        ws[f'I{current_row}'].alignment = center_alignment
        current_row += 1
        
        # Valor de carrera y matrícula
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = f"No Escolarizada (virtual)"  # Formato exacto del original
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].border = thin_border
        
        ws[f'I{current_row}'] = str(matricula_db)
        ws[f'I{current_row}'].font = font_normal
        ws[f'I{current_row}'].alignment = center_alignment
        ws[f'I{current_row}'].border = thin_border
        current_row += 1
        
        # MODALIDAD
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = "MODALIDAD"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        # Campos de nombre - exacto como el original
        ws[f'A{current_row}'] = "APELLIDO PATERNO"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].border = thin_border
        
        ws[f'D{current_row}'] = "APELLIDO MATERNO"
        ws[f'D{current_row}'].font = font_small
        ws[f'D{current_row}'].alignment = center_alignment
        ws[f'D{current_row}'].border = thin_border
        
        ws.merge_cells(f'G{current_row}:I{current_row}')
        ws[f'G{current_row}'] = "NOMBRE(S)"
        ws[f'G{current_row}'].font = font_small
        ws[f'G{current_row}'].alignment = center_alignment
        ws[f'G{current_row}'].border = thin_border
        current_row += 1
        
        # Valores del nombre (dividir el nombre completo)
        nombre_parts = str(nombre_db).split()
        apellido_paterno = nombre_parts[0] if len(nombre_parts) > 0 else ""
        apellido_materno = nombre_parts[1] if len(nombre_parts) > 1 else ""
        nombres = " ".join(nombre_parts[2:]) if len(nombre_parts) > 2 else ""
        
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = apellido_paterno.upper()
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].border = thin_border
        
        ws.merge_cells(f'D{current_row}:F{current_row}')
        ws[f'D{current_row}'] = apellido_materno.upper()
        ws[f'D{current_row}'].font = font_normal
        ws[f'D{current_row}'].alignment = center_alignment
        ws[f'D{current_row}'].border = thin_border
        
        ws.merge_cells(f'G{current_row}:I{current_row}')
        ws[f'G{current_row}'] = nombres.upper()
        ws[f'G{current_row}'].font = font_normal
        ws[f'G{current_row}'].alignment = center_alignment
        ws[f'G{current_row}'].border = thin_border
        current_row += 1
        
        # SEXO y CURP
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = "SEXO"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        
        ws.merge_cells(f'E{current_row}:I{current_row}')
        ws[f'E{current_row}'] = "CURP"
        ws[f'E{current_row}'].font = font_small
        ws[f'E{current_row}'].alignment = center_alignment
        current_row += 1
        
        # Valores vacíos (no disponibles)
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = ""
        ws[f'A{current_row}'].border = thin_border
        
        ws.merge_cells(f'E{current_row}:I{current_row}')
        ws[f'E{current_row}'] = ""
        ws[f'E{current_row}'].border = thin_border
        current_row += 2
        
        # ==================== GENERAR CUATRIMESTRES EXACTOS ====================
        
        # Nombres de cuatrimestres exactos como en el original
        nombres_cuatrimestres = {
            '1': 'PRIMER CUATRIMESTRE',
            '2': 'SEGUNDO CUATRIMESTRE', 
            '3': 'TERCER CUATRIMESTRE',
            '4': 'CUARTO CUATRIMESTRE',
            '5': 'QUINTO CUATRIMESTRE',
            '6': 'SEXTO CUATRIMESTRE',
            '7': 'SÉPTIMO CUATRIMESTRE',
            '8': 'OCTAVO CUATRIMESTRE',
            '9': 'NOVENO CUATRIMESTRE',
            '10': 'DÉCIMO CUATRIMESTRE'
        }
        
        # Generar cada cuatrimestre con formato exacto
        cuatrimestres_ordenados = sorted([int(k) for k in materias_por_cuatrimestre.keys()])
        
        for cuatr_num in cuatrimestres_ordenados:
            cuatr_str = str(cuatr_num)
            if cuatr_str in nombres_cuatrimestres:
                materias_cuatr = materias_por_cuatrimestre[cuatr_str]
                
                # Período Escolar
                ws[f'A{current_row}'] = "Periodo Escolar:"
                ws[f'A{current_row}'].font = font_normal
                current_row += 1
                
                # Encabezados de tabla exactos
                headers = ['Clave', 'Seriación', 'Créditos', 'MATERIAS', 'C.F.', 'Tipo de\nexamen', 'Calif.', 'Fecha', 'Observaciones']
                col_widths = [8, 8, 8, 35, 8, 10, 8, 12, 20]
                
                # Título del cuatrimestre centrado
                ws.merge_cells(f'D{current_row}:E{current_row}')
                ws[f'D{current_row}'] = nombres_cuatrimestres[cuatr_str]
                ws[f'D{current_row}'].font = font_header_table
                ws[f'D{current_row}'].alignment = center_alignment
                
                # Headers de tabla
                header_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
                for i, (header, col_letter) in enumerate(zip(headers, header_cols)):
                    if i != 3:  # Skip MATERIAS que ya está fusionada
                        ws[f'{col_letter}{current_row}'] = header
                        ws[f'{col_letter}{current_row}'].font = font_header_table
                        ws[f'{col_letter}{current_row}'].alignment = center_alignment
                        ws[f'{col_letter}{current_row}'].border = thin_border
                
                current_row += 1
                
                # Filas de materias (exacto como el original - 6 filas por cuatrimestre)
                for i in range(6):
                    if i < len(materias_cuatr):
                        materia_info = materias_cuatr[i]
                        (materia, calificacion, estatus, fecha_materia, fecha_calif, 
                         profesor, cuatr_actual, cuatr_historico, tipo_asig, n_recurs) = materia_info
                        
                        # Datos de la materia
                        ws[f'A{current_row}'] = ""  # Clave (vacío)
                        ws[f'B{current_row}'] = ""  # Seriación (vacío)
                        ws[f'C{current_row}'] = ""  # Créditos (vacío)
                        ws[f'D{current_row}'] = str(materia).upper()
                        ws[f'E{current_row}'] = ""  # C.F. (vacío)
                        ws[f'F{current_row}'] = ""  # Tipo de examen (vacío)
                        
                        # Calificación
                        try:
                            calif_num = float(calificacion)
                            ws[f'G{current_row}'] = f"{calif_num:.1f}"
                        except:
                            ws[f'G{current_row}'] = str(calificacion)
                        
                        # Fecha
                        fecha_mostrar = str(fecha_calif) if fecha_calif and fecha_calif != "None" else str(fecha_materia)
                        if fecha_mostrar and fecha_mostrar not in ["None", "N/A", ""]:
                            try:
                                fecha_dt = pd.to_datetime(fecha_mostrar)
                                ws[f'H{current_row}'] = fecha_dt.strftime('%d/%m/%Y')
                            except:
                                ws[f'H{current_row}'] = fecha_mostrar
                        else:
                            ws[f'H{current_row}'] = ""
                        
                        # Observaciones
                        obs = "RECURSAMIENTO" if tipo_asig == "recursamiento" else ""
                        ws[f'I{current_row}'] = obs
                        
                    else:
                        # Filas vacías
                        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                            ws[f'{col}{current_row}'] = ""
                    
                    # Bordes en todas las celdas de la fila
                    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                        ws[f'{col}{current_row}'].border = thin_border
                        ws[f'{col}{current_row}'].font = font_normal
                    
                    current_row += 1
                
                # Elaboró y Cotejó exacto como el original
                ws[f'A{current_row}'] = "Elaboró:"
                ws[f'A{current_row}'].font = font_small
                
                ws[f'B{current_row}'] = "COORD."
                ws[f'B{current_row}'].font = font_small
                
                ws[f'H{current_row}'] = "Cotejó:"
                ws[f'H{current_row}'].font = font_small
                
                ws[f'I{current_row}'] = "DSE"
                ws[f'I{current_row}'].font = font_small
                
                current_row += 2  # Espacio entre cuatrimestres
        
        # ==================== PIE DE PÁGINA EXACTO ====================
        
        # Calcular promedio general
        calificaciones_num = []
        for materia_info in materias_data:
            try:
                if materia_info[1] is not None and str(materia_info[1]) != "None":
                    cal_float = float(materia_info[1])
                    calificaciones_num.append(cal_float)
            except (ValueError, TypeError):
                continue
        
        promedio_general = sum(calificaciones_num) / len(calificaciones_num) if calificaciones_num else 0.0
        
        current_row += 2
        
        # PROMEDIO GENERAL exacto
        ws.merge_cells(f'A{current_row}:G{current_row}')
        ws[f'A{current_row}'] = "PROMEDIO GENERAL:"
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = right_alignment
        
        ws.merge_cells(f'H{current_row}:I{current_row}')
        ws[f'H{current_row}'] = f"{promedio_general:.1f}"
        ws[f'H{current_row}'].font = font_normal
        ws[f'H{current_row}'].alignment = center_alignment
        ws[f'H{current_row}'].border = thick_border
        
        current_row += 1
        
        # Escala de evaluación exacta
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = "ESCALA DE EVALUACIÓN DEL 0.0 AL 10.0 MÍNIMA APROBATORIA 6.0"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        
        current_row += 3
        
        # Firmas exactas
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = "Vice-Rector Académico"
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = center_alignment
        
        ws.merge_cells(f'F{current_row}:I{current_row}')
        ws[f'F{current_row}'] = "Jefa del Departamento de Servicios Escolares"
        ws[f'F{current_row}'].font = font_normal
        ws[f'F{current_row}'].alignment = center_alignment
        
        # ==================== AJUSTAR DIMENSIONES EXACTAS ====================
        
        # Anchos de columna exactos del formato original
        column_widths = [12, 12, 12, 40, 8, 12, 8, 15, 25]
        column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        
        for i, width in enumerate(column_widths):
            if i < len(column_letters):
                ws.column_dimensions[column_letters[i]].width = width
        
        # ==================== GUARDAR ARCHIVO ====================
        
        # Crear buffer en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M')
        nombre_limpio = str(apellido_paterno + apellido_materno).replace(' ', '_')[:20]
        nombre_archivo = f"KARDEX_OFICIAL_{nombre_limpio}_{matricula}_{fecha_actual}.xlsx"
        
        return {
            'buffer': output.getvalue(),
            'filename': nombre_archivo,
            'estadisticas': {
                'materias_cursadas': len(materias_data),
                'aprobadas': len([c for c in calificaciones_num if c >= 6.0]),
                'reprobadas': len(materias_data) - len([c for c in calificaciones_num if c >= 6.0]),
                'promedio_general': round(promedio_general, 2)
            }
        }
        
    except Exception as e:
        st.error(f"Error al generar kardex oficial: {e}")
        import traceback
        st.error(f"Detalles técnicos: {traceback.format_exc()}")
        return None
    """
    Genera un kardex con formato oficial idéntico al archivo original,
    SIN importar imágenes para evitar errores de dependencias.
    """
    try:
        # Obtener información básica del alumno
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DISTINCT nombre, matricula, carrera, email_personal, usuario_email_nova,
                   fecha_ingreso_original, cuatrimestre
            FROM calificaciones 
            WHERE matricula = ?
            LIMIT 1
        """, (matricula,))
        
        info_alumno = cursor.fetchone()
        if not info_alumno:
            st.error("No se encontró información del alumno en la base de datos")
            return None
            
        nombre_db, matricula_db, carrera, email_personal, email_nova, fecha_ingreso, cuatrimestre_actual = info_alumno
        
        # Obtener todas las materias con calificaciones
        cursor.execute("""
            SELECT 
                materia,
                calificacion,
                estatus,
                fecha_ingreso_materia,
                fecha_calificacion,
                profesor,
                cuatrimestre,
                cuatrimestre_historico,
                tipo_asignacion,
                n_recursamientos
            FROM calificaciones 
            WHERE matricula = ?
            AND materia IS NOT NULL 
            AND materia != ''
            AND calificacion IS NOT NULL
            ORDER BY 
                CAST(COALESCE(cuatrimestre_historico, cuatrimestre, 1) AS INTEGER) ASC,
                fecha_ingreso_materia ASC
        """, (matricula,))
        
        materias_data = cursor.fetchall()
        
        if not materias_data:
            st.warning("No hay materias con calificaciones para generar el kardex")
            return None
        
        # Crear el archivo Excel con formato oficial
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from datetime import datetime
        import io
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "KARDEX"
        
        # ==================== CONFIGURACIÓN DE FORMATO ====================
        
        # Fuentes
        font_titulo = Font(name='Arial', size=16, bold=True)
        font_subtitulo = Font(name='Arial', size=14, bold=True)
        font_normal = Font(name='Arial', size=10)
        font_small = Font(name='Arial', size=9)
        font_encabezado = Font(name='Arial', size=10, bold=True)
        
        # Alineaciones
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        
        # Colores
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        titulo_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # ==================== ENCABEZADO INSTITUCIONAL ====================
        
        # Fila 1: Título principal con fondo azul
        ws.merge_cells('A1:I1')
        ws['A1'] = "UNIVERSIDAD NOVA HISPANA"
        ws['A1'].font = Font(name='Arial', size=18, bold=True, color="FFFFFF")
        ws['A1'].alignment = center_alignment
        ws['A1'].fill = titulo_fill
        ws['A1'].border = thick_border
        
        # Fila 2: Subtítulo
        ws.merge_cells('A2:I2')
        ws['A2'] = "KARDEX ACADÉMICO OFICIAL"
        ws['A2'].font = font_titulo
        ws['A2'].alignment = center_alignment
        ws['A2'].fill = header_fill
        ws['A2'].border = thin_border
        
        # Fila 3: Espacio
        ws.row_dimensions[3].height = 10
        
        # ==================== INFORMACIÓN DEL ESTUDIANTE ====================
        
        # Fila 4: Datos del alumno - Primera línea
        ws['A4'] = "MATRÍCULA:"
        ws['A4'].font = font_encabezado
        ws['A4'].fill = header_fill
        ws['A4'].border = thin_border
        
        ws.merge_cells('B4:C4')
        ws['B4'] = str(matricula_db)
        ws['B4'].font = font_normal
        ws['B4'].border = thin_border
        
        ws['D4'] = "NOMBRE COMPLETO:"
        ws['D4'].font = font_encabezado
        ws['D4'].fill = header_fill
        ws['D4'].border = thin_border
        
        ws.merge_cells('E4:I4')
        ws['E4'] = str(nombre_db).upper()
        ws['E4'].font = font_normal
        ws['E4'].border = thin_border
        
        # Fila 5: Segunda línea de datos
        ws['A5'] = "CARRERA:"
        ws['A5'].font = font_encabezado
        ws['A5'].fill = header_fill
        ws['A5'].border = thin_border
        
        ws.merge_cells('B5:G5')
        ws['B5'] = str(carrera).upper()
        ws['B5'].font = font_normal
        ws['B5'].border = thin_border
        
        ws['H5'] = "CUATRIMESTRE:"
        ws['H5'].font = font_encabezado
        ws['H5'].fill = header_fill
        ws['H5'].border = thin_border
        
        ws['I5'] = str(cuatrimestre_actual)
        ws['I5'].font = font_normal
        ws['I5'].alignment = center_alignment
        ws['I5'].border = thin_border
        
        # Fila 6: Tercera línea de datos
        ws['A6'] = "FECHA INGRESO:"
        ws['A6'].font = font_encabezado
        ws['A6'].fill = header_fill
        ws['A6'].border = thin_border
        
        ws.merge_cells('B6:D6')
        fecha_ing_formatted = str(fecha_ingreso) if fecha_ingreso else "N/A"
        ws['B6'] = fecha_ing_formatted
        ws['B6'].font = font_normal
        ws['B6'].border = thin_border
        
        ws['E6'] = "FECHA EMISIÓN:"
        ws['E6'].font = font_encabezado
        ws['E6'].fill = header_fill
        ws['E6'].border = thin_border
        
        ws.merge_cells('F6:I6')
        ws['F6'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        ws['F6'].font = font_normal
        ws['F6'].border = thin_border
        
        # Fila 7: Email institucional
        ws['A7'] = "EMAIL INSTITUCIONAL:"
        ws['A7'].font = font_encabezado
        ws['A7'].fill = header_fill
        ws['A7'].border = thin_border
        
        ws.merge_cells('B7:I7')
        ws['B7'] = str(email_nova) if email_nova and email_nova != "None" else "N/A"
        ws['B7'].font = font_normal
        ws['B7'].border = thin_border
        
        # Fila 8: Espacio
        ws.row_dimensions[8].height = 15
        
        # ==================== ENCABEZADOS DE TABLA ====================
        
        # Fila 9: Encabezados de la tabla de materias
        encabezados = [
            "CUATRIMESTRE",
            "CÓDIGO",
            "MATERIA",
            "CALIFICACIÓN",
            "CRÉDITOS", 
            "ESTATUS",
            "FECHA",
            "PROFESOR",
            "OBSERVACIONES"
        ]
        
        for col, encabezado in enumerate(encabezados, 1):
            cell = ws.cell(row=9, column=col)
            cell.value = encabezado
            cell.font = font_encabezado
            cell.alignment = center_alignment
            cell.fill = header_fill
            cell.border = thin_border
        
        # ==================== DATOS DE MATERIAS ====================
        
        row_start = 10
        current_row = row_start
        
        # Procesar materias por cuatrimestre
        for materia_info in materias_data:
            (materia, calificacion, estatus, fecha_materia, fecha_calif, 
             profesor, cuatr_actual, cuatr_historico, tipo_asig, n_recurs) = materia_info
            
            # Determinar cuatrimestre a mostrar (histórico si existe, sino actual)
            cuatrimestre_mostrar = str(cuatr_historico) if cuatr_historico and cuatr_historico != "None" else str(cuatr_actual)
            
            # Llenar fila de materia
            ws.cell(row=current_row, column=1).value = cuatrimestre_mostrar
            ws.cell(row=current_row, column=1).alignment = center_alignment
            ws.cell(row=current_row, column=1).border = thin_border
            ws.cell(row=current_row, column=1).font = font_normal
            
            # Código (campo vacío por diseño)
            ws.cell(row=current_row, column=2).value = ""
            ws.cell(row=current_row, column=2).alignment = center_alignment
            ws.cell(row=current_row, column=2).border = thin_border
            ws.cell(row=current_row, column=2).font = font_normal
            
            # Materia
            ws.cell(row=current_row, column=3).value = str(materia).upper()
            ws.cell(row=current_row, column=3).alignment = left_alignment
            ws.cell(row=current_row, column=3).border = thin_border
            ws.cell(row=current_row, column=3).font = font_normal
            
            # Calificación con formato numérico
            try:
                calif_num = float(calificacion)
                calif_formatted = f"{calif_num:.1f}"
            except (ValueError, TypeError):
                calif_formatted = str(calificacion) if calificacion else "0.0"
            
            ws.cell(row=current_row, column=4).value = calif_formatted
            ws.cell(row=current_row, column=4).alignment = center_alignment
            ws.cell(row=current_row, column=4).border = thin_border
            ws.cell(row=current_row, column=4).font = font_normal
            
            # Créditos (campo vacío por no disponible)
            ws.cell(row=current_row, column=5).value = ""
            ws.cell(row=current_row, column=5).alignment = center_alignment
            ws.cell(row=current_row, column=5).border = thin_border
            ws.cell(row=current_row, column=5).font = font_normal
            
            # Estatus
            estatus_map = {
                "aprobado": "APROBADO",
                "reprobado": "REPROBADO", 
                "recursando": "RECURSANDO"
            }
            estatus_formatted = estatus_map.get(str(estatus).lower(), str(estatus).upper()) if estatus else "N/A"
            ws.cell(row=current_row, column=6).value = estatus_formatted
            ws.cell(row=current_row, column=6).alignment = center_alignment
            ws.cell(row=current_row, column=6).border = thin_border
            ws.cell(row=current_row, column=6).font = font_normal
            
            # Fecha de calificación
            fecha_mostrar = str(fecha_calif) if fecha_calif and fecha_calif != "None" else str(fecha_materia)
            if fecha_mostrar and fecha_mostrar not in ["None", "N/A", ""]:
                try:
                    fecha_dt = pd.to_datetime(fecha_mostrar)
                    fecha_formatted = fecha_dt.strftime('%d/%m/%Y')
                except:
                    fecha_formatted = fecha_mostrar
            else:
                fecha_formatted = ""
            
            ws.cell(row=current_row, column=7).value = fecha_formatted
            ws.cell(row=current_row, column=7).alignment = center_alignment
            ws.cell(row=current_row, column=7).border = thin_border
            ws.cell(row=current_row, column=7).font = font_normal
            
            # Profesor
            profesor_formatted = str(profesor).title() if profesor and profesor != "None" else ""
            ws.cell(row=current_row, column=8).value = profesor_formatted
            ws.cell(row=current_row, column=8).alignment = left_alignment
            ws.cell(row=current_row, column=8).border = thin_border
            ws.cell(row=current_row, column=8).font = font_normal
            
            # Observaciones (recursamiento, etc.)
            observaciones = []
            if tipo_asig == "recursamiento":
                observaciones.append("RECURSAMIENTO")
            if n_recurs and n_recurs > 0:
                try:
                    n_rec_int = int(n_recurs)
                    if n_rec_int > 0:
                        observaciones.append(f"RECURS: {n_rec_int}")
                except:
                    pass
            
            obs_text = " | ".join(observaciones)
            ws.cell(row=current_row, column=9).value = obs_text
            ws.cell(row=current_row, column=9).alignment = left_alignment
            ws.cell(row=current_row, column=9).border = thin_border
            ws.cell(row=current_row, column=9).font = font_small
            
            current_row += 1
        
        # ==================== ESTADÍSTICAS FINALES ====================
        
        # Agregar algunas filas vacías
        current_row += 2
        
        # Calcular estadísticas
        materias_cursadas = len(materias_data)
        calificaciones_num = []
        for m in materias_data:
            try:
                if m[1] is not None and str(m[1]) != "None":
                    cal_float = float(m[1])
                    calificaciones_num.append(cal_float)
            except (ValueError, TypeError):
                continue
        
        aprobadas = len([c for c in calificaciones_num if c >= 6.0])
        reprobadas = materias_cursadas - aprobadas
        promedio_general = sum(calificaciones_num) / len(calificaciones_num) if calificaciones_num else 0.0
        
        # Box de estadísticas con bordes
        stats_start_row = current_row
        
        # Título de estadísticas
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = "RESUMEN ACADÉMICO"
        ws[f'A{current_row}'].font = font_encabezado
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].fill = header_fill
        ws[f'A{current_row}'].border = thin_border
        current_row += 1
        
        # Estadísticas en formato de tabla
        stats_data = [
            ("MATERIAS CURSADAS:", materias_cursadas),
            ("MATERIAS APROBADAS:", aprobadas),  
            ("MATERIAS REPROBADAS:", reprobadas),
            ("PROMEDIO GENERAL:", f"{promedio_general:.2f}")
        ]
        
        for stat_label, stat_value in stats_data:
            ws[f'A{current_row}'] = stat_label
            ws[f'A{current_row}'].font = font_encabezado
            ws[f'A{current_row}'].alignment = left_alignment
            ws[f'A{current_row}'].border = thin_border
            
            ws.merge_cells(f'B{current_row}:I{current_row}')
            ws[f'B{current_row}'] = str(stat_value)
            ws[f'B{current_row}'].font = font_normal
            ws[f'B{current_row}'].alignment = center_alignment
            ws[f'B{current_row}'].border = thin_border
            
            current_row += 1
        
        current_row += 1
        
        # ==================== PIE DE PÁGINA OFICIAL ====================
        
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = f"DOCUMENTO GENERADO AUTOMÁTICAMENTE EL {datetime.now().strftime('%d/%m/%Y A LAS %H:%M:%S')}"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = "UNIVERSIDAD NOVA HISPANA - SISTEMA DE CONTROL ESCOLAR"
        ws[f'A{current_row}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = "DOCUMENTO OFICIAL - VÁLIDO PARA TRÁMITES ADMINISTRATIVOS"
        ws[f'A{current_row}'].font = Font(name='Arial', size=9, bold=True, color="FF0000")
        ws[f'A{current_row}'].alignment = center_alignment
        
        # ==================== AJUSTAR DIMENSIONES ====================
        
        # Ajustar anchos de columna para mejor presentación
        column_widths = [12, 8, 40, 12, 10, 15, 12, 25, 20]
        column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        
        for i, width in enumerate(column_widths):
            if i < len(column_letters):
                ws.column_dimensions[column_letters[i]].width = width
        
        # Ajustar alturas de fila importantes
        ws.row_dimensions[1].height = 30  # Título principal
        ws.row_dimensions[2].height = 25  # Subtítulo
        ws.row_dimensions[9].height = 25  # Encabezados de tabla
        
        # Ajustar altura de filas de datos
        for row_num in range(row_start, current_row - 5):
            ws.row_dimensions[row_num].height = 20
        
        # ==================== GUARDAR ARCHIVO ====================
        
        # Crear buffer en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M')
        nombre_limpio = str(nombre_db).replace(' ', '_').replace('/', '_')[:25]
        nombre_archivo = f"KARDEX_OFICIAL_{nombre_limpio}_{matricula}_{fecha_actual}.xlsx"
        
        return {
            'buffer': output.getvalue(),
            'filename': nombre_archivo,
            'estadisticas': {
                'materias_cursadas': materias_cursadas,
                'aprobadas': aprobadas,
                'reprobadas': reprobadas,
                'promedio_general': round(promedio_general, 2)
            }
        }
        
    except Exception as e:
        st.error(f"Error al generar kardex oficial: {e}")
        import traceback
        st.error(f"Detalles técnicos: {traceback.format_exc()}")
        return None
    """
    Genera un kardex con formato oficial idéntico al archivo original,
    SIN importar imágenes para evitar errores de dependencias.
    """
    try:
        # Obtener información básica del alumno
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DISTINCT nombre, matricula, carrera, email_personal, usuario_email_nova,
                   fecha_ingreso_original, cuatrimestre
            FROM calificaciones 
            WHERE matricula = ?
            LIMIT 1
        """, (matricula,))
        
        info_alumno = cursor.fetchone()
        if not info_alumno:
            st.error("No se encontró información del alumno en la base de datos")
            return None
            
        nombre_db, matricula_db, carrera, email_personal, email_nova, fecha_ingreso, cuatrimestre_actual = info_alumno
        
        # Obtener todas las materias con calificaciones
        cursor.execute("""
            SELECT 
                materia,
                calificacion,
                estatus,
                fecha_ingreso_materia,
                fecha_calificacion,
                profesor,
                cuatrimestre,
                cuatrimestre_historico,
                tipo_asignacion,
                n_recursamientos
            FROM calificaciones 
            WHERE matricula = ?
            AND materia IS NOT NULL 
            AND materia != ''
            AND calificacion IS NOT NULL
            ORDER BY 
                CAST(COALESCE(cuatrimestre_historico, cuatrimestre, 1) AS INTEGER) ASC,
                fecha_ingreso_materia ASC
        """, (matricula,))
        
        materias_data = cursor.fetchall()
        
        if not materias_data:
            st.warning("No hay materias con calificaciones para generar el kardex")
            return None
        
        # Crear el archivo Excel con formato oficial
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from datetime import datetime
        import io
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "KARDEX"
        
        # ==================== CONFIGURACIÓN DE FORMATO ====================
        
        # Fuentes
        font_titulo = Font(name='Arial', size=16, bold=True)
        font_subtitulo = Font(name='Arial', size=14, bold=True)
        font_normal = Font(name='Arial', size=10)
        font_small = Font(name='Arial', size=9)
        font_encabezado = Font(name='Arial', size=10, bold=True)
        
        # Alineaciones
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        
        # Colores
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        titulo_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # ==================== ENCABEZADO INSTITUCIONAL ====================
        
        # Fila 1: Título principal con fondo azul
        ws.merge_cells('A1:I1')
        ws['A1'] = "UNIVERSIDAD NOVA HISPANA"
        ws['A1'].font = Font(name='Arial', size=18, bold=True, color="FFFFFF")
        ws['A1'].alignment = center_alignment
        ws['A1'].fill = titulo_fill
        ws['A1'].border = thick_border
        
        # Fila 2: Subtítulo
        ws.merge_cells('A2:I2')
        ws['A2'] = "KARDEX ACADÉMICO OFICIAL"
        ws['A2'].font = font_titulo
        ws['A2'].alignment = center_alignment
        ws['A2'].fill = header_fill
        ws['A2'].border = thin_border
        
        # Fila 3: Espacio
        ws.row_dimensions[3].height = 10
        
        # ==================== INFORMACIÓN DEL ESTUDIANTE ====================
        
        # Fila 4: Datos del alumno - Primera línea
        ws['A4'] = "MATRÍCULA:"
        ws['A4'].font = font_encabezado
        ws['A4'].fill = header_fill
        ws['A4'].border = thin_border
        
        ws.merge_cells('B4:C4')
        ws['B4'] = str(matricula_db)
        ws['B4'].font = font_normal
        ws['B4'].border = thin_border
        
        ws['D4'] = "NOMBRE COMPLETO:"
        ws['D4'].font = font_encabezado
        ws['D4'].fill = header_fill
        ws['D4'].border = thin_border
        
        ws.merge_cells('E4:I4')
        ws['E4'] = str(nombre_db).upper()
        ws['E4'].font = font_normal
        ws['E4'].border = thin_border
        
        # Fila 5: Segunda línea de datos
        ws['A5'] = "CARRERA:"
        ws['A5'].font = font_encabezado
        ws['A5'].fill = header_fill
        ws['A5'].border = thin_border
        
        ws.merge_cells('B5:G5')
        ws['B5'] = str(carrera).upper()
        ws['B5'].font = font_normal
        ws['B5'].border = thin_border
        
        ws['H5'] = "CUATRIMESTRE:"
        ws['H5'].font = font_encabezado
        ws['H5'].fill = header_fill
        ws['H5'].border = thin_border
        
        ws['I5'] = str(cuatrimestre_actual)
        ws['I5'].font = font_normal
        ws['I5'].alignment = center_alignment
        ws['I5'].border = thin_border
        
        # Fila 6: Tercera línea de datos
        ws['A6'] = "FECHA INGRESO:"
        ws['A6'].font = font_encabezado
        ws['A6'].fill = header_fill
        ws['A6'].border = thin_border
        
        ws.merge_cells('B6:D6')
        fecha_ing_formatted = str(fecha_ingreso) if fecha_ingreso else "N/A"
        ws['B6'] = fecha_ing_formatted
        ws['B6'].font = font_normal
        ws['B6'].border = thin_border
        
        ws['E6'] = "FECHA EMISIÓN:"
        ws['E6'].font = font_encabezado
        ws['E6'].fill = header_fill
        ws['E6'].border = thin_border
        
        ws.merge_cells('F6:I6')
        ws['F6'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        ws['F6'].font = font_normal
        ws['F6'].border = thin_border
        
        # Fila 7: Email institucional
        ws['A7'] = "EMAIL INSTITUCIONAL:"
        ws['A7'].font = font_encabezado
        ws['A7'].fill = header_fill
        ws['A7'].border = thin_border
        
        ws.merge_cells('B7:I7')
        ws['B7'] = str(email_nova) if email_nova and email_nova != "None" else "N/A"
        ws['B7'].font = font_normal
        ws['B7'].border = thin_border
        
        # Fila 8: Espacio
        ws.row_dimensions[8].height = 15
        
        # ==================== ENCABEZADOS DE TABLA ====================
        
        # Fila 9: Encabezados de la tabla de materias
        encabezados = [
            "CUATRIMESTRE",
            "CÓDIGO",
            "MATERIA",
            "CALIFICACIÓN",
            "CRÉDITOS", 
            "ESTATUS",
            "FECHA",
            "PROFESOR",
            "OBSERVACIONES"
        ]
        
        for col, encabezado in enumerate(encabezados, 1):
            cell = ws.cell(row=9, column=col)
            cell.value = encabezado
            cell.font = font_encabezado
            cell.alignment = center_alignment
            cell.fill = header_fill
            cell.border = thin_border
        
        # ==================== DATOS DE MATERIAS ====================
        
        row_start = 10
        current_row = row_start
        
        # Procesar materias por cuatrimestre
        for materia_info in materias_data:
            (materia, calificacion, estatus, fecha_materia, fecha_calif, 
             profesor, cuatr_actual, cuatr_historico, tipo_asig, n_recurs) = materia_info
            
            # Determinar cuatrimestre a mostrar (histórico si existe, sino actual)
            cuatrimestre_mostrar = str(cuatr_historico) if cuatr_historico and cuatr_historico != "None" else str(cuatr_actual)
            
            # Llenar fila de materia
            ws.cell(row=current_row, column=1).value = cuatrimestre_mostrar
            ws.cell(row=current_row, column=1).alignment = center_alignment
            ws.cell(row=current_row, column=1).border = thin_border
            ws.cell(row=current_row, column=1).font = font_normal
            
            # Código (campo vacío por diseño)
            ws.cell(row=current_row, column=2).value = ""
            ws.cell(row=current_row, column=2).alignment = center_alignment
            ws.cell(row=current_row, column=2).border = thin_border
            ws.cell(row=current_row, column=2).font = font_normal
            
            # Materia
            ws.cell(row=current_row, column=3).value = str(materia).upper()
            ws.cell(row=current_row, column=3).alignment = left_alignment
            ws.cell(row=current_row, column=3).border = thin_border
            ws.cell(row=current_row, column=3).font = font_normal
            
            # Calificación con formato numérico
            try:
                calif_num = float(calificacion)
                calif_formatted = f"{calif_num:.1f}"
            except (ValueError, TypeError):
                calif_formatted = str(calificacion) if calificacion else "0.0"
            
            ws.cell(row=current_row, column=4).value = calif_formatted
            ws.cell(row=current_row, column=4).alignment = center_alignment
            ws.cell(row=current_row, column=4).border = thin_border
            ws.cell(row=current_row, column=4).font = font_normal
            
            # Créditos (campo vacío por no disponible)
            ws.cell(row=current_row, column=5).value = ""
            ws.cell(row=current_row, column=5).alignment = center_alignment
            ws.cell(row=current_row, column=5).border = thin_border
            ws.cell(row=current_row, column=5).font = font_normal
            
            # Estatus
            estatus_map = {
                "aprobado": "APROBADO",
                "reprobado": "REPROBADO", 
                "recursando": "RECURSANDO"
            }
            estatus_formatted = estatus_map.get(str(estatus).lower(), str(estatus).upper()) if estatus else "N/A"
            ws.cell(row=current_row, column=6).value = estatus_formatted
            ws.cell(row=current_row, column=6).alignment = center_alignment
            ws.cell(row=current_row, column=6).border = thin_border
            ws.cell(row=current_row, column=6).font = font_normal
            
            # Fecha de calificación
            fecha_mostrar = str(fecha_calif) if fecha_calif and fecha_calif != "None" else str(fecha_materia)
            if fecha_mostrar and fecha_mostrar not in ["None", "N/A", ""]:
                try:
                    fecha_dt = pd.to_datetime(fecha_mostrar)
                    fecha_formatted = fecha_dt.strftime('%d/%m/%Y')
                except:
                    fecha_formatted = fecha_mostrar
            else:
                fecha_formatted = ""
            
            ws.cell(row=current_row, column=7).value = fecha_formatted
            ws.cell(row=current_row, column=7).alignment = center_alignment
            ws.cell(row=current_row, column=7).border = thin_border
            ws.cell(row=current_row, column=7).font = font_normal
            
            # Profesor
            profesor_formatted = str(profesor).title() if profesor and profesor != "None" else ""
            ws.cell(row=current_row, column=8).value = profesor_formatted
            ws.cell(row=current_row, column=8).alignment = left_alignment
            ws.cell(row=current_row, column=8).border = thin_border
            ws.cell(row=current_row, column=8).font = font_normal
            
            # Observaciones (recursamiento, etc.)
            observaciones = []
            if tipo_asig == "recursamiento":
                observaciones.append("RECURSAMIENTO")
            if n_recurs and n_recurs > 0:
                try:
                    n_rec_int = int(n_recurs)
                    if n_rec_int > 0:
                        observaciones.append(f"RECURS: {n_rec_int}")
                except:
                    pass
            
            obs_text = " | ".join(observaciones)
            ws.cell(row=current_row, column=9).value = obs_text
            ws.cell(row=current_row, column=9).alignment = left_alignment
            ws.cell(row=current_row, column=9).border = thin_border
            ws.cell(row=current_row, column=9).font = font_small
            
            current_row += 1
        
        # ==================== ESTADÍSTICAS FINALES ====================
        
        # Agregar algunas filas vacías
        current_row += 2
        
        # Calcular estadísticas
        materias_cursadas = len(materias_data)
        calificaciones_num = []
        for m in materias_data:
            try:
                if m[1] is not None and str(m[1]) != "None":
                    cal_float = float(m[1])
                    calificaciones_num.append(cal_float)
            except (ValueError, TypeError):
                continue
        
        aprobadas = len([c for c in calificaciones_num if c >= 6.0])
        reprobadas = materias_cursadas - aprobadas
        promedio_general = sum(calificaciones_num) / len(calificaciones_num) if calificaciones_num else 0.0
        
        # Box de estadísticas con bordes
        stats_start_row = current_row
        
        # Título de estadísticas
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = "RESUMEN ACADÉMICO"
        ws[f'A{current_row}'].font = font_encabezado
        ws[f'A{current_row}'].alignment = center_alignment
        ws[f'A{current_row}'].fill = header_fill
        ws[f'A{current_row}'].border = thin_border
        current_row += 1
        
        # Estadísticas en formato de tabla
        stats_data = [
            ("MATERIAS CURSADAS:", materias_cursadas),
            ("MATERIAS APROBADAS:", aprobadas),  
            ("MATERIAS REPROBADAS:", reprobadas),
            ("PROMEDIO GENERAL:", f"{promedio_general:.2f}")
        ]
        
        for stat_label, stat_value in stats_data:
            ws[f'A{current_row}'] = stat_label
            ws[f'A{current_row}'].font = font_encabezado
            ws[f'A{current_row}'].alignment = left_alignment
            ws[f'A{current_row}'].border = thin_border
            
            ws.merge_cells(f'B{current_row}:I{current_row}')
            ws[f'B{current_row}'] = str(stat_value)
            ws[f'B{current_row}'].font = font_normal
            ws[f'B{current_row}'].alignment = center_alignment
            ws[f'B{current_row}'].border = thin_border
            
            current_row += 1
        
        current_row += 1
        
        # ==================== PIE DE PÁGINA OFICIAL ====================
        
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = f"DOCUMENTO GENERADO AUTOMÁTICAMENTE EL {datetime.now().strftime('%d/%m/%Y A LAS %H:%M:%S')}"
        ws[f'A{current_row}'].font = font_small
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = "UNIVERSIDAD NOVA HISPANA - SISTEMA DE CONTROL ESCOLAR"
        ws[f'A{current_row}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'] = "DOCUMENTO OFICIAL - VÁLIDO PARA TRÁMITES ADMINISTRATIVOS"
        ws[f'A{current_row}'].font = Font(name='Arial', size=9, bold=True, color="FF0000")
        ws[f'A{current_row}'].alignment = center_alignment
        
        # ==================== AJUSTAR DIMENSIONES ====================
        
        # Ajustar anchos de columna para mejor presentación
        column_widths = [12, 8, 40, 12, 10, 15, 12, 25, 20]
        for i, width in enumerate(column_widths, 1):
            column_letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[column_letter].width = width
        
        # Ajustar alturas de fila importantes
        ws.row_dimensions[1].height = 30  # Título principal
        ws.row_dimensions[2].height = 25  # Subtítulo
        ws.row_dimensions[9].height = 25  # Encabezados de tabla
        
        # Ajustar altura de filas de datos
        for row_num in range(row_start, current_row - 5):
            ws.row_dimensions[row_num].height = 20
        
        # ==================== GUARDAR ARCHIVO ====================
        
        # Crear buffer en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M')
        nombre_limpio = str(nombre_db).replace(' ', '_').replace('/', '_')[:25]
        nombre_archivo = f"KARDEX_OFICIAL_{nombre_limpio}_{matricula}_{fecha_actual}.xlsx"
        
        return {
            'buffer': output.getvalue(),
            'filename': nombre_archivo,
            'estadisticas': {
                'materias_cursadas': materias_cursadas,
                'aprobadas': aprobadas,
                'reprobadas': reprobadas,
                'promedio_general': round(promedio_general, 2)
            }
        }
        
    except Exception as e:
        st.error(f"Error al generar kardex oficial: {e}")
        import traceback
        st.error(f"Detalles técnicos: {traceback.format_exc()}")
        return None


def normalizar_cuatrimestre(valor):
    """
    Normaliza valores de cuatrimestre a string, manejando None y tipos mixtos
    """
    if valor is None or valor == '' or pd.isna(valor):
        return "1"  # Valor por defecto
    
    try:
        # Convertir a string y limpiar espacios
        str_valor = str(valor).strip()
        if str_valor == '' or str_valor.lower() == 'none' or str_valor.lower() == 'null':
            return "1"
        
        # Validar que sea un número válido
        int(str_valor)  # Esto lanzará excepción si no es número
        return str_valor
    
    except (ValueError, TypeError):
        return "1"  # Valor por defecto si no se puede convertir

def verificar_dependencias_excel():
    """
    Función para verificar que todas las dependencias estén instaladas
    EJECUTA ESTA FUNCIÓN PRIMERO para diagnosticar problemas
    """
    st.subheader("🔍 Verificación de Dependencias para Excel")
    
    dependencias = {
        'pandas': False,
        'openpyxl': False,
        'io': False,
        'datetime': False
    }
    
    try:
        import pandas
        dependencias['pandas'] = True
        st.success("✅ pandas: Disponible")
    except ImportError:
        st.error("❌ pandas: NO disponible - Ejecuta: pip install pandas")
    
    try:
        import openpyxl
        dependencias['openpyxl'] = True
        st.success("✅ openpyxl: Disponible")
    except ImportError:
        st.error("❌ openpyxl: NO disponible - Ejecuta: pip install openpyxl")
    
    try:
        import io
        dependencias['io'] = True
        st.success("✅ io: Disponible (módulo estándar)")
    except ImportError:
        st.error("❌ io: NO disponible - Error crítico del sistema")
    
    try:
        from datetime import datetime
        dependencias['datetime'] = True
        st.success("✅ datetime: Disponible (módulo estándar)")
    except ImportError:
        st.error("❌ datetime: NO disponible - Error crítico del sistema")
    
    # Resumen
    disponibles = sum(dependencias.values())
    total = len(dependencias)
    
    if disponibles == total:
        st.success(f"🎉 TODAS LAS DEPENDENCIAS DISPONIBLES ({disponibles}/{total})")
        st.info("✅ El botón de exportación debería funcionar correctamente")
    else:
        st.warning(f"⚠️ FALTAN DEPENDENCIAS ({disponibles}/{total})")
        st.info("🔧 Instala las dependencias faltantes y reinicia la aplicación")

# PASO 1: AGREGAR LA FUNCIÓN HELPER AL ARCHIVO
# Agregar esta función en cualquier parte del archivo (recomendado después de las funciones de DB)

def crear_nuevo_registro_con_historial(estudiante_base, materia, fecha_materia, fecha_manual, calificacion, profesor, tipo="regular"):
    """
    Versión corregida que maneja tipos correctamente
    """
    # Normalizar el cuatrimestre del estudiante
    cuatrimestre_actual = normalizar_cuatrimestre(estudiante_base[4])
    
    nuevo_registro = {
        "matricula": estudiante_base[0],
        "nombre": estudiante_base[1],
        "grupo": estudiante_base[2],
        "materia": materia,
        "fecha_ingreso_materia": fecha_materia.strftime("%Y-%m-%d"),
        "fecha_ingreso_original": estudiante_base[8] if estudiante_base[8] else fecha_materia.strftime("%Y-%m-%d"),
        "fecha_recursamiento": None,
        "cuatrimestre": cuatrimestre_actual,  # Ya normalizado como string
        "carrera": estudiante_base[3],
        "email_personal": estudiante_base[5],
        "usuario_email_nova": estudiante_base[6],
        "contraseña": estudiante_base[7],
        "tipo_asignacion": tipo,
        "n_recursamientos": 0,
        "calificacion": calificacion,
        "fecha_calificacion": fecha_manual.strftime("%Y-%m-%d"),
        "profesor": profesor,
        "estatus": "aprobado" if calificacion >= 6 else "reprobado",
        "origen_asignacion": tipo,
        "cuatrimestre_historico": cuatrimestre_actual  # También normalizado
    }
    
    return nuevo_registro

# 1. MODIFICAR LA FUNCIÓN inicializar_base() (línea aproximada 20-60)
# Agregar nueva columna para conservar el cuatrimestre al momento de la materia

def inicializar_base():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='calificaciones'")
    if cursor.fetchone() is None:
        # Crear tabla nueva con todas las columnas incluyendo cuatrimestre_historico
        columnas_requeridas = [
            "matricula TEXT", "nombre TEXT", "grupo TEXT", "materia TEXT",
            "fecha_ingreso_materia TEXT", "fecha_ingreso_original TEXT",
            "fecha_recursamiento TEXT", "cuatrimestre TEXT", "carrera TEXT",
            "email_personal TEXT", "usuario_email_nova TEXT", "contraseña TEXT",
            "tipo_asignacion TEXT", "n_recursamientos INTEGER", "calificacion REAL",
            "fecha_calificacion TEXT", "profesor TEXT", "estatus TEXT", "origen_asignacion TEXT",
            "cuatrimestre_historico TEXT"  # Nueva columna incluida desde el inicio
        ]
        cursor.execute(
            f"CREATE TABLE calificaciones (id INTEGER PRIMARY KEY AUTOINCREMENT, {', '.join(columnas_requeridas)})"
        )
    else:
        # Verificar qué columnas existen actualmente
        cursor.execute("PRAGMA table_info(calificaciones)")
        columnas_actuales = [x[1] for x in cursor.fetchall()]
        
        # Lista de todas las columnas que deberían existir
        columnas_esperadas = [
            ("matricula", "TEXT"), ("nombre", "TEXT"), ("grupo", "TEXT"), ("materia", "TEXT"),
            ("fecha_ingreso_materia", "TEXT"), ("fecha_ingreso_original", "TEXT"),
            ("fecha_recursamiento", "TEXT"), ("cuatrimestre", "TEXT"), ("carrera", "TEXT"),
            ("email_personal", "TEXT"), ("usuario_email_nova", "TEXT"), ("contraseña", "TEXT"),
            ("tipo_asignacion", "TEXT"), ("n_recursamientos", "INTEGER"), ("calificacion", "REAL"),
            ("fecha_calificacion", "TEXT"), ("profesor", "TEXT"), ("estatus", "TEXT"), 
            ("origen_asignacion", "TEXT"), ("cuatrimestre_historico", "TEXT")
        ]
        
        # Agregar solo las columnas que NO existen
        for col_name, col_type in columnas_esperadas:
            if col_name not in columnas_actuales:
                try:
                    cursor.execute(f"ALTER TABLE calificaciones ADD COLUMN {col_name} {col_type}")
                    print(f"✅ Columna agregada: {col_name}")
                except sqlite3.OperationalError as e:
                    if "duplicate column name" in str(e).lower():
                        print(f"⚠️ La columna {col_name} ya existe, saltando...")
                    else:
                        print(f"❌ Error agregando columna {col_name}: {e}")
        
        # MIGRACIÓN DE DATOS: Si cuatrimestre_historico existe pero está vacía, migrar datos
        cursor.execute("SELECT COUNT(*) FROM calificaciones WHERE cuatrimestre_historico IS NOT NULL AND cuatrimestre_historico != ''")
        registros_con_historico = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM calificaciones")
        total_registros = cursor.fetchone()[0]
        
        # Si hay registros sin historial, migrar
        if total_registros > 0 and registros_con_historico == 0:
            print("🔄 Migrando datos existentes a cuatrimestre_historico...")
            cursor.execute("""
                UPDATE calificaciones 
                SET cuatrimestre_historico = COALESCE(cuatrimestre, '1')
                WHERE cuatrimestre_historico IS NULL OR cuatrimestre_historico = ''
            """)
            cursor.execute("SELECT changes()")
            registros_migrados = cursor.fetchone()[0]
            print(f"✅ {registros_migrados} registros migrados al historial")

    conn.commit()
    conn.close()
    print("🏁 Base de datos inicializada correctamente")

def cargar_datos_db():
    conn = sqlite3.connect(DB_FILE)
    df = pd.read_sql_query("SELECT * FROM calificaciones", conn)
    conn.close()
    return df

def limpiar_tipos_cuatrimestre():
    """
    Función para limpiar y normalizar todos los tipos de cuatrimestre en la BD
    """
    st.subheader("🧹 Limpieza de Tipos de Cuatrimestre")
    
    if st.button("🔧 Ejecutar Limpieza de Tipos"):
        try:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            
            with st.spinner("Limpiando tipos de datos..."):
                
                # Obtener todos los registros
                cursor.execute("SELECT id, cuatrimestre, cuatrimestre_historico FROM calificaciones")
                registros = cursor.fetchall()
                
                actualizaciones = 0
                
                for registro_id, cuatr, cuatr_hist in registros:
                    # Normalizar ambos campos
                    cuatr_nuevo = normalizar_cuatrimestre(cuatr)
                    cuatr_hist_nuevo = normalizar_cuatrimestre(cuatr_hist) if cuatr_hist else cuatr_nuevo
                    
                    # Actualizar solo si hay cambios
                    if str(cuatr) != cuatr_nuevo or str(cuatr_hist) != cuatr_hist_nuevo:
                        cursor.execute("""
                            UPDATE calificaciones 
                            SET cuatrimestre = ?, cuatrimestre_historico = ?
                            WHERE id = ?
                        """, (cuatr_nuevo, cuatr_hist_nuevo, registro_id))
                        actualizaciones += 1
                
                conn.commit()
                conn.close()
                
                st.success(f"✅ Limpieza completada: {actualizaciones} registros normalizados")
                st.info("💡 Ahora puedes intentar cargar calificaciones nuevamente")
                
        except Exception as e:
            st.error(f"Error durante la limpieza: {e}")



# ============================================================================
# FUNCIÓN DE EXPORTACIÓN ADAPTADA PARA HISTORIAL DE CUATRIMESTRE
# ============================================================================
def exportar_historial_alumno_excel_desde_cuatrimestre_robusto(matricula, nombre_alumno, conn):
    """
    Función de exportación ROBUSTA que maneja correctamente valores None y tipos mixtos
    """
    try:
        # Obtener información básica del alumno
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DISTINCT nombre, matricula, carrera 
            FROM calificaciones 
            WHERE matricula = ?
            LIMIT 1
        """, (matricula,))
        
        info_alumno = cursor.fetchone()
        if not info_alumno:
            st.error("No se encontró información del alumno en la base de datos")
            return None
            
        nombre_db, matricula_db, carrera = info_alumno
        
        # ✅ CONSULTA MEJORADA CON MANEJO ROBUSTO DE VALORES None
        cursor.execute("""
            SELECT 
                COALESCE(CAST(cuatrimestre_historico AS TEXT), CAST(cuatrimestre AS TEXT), '1') as cuatrimestre_efectivo,
                COALESCE(materia, 'Sin Materia') as materia_segura,
                COALESCE(calificacion, 0.0) as calificacion_segura,
                COALESCE(estatus, 'Sin Estatus') as estatus_seguro,
                COALESCE(fecha_ingreso_materia, 'Sin Fecha') as fecha_segura,
                COALESCE(profesor, 'Sin Profesor') as profesor_seguro,
                COALESCE(CAST(cuatrimestre AS TEXT), '1') as cuatrimestre_actual,
                COALESCE(CAST(cuatrimestre_historico AS TEXT), 'N/A') as cuatrimestre_historico
            FROM calificaciones 
            WHERE matricula = ?
            AND materia IS NOT NULL 
            AND materia != ''
            AND calificacion IS NOT NULL
            ORDER BY 
                CAST(COALESCE(cuatrimestre_historico, cuatrimestre, 1) AS INTEGER) ASC,
                materia ASC,
                fecha_ingreso_materia DESC
        """, (matricula,))
        
        datos_raw = cursor.fetchall()
        
        if not datos_raw:
            st.warning("No hay calificaciones válidas para exportar")
            return None
        
        # ✅ PROCESAR DATOS CON VALIDACIÓN ROBUSTA
        datos_procesados = []
        for row in datos_raw:
            cuatr, materia, calif, estatus, fecha, profesor, cuatr_actual, cuatr_hist = row
            
            # Validar y convertir calificación
            try:
                calificacion_float = float(calif) if calif is not None else 0.0
            except (ValueError, TypeError):
                calificacion_float = 0.0
            
            # Determinar resultado con validación
            resultado = 'Aprobada' if calificacion_float >= 6.0 else 'Reprobada'
            
            datos_procesados.append((
                str(cuatr),
                str(materia),
                calificacion_float,
                str(estatus),
                str(fecha),
                str(profesor),
                str(cuatr_actual),
                str(cuatr_hist),
                resultado
            ))
        
        if not datos_procesados:
            st.warning("No hay datos válidos después del procesamiento")
            return None
        
        # ✅ CREAR BUFFER EXCEL CON MANEJO ROBUSTO
        output = io.BytesIO()
        
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                
                # ===== HOJA 1: INFORMACIÓN DEL ESTUDIANTE =====
                info_estudiante = {
                    'Campo': [
                        'Nombre Completo', 
                        'Matrícula', 
                        'Carrera', 
                        'Fecha de Exportación',
                        'Total de Registros Válidos',
                        'Fuente de Datos'
                    ],
                    'Valor': [
                        str(nombre_db), 
                        str(matricula_db), 
                        str(carrera) if carrera else 'No especificada', 
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        len(datos_procesados),
                        'Sistema de Calificaciones - Historial por Cuatrimestre'
                    ]
                }
                df_info = pd.DataFrame(info_estudiante)
                df_info.to_excel(writer, sheet_name='Información del Estudiante', index=False)
                
                # ===== HOJA 2: HISTORIAL COMPLETO =====
                df_completo = pd.DataFrame(datos_procesados, columns=[
                    'Cuatrimestre', 'Materia', 'Calificación', 'Estatus', 
                    'Fecha Ingreso', 'Profesor', 'Cuatrimestre Actual', 
                    'Cuatrimestre Histórico', 'Resultado'
                ])
                df_completo.to_excel(writer, sheet_name='Historial Completo', index=False)
                
                # ===== HOJAS POR CUATRIMESTRE CON VALIDACIÓN =====
                cuatrimestres_unicos = sorted(set([row[0] for row in datos_procesados if row[0]]), 
                                            key=lambda x: int(x) if x.isdigit() else 999)
                
                estadisticas_cuatrimestre = []
                
                for cuatrimestre in cuatrimestres_unicos:
                    # Filtrar datos del cuatrimestre
                    datos_cuatrimestre = [row for row in datos_procesados if str(row[0]) == str(cuatrimestre)]
                    
                    if datos_cuatrimestre:
                        df_cuatr = pd.DataFrame(datos_cuatrimestre, columns=[
                            'Cuatrimestre', 'Materia', 'Calificación', 'Estatus', 
                            'Fecha Ingreso', 'Profesor', 'Cuatrimestre Actual',
                            'Cuatrimestre Histórico', 'Resultado'
                        ])
                        
                        # Crear hoja para el cuatrimestre
                        sheet_name = f'Cuatrimestre {cuatrimestre}'[:31]
                        df_cuatr.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # ✅ CALCULAR ESTADÍSTICAS CON VALIDACIÓN ROBUSTA
                        calificaciones_validas = [row[2] for row in datos_cuatrimestre if isinstance(row[2], (int, float)) and row[2] is not None]
                        
                        total_materias = len(datos_cuatrimestre)
                        aprobadas = len([calif for calif in calificaciones_validas if calif >= 6.0])
                        reprobadas = total_materias - aprobadas
                        promedio = sum(calificaciones_validas) / len(calificaciones_validas) if calificaciones_validas else 0.0
                        
                        estadisticas_cuatrimestre.append({
                            'Cuatrimestre': cuatrimestre,
                            'Total Materias': total_materias,
                            'Aprobadas': aprobadas,
                            'Reprobadas': reprobadas,
                            'Promedio': round(promedio, 2),
                            'Porcentaje Aprobación': f"{round((aprobadas/total_materias)*100, 1)}%" if total_materias > 0 else "0%"
                        })
                
                # ===== HOJA 3: ESTADÍSTICAS POR CUATRIMESTRE =====
                if estadisticas_cuatrimestre:
                    df_estadisticas = pd.DataFrame(estadisticas_cuatrimestre)
                    df_estadisticas.to_excel(writer, sheet_name='Estadísticas por Cuatrimestre', index=False)
                
                # ===== HOJA 4: ANÁLISIS DE MATERIAS ÚNICAS =====
                materias_unicas = {}
                for row in datos_procesados:
                    materia = row[1]
                    calificacion = row[2]
                    cuatrimestre_reg = row[0]
                    
                    # ✅ VALIDAR CALIFICACIÓN ANTES DE COMPARAR
                    try:
                        calif_float = float(calificacion) if calificacion is not None else 0.0
                    except (ValueError, TypeError):
                        calif_float = 0.0
                    
                    # Solo actualizar si es mejor calificación
                    if materia not in materias_unicas or calif_float > materias_unicas[materia]['calificacion']:
                        materias_unicas[materia] = {
                            'calificacion': calif_float,
                            'cuatrimestre': cuatrimestre_reg,
                            'estatus': row[3],
                            'resultado': 'Aprobada' if calif_float >= 6.0 else 'Reprobada'
                        }
                
                if materias_unicas:
                    analisis_materias = []
                    for materia, info in materias_unicas.items():
                        analisis_materias.append({
                            'Materia': materia,
                            'Mejor Calificación': info['calificacion'],
                            'Cuatrimestre': info['cuatrimestre'],
                            'Estatus Final': info['estatus'],
                            'Resultado': info['resultado']
                        })
                    
                    df_analisis = pd.DataFrame(analisis_materias)
                    df_analisis.to_excel(writer, sheet_name='Análisis de Materias Únicas', index=False)
        
        except Exception as excel_error:
            st.error(f"❌ Error creando el archivo Excel: {excel_error}")
            import traceback
            st.error(f"Detalle del error: {traceback.format_exc()}")
            return None
        
        # Posicionar buffer al inicio
        output.seek(0)
        
        # ✅ CALCULAR ESTADÍSTICAS GENERALES CON VALIDACIÓN
        calificaciones_todas = [row[2] for row in datos_procesados if isinstance(row[2], (int, float)) and row[2] is not None]
        total_materias_unicas = len(materias_unicas)
        aprobadas_unicas = len([m for m in materias_unicas.values() if m['calificacion'] >= 6.0])
        promedio_general = sum(calificaciones_todas) / len(calificaciones_todas) if calificaciones_todas else 0.0
        
        # Nombre del archivo
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M')
        nombre_limpio = str(nombre_db).replace(' ', '_').replace('/', '_')[:30]
        nombre_archivo = f"Historial_{nombre_limpio}_{matricula}_{fecha_actual}.xlsx"
        
        return {
            'buffer': output.getvalue(),
            'filename': nombre_archivo,
            'estadisticas': {
                'total_materias_cursadas': len(datos_procesados),
                'total_materias_unicas': total_materias_unicas,
                'aprobadas_unicas': aprobadas_unicas,
                'promedio_general': round(promedio_general, 2),
                'cuatrimestres_cursados': len(cuatrimestres_unicos)
            }
        }
        
    except Exception as e:
        st.error(f"❌ Error general: {e}")
        import traceback
        st.error(f"Detalles técnicos: {traceback.format_exc()}")
        return None

def insertar_desde_excel(df):
    columnas_comunes = [col for col in df.columns if col in COLUMNAS_DB]
    if not columnas_comunes:
        st.error("❌ El archivo no contiene columnas válidas para insertar en la base de datos.")
        return
    df = df[columnas_comunes]
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        insertados = 0
        duplicados = 0

        for _, row in df.iterrows():
            matricula = row.get("matricula", "").strip()
            grupo = row.get("grupo", "").strip()
            fecha_ingreso_materia = row.get("fecha_ingreso_materia", "").strip()
            materia = row.get("materia", "").strip() if "materia" in row and row.get("materia") else ""

            if materia:
                cursor.execute("""
                    SELECT COUNT(*) FROM calificaciones
                    WHERE matricula = ? AND materia = ? AND grupo = ? AND fecha_ingreso_materia = ?
                """, (matricula, materia, grupo, fecha_ingreso_materia))
            else:
                cursor.execute("""
                    SELECT COUNT(*) FROM calificaciones
                    WHERE matricula = ? AND grupo = ? AND fecha_ingreso_materia = ? AND (materia IS NULL OR materia = '')
                """, (matricula, grupo, fecha_ingreso_materia))

            existe = cursor.fetchone()[0]

            if existe == 0:
                valores = [row.get(col, None) for col in COLUMNAS_DB]
                cursor.execute(f"""
                    INSERT INTO calificaciones ({', '.join(COLUMNAS_DB)})
                    VALUES ({', '.join(['?'] * len(COLUMNAS_DB))})
                """, valores)
                insertados += 1
            else:
                duplicados += 1

        conn.commit()
        conn.close()
        st.success(f"✅ Insertados: {insertados} | ⛔ Duplicados ignorados: {duplicados}")
    except Exception as e:
        st.error(f"❌ Error al insertar en la base de datos: {e}")

def aplicar_filtros(df):
    carrera = st.sidebar.selectbox("Carrera", ["Todos"] + sorted(df["carrera"].dropna().unique().tolist()), key="filtro_sidebar_carrera")
    grupo = st.sidebar.selectbox("Grupo", ["Todos"] + sorted(df["grupo"].dropna().unique().tolist()), key="filtro_sidebar_grupo")
    materia = st.sidebar.selectbox("Materia", ["Todos"] + sorted(df["materia"].dropna().unique().tolist()), key="filtro_sidebar_materia")
    profesor = st.sidebar.selectbox("Profesor", ["Todos"] + sorted(df["profesor"].dropna().unique().tolist()), key="filtro_sidebar_profesor")
    estatus = st.sidebar.selectbox("Estatus", ["Todos"] + sorted(df["estatus"].dropna().unique().tolist()), key="filtro_sidebar_estatus")

    if carrera != "Todos":
        df = df[df["carrera"] == carrera]
    if grupo != "Todos":
        df = df[df["grupo"] == grupo]
    if materia != "Todos":
        df = df[df["materia"] == materia]
    if profesor != "Todos":
        df = df[df["profesor"] == profesor]
    if estatus != "Todos":
        df = df[df["estatus"] == estatus]
    return df

def normalizar_columna(col):
    """Función para normalizar nombres de columnas"""
    col = col.strip().lower()
    col = ''.join(c for c in unicodedata.normalize('NFD', col) if unicodedata.category(c) != 'Mn')
    return col.replace(" ", "_")

def procesar_calificaciones(df_excel, materia, profesor, fecha_materia, fecha_manual):
    """Función mejorada para procesar calificaciones con normalización de columnas"""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        actualizados = 0
        creados = 0
        no_encontrados = 0
        ya_con_calificacion = 0
        errores = []
        
        st.write(f"📊 Procesando {len(df_excel)} registros...")
        
        # NUEVA FUNCIÓN: Normalizar y encontrar columna de calificación
        def encontrar_columna_calificacion(df):
            """Encuentra la columna de calificación independientemente de su nombre exacto"""
            columnas_posibles = [
                "total_del_curso",
                "total del curso", 
                "total_del_curso_(real)",
                "total del curso (real)",
                "total_del_curso_real",
                "total curso",
                "calificacion",
                "calificacion_final",
                "nota_final",
                "resultado"
            ]
            
            # Normalizar nombres de columnas del DataFrame
            columnas_df_normalizadas = {}
            for col in df.columns:
                col_normalizada = col.lower().strip()
                # Remover caracteres especiales y espacios extra
                col_normalizada = col_normalizada.replace("(", "").replace(")", "").replace("-", "_")
                col_normalizada = col_normalizada.replace(" ", "_").replace("__", "_")
                columnas_df_normalizadas[col_normalizada] = col
            
            # Buscar coincidencias
            for posible in columnas_posibles:
                posible_normalizada = posible.lower().replace(" ", "_").replace("(", "").replace(")", "")
                if posible_normalizada in columnas_df_normalizadas:
                    columna_original = columnas_df_normalizadas[posible_normalizada]
                    st.success(f"✅ Columna de calificación encontrada: '{columna_original}'")
                    return columna_original
            
            # Si no encuentra ninguna, mostrar columnas disponibles
            st.error("❌ No se encontró columna de calificación.")
            st.write("🔍 **Columnas disponibles en el archivo:**")
            for col in df.columns:
                st.write(f"• {col}")
            return None
        
        # Encontrar la columna de calificación
        columna_calificacion = encontrar_columna_calificacion(df_excel)
        
        if not columna_calificacion:
            st.error("❌ No se puede procesar el archivo sin una columna de calificación válida.")
            st.info("💡 **Columnas esperadas:** total_del_curso, total del curso, total del curso (real), calificacion, etc.")
            return 0
        
        # NUEVA FUNCIÓN: Limpiar y extraer calificación
        def extraer_calificacion_limpia(valor_raw):
            """Extrae y limpia la calificación de cualquier formato"""
            if pd.isna(valor_raw) or valor_raw == "":
                return 0.0
            
            # Convertir a string y limpiar
            valor_str = str(valor_raw).strip()
            
            # Remover texto común que no es numérico
            texto_a_remover = [
                "(real)", "(Real)", "(REAL)",
                "real", "Real", "REAL",
                "pts", "puntos", "punto",
                "%", "porciento"
            ]
            
            for texto in texto_a_remover:
                valor_str = valor_str.replace(texto, "")
            
            # Limpiar espacios y caracteres especiales
            valor_str = valor_str.strip().replace(" ", "").replace(",", ".")
            
            # Reemplazar guiones por cero (común en archivos de Excel)
            if valor_str in ["-", "--", "---", "N/A", "n/a", "NA", "na"]:
                return 0.0
            
            # Extraer solo números y punto decimal
            import re
            patron_numero = r'(\d+\.?\d*)'
            coincidencias = re.findall(patron_numero, valor_str)
            
            if coincidencias:
                try:
                    calificacion = float(coincidencias[0])
                    # Validar rango lógico (0-10 o 0-100)
                    if calificacion > 100:
                        st.warning(f"⚠️ Calificación fuera de rango: {calificacion} - se mantendrá el valor")
                    return calificacion
                except ValueError:
                    return 0.0
            
            return 0.0
        
        # MOSTRAR VISTA PREVIA DE CALIFICACIONES ENCONTRADAS
        st.markdown("### 👀 Vista previa de calificaciones detectadas")
        
        # Crear muestra de las primeras 5 calificaciones para verificación
        muestra_calificaciones = []
        for i in range(min(5, len(df_excel))):
            valor_original = df_excel.iloc[i][columna_calificacion]
            valor_limpio = extraer_calificacion_limpia(valor_original)
            muestra_calificaciones.append({
                "Fila": i + 1,
                "Valor Original": str(valor_original),
                "Valor Procesado": valor_limpio,
                "Estado": "✅ OK" if valor_limpio > 0 else "⚠️ Cero/Vacío"
            })
        
        df_muestra = pd.DataFrame(muestra_calificaciones)
        st.dataframe(df_muestra, use_container_width=True)
        
        # Estadísticas de la columna
        calificaciones_procesadas = [extraer_calificacion_limpia(val) for val in df_excel[columna_calificacion]]
        calificaciones_validas = [cal for cal in calificaciones_procesadas if cal > 0]
        
        col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
        with col_stat1:
            st.metric("📊 Total registros", len(calificaciones_procesadas))
        with col_stat2:
            st.metric("✅ Con calificación", len(calificaciones_validas))
        with col_stat3:
            st.metric("❌ Sin calificación", len(calificaciones_procesadas) - len(calificaciones_validas))
        with col_stat4:
            if calificaciones_validas:
                promedio = sum(calificaciones_validas) / len(calificaciones_validas)
                st.metric("📈 Promedio", f"{promedio:.2f}")
            else:
                st.metric("📈 Promedio", "N/A")
        
        # Procesar cada registro
        for idx, row in df_excel.iterrows():
            try:
                # Obtener email (sin cambios)
                email = ""
                if "direccion_email" in row and pd.notna(row["direccion_email"]):
                    email = str(row["direccion_email"]).strip().lower()
                elif "usuario_email_nova" in row and pd.notna(row["usuario_email_nova"]):
                    email = str(row["usuario_email_nova"]).strip().lower()
                
                # Obtener carrera/institución (sin cambios)
                carrera = ""
                if "institucion" in row and pd.notna(row["institucion"]):
                    carrera = str(row["institucion"]).strip().upper()
                elif "carrera" in row and pd.notna(row["carrera"]):
                    carrera = str(row["carrera"]).strip().upper()
                
                # NUEVA LÓGICA: Obtener calificación usando la función mejorada
                calif_raw = row.get(columna_calificacion, "")
                calificacion = extraer_calificacion_limpia(calif_raw)
                
                # Validar datos obligatorios (sin cambios)
                if not email:
                    errores.append(f"Fila {idx + 1}: Email faltante")
                    continue
                
                if not carrera:
                    errores.append(f"Fila {idx + 1}: Carrera/Institución faltante")
                    continue
                
                # Buscar estudiante base (sin cambios)
                cursor.execute("""
                    SELECT matricula, nombre, grupo, carrera, cuatrimestre, email_personal, usuario_email_nova, 
                           contraseña, fecha_ingreso_original
                    FROM calificaciones
                    WHERE LOWER(TRIM(usuario_email_nova)) = ? 
                    AND UPPER(TRIM(carrera)) = ?
                    AND (materia IS NULL OR materia = '')
                    LIMIT 1
                """, (email, carrera))
                
                estudiante_base = cursor.fetchone()
                
                if not estudiante_base:
                    no_encontrados += 1
                    errores.append(f"Fila {idx + 1}: No se encontró estudiante base con email '{email}' y carrera '{carrera}'")
                    continue
                
                # Verificar si ya existe un registro (sin cambios)
                cursor.execute("""
                    SELECT id FROM calificaciones
                    WHERE matricula = ? 
                    AND materia = ? 
                    AND fecha_ingreso_materia = ?
                    AND calificacion IS NOT NULL
                """, (estudiante_base[0], materia, fecha_materia.strftime("%Y-%m-%d")))
                
                registro_existente = cursor.fetchone()
                
                if registro_existente:
                    ya_con_calificacion += 1
                    errores.append(f"Fila {idx + 1}: Ya existe calificación para {estudiante_base[0]} en {materia} con fecha {fecha_materia.strftime('%Y-%m-%d')}")
                    continue
                
                # Buscar si existe un registro sin calificación (sin cambios)
                cursor.execute("""
                    SELECT id FROM calificaciones
                    WHERE matricula = ? 
                    AND materia = ? 
                    AND fecha_ingreso_materia = ?
                    AND calificacion IS NULL
                """, (estudiante_base[0], materia, fecha_materia.strftime("%Y-%m-%d")))
                
                registro_sin_calificacion = cursor.fetchone()
                
                if registro_sin_calificacion:
                    # Actualizar registro existente sin calificación
                    cursor.execute("""
                        UPDATE calificaciones
                        SET calificacion = ?, 
                            fecha_calificacion = ?, 
                            estatus = ?, 
                            profesor = ?
                        WHERE id = ?
                    """, (
                        calificacion,
                        fecha_manual.strftime("%Y-%m-%d"),
                        "aprobado" if calificacion >= 6 else "reprobado",
                        profesor,
                        registro_sin_calificacion[0]
                    ))
                    actualizados += 1
                else:
                    # Crear nuevo registro para esta materia
                    nuevo_registro = crear_nuevo_registro_con_historial(estudiante_base, materia, fecha_materia, fecha_manual, calificacion, profesor, "regular") 
                    {
                        "matricula": estudiante_base[0],
                        "nombre": estudiante_base[1],
                        "grupo": estudiante_base[2],
                        "materia": materia,
                        "fecha_ingreso_materia": fecha_materia.strftime("%Y-%m-%d"),
                        "fecha_ingreso_original": estudiante_base[8] if estudiante_base[8] else fecha_materia.strftime("%Y-%m-%d"),
                        "fecha_recursamiento": None,
                        "cuatrimestre": estudiante_base[4],
                        "carrera": estudiante_base[3],
                        "email_personal": estudiante_base[5],
                        "usuario_email_nova": estudiante_base[6],
                        "contraseña": estudiante_base[7],
                        "tipo_asignacion": "regular",
                        "n_recursamientos": 0,
                        "calificacion": calificacion,
                        "fecha_calificacion": fecha_manual.strftime("%Y-%m-%d"),
                        "profesor": profesor,
                        "estatus": "aprobado" if calificacion >= 6 else "reprobado",
                        "origen_asignacion": "regular",
                        "cuatrimestre_historico": estudiante_base[4]  # ← AGREGAR ESTA LÍNEA
                    }
                    
                    valores = [nuevo_registro.get(col, None) for col in COLUMNAS_DB]
                    cursor.execute(f"""
                        INSERT INTO calificaciones ({', '.join(COLUMNAS_DB)})
                        VALUES ({', '.join(['?'] * len(COLUMNAS_DB))})
                    """, valores)
                    creados += 1
                    
                    # Actualizar cuatrimestre si se aprueban múltiplos de 5 materias
                    cursor.execute("""
                        SELECT COUNT(*) FROM calificaciones
                        WHERE matricula = ? AND calificacion >= 6
                    """, (estudiante_base[0],))
                    materias_aprobadas = cursor.fetchone()[0]
                    nuevo_cuatrimestre = str(1 + (materias_aprobadas // 5))

                    # Obtener cuatrimestre actual
                    cursor.execute("""
                        SELECT cuatrimestre FROM calificaciones
                        WHERE matricula = ? ORDER BY id DESC LIMIT 1
                    """, (estudiante_base[0],))
                    actual = cursor.fetchone()
                    cuatrimestre_actual = actual[0] if actual and actual[0] else "1"

                    if nuevo_cuatrimestre != cuatrimestre_actual:
                        cursor.execute("""
                            UPDATE calificaciones
                            SET cuatrimestre = ?
                            WHERE matricula = ?
                        """, (nuevo_cuatrimestre, estudiante_base[0]))
                    
            except Exception as e:
                errores.append(f"Fila {idx + 1}: Error procesando registro - {str(e)}")
        
        conn.commit()
        conn.close()
        
        # Mostrar resultados
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("✅ Actualizados", actualizados)
        with col2:
            st.metric("🆕 Creados", creados)
        with col3:
            st.metric("❌ No encontrados", no_encontrados)
        with col4:
            st.metric("⚠️ Ya con calificación", ya_con_calificacion)
        
        total_procesados = actualizados + creados
        if total_procesados > 0:
            st.success(f"✅ {total_procesados} registros procesados exitosamente ({actualizados} actualizados, {creados} creados)")
            
            # NUEVO: Mostrar estadísticas de calificaciones procesadas
            if calificaciones_validas:
                aprobados_nuevos = len([cal for cal in calificaciones_validas if cal >= 6])
                reprobados_nuevos = len(calificaciones_validas) - aprobados_nuevos
                tasa_aprobacion = (aprobados_nuevos / len(calificaciones_validas)) * 100
                
                st.markdown("### 📊 Estadísticas de Calificaciones Procesadas")
                col_stats1, col_stats2, col_stats3 = st.columns(3)
                with col_stats1:
                    st.metric("✅ Aprobados", f"{aprobados_nuevos} ({tasa_aprobacion:.1f}%)")
                with col_stats2:
                    st.metric("❌ Reprobados", reprobados_nuevos)
                with col_stats3:
                    st.metric("📈 Promedio General", f"{sum(calificaciones_validas) / len(calificaciones_validas):.2f}")
        
        if errores:
            st.warning(f"⚠️ Se encontraron {len(errores)} problemas:")
            with st.expander("Ver detalles de errores"):
                for error in errores[:15]:
                    st.write(f"• {error}")
                if len(errores) > 15:
                    st.write(f"... y {len(errores) - 15} errores más")
        
        return total_procesados
        
    except Exception as e:
        st.error(f"❌ Error general al procesar calificaciones: {e}")
        return 0
        
def procesar_calificaciones_con_grupo(df_excel, materia, profesor, fecha_materia, fecha_manual, nuevo_grupo):
    """Función CORREGIDA para procesar calificaciones CON asignación de grupo - PRESERVA HISTÓRICO"""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        actualizados = 0
        creados = 0
        no_encontrados = 0
        ya_con_calificacion = 0
        registros_con_nuevo_grupo = 0  # Cambiado el nombre
        errores = []
        
        st.write(f"📊 Procesando {len(df_excel)} registros con grupo: **{nuevo_grupo}**")
        
        # Reutilizar funciones de búsqueda de calificación del código original
        def encontrar_columna_calificacion(df):
            """Encuentra la columna de calificación independientemente de su nombre exacto"""
            columnas_posibles = [
                "total_del_curso",
                "total del curso", 
                "total_del_curso_(real)",
                "total del curso (real)",
                "total_del_curso_real",
                "total curso",
                "calificacion",
                "calificacion_final",
                "nota_final",
                "resultado"
            ]
            
            # Normalizar nombres de columnas del DataFrame
            columnas_df_normalizadas = {}
            for col in df.columns:
                col_normalizada = col.lower().strip()
                col_normalizada = col_normalizada.replace("(", "").replace(")", "").replace("-", "_")
                col_normalizada = col_normalizada.replace(" ", "_").replace("__", "_")
                columnas_df_normalizadas[col_normalizada] = col
            
            # Buscar coincidencias
            for posible in columnas_posibles:
                posible_normalizada = posible.lower().replace(" ", "_").replace("(", "").replace(")", "")
                if posible_normalizada in columnas_df_normalizadas:
                    columna_original = columnas_df_normalizadas[posible_normalizada]
                    st.success(f"✅ Columna de calificación encontrada: '{columna_original}'")
                    return columna_original
            
            st.error("❌ No se encontró columna de calificación.")
            return None
        
        def extraer_calificacion_limpia(valor_raw):
            """Extrae y limpia la calificación de cualquier formato"""
            if pd.isna(valor_raw) or valor_raw == "":
                return 0.0
            
            valor_str = str(valor_raw).strip()
            
            texto_a_remover = [
                "(real)", "(Real)", "(REAL)",
                "real", "Real", "REAL",
                "pts", "puntos", "punto",
                "%", "porciento"
            ]
            
            for texto in texto_a_remover:
                valor_str = valor_str.replace(texto, "")
            
            valor_str = valor_str.strip().replace(" ", "").replace(",", ".")
            
            if valor_str in ["-", "--", "---", "N/A", "n/a", "NA", "na"]:
                return 0.0
            
            import re
            patron_numero = r'(\d+\.?\d*)'
            coincidencias = re.findall(patron_numero, valor_str)
            
            if coincidencias:
                try:
                    calificacion = float(coincidencias[0])
                    if calificacion > 100:
                        st.warning(f"⚠️ Calificación fuera de rango: {calificacion} - se mantendrá el valor")
                    return calificacion
                except ValueError:
                    return 0.0
            
            return 0.0
        
        # Encontrar la columna de calificación
        columna_calificacion = encontrar_columna_calificacion(df_excel)
        
        if not columna_calificacion:
            st.error("❌ No se puede procesar el archivo sin una columna de calificación válida.")
            return 0
        
        # Mostrar vista previa de calificaciones
        st.markdown("### 👀 Vista previa de calificaciones detectadas")
        
        muestra_calificaciones = []
        for i in range(min(5, len(df_excel))):
            valor_original = df_excel.iloc[i][columna_calificacion]
            valor_limpio = extraer_calificacion_limpia(valor_original)
            muestra_calificaciones.append({
                "Fila": i + 1,
                "Valor Original": str(valor_original),
                "Valor Procesado": valor_limpio,
                "Estado": "✅ OK" if valor_limpio > 0 else "⚠️ Cero/Vacío"
            })
        
        df_muestra = pd.DataFrame(muestra_calificaciones)
        st.dataframe(df_muestra, use_container_width=True)
        
        # Procesar cada registro
        for idx, row in df_excel.iterrows():
            try:
                # Obtener email
                email = ""
                if "direccion_email" in row and pd.notna(row["direccion_email"]):
                    email = str(row["direccion_email"]).strip().lower()
                elif "usuario_email_nova" in row and pd.notna(row["usuario_email_nova"]):
                    email = str(row["usuario_email_nova"]).strip().lower()
                
                # Obtener carrera/institución
                carrera = ""
                if "institucion" in row and pd.notna(row["institucion"]):
                    carrera = str(row["institucion"]).strip().upper()
                elif "carrera" in row and pd.notna(row["carrera"]):
                    carrera = str(row["carrera"]).strip().upper()
                
                # Obtener calificación
                calif_raw = row.get(columna_calificacion, "")
                calificacion = extraer_calificacion_limpia(calif_raw)
                
                # Validar datos obligatorios
                if not email:
                    errores.append(f"Fila {idx + 1}: Email faltante")
                    continue
                
                if not carrera:
                    errores.append(f"Fila {idx + 1}: Carrera/Institución faltante")
                    continue
                
                # Buscar estudiante base
                cursor.execute("""
                    SELECT matricula, nombre, grupo, carrera, cuatrimestre, email_personal, usuario_email_nova, 
                           contraseña, fecha_ingreso_original
                    FROM calificaciones
                    WHERE LOWER(TRIM(usuario_email_nova)) = ? 
                    AND UPPER(TRIM(carrera)) = ?
                    AND (materia IS NULL OR materia = '')
                    LIMIT 1
                """, (email, carrera))
                
                estudiante_base = cursor.fetchone()
                
                if not estudiante_base:
                    no_encontrados += 1
                    errores.append(f"Fila {idx + 1}: No se encontró estudiante base con email '{email}' y carrera '{carrera}'")
                    continue
                
                matricula_estudiante = estudiante_base[0]
                
                # Verificar si el alumno ya tiene la materia
                cursor.execute("""
                    SELECT id, calificacion FROM calificaciones
                    WHERE matricula = ? 
                    AND materia = ? 
                    AND fecha_ingreso_materia = ?
                """, (matricula_estudiante, materia, fecha_materia.strftime("%Y-%m-%d")))
                
                registro_existente = cursor.fetchone()
                
                if registro_existente:
                    # CASO 1: Ya existe registro para esta materia/fecha
                    if registro_existente[1] is not None:  # Ya tiene calificación
                        ya_con_calificacion += 1
                        errores.append(f"Fila {idx + 1}: {matricula_estudiante} ya tiene calificación en {materia}")
                        continue
                    else:
                        # CASO 2: Existe pero sin calificación - actualizar SOLO EL REGISTRO ESPECÍFICO
                        cursor.execute("""
                            UPDATE calificaciones
                            SET calificacion = ?, 
                                fecha_calificacion = ?, 
                                estatus = ?, 
                                profesor = ?,
                                grupo = ?
                            WHERE id = ?
                        """, (
                            calificacion,
                            fecha_manual.strftime("%Y-%m-%d"),
                            "aprobado" if calificacion >= 6 else "reprobado",
                            profesor,
                            nuevo_grupo,
                            registro_existente[0]
                        ))
                        actualizados += 1
                        registros_con_nuevo_grupo += 1
                        
                        # ✅ ELIMINADO: La actualización masiva de todos los registros del alumno
                        # ❌ CÓDIGO PROBLEMÁTICO ELIMINADO:
                        # if grupo_anterior != nuevo_grupo:
                        #     cursor.execute("""
                        #         UPDATE calificaciones
                        #         SET grupo = ?
                        #         WHERE matricula = ?
                        #     """, (nuevo_grupo, matricula_estudiante))
                        #     grupos_actualizados += 1
                        
                else:
                    # CASO 3: No existe registro - crear nuevo CON EL NUEVO GRUPO
                    nuevo_registro = crear_nuevo_registro_con_historial(estudiante_base, materia, fecha_materia, fecha_manual, calificacion, profesor, "regular") 
                    {
                        "matricula": estudiante_base[0],
                        "nombre": estudiante_base[1],
                        "grupo": nuevo_grupo,  # USAR EL NUEVO GRUPO SOLO PARA ESTE REGISTRO
                        "materia": materia,
                        "fecha_ingreso_materia": fecha_materia.strftime("%Y-%m-%d"),
                        "fecha_ingreso_original": estudiante_base[8] if estudiante_base[8] else fecha_materia.strftime("%Y-%m-%d"),
                        "fecha_recursamiento": None,
                        "cuatrimestre": estudiante_base[4],
                        "carrera": estudiante_base[3],
                        "email_personal": estudiante_base[5],
                        "usuario_email_nova": estudiante_base[6],
                        "contraseña": estudiante_base[7],
                        "tipo_asignacion": "regular",
                        "n_recursamientos": 0,
                        "calificacion": calificacion,
                        "fecha_calificacion": fecha_manual.strftime("%Y-%m-%d"),
                        "profesor": profesor,
                        "estatus": "aprobado" if calificacion >= 6 else "reprobado",
                        "origen_asignacion": "regular",
                        "cuatrimestre_historico": estudiante_base[4]  # ← AGREGAR ESTA LÍNEA
                    }
                    
                    valores = [nuevo_registro.get(col, None) for col in COLUMNAS_DB]
                    cursor.execute(f"""
                        INSERT INTO calificaciones ({', '.join(COLUMNAS_DB)})
                        VALUES ({', '.join(['?'] * len(COLUMNAS_DB))})
                    """, valores)
                    creados += 1
                    registros_con_nuevo_grupo += 1
                    
                    # ✅ ELIMINADO: La actualización masiva de otros registros del alumno
                    # ❌ CÓDIGO PROBLEMÁTICO ELIMINADO:
                    # if grupo_anterior != nuevo_grupo:
                    #     cursor.execute("""
                    #         UPDATE calificaciones
                    #         SET grupo = ?
                    #         WHERE matricula = ?
                    #     """, (nuevo_grupo, matricula_estudiante))
                    #     grupos_actualizados += 1
                    
            except Exception as e:
                errores.append(f"Fila {idx + 1}: Error procesando registro - {str(e)}")
        
        conn.commit()
        conn.close()
        
        # Mostrar resultados CORREGIDOS
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            st.metric("✅ Actualizados", actualizados)
        with col2:
            st.metric("🆕 Creados", creados)
        with col3:
            st.metric("❌ No encontrados", no_encontrados)
        with col4:
            st.metric("⚠️ Ya con calificación", ya_con_calificacion)
        with col5:
            st.metric("👥 Con nuevo grupo", registros_con_nuevo_grupo)
        
        total_procesados = actualizados + creados
        if total_procesados > 0:
            st.success(f"✅ {total_procesados} registros procesados exitosamente")
            
            # ✅ MENSAJE CORREGIDO sobre preservación del histórico
            if registros_con_nuevo_grupo > 0:
                st.success(f"🔒 **HISTÓRICO PRESERVADO:** {registros_con_nuevo_grupo} registros ahora tienen grupo **{nuevo_grupo}** SOLO para esta materia")
                st.info("💡 **Importante:** Se preservó el grupo original en registros de materias anteriores")
        
        if errores:
            st.warning(f"⚠️ Se encontraron {len(errores)} problemas:")
            with st.expander("Ver detalles de errores"):
                for error in errores[:15]:
                    st.write(f"• {error}")
                if len(errores) > 15:
                    st.write(f"... y {len(errores) - 15} errores más")
        
        return total_procesados
        
    except Exception as e:
        st.error(f"❌ Error general al procesar calificaciones: {e}")
        return 0  
# 🔹 FUNCIÓN 1: Calcular cuatrimestre por materias aprobadas

 
# FUNCIÓN PRINCIPAL: Calcular y actualizar cuatrimestre automáticamente
def calcular_cuatrimestre_por_materias_aprobadas(matricula, conn):
    """
    Calcula el cuatrimestre basado en materias aprobadas (≥6.0)
    Cada 5 materias aprobadas = +1 cuatrimestre
    Cuatrimestre inicial = 1
    """
    cursor = conn.cursor()
    
    # Contar materias aprobadas ÚNICAS por alumno (evitar duplicados por recursamientos)
    cursor.execute("""
        SELECT COUNT(DISTINCT materia) as materias_aprobadas
        FROM calificaciones 
        WHERE matricula = ? 
        AND materia IS NOT NULL 
        AND materia != '' 
        AND calificacion >= 6.0
    """, (matricula,))
    
    resultado = cursor.fetchone()
    materias_aprobadas = resultado[0] if resultado else 0
    
    # Calcular cuatrimestre: base 1 + cada 5 materias aprobadas
    cuatrimestre_calculado = 1 + (materias_aprobadas // 5)
    
    # ✅ RETORNAR COMO STRING para evitar errores de tipo
    return str(cuatrimestre_calculado), materias_aprobadas
    
    
def actualizar_cuatrimestre_todos_alumnos():
    """Actualiza el cuatrimestre de TODOS los alumnos basado en materias aprobadas"""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        # Obtener todas las matrículas únicas
        cursor.execute("SELECT DISTINCT matricula FROM calificaciones WHERE matricula IS NOT NULL")
        matriculas = [row[0] for row in cursor.fetchall()]
        
        actualizaciones = 0
        log_cambios = []
        
        for matricula in matriculas:
            # Calcular nuevo cuatrimestre
            cuatrimestre_nuevo, materias_aprobadas = calcular_cuatrimestre_por_materias_aprobadas(matricula, conn)
            
            # Obtener cuatrimestre actual
            cursor.execute("""
                SELECT cuatrimestre, nombre FROM calificaciones 
                WHERE matricula = ? LIMIT 1
            """, (matricula,))
            resultado = cursor.fetchone()
            
            if resultado:
                cuatrimestre_actual = resultado[0] or 1
                nombre_alumno = resultado[1]
                
                # Solo actualizar si hay cambio
                if cuatrimestre_nuevo != cuatrimestre_actual:
                    cursor.execute("""
                        UPDATE calificaciones 
                        SET cuatrimestre = ? 
                        WHERE matricula = ?
                    """, (cuatrimestre_nuevo, matricula))
                    
                    actualizaciones += 1
                    log_cambios.append({
                        "matricula": matricula,
                        "nombre": nombre_alumno,
                        "cuatrimestre_anterior": cuatrimestre_actual,
                        "cuatrimestre_nuevo": cuatrimestre_nuevo,
                        "materias_aprobadas": materias_aprobadas
                    })
        
        conn.commit()
        conn.close()
        
        return actualizaciones, log_cambios
        
    except Exception as e:
        st.error(f"Error al actualizar cuatrimestres: {e}")
        return 0, []

# ============================================================================
# 3. NUEVA FUNCIÓN PARA ACTUALIZAR CUATRIMESTRE CON HISTORIAL
# Agregar después de la línea 800 (después de actualizar_cuatrimestre_todos_alumnos)
# ============================================================================
def actualizar_cuatrimestre_con_historial(matricula, nuevo_cuatrimestre, conn):
    """
    FUNCIÓN CORREGIDA: Actualiza cuatrimestre manejando tipos correctamente
    """
    cursor = conn.cursor()
    
    # Normalizar el nuevo cuatrimestre
    nuevo_cuatrimestre_norm = normalizar_cuatrimestre(nuevo_cuatrimestre)
    
    # Obtener cuatrimestre actual
    cursor.execute("""
        SELECT cuatrimestre FROM calificaciones 
        WHERE matricula = ? LIMIT 1
    """, (matricula,))
    resultado = cursor.fetchone()
    
    if resultado and resultado[0] is not None:
        cuatrimestre_actual = normalizar_cuatrimestre(resultado[0])
    else:
        cuatrimestre_actual = "1"
    
    # Comparar strings normalizados (NO int vs str)
    if nuevo_cuatrimestre_norm != cuatrimestre_actual:
        
        # PASO 1: Establecer historial para registros sin él
        cursor.execute("""
            UPDATE calificaciones 
            SET cuatrimestre_historico = cuatrimestre
            WHERE matricula = ? 
            AND (cuatrimestre_historico IS NULL OR cuatrimestre_historico = '' OR cuatrimestre_historico = 'None')
        """, (matricula,))
        
        # PASO 2: Actualizar cuatrimestre actual
        cursor.execute("""
            UPDATE calificaciones 
            SET cuatrimestre = ? 
            WHERE matricula = ?
        """, (nuevo_cuatrimestre_norm, matricula))
        
        return True
    return False  
    
     
def actualizar_cuatrimestre_todos_alumnos_con_historial():
    """
    FUNCIÓN CORREGIDA: Actualiza todos los cuatrimestres manejando tipos correctamente
    """
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        # PASO 1: Normalizar todos los cuatrimestres existentes
        st.info("🔧 Normalizando cuatrimestres existentes...")
        
        cursor.execute("SELECT id, cuatrimestre, cuatrimestre_historico FROM calificaciones")
        todos_registros = cursor.fetchall()
        
        for registro_id, cuatr, cuatr_hist in todos_registros:
            cuatr_normalizado = normalizar_cuatrimestre(cuatr)
            cuatr_hist_normalizado = normalizar_cuatrimestre(cuatr_hist) if cuatr_hist else cuatr_normalizado
            
            cursor.execute("""
                UPDATE calificaciones 
                SET cuatrimestre = ?, cuatrimestre_historico = ?
                WHERE id = ?
            """, (cuatr_normalizado, cuatr_hist_normalizado, registro_id))
        
        st.info("✅ Normalización completada")
        
        # PASO 2: Establecer cuatrimestre_historico para registros que no lo tienen
        cursor.execute("""
            UPDATE calificaciones 
            SET cuatrimestre_historico = cuatrimestre
            WHERE cuatrimestre_historico IS NULL 
            OR cuatrimestre_historico = '' 
            OR cuatrimestre_historico = 'None'
        """)
        
        # PASO 3: Obtener todas las matrículas
        cursor.execute("SELECT DISTINCT matricula FROM calificaciones WHERE matricula IS NOT NULL")
        matriculas = [row[0] for row in cursor.fetchall()]
        
        actualizaciones = 0
        log_cambios = []
        
        for matricula in matriculas:
            # Obtener datos actuales del alumno
            cursor.execute("""
                SELECT nombre, cuatrimestre FROM calificaciones 
                WHERE matricula = ? LIMIT 1
            """, (matricula,))
            resultado = cursor.fetchone()
            
            if not resultado:
                continue
                
            nombre_alumno = resultado[0]
            cuatrimestre_actual = normalizar_cuatrimestre(resultado[1])
            
            # Calcular nuevo cuatrimestre usando la función CORREGIDA
            cuatrimestre_nuevo, materias_aprobadas = calcular_cuatrimestre_por_materias_aprobadas(matricula, conn)
            
            # Comparar strings normalizados (NO int vs str)
            if cuatrimestre_nuevo != cuatrimestre_actual:
                
                # Actualizar el cuatrimestre actual
                cursor.execute("""
                    UPDATE calificaciones 
                    SET cuatrimestre = ? 
                    WHERE matricula = ?
                """, (cuatrimestre_nuevo, matricula))
                
                actualizaciones += 1
                log_cambios.append({
                    "matricula": matricula,
                    "nombre": nombre_alumno,
                    "cuatrimestre_anterior": cuatrimestre_actual,
                    "cuatrimestre_nuevo": cuatrimestre_nuevo,
                    "materias_aprobadas": materias_aprobadas
                })
        
        conn.commit()
        conn.close()
        
        return actualizaciones, log_cambios
        
    except Exception as e:
        st.error(f"Error al actualizar cuatrimestres: {e}")
        import traceback
        st.error(f"Detalles del error: {traceback.format_exc()}")
        return 0, []
        
# MODIFICACIÓN A LA FUNCIÓN DE PROCESAMIENTO DE CALIFICACIONES
def procesar_calificaciones_con_promocion_automatica(df_excel, materia, profesor, fecha_materia, fecha_manual, columna_grupo):
    """
    Versión modificada que actualiza automáticamente el cuatrimestre 
    después de procesar las calificaciones
    """
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        actualizados = 0
        creados = 0
        no_encontrados = 0
        ya_con_calificacion = 0
        registros_con_grupo = 0
        promociones_cuatrimestre = 0
        creados_base = 0  # Nueva variable para contar registros base creados
        errores = []
        matriculas_procesadas = set()  # Para tracking de promociones
        
        st.write(f"📊 Procesando {len(df_excel)} registros con promoción automática de cuatrimestre")
        
        # [AQUÍ VA TODA LA LÓGICA ORIGINAL DE PROCESAMIENTO...]
        # (Funciones auxiliares y procesamiento de calificaciones igual que antes)
        
        def encontrar_columna_calificacion(df):
            columnas_posibles = [
                "total_del_curso", "total del curso", "total_del_curso_(real)",
                "total del curso (real)", "total_del_curso_real", "total curso",
                "calificacion", "calificacion_final", "nota_final", "resultado"
            ]
            
            columnas_df_normalizadas = {}
            for col in df.columns:
                col_normalizada = col.lower().strip().replace("(", "").replace(")", "").replace("-", "_")
                col_normalizada = col_normalizada.replace(" ", "_").replace("__", "_")
                columnas_df_normalizadas[col_normalizada] = col
            
            for posible in columnas_posibles:
                posible_normalizada = posible.lower().replace(" ", "_").replace("(", "").replace(")", "")
                if posible_normalizada in columnas_df_normalizadas:
                    columna_original = columnas_df_normalizadas[posible_normalizada]
                    st.success(f"✅ Columna de calificación encontrada: '{columna_original}'")
                    return columna_original
            
            st.error("❌ No se encontró columna de calificación.")
            return None
        
        def extraer_calificacion_limpia(valor_raw):
            if pd.isna(valor_raw) or valor_raw == "":
                return 0.0
            
            valor_str = str(valor_raw).strip()
            texto_a_remover = ["(real)", "(Real)", "(REAL)", "real", "Real", "REAL", "pts", "puntos", "punto", "%", "porciento"]
            
            for texto in texto_a_remover:
                valor_str = valor_str.replace(texto, "")
            
            valor_str = valor_str.strip().replace(" ", "").replace(",", ".")
            
            if valor_str in ["-", "--", "---", "N/A", "n/a", "NA", "na"]:
                return 0.0
            
            import re
            patron_numero = r'(\d+\.?\d*)'
            coincidencias = re.findall(patron_numero, valor_str)
            
            if coincidencias:
                try:
                    calificacion = float(coincidencias[0])
                    return calificacion
                except ValueError:
                    return 0.0
            return 0.0
        
        # Encontrar columna de calificación
        columna_calificacion = encontrar_columna_calificacion(df_excel)
        if not columna_calificacion:
            st.error("❌ No se puede procesar sin columna de calificación válida.")
            return 0
        
        # Procesar cada registro (lógica original)
        for idx, row in df_excel.iterrows():
            try:
                # Obtener email, carrera, grupo, calificación (igual que antes)
                email = ""
                if "direccion_email" in row and pd.notna(row["direccion_email"]):
                    email = str(row["direccion_email"]).strip().lower()
                elif "usuario_email_nova" in row and pd.notna(row["usuario_email_nova"]):
                    email = str(row["usuario_email_nova"]).strip().lower()
                
                carrera = ""
                if "institucion" in row and pd.notna(row["institucion"]):
                    carrera = str(row["institucion"]).strip().upper()
                elif "carrera" in row and pd.notna(row["carrera"]):
                    carrera = str(row["carrera"]).strip().upper()
                
                grupo_alumno = str(row[columna_grupo]).strip() if pd.notna(row[columna_grupo]) else ""
                calif_raw = row.get(columna_calificacion, "")
                calificacion = extraer_calificacion_limpia(calif_raw)
                
                # Validaciones básicas
                if not email or not carrera or not grupo_alumno:
                    errores.append(f"Fila {idx + 1}: Datos obligatorios faltantes")
                    continue
                
                # Buscar estudiante base
                cursor.execute("""
                    SELECT matricula, nombre, grupo, carrera, cuatrimestre, email_personal, usuario_email_nova, 
                           contraseña, fecha_ingreso_original
                    FROM calificaciones
                    WHERE LOWER(TRIM(usuario_email_nova)) = ? 
                    AND UPPER(TRIM(carrera)) = ?
                    AND (materia IS NULL OR materia = '')
                    LIMIT 1
                """, (email, carrera))
                
                estudiante_base = cursor.fetchone()
                
                # Si no existe el registro base, crearlo automáticamente
                if not estudiante_base:
                    # Extraer datos del alumno del Excel
                    matricula_nueva = ""
                    if "usuario" in row and pd.notna(row["usuario"]):
                        matricula_nueva = str(row["usuario"]).strip()
                    elif "matricula" in row and pd.notna(row["matricula"]):
                        matricula_nueva = str(row["matricula"]).strip()
                    
                    nombre = ""
                    if "nombre" in row and pd.notna(row["nombre"]):
                        nombre = str(row["nombre"]).strip()
                    
                    apellidos = ""
                    if "apellido(s)" in row and pd.notna(row["apellido(s)"]):
                        apellidos = str(row["apellido(s)"]).strip()
                    elif "apellidos" in row and pd.notna(row["apellidos"]):
                        apellidos = str(row["apellidos"]).strip()
                    
                    nombre_completo = f"{nombre} {apellidos}".strip()
                    
                    email_personal = ""
                    if "correo_personal" in row and pd.notna(row["correo_personal"]):
                        email_personal = str(row["correo_personal"]).strip()
                    
                    if not matricula_nueva or not nombre_completo:
                        errores.append(f"Fila {idx + 1}: No se pudo crear registro base - faltan matrícula o nombre")
                        no_encontrados += 1
                        continue
                    
                    # Crear el registro base del alumno (sin materia)
                    try:
                        cursor.execute(f"""
                            INSERT INTO calificaciones ({', '.join(COLUMNAS_DB)})
                            VALUES ({', '.join(['?'] * len(COLUMNAS_DB))})
                        """, (
                            matricula_nueva,           # matricula
                            nombre_completo,           # nombre
                            "",                        # grupo (vacío en registro base)
                            "",                        # materia (vacío en registro base)
                            "",                        # fecha_ingreso_materia (vacío)
                            fecha_materia.strftime("%Y-%m-%d"),  # fecha_ingreso_original
                            None,                      # fecha_recursamiento
                            1,                         # cuatrimestre (inicia en 1)
                            carrera,                   # carrera
                            email_personal,            # email_personal
                            email,                     # usuario_email_nova
                            "",                        # contraseña
                            "regular",                 # tipo_asignacion
                            0,                         # n_recursamientos
                            None,                      # calificacion (vacío)
                            None,                      # fecha_calificacion (vacío)
                            "",                        # profesor (vacío)
                            "",                        # estatus (vacío)
                            "regular",                 # origen_asignacion
                            1                          # cuatrimestre_historico
                        ))
                        creados_base += 1
                        
                        # Volver a buscar el registro base recién creado
                        cursor.execute("""
                            SELECT matricula, nombre, grupo, carrera, cuatrimestre, email_personal, usuario_email_nova, 
                                   contraseña, fecha_ingreso_original
                            FROM calificaciones
                            WHERE LOWER(TRIM(usuario_email_nova)) = ? 
                            AND UPPER(TRIM(carrera)) = ?
                            AND (materia IS NULL OR materia = '')
                            LIMIT 1
                        """, (email, carrera))
                        
                        estudiante_base = cursor.fetchone()
                        
                        if not estudiante_base:
                            errores.append(f"Fila {idx + 1}: Error al recuperar registro base creado")
                            no_encontrados += 1
                            continue
                            
                    except Exception as e:
                        errores.append(f"Fila {idx + 1}: Error creando registro base - {str(e)}")
                        no_encontrados += 1
                        continue
                
                matricula_estudiante = estudiante_base[0]
                matriculas_procesadas.add(matricula_estudiante)  # Para tracking
                
                # Verificar registro existente
                cursor.execute("""
                    SELECT id, calificacion FROM calificaciones
                    WHERE matricula = ? AND materia = ? AND fecha_ingreso_materia = ?
                """, (matricula_estudiante, materia, fecha_materia.strftime("%Y-%m-%d")))
                
                registro_existente = cursor.fetchone()
                
                if registro_existente:
                    if registro_existente[1] is not None:
                        ya_con_calificacion += 1
                        continue
                    else:
                        # Actualizar registro existente
                        cursor.execute("""
                            UPDATE calificaciones
                            SET calificacion = ?, fecha_calificacion = ?, estatus = ?, profesor = ?, grupo = ?
                            WHERE id = ?
                        """, (
                            calificacion,
                            fecha_manual.strftime("%Y-%m-%d"),
                            "aprobado" if calificacion >= 6 else "reprobado",
                            profesor,
                            grupo_alumno,
                            registro_existente[0]
                        ))
                        actualizados += 1
                        registros_con_grupo += 1
                else:
                    # Crear nuevo registro
                    nuevo_registro = crear_nuevo_registro_con_historial(estudiante_base, materia, fecha_materia, fecha_manual, calificacion, profesor, "regular") 
                    {
                        "matricula": estudiante_base[0],
                        "nombre": estudiante_base[1],
                        "grupo": grupo_alumno,
                        "materia": materia,
                        "fecha_ingreso_materia": fecha_materia.strftime("%Y-%m-%d"),
                        "fecha_ingreso_original": estudiante_base[8] if estudiante_base[8] else fecha_materia.strftime("%Y-%m-%d"),
                        "fecha_recursamiento": None,
                        "cuatrimestre": estudiante_base[4],
                        "carrera": estudiante_base[3],
                        "email_personal": estudiante_base[5],
                        "usuario_email_nova": estudiante_base[6],
                        "contraseña": estudiante_base[7],
                        "tipo_asignacion": "regular",
                        "n_recursamientos": 0,
                        "calificacion": calificacion,
                        "fecha_calificacion": fecha_manual.strftime("%Y-%m-%d"),
                        "profesor": profesor,
                        "estatus": "aprobado" if calificacion >= 6 else "reprobado",
                        "origen_asignacion": "regular"
                    }
                    
                    valores = [nuevo_registro.get(col, None) for col in COLUMNAS_DB]
                    cursor.execute(f"""
                        INSERT INTO calificaciones ({', '.join(COLUMNAS_DB)})
                        VALUES ({', '.join(['?'] * len(COLUMNAS_DB))})
                    """, valores)
                    creados += 1
                    registros_con_grupo += 1
                    
            except Exception as e:
                errores.append(f"Fila {idx + 1}: Error - {str(e)}")
        
        # ✅ NUEVA FUNCIONALIDAD: ACTUALIZAR CUATRIMESTRES AUTOMÁTICAMENTE
        st.markdown("### 🎓 Verificando promociones de cuatrimestre...")

        promociones_log = []
        for matricula in matriculas_procesadas:
            try:
        # Calcular nuevo cuatrimestre basado en materias aprobadas
                cuatrimestre_nuevo, materias_aprobadas = calcular_cuatrimestre_por_materias_aprobadas(matricula, conn)
        
        # Obtener cuatrimestre actual
                cursor.execute("""
                    SELECT cuatrimestre, nombre FROM calificaciones 
                    WHERE matricula = ? LIMIT 1
                """, (matricula,))
                resultado = cursor.fetchone()
        
                if resultado:
            # ✅ NORMALIZAR AMBOS VALORES ANTES DE COMPARAR
                    cuatrimestre_actual_raw = resultado[0] or 1
                    cuatrimestre_actual = normalizar_cuatrimestre(cuatrimestre_actual_raw)
                    cuatrimestre_nuevo_norm = normalizar_cuatrimestre(cuatrimestre_nuevo)
                    nombre_alumno = resultado[1]
            
            # ✅ CONVERTIR A ENTEROS PARA COMPARACIÓN SEGURA
                    try:
                        cuatr_actual_int = int(cuatrimestre_actual)
                        cuatr_nuevo_int = int(cuatrimestre_nuevo_norm)
                
                # Actualizar si hay promoción
                        if cuatr_nuevo_int > cuatr_actual_int:
                            cursor.execute("""
                                UPDATE calificaciones 
                                SET cuatrimestre = ? 
                                WHERE matricula = ?
                            """, (cuatrimestre_nuevo_norm, matricula))
                    
                            promociones_cuatrimestre += 1
                            promociones_log.append({
                                "nombre": nombre_alumno,
                                "matricula": matricula,
                                "cuatrimestre_anterior": cuatrimestre_actual,
                                "cuatrimestre_nuevo": cuatrimestre_nuevo_norm,
                                "materias_aprobadas": materias_aprobadas
                            })
            
                    except (ValueError, TypeError) as e:
                        st.warning(f"Error comparando cuatrimestres para {matricula}: {e}")
                        continue
    
            except Exception as e:
                st.warning(f"Error procesando promoción para {matricula}: {e}")
                continue
        
        conn.commit()
        conn.close()
        
        # Mostrar resultados
        if creados_base > 0:
            st.info(f"🎉 {creados_base} alumnos nuevos fueron agregados a la base de datos")
        
        col1, col2, col3, col4, col5, col6, col7 = st.columns(7)
        with col1:
            st.metric("✅ Actualizados", actualizados)
        with col2:
            st.metric("🆕 Creados", creados)
        with col3:
            st.metric("👤 Alumnos nuevos", creados_base)
        with col4:
            st.metric("❌ No encontrados", no_encontrados)
        with col5:
            st.metric("⚠️ Ya con calificación", ya_con_calificacion)
        with col6:
            st.metric("👥 Con grupo asignado", registros_con_grupo)
        with col7:
            st.metric("🎓 Promociones cuatrimestre", promociones_cuatrimestre)
        
        total_procesados = actualizados + creados
        
        # Mostrar promociones si las hay
        if promociones_cuatrimestre > 0:
            st.balloons()
            st.success(f"🎉 {promociones_cuatrimestre} alumnos fueron promovidos de cuatrimestre automáticamente")
            
            with st.expander("🎓 Ver detalles de promociones"):
                for promocion in promociones_log:
                    st.success(f"""
                    **{promocion['nombre']}** ({promocion['matricula']})
                    - Del cuatrimestre **{promocion['cuatrimestre_anterior']}** al **{promocion['cuatrimestre_nuevo']}**
                    - Total materias aprobadas: **{promocion['materias_aprobadas']}**
                    """)
        
        if total_procesados > 0:
            st.success(f"✅ {total_procesados} registros procesados exitosamente")
        
        if errores:
            st.warning(f"⚠️ {len(errores)} problemas encontrados")
            with st.expander("Ver errores"):
                for error in errores[:10]:
                    st.write(f"• {error}")
        
        return total_procesados
        
    except Exception as e:
        st.error(f"❌ Error general: {e}")
        return 0


# ============================================================================
# 5. MODIFICAR LA FUNCIÓN DE PROMOCIÓN MASIVA (línea ~850)
# Cambiar ejecutar_promocion_cuatrimestre_masiva() por:
# ============================================================================

def ejecutar_promocion_cuatrimestre_masiva():
    """Botón independiente para recalcular cuatrimestres conservando historial"""
    st.markdown("### 🎓 Promoción Automática de Cuatrimestre (Con Historial)")
    st.info("""
    **Sistema de Promoción con Historial:**
    - **Cuatrimestre inicial:** 1
    - **Por cada 5 materias aprobadas (≥6.0):** +1 cuatrimestre
    - **NUEVO: Conservación de historial:** Las materias ya cursadas conservan su cuatrimestre original
    - **Ejemplo:** 
      * 0-4 materias aprobadas → Cuatrimestre 1
      * 5-9 materias aprobadas → Cuatrimestre 2
      * Las materias del cuatrimestre 1 siguen mostrándose como cuatrimestre 1 en el historial
    """)
    
    if st.button("🔄 Recalcular cuatrimestres (conservando historial)", type="primary"):
        with st.spinner("Actualizando cuatrimestres con historial..."):
            actualizaciones, log_cambios = actualizar_cuatrimestre_todos_alumnos_con_historial()
        
        if actualizaciones > 0:
            st.balloons()
            st.success(f"✅ {actualizaciones} alumnos actualizados correctamente (historial conservado)")
            
            with st.expander(f"📋 Ver {len(log_cambios)} cambios realizados"):
                for cambio in log_cambios:
                    st.info(f"""
                    **{cambio['nombre']}** ({cambio['matricula']})
                    - Cuatrimestre: {cambio['cuatrimestre_anterior']} → **{cambio['cuatrimestre_nuevo']}**
                    - Materias aprobadas: **{cambio['materias_aprobadas']}**
                    - ✅ Historial de materias anteriores conservado
                    """)
        else:
            st.info("ℹ️ Todos los cuatrimestres ya están actualizados correctamente")        
            
# FUNCIÓN PRINCIPAL: procesar_calificaciones_con_grupo_excel
def procesar_calificaciones_con_grupo_excel(df_excel, materia, profesor, fecha_materia, fecha_manual, columna_grupo):
    """Procesa calificaciones usando los grupos del Excel para cada alumno individualmente"""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        actualizados = 0
        creados = 0
        no_encontrados = 0
        ya_con_calificacion = 0
        registros_con_grupo = 0
        errores = []
        
        st.write(f"📊 Procesando {len(df_excel)} registros con grupos desde Excel (columna: {columna_grupo})")
        
        # Funciones auxiliares para calificaciones
        def encontrar_columna_calificacion(df):
            """Encuentra la columna de calificación independientemente de su nombre exacto"""
            columnas_posibles = [
                "total_del_curso",
                "total del curso", 
                "total_del_curso_(real)",
                "total del curso (real)",
                "total_del_curso_real",
                "total curso",
                "calificacion",
                "calificacion_final",
                "nota_final",
                "resultado"
            ]
            
            # Normalizar nombres de columnas del DataFrame
            columnas_df_normalizadas = {}
            for col in df.columns:
                col_normalizada = col.lower().strip()
                col_normalizada = col_normalizada.replace("(", "").replace(")", "").replace("-", "_")
                col_normalizada = col_normalizada.replace(" ", "_").replace("__", "_")
                columnas_df_normalizadas[col_normalizada] = col
            
            # Buscar coincidencias
            for posible in columnas_posibles:
                posible_normalizada = posible.lower().replace(" ", "_").replace("(", "").replace(")", "")
                if posible_normalizada in columnas_df_normalizadas:
                    columna_original = columnas_df_normalizadas[posible_normalizada]
                    st.success(f"✅ Columna de calificación encontrada: '{columna_original}'")
                    return columna_original
            
            st.error("❌ No se encontró columna de calificación.")
            return None
        
        def extraer_calificacion_limpia(valor_raw):
            """Extrae y limpia la calificación de cualquier formato"""
            if pd.isna(valor_raw) or valor_raw == "":
                return 0.0
            
            valor_str = str(valor_raw).strip()
            
            texto_a_remover = [
                "(real)", "(Real)", "(REAL)",
                "real", "Real", "REAL",
                "pts", "puntos", "punto",
                "%", "porciento"
            ]
            
            for texto in texto_a_remover:
                valor_str = valor_str.replace(texto, "")
            
            valor_str = valor_str.strip().replace(" ", "").replace(",", ".")
            
            if valor_str in ["-", "--", "---", "N/A", "n/a", "NA", "na"]:
                return 0.0
            
            import re
            patron_numero = r'(\d+\.?\d*)'
            coincidencias = re.findall(patron_numero, valor_str)
            
            if coincidencias:
                try:
                    calificacion = float(coincidencias[0])
                    if calificacion > 100:
                        st.warning(f"⚠️ Calificación fuera de rango: {calificacion} - se mantendrá el valor")
                    return calificacion
                except ValueError:
                    return 0.0
            
            return 0.0
        
        # Encontrar columna de calificación
        columna_calificacion = encontrar_columna_calificacion(df_excel)
        if not columna_calificacion:
            st.error("❌ No se puede procesar el archivo sin una columna de calificación válida.")
            return 0
        
        # Mostrar resumen de grupos detectados
        grupos_en_archivo = df_excel[columna_grupo].value_counts()
        st.markdown("#### 📊 Distribución de grupos en el archivo:")
        for grupo, cantidad in grupos_en_archivo.items():
            if str(grupo).strip():
                st.write(f"• **{grupo}**: {cantidad} alumnos")
        
        # Mostrar vista previa de calificaciones
        st.markdown("### 👀 Vista previa de calificaciones detectadas")
        muestra_calificaciones = []
        for i in range(min(5, len(df_excel))):
            valor_original = df_excel.iloc[i][columna_calificacion]
            valor_limpio = extraer_calificacion_limpia(valor_original)
            grupo_alumno = str(df_excel.iloc[i][columna_grupo]).strip() if pd.notna(df_excel.iloc[i][columna_grupo]) else "Sin grupo"
            
            muestra_calificaciones.append({
                "Fila": i + 1,
                "Grupo": grupo_alumno,
                "Valor Original": str(valor_original),
                "Valor Procesado": valor_limpio,
                "Estado": "✅ OK" if valor_limpio > 0 else "⚠️ Cero/Vacío"
            })
        
        df_muestra = pd.DataFrame(muestra_calificaciones)
        st.dataframe(df_muestra, use_container_width=True)
        
        # Procesar cada registro
        for idx, row in df_excel.iterrows():
            try:
                # Obtener email
                email = ""
                if "direccion_email" in row and pd.notna(row["direccion_email"]):
                    email = str(row["direccion_email"]).strip().lower()
                elif "usuario_email_nova" in row and pd.notna(row["usuario_email_nova"]):
                    email = str(row["usuario_email_nova"]).strip().lower()
                
                # Obtener carrera/institución
                carrera = ""
                if "institucion" in row and pd.notna(row["institucion"]):
                    carrera = str(row["institucion"]).strip().upper()
                elif "carrera" in row and pd.notna(row["carrera"]):
                    carrera = str(row["carrera"]).strip().upper()
                
                # OBTENER GRUPO DEL EXCEL
                grupo_alumno = str(row[columna_grupo]).strip() if pd.notna(row[columna_grupo]) else ""
                
                # Obtener calificación
                calif_raw = row.get(columna_calificacion, "")
                calificacion = extraer_calificacion_limpia(calif_raw)
                
                # Validar datos obligatorios
                if not email:
                    errores.append(f"Fila {idx + 1}: Email faltante")
                    continue
                
                if not carrera:
                    errores.append(f"Fila {idx + 1}: Carrera/Institución faltante")
                    continue
                
                if not grupo_alumno:
                    errores.append(f"Fila {idx + 1}: Grupo faltante en columna '{columna_grupo}'")
                    continue
                
                # Buscar estudiante base
                cursor.execute("""
                    SELECT matricula, nombre, grupo, carrera, cuatrimestre, email_personal, usuario_email_nova, 
                           contraseña, fecha_ingreso_original
                    FROM calificaciones
                    WHERE LOWER(TRIM(usuario_email_nova)) = ? 
                    AND UPPER(TRIM(carrera)) = ?
                    AND (materia IS NULL OR materia = '')
                    LIMIT 1
                """, (email, carrera))
                
                estudiante_base = cursor.fetchone()
                
                if not estudiante_base:
                    no_encontrados += 1
                    errores.append(f"Fila {idx + 1}: No se encontró estudiante base con email '{email}' y carrera '{carrera}'")
                    continue
                
                matricula_estudiante = estudiante_base[0]
                
                # Verificar si el alumno ya tiene la materia
                cursor.execute("""
                    SELECT id, calificacion FROM calificaciones
                    WHERE matricula = ? 
                    AND materia = ? 
                    AND fecha_ingreso_materia = ?
                """, (matricula_estudiante, materia, fecha_materia.strftime("%Y-%m-%d")))
                
                registro_existente = cursor.fetchone()
                
                if registro_existente:
                    # Ya existe registro para esta materia/fecha
                    if registro_existente[1] is not None:  # Ya tiene calificación
                        ya_con_calificacion += 1
                        errores.append(f"Fila {idx + 1}: {matricula_estudiante} ya tiene calificación en {materia}")
                        continue
                    else:
                        # Existe pero sin calificación - actualizar
                        cursor.execute("""
                            UPDATE calificaciones
                            SET calificacion = ?, 
                                fecha_calificacion = ?, 
                                estatus = ?, 
                                profesor = ?,
                                grupo = ?
                            WHERE id = ?
                        """, (
                            calificacion,
                            fecha_manual.strftime("%Y-%m-%d"),
                            "aprobado" if calificacion >= 6 else "reprobado",
                            profesor,
                            grupo_alumno,  # USAR GRUPO DEL EXCEL
                            registro_existente[0]
                        ))
                        actualizados += 1
                        registros_con_grupo += 1
                else:
                    # No existe registro - crear nuevo
                    nuevo_registro = crear_nuevo_registro_con_historial(estudiante_base, materia, fecha_materia, fecha_manual, calificacion, profesor, "regular") 
                    {
                        "matricula": estudiante_base[0],
                        "nombre": estudiante_base[1],
                        "grupo": grupo_alumno,  # USAR GRUPO DEL EXCEL
                        "materia": materia,
                        "fecha_ingreso_materia": fecha_materia.strftime("%Y-%m-%d"),
                        "fecha_ingreso_original": estudiante_base[8] if estudiante_base[8] else fecha_materia.strftime("%Y-%m-%d"),
                        "fecha_recursamiento": None,
                        "cuatrimestre": estudiante_base[4],
                        "carrera": estudiante_base[3],
                        "email_personal": estudiante_base[5],
                        "usuario_email_nova": estudiante_base[6],
                        "contraseña": estudiante_base[7],
                        "tipo_asignacion": "regular",
                        "n_recursamientos": 0,
                        "calificacion": calificacion,
                        "fecha_calificacion": fecha_manual.strftime("%Y-%m-%d"),
                        "profesor": profesor,
                        "estatus": "aprobado" if calificacion >= 6 else "reprobado",
                        "origen_asignacion": "regular",
                        "cuatrimestre_historico": estudiante_base[4]  # ← AGREGAR ESTA LÍNEA
                    }
                    
                    valores = [nuevo_registro.get(col, None) for col in COLUMNAS_DB]
                    cursor.execute(f"""
                        INSERT INTO calificaciones ({', '.join(COLUMNAS_DB)})
                        VALUES ({', '.join(['?'] * len(COLUMNAS_DB))})
                    """, valores)
                    creados += 1
                    registros_con_grupo += 1
                    
            except Exception as e:
                errores.append(f"Fila {idx + 1}: Error procesando registro - {str(e)}")
        
        conn.commit()
        conn.close()
        
        # Mostrar resultados
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1:
            st.metric("✅ Actualizados", actualizados)
        with col2:
            st.metric("🆕 Creados", creados)
        with col3:
            st.metric("❌ No encontrados", no_encontrados)
        with col4:
            st.metric("⚠️ Ya con calificación", ya_con_calificacion)
        with col5:
            st.metric("👥 Con grupo del Excel", registros_con_grupo)
        
        total_procesados = actualizados + creados
        if total_procesados > 0:
            st.success(f"✅ {total_procesados} registros procesados exitosamente")
            st.success(f"🎯 **Grupos asignados automáticamente desde el Excel**")
            
            # Mostrar resumen final de grupos procesados
            if registros_con_grupo > 0:
                st.info(f"📊 Se procesaron alumnos en {len(grupos_en_archivo)} grupos diferentes")
        
        if errores:
            st.warning(f"⚠️ Se encontraron {len(errores)} problemas:")
            with st.expander("Ver detalles de errores"):
                for error in errores[:15]:
                    st.write(f"• {error}")
                if len(errores) > 15:
                    st.write(f"... y {len(errores) - 15} errores más")
        
        return total_procesados
        
    except Exception as e:
        st.error(f"❌ Error general al procesar calificaciones con grupos del Excel: {e}")
        return 0

# ============================================================================
# 6. NUEVA FUNCIONALIDAD: VISUALIZAR HISTORIAL DE CUATRIMESTRE
# Agregar después de la línea 1000 (en la sección de reportes)
# ============================================================================
def mostrar_historial_cuatrimestre_alumno():
    """
    Versión DEFINITIVAMENTE corregida con manejo robusto de errores
    """
    st.subheader("🎓 Historial de Cuatrimestre por Alumno")
      
    df = cargar_datos_db()
    if df.empty:
        st.warning("No hay datos disponibles.")
        return
    
    # Selector de alumno
    alumnos = sorted(df[df["nombre"].notna()]["nombre"].unique().tolist())
    if not alumnos:
        st.warning("No hay alumnos disponibles.")
        return
        
    alumno_seleccionado = st.selectbox("👤 Seleccionar alumno:", alumnos)
    
    if alumno_seleccionado:
        # Obtener datos del alumno
        df_alumno = df[df["nombre"] == alumno_seleccionado].copy()
        
        if not df_alumno.empty:
            matricula = df_alumno.iloc[0]["matricula"]
            carrera = df_alumno.iloc[0]["carrera"]
            cuatrimestre_actual = df_alumno.iloc[0]["cuatrimestre"]
            
            # ==================== SECCIÓN DE EXPORTACIÓN CORREGIDA ====================
            st.markdown("---")

            # Título centrado
            st.markdown("""
            <div style="text-align: center; margin-bottom: 20px;">
                <h3 style="color: #FDFD96; margin: 0;">Información del Estudiante por Cuatrimestre</h3>
            </div>
            """, unsafe_allow_html=True)

            col_info, col_export = st.columns([2, 1])

            with col_info:
                # Contenedor centrado para las subcolunas
                subcol1, subcol2, subcol3 = st.columns(3)

                with subcol1:
                    st.markdown(f"""
                    <div style="text-align: center; padding: 8px;">
                        <div style="font-size: 15px; color: #FDFD96; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 3px;">👤 Matrícula</div>
                        <div style="font-size: 20px; font-weight: 500; color: #888;">{matricula}</div>
                    </div>
                    """, unsafe_allow_html=True)

                with subcol2:
                    st.markdown(f"""
                    <div style="text-align: center; padding: 8px;">
                        <div style="font-size: 15px; color: #FDFD96; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 3px;">🎓 Carrera</div>
                        <div style="font-size: 14px; font-weight: 500; color: #888;">{carrera}</div>
                    </div>
                    """, unsafe_allow_html=True)

                with subcol3:
                    st.markdown(f"""
                    <div style="text-align: center; padding: 8px;">
                        <div style="font-size: 15px; color: #FDFD96; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 3px;">📚 Cuatrimestre</div>
                        <div style="font-size: 14px; font-weight: 500; color: #888;">{cuatrimestre_actual}</div>
                    </div>
                    """, unsafe_allow_html=True)
            
            with col_export:
                st.markdown("### 📊 Exportar")
                if st.button("📥 Descargar Excel", type="primary", use_container_width=True):
                   conn = sqlite3.connect(DB_FILE)
        
                   with st.spinner("🔄 Generando Excel con detalle por cuatrimestre..."):
                       resultado_export = exportar_detalle_por_cuatrimestre_excel(
                           matricula, alumno_seleccionado, conn
                       )
            
                       if resultado_export:
                           st.success("✅ ¡Excel generado exitosamente!")
                
                           st.info(f"""
                           📈 **Archivo generado:**
                           • Total de cuatrimestres: {resultado_export['total_cuatrimestres']}
                           • Total de materias: {resultado_export['total_materias']}
                           • Incluye estadísticas por cuatrimestre
                           • Una hoja por cada cuatrimestre con el detalle completo
                           """)
                
                # Botón de descarga
                           st.download_button(
                               label="💾 Descargar Detalle por Cuatrimestre",
                               data=resultado_export['buffer'],
                               file_name=resultado_export['filename'],
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True,
                               help="Descarga el historial organizado por cuatrimestre, igual que en la interfaz"
                           )
                       else:
                           st.error("❌ No se pudo generar el archivo. Revisa los datos.")
        
                   conn.close()
            # ==================== FIN NUEVA SECCIÓN ====================
            
            # Filtrar materias con calificación
            materias_con_calificacion = df_alumno[
                (df_alumno["materia"].notna()) & 
                (df_alumno["materia"] != "") &
                (df_alumno["calificacion"].notna())
            ].copy()
            
            if not materias_con_calificacion.empty:
                
                # CORRECCIÓN CLAVE: Lógica mejorada para determinar el cuatrimestre
                def determinar_cuatrimestre_historico(row):
                    """Determina el cuatrimestre histórico correcto"""
                    
                    # Si ya tiene cuatrimestre_historico y no está vacío, usarlo
                    if pd.notna(row.get("cuatrimestre_historico")) and str(row.get("cuatrimestre_historico")).strip():
                        return str(row["cuatrimestre_historico"]).strip()
                    
                    # Si no tiene cuatrimestre_historico, intentar calcularlo por fecha
                    if pd.notna(row.get("fecha_ingreso_materia")):
                        try:
                            fecha_ingreso = pd.to_datetime(row["fecha_ingreso_materia"])
                            
                            # Lógica simple: asumir que materias más antiguas son de cuatrimestres anteriores
                            # Esta es una aproximación que deberás ajustar según tu lógica de negocio
                            año_ingreso = fecha_ingreso.year
                            mes_ingreso = fecha_ingreso.month
                            
                            # Ejemplo de lógica (ajustar según tus necesidades):
                            # - Enero-Abril: Cuatrimestre de inicio del año
                            # - Mayo-Agosto: Cuatrimestre medio
                            # - Septiembre-Diciembre: Cuatrimestre final
                            
                            if mes_ingreso <= 4:
                                cuatr_base = 1
                            elif mes_ingreso <= 8:
                                cuatr_base = 2
                            else:
                                cuatr_base = 3
                                
                            # Ajustar por año si es necesario
                            año_base = 2025  # Ajustar según tu año base
                            diferencia_años = año_ingreso - año_base
                            
                            cuatrimestre_calculado = cuatr_base + (diferencia_años * 3)
                            
                            return str(max(1, cuatrimestre_calculado))
                            
                        except:
                            pass
                    
                    # Como último recurso, usar el cuatrimestre actual
                    return str(row.get("cuatrimestre", "1"))
                
                # Aplicar la lógica de cuatrimestre histórico
                materias_con_calificacion["cuatrimestre_display"] = materias_con_calificacion.apply(
                    determinar_cuatrimestre_historico, axis=1
                )
                
                # Mostrar distribución
                st.markdown("### 📊 Distribución de Materias por Cuatrimestre")
                
                try:
                    distribucion_cuatrimestre = materias_con_calificacion.groupby("cuatrimestre_display").agg({
                        "materia": "count",
                        "calificacion": ["mean", lambda x: sum(x >= 6)]
                    }).round(2)
                    
                    distribucion_cuatrimestre.columns = ["Total_Materias", "Promedio", "Aprobadas"]
                    distribucion_cuatrimestre["Reprobadas"] = distribucion_cuatrimestre["Total_Materias"] - distribucion_cuatrimestre["Aprobadas"]
                    distribucion_cuatrimestre["Tasa_Aprobacion"] = (
                        distribucion_cuatrimestre["Aprobadas"] / distribucion_cuatrimestre["Total_Materias"] * 100
                    ).round(1)
                    
                    st.dataframe(distribucion_cuatrimestre, use_container_width=True)
                    
                    # Detalle por cuatrimestre
                    st.markdown("### 📚 Detalle por Cuatrimestre")
                    
                    cuatrimestres_disponibles = sorted(materias_con_calificacion["cuatrimestre_display"].unique())
                    
                    for cuatr in cuatrimestres_disponibles:
                        materias_cuatr = materias_con_calificacion[
                            materias_con_calificacion["cuatrimestre_display"] == cuatr
                        ]
                        
                        with st.expander(f"📚 Cuatrimestre {cuatr} ({len(materias_cuatr)} materias)"):
                            
                            # Métricas del cuatrimestre
                            aprobadas = len(materias_cuatr[materias_cuatr["calificacion"] >= 6])
                            promedio = materias_cuatr["calificacion"].mean()
                            
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("✅ Aprobadas", aprobadas)
                            with col2:
                                st.metric("❌ Reprobadas", len(materias_cuatr) - aprobadas)
                            with col3:
                                st.metric("📊 Promedio", f"{promedio:.2f}")
                            
                            # Tabla de materias
                            columnas_mostrar = ["materia", "calificacion", "estatus", "fecha_ingreso_materia"]
                            if "profesor" in materias_cuatr.columns:
                                columnas_mostrar.append("profesor")
                            
                            st.dataframe(
                                materias_cuatr[columnas_mostrar],
                                use_container_width=True
                            )
                
                except Exception as e:
                    st.error(f"Error al procesar datos: {e}")
                    st.info("Intenta ejecutar el diagnóstico para identificar problemas específicos.")
                    
                    
def corregir_historial_automaticamente():
    """
    Función para corregir automáticamente el historial basado en fechas
    """
    st.subheader("🔧 Corrección Automática del Historial")
    
    st.info("""
    **Esta función intentará corregir el historial basándose en:**
    - Fechas de ingreso de materias
    - Orden cronológico de las materias
    - Patrones de progresión académica
    """)
    
    if st.button("🚀 Ejecutar Corrección Automática", type="primary"):
        
        try:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            
            with st.spinner("Corrigiendo historial..."):
                
                # Obtener todos los alumnos
                cursor.execute("SELECT DISTINCT matricula FROM calificaciones WHERE matricula IS NOT NULL")
                matriculas = [row[0] for row in cursor.fetchall()]
                
                correcciones = 0
                
                for matricula in matriculas:
                    
                    # Obtener materias del alumno ordenadas por fecha
                    cursor.execute("""
                        SELECT id, materia, fecha_ingreso_materia, cuatrimestre_historico
                        FROM calificaciones 
                        WHERE matricula = ? 
                        AND materia IS NOT NULL 
                        AND fecha_ingreso_materia IS NOT NULL
                        ORDER BY fecha_ingreso_materia ASC
                    """, (matricula,))
                    
                    materias_alumno = cursor.fetchall()
                    
                    if materias_alumno:
                        # Asignar cuatrimestre histórico basado en orden cronológico
                        # Cada 5-7 materias = nuevo cuatrimestre
                        materias_por_cuatrimestre = 6  # Ajustable
                        
                        for i, (id_registro, materia, fecha, historico_actual) in enumerate(materias_alumno):
                            
                            # Calcular cuatrimestre basado en posición cronológica
                            cuatrimestre_historico_calculado = (i // materias_por_cuatrimestre) + 1
                            
                            # Solo actualizar si no tiene historial o si es diferente
                            if not historico_actual or str(historico_actual) != str(cuatrimestre_historico_calculado):
                                cursor.execute("""
                                    UPDATE calificaciones 
                                    SET cuatrimestre_historico = ? 
                                    WHERE id = ?
                                """, (str(cuatrimestre_historico_calculado), id_registro))
                                
                                correcciones += 1
                
                conn.commit()
                conn.close()
                
                st.success(f"✅ Corrección completada: {correcciones} registros actualizados")
                st.info("💡 Puedes ahora verificar el historial para confirmar los cambios")
                
        except Exception as e:
            st.error(f"Error durante la corrección: {e}")

# FUNCIÓN AUXILIAR: procesar_calificaciones_con_grupo_manual
def procesar_calificaciones_con_grupo_manual(df_excel, materia, profesor, fecha_materia, fecha_manual, grupo_manual):
    """Procesa calificaciones asignando el mismo grupo manual a todos los alumnos"""
    # Reutiliza la función original pero con el parámetro renombrado para claridad
    return procesar_calificaciones_con_grupo(df_excel, materia, profesor, fecha_materia, fecha_manual, grupo_manual)
# NUEVA SECCIÓN: Análisis Aprobados vs Recursadores (CON FILTROS DE FECHA)
def analisis_aprobados_vs_recursadores():
    """Nueva función para analizar aprobados vs recursadores con filtros de fecha"""
    st.subheader("📊 Análisis: Aprobados vs Recursadores")
    
    # Header explicativo
    st.markdown("""
    <div style="background: linear-gradient(135deg, #ff6b6b, #ffa726); padding: 20px; border-radius: 10px; margin: 20px 0;">
        <h3 style="color: white; margin: 0; text-align: center;">🔍 Análisis Aprobados vs Recursadores</h3>
        <p style="color: #fff8e1; margin: 5px 0 0 0; text-align: center;">Compara grupos originales con recursadores integrados</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Información sobre la funcionalidad
    with st.expander("💡 ¿Cómo funciona esta funcionalidad?", expanded=False):
        st.markdown("""
        **🎯 Objetivo:**
        Esta función te permite comparar el rendimiento académico entre diferentes tipos de estudiantes:
        
        **📊 Tipos de análisis:**
        - **Grupos Originales:** Estudiantes que tomaron la materia por primera vez
        - **Recursadores:** Estudiantes que están repitiendo la materia
        - **Mixtos:** Grupos que combinan ambos tipos
        
        **📈 Métricas comparadas:**
        - Promedio de calificaciones
        - Tasa de aprobación
        - Distribución de notas
        - Cantidad de estudiantes por tipo
        
        **🔍 Filtros disponibles:**
        - Por carrera
        - Por materia específica
        - **Por fecha de ingreso a materia** (NUEVO)
        - Vista de fechas y materia en resultados
        """)
    
    df = cargar_datos_db()
    if df.empty:
        st.warning("No hay datos disponibles para análisis.")
        return
    
    # FILTROS PRINCIPALES CON FECHAS
    st.markdown("### 🎯 Filtros de Búsqueda")
    
    # Primera fila de filtros
    col1, col2 = st.columns(2)
    
    with col1:
        carreras = sorted(df["carrera"].dropna().unique().tolist())
        carrera_analisis = st.selectbox("🎓 Carrera", carreras, key="analisis_carrera")
    
    with col2:
        # Filtrar materias por carrera
        materias_carrera = sorted(df[df["carrera"] == carrera_analisis]["materia"].dropna().unique().tolist())
        materia_analisis = st.selectbox("📘 Materia", materias_carrera, key="analisis_materia")
    
    # Segunda fila de filtros - NUEVOS FILTROS DE FECHA
    col3, col4 = st.columns(2)
    
    with col3:
        # Filtro por fecha de ingreso a materia
        df_temp = df[
            (df["carrera"] == carrera_analisis) & 
            (df["materia"] == materia_analisis)
        ]
        
        fechas_materia = sorted(df_temp["fecha_ingreso_materia"].dropna().unique().tolist())
        
        if fechas_materia:
            fecha_materia_analisis = st.selectbox(
                "📅 Fecha de ingreso a materia", 
                ["Todas las fechas"] + fechas_materia, 
                key="analisis_fecha_materia",
                help="Filtra por una fecha específica de ingreso a la materia"
            )
        else:
            st.warning("No hay fechas de materia disponibles")
            fecha_materia_analisis = "Todas las fechas"
    
    with col4:
        # Información de fechas disponibles
        if fechas_materia:
            st.markdown("**📊 Fechas disponibles:**")
            for fecha in fechas_materia[:3]:  # Mostrar solo las primeras 3
                count_fecha = len(df_temp[df_temp["fecha_ingreso_materia"] == fecha])
                st.write(f"• {fecha}: {count_fecha} registros")
            if len(fechas_materia) > 3:
                st.write(f"... y {len(fechas_materia) - 3} fechas más")
        else:
            st.info("No hay fechas de materia para mostrar")
    
    # APLICAR FILTROS
    df_filtrado = df[
        (df["carrera"] == carrera_analisis) & 
        (df["materia"] == materia_analisis) &
        (df["calificacion"].notna())  # Solo registros con calificación
    ].copy()
    
    # Aplicar filtro de fecha si está seleccionado
    if fecha_materia_analisis != "Todas las fechas":
        df_filtrado = df_filtrado[df_filtrado["fecha_ingreso_materia"] == fecha_materia_analisis]
    
    if df_filtrado.empty:
        st.warning("No hay datos con calificaciones para los filtros seleccionados.")
        return
    
    # Convertir calificaciones a numérico de forma segura
    df_filtrado["calificacion"] = pd.to_numeric(df_filtrado["calificacion"], errors="coerce")
    df_filtrado = df_filtrado[df_filtrado["calificacion"].notna()]
    
    if df_filtrado.empty:
        st.warning("No hay calificaciones numéricas válidas para analizar.")
        return
    
    # INFORMACIÓN DE CONTEXTO - NUEVA SECCIÓN
    st.markdown("### 📋 Información del Análisis")
    
    # Mostrar información contextual con las fechas
    col_info1, col_info2, col_info3 = st.columns(3)
    
    with col_info1:
        st.markdown(f"""
        <div style="background: #e3f2fd; padding: 15px; border-radius: 10px; border-left: 4px solid #2196f3;">
            <h4 style="margin: 0; color: #1565c0;">🎓 Carrera</h4>
            <p style="margin: 5px 0 0 0; font-weight: bold;">{carrera_analisis}</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col_info2:
        st.markdown(f"""
        <div style="background: #f3e5f5; padding: 15px; border-radius: 10px; border-left: 4px solid #9c27b0;">
            <h4 style="margin: 0; color: #7b1fa2;">📘 Materia</h4>
            <p style="margin: 5px 0 0 0; font-weight: bold;">{materia_analisis}</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col_info3:
        fecha_display = fecha_materia_analisis if fecha_materia_analisis != "Todas las fechas" else f"{len(fechas_materia)} fechas"
        st.markdown(f"""
        <div style="background: #e8f5e8; padding: 15px; border-radius: 10px; border-left: 4px solid #4caf50;">
            <h4 style="margin: 0; color: #2e7d32;">📅 Fecha Análisis</h4>
            <p style="margin: 5px 0 0 0; font-weight: bold;">{fecha_display}</p>
        </div>
        """, unsafe_allow_html=True)
    
    # Clasificar estudiantes por tipo
    def clasificar_estudiante(row):
        """Clasifica al estudiante según su tipo"""
        try:
            tipo_asignacion = str(row.get("tipo_asignacion", "")).lower()
            estatus = str(row.get("estatus", "")).lower()
            n_recursamientos = row.get("n_recursamientos", 0)
            
            # Convertir recursamientos a número de forma segura
            try:
                n_recursamientos = int(float(n_recursamientos)) if pd.notna(n_recursamientos) else 0
            except:
                n_recursamientos = 0
            
            # Lógica de clasificación
            if tipo_asignacion == "recursamiento" or estatus == "recursando" or n_recursamientos > 0:
                return "Recursador"
            else:
                return "Original"
        except:
            return "Original"  # Default
    
    # Aplicar clasificación
    df_filtrado["tipo_estudiante"] = df_filtrado.apply(clasificar_estudiante, axis=1)
    
    # Contar tipos
    conteo_tipos = df_filtrado["tipo_estudiante"].value_counts()
    originales = conteo_tipos.get("Original", 0)
    recursadores = conteo_tipos.get("Recursador", 0)
    total = originales + recursadores
    
    # Separar datos desde el inicio
    df_originales = df_filtrado[df_filtrado["tipo_estudiante"] == "Original"]
    df_recursadores = df_filtrado[df_filtrado["tipo_estudiante"] == "Recursador"]
    
    # Calcular estadísticas desde el principio (con valores por defecto)
    promedio_originales = df_originales["calificacion"].mean() if len(df_originales) > 0 else 0.0
    promedio_recursadores = df_recursadores["calificacion"].mean() if len(df_recursadores) > 0 else 0.0
    
    aprobados_originales = len(df_originales[df_originales["calificacion"] >= 6]) if len(df_originales) > 0 else 0
    aprobados_recursadores = len(df_recursadores[df_recursadores["calificacion"] >= 6]) if len(df_recursadores) > 0 else 0
    
    tasa_aprobacion_originales = (aprobados_originales / originales) * 100 if originales > 0 else 0
    tasa_aprobacion_recursadores = (aprobados_recursadores / recursadores) * 100 if recursadores > 0 else 0
    
    # Mostrar métricas generales
    st.markdown("### 📊 Resumen General")
    
    col_met1, col_met2, col_met3, col_met4 = st.columns(4)
    with col_met1:
        st.metric("👥 Total Estudiantes", total)
    with col_met2:
        st.metric("📚 Estudiantes Originales", originales)
    with col_met3:
        st.metric("🔁 Recursadores", recursadores)
    with col_met4:
        if total > 0:
            porcentaje_recursadores = (recursadores / total) * 100
            st.metric("📈 % Recursadores", f"{porcentaje_recursadores:.1f}%")
        else:
            st.metric("📈 % Recursadores", "0%")
    
    # Análisis comparativo
    if originales > 0 and recursadores > 0:
        st.markdown("### 🔍 Análisis Comparativo")
        
        # Mostrar comparación en tabs
        tab1, tab2, tab3 = st.tabs(["📊 Promedios", "✅ Tasas de Aprobación", "📈 Distribución Detallada"])
        
        with tab1:
            st.markdown("#### 📊 Comparación de Promedios")
            
            col_prom1, col_prom2, col_prom3 = st.columns(3)
            
            with col_prom1:
                # Color según promedio
                color_orig = "#27ae60" if promedio_originales >= 7 else "#e74c3c" if promedio_originales < 6 else "#f39c12"
                st.markdown(f"""
                <div style="background: {color_orig}; padding: 15px; border-radius: 10px; text-align: center;">
                    <h4 style="color: white; margin: 0;">📚 Originales</h4>
                    <h2 style="color: white; margin: 5px 0 0 0;">{promedio_originales:.2f}</h2>
                </div>
                """, unsafe_allow_html=True)
            
            with col_prom2:
                color_rec = "#27ae60" if promedio_recursadores >= 7 else "#e74c3c" if promedio_recursadores < 6 else "#f39c12"
                st.markdown(f"""
                <div style="background: {color_rec}; padding: 15px; border-radius: 10px; text-align: center;">
                    <h4 style="color: white; margin: 0;">🔁 Recursadores</h4>
                    <h2 style="color: white; margin: 5px 0 0 0;">{promedio_recursadores:.2f}</h2>
                </div>
                """, unsafe_allow_html=True)
            
            with col_prom3:
                diferencia = promedio_originales - promedio_recursadores
                color_dif = "#27ae60" if diferencia > 0 else "#e74c3c" if diferencia < 0 else "#95a5a6"
                simbolo = "↑" if diferencia > 0 else "↓" if diferencia < 0 else "="
                st.markdown(f"""
                <div style="background: {color_dif}; padding: 15px; border-radius: 10px; text-align: center;">
                    <h4 style="color: white; margin: 0;">{simbolo} Diferencia</h4>
                    <h2 style="color: white; margin: 5px 0 0 0;">{abs(diferencia):.2f}</h2>
                </div>
                """, unsafe_allow_html=True)
            
            # Interpretación
            if diferencia > 0.5:
                st.success(f"✅ Los estudiantes originales tienen un promedio significativamente mayor ({diferencia:.2f} puntos)")
            elif diferencia < -0.5:
                st.warning(f"⚠️ Los recursadores tienen un promedio mayor ({abs(diferencia):.2f} puntos)")
            else:
                st.info(f"📊 Los promedios son similares (diferencia de {abs(diferencia):.2f} puntos)")
        
        with tab2:
            st.markdown("#### ✅ Comparación de Tasas de Aprobación")
            
            col_tasa1, col_tasa2, col_tasa3 = st.columns(3)
            
            with col_tasa1:
                color_tasa_orig = "#27ae60" if tasa_aprobacion_originales >= 80 else "#f39c12" if tasa_aprobacion_originales >= 60 else "#e74c3c"
                st.markdown(f"""
                <div style="background: {color_tasa_orig}; padding: 15px; border-radius: 10px; text-align: center;">
                    <h4 style="color: white; margin: 0;">📚 Originales</h4>
                    <h2 style="color: white; margin: 5px 0 0 0;">{tasa_aprobacion_originales:.1f}%</h2>
                    <p style="color: white; margin: 0;">({aprobados_originales}/{originales})</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col_tasa2:
                color_tasa_rec = "#27ae60" if tasa_aprobacion_recursadores >= 80 else "#f39c12" if tasa_aprobacion_recursadores >= 60 else "#e74c3c"
                st.markdown(f"""
                <div style="background: {color_tasa_rec}; padding: 15px; border-radius: 10px; text-align: center;">
                    <h4 style="color: white; margin: 0;">🔁 Recursadores</h4>
                    <h2 style="color: white; margin: 5px 0 0 0;">{tasa_aprobacion_recursadores:.1f}%</h2>
                    <p style="color: white; margin: 0;">({aprobados_recursadores}/{recursadores})</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col_tasa3:
                dif_tasa = tasa_aprobacion_originales - tasa_aprobacion_recursadores
                color_dif_tasa = "#27ae60" if dif_tasa > 0 else "#e74c3c" if dif_tasa < 0 else "#95a5a6"
                simbolo_tasa = "↑" if dif_tasa > 0 else "↓" if dif_tasa < 0 else "="
                st.markdown(f"""
                <div style="background: {color_dif_tasa}; padding: 15px; border-radius: 10px; text-align: center;">
                    <h4 style="color: white; margin: 0;">{simbolo_tasa} Diferencia</h4>
                    <h2 style="color: white; margin: 5px 0 0 0;">{abs(dif_tasa):.1f}%</h2>
                </div>
                """, unsafe_allow_html=True)
        
        with tab3:
            st.markdown("#### 📈 Distribución Detallada por Rangos")
            
            # Definir rangos de calificación
            rangos = {
                "Excelente (9-10)": (9, 10),
                "Muy Bueno (8-8.9)": (8, 8.9),
                "Bueno (7-7.9)": (7, 7.9),
                "Suficiente (6-6.9)": (6, 6.9),
                "Insuficiente (0-5.9)": (0, 5.9)
            }
            
            # Crear tabla de distribución
            distribucion_data = []
            
            for rango_nombre, (min_val, max_val) in rangos.items():
                # Contar originales en este rango
                count_orig = len(df_originales[
                    (df_originales["calificacion"] >= min_val) & 
                    (df_originales["calificacion"] <= max_val)
                ])
                
                # Contar recursadores en este rango
                count_rec = len(df_recursadores[
                    (df_recursadores["calificacion"] >= min_val) & 
                    (df_recursadores["calificacion"] <= max_val)
                ])
                
                # Calcular porcentajes
                pct_orig = (count_orig / originales * 100) if originales > 0 else 0
                pct_rec = (count_rec / recursadores * 100) if recursadores > 0 else 0
                
                distribucion_data.append({
                    "Rango": rango_nombre,
                    "Originales": f"{count_orig} ({pct_orig:.1f}%)",
                    "Recursadores": f"{count_rec} ({pct_rec:.1f}%)",
                    "Total": count_orig + count_rec
                })
            
            # Mostrar tabla
            df_distribucion = pd.DataFrame(distribucion_data)
            st.dataframe(df_distribucion, use_container_width=True)
            
            # Resumen de insights
            st.markdown("#### 💡 Insights del Análisis")
            
            if len(df_originales) > 0:
                mejor_rango_orig = df_originales["calificacion"].quantile(0.75)
            else:
                mejor_rango_orig = 0
                
            if len(df_recursadores) > 0:
                mejor_rango_rec = df_recursadores["calificacion"].quantile(0.75)
            else:
                mejor_rango_rec = 0
            
            insights = []
            
            if promedio_originales > promedio_recursadores and originales > 0 and recursadores > 0:
                insights.append(f"📈 Los estudiantes originales tienen mejor rendimiento promedio (+{(promedio_originales - promedio_recursadores):.2f} puntos)")
            elif promedio_recursadores > promedio_originales and originales > 0 and recursadores > 0:
                insights.append(f"📈 Los recursadores muestran mejor rendimiento promedio (+{(promedio_recursadores - promedio_originales):.2f} puntos)")
            
            if tasa_aprobacion_originales > tasa_aprobacion_recursadores and originales > 0 and recursadores > 0:
                insights.append(f"✅ Mayor tasa de aprobación en originales (+{(tasa_aprobacion_originales - tasa_aprobacion_recursadores):.1f}%)")
            elif tasa_aprobacion_recursadores > tasa_aprobacion_originales and originales > 0 and recursadores > 0:
                insights.append(f"✅ Mayor tasa de aprobación en recursadores (+{(tasa_aprobacion_recursadores - tasa_aprobacion_originales):.1f}%)")
            
            if mejor_rango_orig > mejor_rango_rec and originales > 0 and recursadores > 0:
                insights.append("🎯 Los estudiantes originales muestran mejor consistencia en calificaciones altas")
            elif mejor_rango_rec > mejor_rango_orig and originales > 0 and recursadores > 0:
                insights.append("🎯 Los recursadores muestran mejor consistencia en calificaciones altas")
            
            for insight in insights:
                st.write(f"• {insight}")
            
            if not insights:
                st.write("• 📊 Los grupos muestran rendimiento similar o no hay suficientes datos para comparar")
    
    elif originales > 0 and recursadores == 0:
        st.info("📋 **Grupo homogéneo:** Solo hay estudiantes originales en esta materia")
        
        # Mostrar estadísticas básicas del grupo original
        promedio_grupo = df_filtrado["calificacion"].mean()
        aprobados_grupo = len(df_filtrado[df_filtrado["calificacion"] >= 6])
        tasa_grupo = (aprobados_grupo / originales) * 100
        
        col_grupo1, col_grupo2, col_grupo3 = st.columns(3)
        with col_grupo1:
            st.metric("📊 Promedio", f"{promedio_grupo:.2f}")
        with col_grupo2:
            st.metric("✅ Aprobados", f"{aprobados_grupo}/{originales}")
        with col_grupo3:
            st.metric("📈 Tasa Aprobación", f"{tasa_grupo:.1f}%")
    
    elif originales == 0 and recursadores > 0:
        st.info("📋 **Grupo de recursamiento:** Solo hay recursadores en esta materia")
        
        # Mostrar estadísticas básicas del grupo de recursamiento
        promedio_grupo = df_filtrado["calificacion"].mean()
        aprobados_grupo = len(df_filtrado[df_filtrado["calificacion"] >= 6])
        tasa_grupo = (aprobados_grupo / recursadores) * 100
        
        col_grupo1, col_grupo2, col_grupo3 = st.columns(3)
        with col_grupo1:
            st.metric("📊 Promedio", f"{promedio_grupo:.2f}")
        with col_grupo2:
            st.metric("✅ Aprobados", f"{aprobados_grupo}/{recursadores}")
        with col_grupo3:
            st.metric("📈 Tasa Aprobación", f"{tasa_grupo:.1f}%")
    
    # Tabla de datos completa CON FECHAS
    st.markdown("### 📋 Datos Detallados")
    
    # Preparar tabla para mostrar (AHORA INCLUYE FECHAS)
    columnas_mostrar = [
        "matricula", "nombre", "grupo", "materia", "fecha_ingreso_materia", 
        "calificacion", "estatus", "tipo_asignacion", "n_recursamientos", 
        "tipo_estudiante", "profesor"
    ]
    columnas_existentes = [col for col in columnas_mostrar if col in df_filtrado.columns]
    
    # Ordenar por tipo y luego por calificación
    df_mostrar = df_filtrado[columnas_existentes].sort_values(
        ["tipo_estudiante", "calificacion"], 
        ascending=[True, False]
    )
    
    st.dataframe(df_mostrar, use_container_width=True)
    
    # INFORMACIÓN ADICIONAL DE FECHAS
    if fecha_materia_analisis == "Todas las fechas" and len(fechas_materia) > 1:
        st.markdown("### 📅 Información de Fechas Incluidas")
        
        # Mostrar distribución por fechas
        distribucion_fechas = df_filtrado.groupby(['fecha_ingreso_materia', 'tipo_estudiante']).size().unstack(fill_value=0)
        
        if not distribucion_fechas.empty:
            st.markdown("**📊 Distribución por fecha:**")
            st.dataframe(distribucion_fechas, use_container_width=True)
    
    # Botón de descarga (MEJORADO CON FECHAS)
    st.markdown("### 📥 Descargar Análisis")
    
    import io
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_mostrar.to_excel(writer, index=False, sheet_name='Analisis_Completo')
        
        # Hoja adicional con resumen (INCLUYE FECHAS)
        resumen_data = {
            "Métrica": [
                "Carrera", "Materia", "Fecha Analizada",
                "Total Estudiantes", "Estudiantes Originales", "Recursadores",
                "Promedio Originales", "Promedio Recursadores", 
                "Tasa Aprobación Originales", "Tasa Aprobación Recursadores"
            ],
            "Valor": [
                carrera_analisis, materia_analisis, fecha_materia_analisis,
                total, originales, recursadores,
                f"{promedio_originales:.2f}" if originales > 0 else "N/A",
                f"{promedio_recursadores:.2f}" if recursadores > 0 else "N/A",
                f"{tasa_aprobacion_originales:.1f}%" if originales > 0 else "N/A",
                f"{tasa_aprobacion_recursadores:.1f}%" if recursadores > 0 else "N/A"
            ]
        }
        df_resumen = pd.DataFrame(resumen_data)
        df_resumen.to_excel(writer, index=False, sheet_name='Resumen')
        
        # Hoja adicional con distribución por fechas si aplica
        if fecha_materia_analisis == "Todas las fechas" and len(fechas_materia) > 1:
            distribucion_fechas = df_filtrado.groupby(['fecha_ingreso_materia', 'tipo_estudiante']).size().unstack(fill_value=0)
            if not distribucion_fechas.empty:
                distribucion_fechas.to_excel(writer, sheet_name='Distribucion_Fechas')
    
    output.seek(0)
    
    # Nombre de archivo que incluye la fecha
    fecha_sufijo = fecha_materia_analisis.replace("-", "") if fecha_materia_analisis != "Todas las fechas" else "todas_fechas"
    nombre_archivo = f"analisis_aprobados_vs_recursadores_{carrera_analisis}_{materia_analisis}_{fecha_sufijo}.xlsx"
    nombre_archivo = nombre_archivo.replace(" ", "_").replace("/", "-")
    
    st.download_button(
        label="📥 Descargar Análisis Completo",
        data=output,
        file_name=nombre_archivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )       

# App principal con múltiples páginas
st.set_page_config(page_title="Sistema de Calificaciones")
inicializar_base()

menu = st.sidebar.radio("Menú", [
    "📂 Cargar datos desde Excel",
    "📚 Generación de materia",
    "📥 Captura de calificaciones",
    "📌 Gestión de recursamiento",
   # "🎓 Promoción de cuatrimestre",
    "📋 Ver datos y filtrar",
    "📊 Reportes",
    "📊 Resumen por grupo",
    "🧾 Historial por alumno",
    "🎓 Historial de cuatrimestre",  # ← AGREGAR ESTA LÍNEA
   # "🎓 Limpiar Base de Datos",
    "👥 Estatus de alumnos",
    "✏️ Editar registros",
    "🔄 Mover materias de cuatrimestre",
    "🗑️ Eliminar registro",
    "🗑️ Eliminar recursamientos con calificación",
    "📋 Ver integraciones",
    "🔍 Análisis Aprobados vs Recursadores",  # ← ESTA LÍNEA ES NUEVA
    "👨‍🏫 Tutorías"
])

if menu == "📂 Cargar datos desde Excel":
    fecha_ingreso_materia = st.date_input("📅 Fecha de ingreso de los alumnos")
    archivo = st.file_uploader("Selecciona un archivo Excel con datos de alumnos", type=["xlsx"])
    if archivo:
        try:
            df_excel = pd.read_excel(archivo, dtype=str).fillna("")
            df_excel.columns = [normalizar_columna(col) for col in df_excel.columns]
            if "direccion_email" in df_excel.columns:
                df_excel["usuario_email_nova"] = df_excel["direccion_email"]

            columnas_esperadas = {
                "matricula": ["matricula", "id"],
                "nombre": ["nombre"],
                "grupo": ["grupo"],
                "cuatrimestre": ["cuatrimestre"],
                "carrera": ["carrera"],
                "email_personal": ["email", "correo"],
                "usuario_email_nova": ["usuario_email_nova", "correo_institucional"]
            }

            columnas_renombradas = {}
            for clave, alternativas in columnas_esperadas.items():
                for alt in alternativas:
                    if alt in df_excel.columns:
                        columnas_renombradas[alt] = clave
                        break

            df_excel.rename(columns=columnas_renombradas, inplace=True)

            if "materia" not in df_excel.columns:
                df_excel["materia"] = ""

            columnas_obligatorias = ["matricula", "nombre", "grupo", "carrera", "email_personal", "usuario_email_nova"]
            faltantes = [col for col in columnas_obligatorias if col not in df_excel.columns]
            if faltantes:
                st.error(f"❌ Faltan columnas requeridas: {faltantes}")
                st.stop()

            df_excel["fecha_ingreso_materia"] = fecha_ingreso_materia.strftime("%Y-%m-%d")
            df_excel["fecha_ingreso_original"] = fecha_ingreso_materia.strftime("%Y-%m-%d")

            st.dataframe(df_excel)

            if st.button("📥 Insertar en base de datos"):
                insertar_desde_excel(df_excel)
        except Exception as e:
            st.error(f"❌ Error al procesar el archivo: {e}")

elif menu == "📋 Ver datos y filtrar":
    st.subheader("📋 Vista completa de registros filtrados")
    df = cargar_datos_db()

    if not df.empty:
        columnas_visibles = [
            "matricula", "nombre", "grupo", "materia", "calificacion", "fecha_ingreso_original",
            "estatus", "profesor", "origen_asignacion", "fecha_calificacion",
            "cuatrimestre", "carrera", "email_personal", "usuario_email_nova", "fecha_ingreso_materia"
        ]
        columnas_existentes = [col for col in columnas_visibles if col in df.columns]

        incluir_vacios = st.checkbox("🔄 Incluir registros sin calificación", value=False)

        if not incluir_vacios and "calificacion" in df.columns:
            df = df[df["calificacion"].notnull()]

        df_filtrado = aplicar_filtros(df)

        st.markdown(f"**Resultados encontrados:** {len(df_filtrado)} registros")
        df_visibles = df_filtrado[df_filtrado['estatus'].isin(['aprobado', 'recursando'])].drop_duplicates(subset=['matricula'])
        st.markdown(f"🧮 **TOTAL ACTIVOS: {len(df_visibles)} alumnos**")
        st.dataframe(df_filtrado[columnas_existentes], use_container_width=True)

        st.markdown("### 🗑️ Eliminar registros seleccionados")

        df_filtrado = df_filtrado.reset_index(drop=True)
        seleccionados = st.multiselect(
            "Selecciona los registros a eliminar:",
            options=df_filtrado.index,
            format_func=lambda i: f"{df_filtrado.at[i, 'matricula']} - {df_filtrado.at[i, 'materia']} - {df_filtrado.at[i, 'fecha_ingreso_materia']}"
        )

        if seleccionados and st.button("❌ Eliminar registros seleccionados de la base de datos"):
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            eliminados = 0
            bloqueados = 0
            for i in seleccionados:
                row = df_filtrado.loc[i]
                if pd.notna(row.get("calificacion")):
                    bloqueados += 1
                    continue
                if not row["materia"]:
                    cursor.execute(
                        """
                        DELETE FROM calificaciones
                        WHERE matricula = ? AND grupo = ? AND fecha_ingreso_materia = ? AND (materia IS NULL OR materia = '')
                        """,
                        (row["matricula"], row["grupo"], row["fecha_ingreso_materia"])
                    )
                else:
                    cursor.execute(
                        """
                        DELETE FROM calificaciones
                        WHERE matricula = ? AND materia = ? AND grupo = ? AND fecha_ingreso_materia = ?
                        """,
                        (row["matricula"], row["materia"], row["grupo"], row["fecha_ingreso_materia"])
                    )
                eliminados += cursor.rowcount
            conn.commit()
            conn.close()
            if eliminados:
                st.success(f"✅ Registros eliminados: {eliminados}")
            if bloqueados:
                st.warning(f"⚠️ {bloqueados} registro(s) no se eliminaron porque tienen calificación asignada.")

    else:
        st.info("No hay datos disponibles.")

elif menu == "📥 Captura de calificaciones":
    st.header("🧾 Carga de calificaciones finales")
    archivo = st.file_uploader("Carga el archivo Excel con calificaciones finales", type=["xlsx"])
    materia = st.text_input("📝 Nombre de la materia")
    profesor = st.text_input("👨‍🏫 Nombre del profesor que calificó")
    fecha_materia = st.date_input("📅 Fecha de ingreso a la materia")
    fecha_manual = st.date_input("🗓 Fecha de la calificación")
    
    # MODIFICADO: Verificar si existe columna grupo en Excel
    grupo_desde_excel = False
    grupos_detectados = []
    modo_grupo = None
    
    if archivo:
        try:
            # Leer archivo para detectar columna grupo
            df_temp = pd.read_excel(archivo, dtype=str).fillna("")
            df_temp.columns = [normalizar_columna(col) for col in df_temp.columns]
            
            # Buscar columna de grupo
            columnas_grupo = ["grupo", "group", "seccion", "grupo_asignado"]
            columna_grupo_encontrada = None
            
            for col_posible in columnas_grupo:
                if col_posible in df_temp.columns:
                    columna_grupo_encontrada = col_posible
                    grupos_detectados = list(df_temp[col_posible].dropna().unique())
                    grupos_detectados = [str(g).strip() for g in grupos_detectados if str(g).strip()]
                    break
            
            if columna_grupo_encontrada and grupos_detectados:
                grupo_desde_excel = True
                st.success(f"✅ Columna de grupo detectada: '{columna_grupo_encontrada}'")
                st.info(f"🎯 Grupos encontrados en el archivo: {', '.join(grupos_detectados)}")
                modo_grupo = "excel"
            else:
                st.warning("⚠️ No se detectó columna de grupo en el Excel")
                modo_grupo = "manual"
        except Exception as e:
            st.error(f"Error al leer el archivo: {e}")
            modo_grupo = "manual"
    
    # Mostrar opciones según si hay grupo en Excel o no
    st.markdown("### 👥 Asignación de Grupo")
    
    if grupo_desde_excel:
        st.success("🎯 **Modo automático:** Se usarán los grupos del archivo Excel")
        st.info(f"Se procesarán {len(grupos_detectados)} grupos diferentes: **{', '.join(grupos_detectados)}**")
        
        # Mostrar vista previa de grupos por alumno
        if archivo:
            st.markdown("#### 👀 Vista previa de grupos por alumno:")
            df_preview = pd.read_excel(archivo, dtype=str).fillna("")
            df_preview.columns = [normalizar_columna(col) for col in df_preview.columns]
            
            # Mostrar solo columnas relevantes para verificación
            columnas_mostrar = []
            if "nombre" in df_preview.columns:
                columnas_mostrar.append("nombre")
            if "usuario_email_nova" in df_preview.columns:
                columnas_mostrar.append("usuario_email_nova")
            elif "direccion_email" in df_preview.columns:
                columnas_mostrar.append("direccion_email")
            columnas_mostrar.append(columna_grupo_encontrada)
            
            st.dataframe(df_preview[columnas_mostrar].head(10), use_container_width=True)
        
        usar_grupo_manual = st.checkbox("🔧 Sobrescribir con grupo manual (opcional)")
        if usar_grupo_manual:
            nuevo_grupo = st.text_input(
                "🎯 Grupo manual para TODOS los alumnos", 
                placeholder="Ej: 1, 2, G1A, etc.",
                help="Este grupo sobrescribirá todos los grupos del Excel"
            )
            modo_grupo = "manual"
        else:
            nuevo_grupo = None
    else:
        st.info("💡 **Modo manual:** Debes asignar un grupo para todos los alumnos")
        nuevo_grupo = st.text_input(
            "🎯 Grupo a asignar", 
            placeholder="Ej: 1, 2, G1A, etc.",
            help="Este grupo se asignará a todos los alumnos procesados"
        )
        if nuevo_grupo:
            st.success(f"✅ Grupo seleccionado: **{nuevo_grupo}**")
            st.info("Todos los alumnos del archivo serán asignados a este grupo")

    # Validación de campos requeridos
    campos_completos = archivo and materia and profesor
    grupo_definido = (grupo_desde_excel and modo_grupo == "excel") or (nuevo_grupo and modo_grupo == "manual")

    if campos_completos and grupo_definido:
        try:
            # Crear respaldo antes de procesar
            if 'respaldo_captura' not in st.session_state:
                st.session_state["respaldo_captura"] = cargar_datos_db()
            
            df_excel = pd.read_excel(archivo, dtype=str).fillna("")
            df_excel.columns = [normalizar_columna(col) for col in df_excel.columns]
            
            # Mapear columnas de email
            if "direccion_email" in df_excel.columns and "usuario_email_nova" not in df_excel.columns:
                df_excel["usuario_email_nova"] = df_excel["direccion_email"]

            st.markdown("### 📊 Vista previa del archivo:")
            st.dataframe(df_excel.head())

            # Verificar columnas necesarias
            if "usuario_email_nova" not in df_excel.columns and "direccion_email" not in df_excel.columns:
                st.error("❌ No se encontró columna de email (usuario_email_nova o direccion_email)")
                st.stop()
            
            if "institucion" not in df_excel.columns and "carrera" not in df_excel.columns:
                st.error("❌ No se encontró columna de carrera/institución")
                st.stop()

            # Información del proceso
            info_grupo = ""
            if modo_grupo == "excel":
                info_grupo = f"**Grupos desde Excel:** {', '.join(grupos_detectados)}"
            else:
                info_grupo = f"**Grupo manual:** {nuevo_grupo}"

            st.info(f"""
            📝 **Información del proceso:**
            - **Materia:** {materia}
            - **Profesor:** {profesor}
            - **Fecha de materia:** {fecha_materia}
            - **Fecha de calificación:** {fecha_manual}
            - {info_grupo}
            - **Registros a procesar:** {len(df_excel)}
            
            ⚠️ **IMPORTANTE:** 
            - No se sobrescribirán calificaciones existentes
            - Se crearán nuevos registros para materias diferentes
            - Se preservará el historial académico completo
            """)

            if st.button("📥 Registrar calificaciones", type="primary"):
                with st.spinner("Procesando calificaciones..."):
                    # NUEVO CÓDIGO CON PROMOCIÓN AUTOMÁTICA:
                    if modo_grupo == "excel":
                        total_procesados = procesar_calificaciones_con_promocion_automatica(
                            df_excel, materia, profesor, fecha_materia, fecha_manual, columna_grupo_encontrada
                        )
                    else:
                        total_procesados = procesar_calificaciones_con_promocion_automatica(
                            df_excel, materia, profesor, fecha_materia, fecha_manual, nuevo_grupo
                        )
                
                if total_procesados > 0:
                    st.balloons()
                    
        except Exception as e:
            st.error(f"❌ Error al procesar el archivo: {e}")
            st.write(f"Detalles del error: {str(e)}")
    
    elif archivo:
        if not grupo_definido:
            if grupo_desde_excel:
                st.warning("⚠️ Los grupos se tomarán automáticamente del Excel")
            else:
                st.warning("⚠️ Por favor asigna un grupo para los alumnos")
        else:
            st.warning("⚠️ Por favor completa todos los campos obligatorios (materia y profesor)")
    
    # Información sobre la asignación de grupo
    st.markdown("---")
    st.markdown("### ℹ️ Sobre la Asignación de Grupo")
    st.info("""
    **🎯 Modos de asignación de grupo:**
    
    **🤖 Automático (desde Excel):**
    - Se detecta automáticamente la columna 'grupo' en el Excel
    - Cada alumno mantiene su grupo específico del archivo
    - Permite grupos heterogéneos en un solo procesamiento
    
    **✋ Manual:**
    - Asignas un grupo único para todos los alumnos
    - Útil cuando el Excel no tiene columna de grupo
    - Todos los procesados quedarán en el mismo grupo
    
    **✅ Para alumnos nuevos en la materia:**
    - Se crea un nuevo registro con el grupo asignado
    - Se preserva su grupo original en otros registros
    
    **🔄 Para repetidores (ya tienen la materia):**
    - Se actualiza su grupo actual solo para esta materia
    - Se preserva todo su historial académico
    """)

# Botón para deshacer cambios en captura de calificaciones
if menu == "📥 Captura de calificaciones" and "respaldo_captura" in st.session_state:
    st.markdown("---")
    if st.button("🔙 Deshacer último cambio en calificaciones", type="secondary"):
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM calificaciones")
        for _, row in st.session_state["respaldo_captura"].iterrows():
            valores = [row.get(col, None) for col in COLUMNAS_DB]
            cursor.execute(f"""
                INSERT INTO calificaciones ({', '.join(COLUMNAS_DB)})
                VALUES ({', '.join(['?'] * len(COLUMNAS_DB))})
            """, valores)
        conn.commit()
        conn.close()
        st.success("✅ Último cambio en calificaciones revertido.")
        del st.session_state["respaldo_captura"]
        st.rerun()

elif menu == "📊 Reportes":
    st.subheader("📊 Reportes de calificaciones")
    df = cargar_datos_db()
    df["calificacion"] = pd.to_numeric(df["calificacion"], errors="coerce")
    df["fecha_ingreso_materia"] = pd.to_datetime(df["fecha_ingreso_materia"], errors="coerce")

    fechas = sorted(df["fecha_ingreso_materia"].dropna().dt.strftime("%Y-%m-%d").unique().tolist())
    fecha_objetivo = st.selectbox("📅 Selecciona la fecha de ingreso a materia", fechas)
    df = df[df["fecha_ingreso_materia"].dt.strftime("%Y-%m-%d") == fecha_objetivo]
    grupos = sorted(df["grupo"].dropna().unique().tolist())
    grupo_objetivo = st.selectbox("👥 Selecciona el grupo", ["Todos"] + grupos)
    if grupo_objetivo != "Todos":
        df = df[df["grupo"] == grupo_objetivo]

    df = df[df["fecha_ingreso_materia"].dt.strftime("%Y-%m-%d") == fecha_objetivo]

    carreras = sorted(df["carrera"].dropna().unique().tolist())
    carrera_objetivo = st.selectbox("🎓 Selecciona la carrera", ["Todas"] + carreras)
    if carrera_objetivo != "Todas":
        df = df[df["carrera"] == carrera_objetivo]

    materias = sorted(df["materia"].dropna().unique().tolist())
    materia_objetivo = st.selectbox("📘 Selecciona la materia", ["Todas"] + materias)
    if materia_objetivo != "Todas":
        df = df[df["materia"] == materia_objetivo]

    profesores = sorted(df["profesor"].dropna().unique().tolist())
    profesor_objetivo = st.selectbox("👨‍🏫 Selecciona el profesor", ["Todos"] + profesores)
    if profesor_objetivo != "Todos":
        df = df[df["profesor"] == profesor_objetivo]

    columnas_visibles = [
        "matricula", "nombre", "grupo", "materia", "calificacion", "fecha_ingreso_original",
        "estatus", "profesor", "origen_asignacion", "fecha_calificacion",
        "n_recursamientos", "cuatrimestre", "carrera", "email_personal", "usuario_email_nova", "fecha_ingreso_materia"
    ]
    columnas_existentes = [col for col in columnas_visibles if col in df.columns]

    incluir_vacios = st.checkbox("🔄 Incluir registros sin calificación (vista general)", value=False)

    if not incluir_vacios and "calificacion" in df.columns:
        df = df[df["calificacion"].notnull()]

    st.markdown("### Vista general de todos los registros filtrados")
    st.dataframe(df[columnas_existentes], use_container_width=True)

    filtro = st.radio("Selecciona el tipo de reporte", ["Aprobados", "Reprobados", "Recursamiento", "Aprobados + Recursando (misma fecha)"])

    if filtro == "Aprobados":
        df_reporte = df[df["calificacion"] >= 6]
    elif filtro == "Reprobados":
        df_reporte = df[(df["calificacion"] < 6) & (df["origen_asignacion"] != "recursamiento")]
    elif filtro == "Recursamiento":
        df_reporte = df[
          (df["tipo_asignacion"] == "recursamiento") | 
          (df["origen_asignacion"] == "recursamiento") | 
          (df["estatus"] == "recursando")
    ]
    elif filtro == "Aprobados + Recursando (misma fecha)":
        df_reporte = df[df["estatus"].isin(["aprobado", "recursando"])]

    st.write(f"Total encontrados: {len(df_reporte)}")
    st.dataframe(df_reporte[columnas_existentes], use_container_width=True)

    nombre_archivo = f"reporte_{filtro.lower().replace(' ', '_')}_{fecha_objetivo}"
    if carrera_objetivo != "Todas":
        nombre_archivo += f"_{carrera_objetivo}"
    if materia_objetivo != "Todas":
        nombre_archivo += f"_{materia_objetivo}"
    if profesor_objetivo != "Todos":
        nombre_archivo += f"_{profesor_objetivo}"
    nombre_archivo += ".xlsx"

    if not df_reporte.empty:
        import io
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_reporte[columnas_existentes].to_excel(writer, index=False)
        output.seek(0)

        st.download_button(
            label="📅 Descargar reporte Excel",
            data=output,
            file_name=nombre_archivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ========= MÓDULO MEJORADO: GESTIÓN DE RECURSAMIENTO =========
elif menu == "📌 Gestión de recursamiento":
    st.subheader("📌 Gestión de alumnos para recursamiento")
    
    # Crear tabs para organizar el flujo
    tab1, tab2, tab3 = st.tabs(["🔍 Paso 1: Buscar Reprobados", "📋 Paso 2: Seleccionar Alumnos", "✅ Paso 3: Confirmar Recursamiento"])
    
    df = cargar_datos_db()
    df["calificacion"] = pd.to_numeric(df["calificacion"], errors="coerce")
    df = df[df["calificacion"].notnull()]
    
    # MEJORADO: Buscar TODOS los alumnos reprobados (incluyendo múltiples reprobaciones)
    df_reprobados = df[
        (df["calificacion"] < 6) & 
        (df["estatus"] != "recursando")  # Solo excluir los que están actualmente recursando (sin calificación)
    ]
    
    # MEJORADO: Manejo inteligente de múltiples reprobaciones
    # Opción 1: Mostrar solo el registro más reciente por alumno-materia
    df_reprobados_unicos = df_reprobados.sort_values(['matricula', 'materia', 'fecha_ingreso_materia'], ascending=[True, True, False])
    df_reprobados_unicos = df_reprobados_unicos.drop_duplicates(subset=['matricula', 'materia'], keep='first')

    with tab1:
        st.markdown("### 🔍 Buscar Alumnos Reprobados")
        
        # NUEVO: Selector de modo de visualización
        modo_visualizacion = st.radio(
            "📊 ¿Cómo mostrar alumnos con múltiples reprobaciones?",
            [
                "Solo registro más reciente por materia",
                "Todos los registros (incluyendo múltiples reprobaciones)",
                "Solo alumnos con múltiples reprobaciones"
            ],
            help="Elige cómo manejar alumnos que han reprobado la misma materia múltiples veces"
        )
        
        # Aplicar el modo seleccionado
        if modo_visualizacion == "Solo registro más reciente por materia":
            df_reprobados_final = df_reprobados_unicos
        elif modo_visualizacion == "Todos los registros (incluyendo múltiples reprobaciones)":
            df_reprobados_final = df_reprobados
        else:  # Solo múltiples reprobaciones
            # Identificar alumnos con múltiples reprobaciones en la misma materia
            conteo_reprobaciones = df_reprobados.groupby(['matricula', 'materia']).size()
            multiples_reprobaciones = conteo_reprobaciones[conteo_reprobaciones > 1]
            
            if not multiples_reprobaciones.empty:
                matriculas_multiples = []
                materias_multiples = []
                for (matricula, materia), count in multiples_reprobaciones.items():
                    matriculas_multiples.append(matricula)
                    materias_multiples.append(materia)
                
                df_reprobados_final = df_reprobados[
                    (df_reprobados['matricula'].isin(matriculas_multiples)) &
                    (df_reprobados['materia'].isin(materias_multiples))
                ]
            else:
                df_reprobados_final = pd.DataFrame()  # Vacío si no hay múltiples reprobaciones
        
        if df_reprobados_final.empty:
            if modo_visualizacion == "Solo alumnos con múltiples reprobaciones":
                st.success("✅ No hay alumnos con múltiples reprobaciones en la misma materia.")
            else:
                st.success("✅ No hay alumnos reprobados pendientes de recursar.")
            st.info("🎉 ¡Excelente! Todos los alumnos han aprobado o ya están en recursamiento.")
        else:
            # Métricas visuales MEJORADAS
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("📊 Total Registros", len(df_reprobados_final))
            
            with col2:
                alumnos_unicos = df_reprobados_final['matricula'].nunique()
                st.metric("👥 Alumnos Únicos", alumnos_unicos)
            
            with col3:
                materias_reprobados = df_reprobados_final['materia'].nunique()
                st.metric("📚 Materias", materias_reprobados)
            
            with col4:
                promedio_reprobados = df_reprobados_final['calificacion'].mean()
                st.metric("📉 Promedio", f"{promedio_reprobados:.1f}")
            
            # Filtros principales
            st.markdown("#### 🎯 Aplicar Filtros")
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                carreras = ["Todos"] + sorted(df_reprobados_final["carrera"].dropna().unique().tolist())
                carrera_filtro = st.selectbox("🎓 Filtrar por carrera", carreras, key="recursamiento_carrera")
            
            with col2:
                fechas_ingreso_original = ["Todos"] + sorted(df_reprobados_final["fecha_ingreso_original"].dropna().unique().tolist())
                fecha_ingreso_filtro = st.selectbox("📅 Filtrar por fecha de ingreso original", fechas_ingreso_original, key="recursamiento_fecha_ingreso")
            
            with col3:
                fechas_ingreso_materia = ["Todos"] + sorted(df_reprobados_final["fecha_ingreso_materia"].dropna().unique().tolist())
                fecha_materia_filtro = st.selectbox("📚 Filtrar por fecha de ingreso a materia", fechas_ingreso_materia, key="recursamiento_fecha_materia")
            
            with col4:
                materias = ["Todos"] + sorted(df_reprobados_final["materia"].dropna().unique().tolist())
                materia_filtro = st.selectbox("📘 Filtrar por materia", materias, key="recursamiento_materia")

            # Aplicar filtros
            df_filtrado = df_reprobados_final.copy()
            if carrera_filtro != "Todos":
                df_filtrado = df_filtrado[df_filtrado["carrera"] == carrera_filtro]
            if fecha_ingreso_filtro != "Todos":
                df_filtrado = df_filtrado[df_filtrado["fecha_ingreso_original"] == fecha_ingreso_filtro]
            if fecha_materia_filtro != "Todos":
                df_filtrado = df_filtrado[df_filtrado["fecha_ingreso_materia"] == fecha_materia_filtro]
            if materia_filtro != "Todos":
                df_filtrado = df_filtrado[df_filtrado["materia"] == materia_filtro]

            if df_filtrado.empty:
                st.warning("⚠️ No hay registros que coincidan con los filtros seleccionados.")
                st.info("💡 Prueba con filtros menos restrictivos o cambia el modo de visualización.")
            else:
                st.success(f"✅ Se encontraron {len(df_filtrado)} registros con los filtros aplicados.")
                
                # Guardar datos filtrados en session state
                st.session_state['df_reprobados_filtrado'] = df_filtrado
                
                # Vista previa MEJORADA
                with st.expander("👀 Vista previa de alumnos reprobados", expanded=True):
                    columnas_preview = ['matricula', 'nombre', 'carrera', 'materia', 'grupo', 'calificacion', 
                                      'fecha_ingreso_materia', 'fecha_ingreso_original', 'estatus', 'tipo_asignacion', 'n_recursamientos']
                    columnas_existentes = [col for col in columnas_preview if col in df_filtrado.columns]
                    
                    # Ordenar por matrícula y fecha para mejor visualización
                    df_preview = df_filtrado[columnas_existentes].sort_values(['matricula', 'fecha_ingreso_materia'], ascending=[True, False])
                    st.dataframe(df_preview, use_container_width=True)

    with tab2:
        st.markdown("### 📋 Seleccionar Alumnos para Recursamiento")
        
        if 'df_reprobados_filtrado' not in st.session_state:
            st.info("👈 Primero ve al Paso 1 para buscar alumnos reprobados.")
        else:
            df_filtrado = st.session_state['df_reprobados_filtrado'].copy()
            
            # Fecha de recursamiento destacada
            st.markdown("""
            <div style="background: linear-gradient(90deg, #ff6b6b, #ffa726); padding: 20px; border-radius: 10px; margin: 20px 0;">
                <h3 style="color: white; margin: 0; text-align: center;">📅 FECHA DE INGRESO PARA RECURSAMIENTO</h3>
                <p style="color: white; margin: 5px 0 0 0; text-align: center;">Esta será la nueva fecha de inicio para los alumnos seleccionados</p>
            </div>
            """, unsafe_allow_html=True)
            
            fecha_recursamiento = st.date_input("Selecciona la fecha de recursamiento", key="fecha_recursamiento_input")
            
            if fecha_recursamiento:
                st.markdown("#### 👥 Seleccionar Alumnos")
                st.info("💡 Selecciona los alumnos que confirman su deseo de recursar la materia.")
                
                # Resetear índices y crear un mapeo
                df_filtrado = df_filtrado.reset_index(drop=True)
                df_filtrado['idx_original'] = df_filtrado.index
                
                # Crear checkboxes organizados
                seleccionados = []
                
                # Agrupar por carrera para mejor organización
                carreras_unicas = df_filtrado['carrera'].unique()
                
                for carrera in carreras_unicas:
                    st.markdown(f"**🎓 {carrera}**")
                    df_carrera = df_filtrado[df_filtrado['carrera'] == carrera]
                    
                    # Checkbox para seleccionar todos de la carrera
                    select_all_carrera = st.checkbox(f"Seleccionar todos de {carrera}", key=f"select_all_{carrera}")
                    
                    for idx in df_carrera.index:
                        row = df_carrera.loc[idx]
                        # Determinar si debe estar seleccionado
                        default_selected = select_all_carrera
                        
                        selected = st.checkbox(
                            f"📝 {row['matricula']} - {row['nombre']} | {row['materia']} | Calif: {row['calificacion']}",
                            value=default_selected,
                            key=f"alumno_{idx}"
                        )
                        if selected:
                            seleccionados.append(idx)
                
                # Guardar seleccionados en session state
                if seleccionados:
                    st.session_state['alumnos_seleccionados'] = seleccionados
                    st.session_state['fecha_recursamiento'] = fecha_recursamiento
                    st.session_state['df_filtrado_para_recursamiento'] = df_filtrado
                    
                    st.success(f"✅ Has seleccionado {len(seleccionados)} alumnos para recursamiento.")
                    
                    # Vista previa de seleccionados
                    with st.expander("👀 Vista previa de alumnos seleccionados", expanded=True):
                        df_seleccionados = df_filtrado.loc[seleccionados]
                        st.dataframe(
                            df_seleccionados[['matricula', 'nombre', 'carrera', 'materia', 'calificacion']],
                            use_container_width=True
                        )

    with tab3:
        st.markdown("### ✅ Confirmar y Procesar Recursamiento")
        
        if ('alumnos_seleccionados' not in st.session_state or 
            'fecha_recursamiento' not in st.session_state or 
            'df_filtrado_para_recursamiento' not in st.session_state or 
            len(st.session_state.get('alumnos_seleccionados', [])) == 0):
            st.info("👈 Primero completa los pasos 1 y 2 para seleccionar alumnos.")
        else:
            df_filtrado = st.session_state['df_filtrado_para_recursamiento']
            seleccionados = st.session_state['alumnos_seleccionados']
            fecha_recursamiento = st.session_state['fecha_recursamiento']
            
            # Resumen final
            st.markdown("""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 20px; border-radius: 10px; margin: 20px 0;">
                <h3 style="color: white; margin: 0; text-align: center;">📋 RESUMEN FINAL</h3>
            </div>
            """, unsafe_allow_html=True)
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("👥 Alumnos Seleccionados", len(seleccionados))
            with col2:
                st.metric("📅 Fecha Recursamiento", fecha_recursamiento.strftime("%d/%m/%Y"))
            with col3:
                # Validar que los índices existen en el DataFrame
                indices_validos = [idx for idx in seleccionados if idx in df_filtrado.index]
                if indices_validos:
                    carreras_afectadas = df_filtrado.loc[indices_validos]['carrera'].nunique()
                    st.metric("🎓 Carreras Afectadas", carreras_afectadas)
                else:
                    st.metric("🎓 Carreras Afectadas", 0)
            
            # Tabla de confirmación - usar solo índices válidos
            st.markdown("#### 📝 Alumnos que serán registrados en recursamiento:")
            indices_validos = [idx for idx in seleccionados if idx in df_filtrado.index]
            
            if indices_validos:
                df_confirmacion = df_filtrado.loc[indices_validos].copy()
                df_confirmacion['nuevo_estatus'] = 'recursando'
                df_confirmacion['nueva_fecha'] = fecha_recursamiento.strftime("%Y-%m-%d")
                
                st.dataframe(
                    df_confirmacion[['matricula', 'nombre', 'carrera', 'materia', 'calificacion', 'nueva_fecha', 'nuevo_estatus']],
                    use_container_width=True
                )
                
                # Confirmación final
                st.markdown("#### ⚠️ Confirmación Final")
                st.warning("Esta acción creará nuevos registros de recursamiento en la base de datos. ¿Estás seguro?")
                
                confirmacion = st.checkbox("✅ Confirmo que quiero proceder con el registro de recursamiento")
                
                if confirmacion and st.button("💾 PROCESAR RECURSAMIENTO", type="primary"):
                    try:
                        conn = sqlite3.connect(DB_FILE)
                        cursor = conn.cursor()
                        nuevos_registros = []
                        
                        for i in indices_validos:
                            row = df_filtrado.loc[i]
                            nuevos_registros.append({
                                "matricula": row["matricula"],
                                "nombre": row["nombre"],
                                "grupo": row["grupo"],
                                "materia": row["materia"],
                                "fecha_ingreso_materia": fecha_recursamiento.strftime("%Y-%m-%d"),
                                "fecha_ingreso_original": row.get("fecha_ingreso_original", row["fecha_ingreso_materia"]),
                                "fecha_recursamiento": fecha_recursamiento.strftime("%Y-%m-%d"),
                                "cuatrimestre": row.get("cuatrimestre", ""),
                                "carrera": row["carrera"],
                                "email_personal": row["email_personal"],
                                "usuario_email_nova": row.get("usuario_email_nova", ""),
                                "contraseña": row.get("contraseña", ""),
                                "tipo_asignacion": "recursamiento",
                                "n_recursamientos": int(0 if pd.isna(row.get("n_recursamientos")) else row["n_recursamientos"]) + 1,
                                "calificacion": None,
                                "fecha_calificacion": None,
                                "profesor": None,
                                "estatus": "recursando",
                                "origen_asignacion": "recursamiento"
                            })

                        for nuevo in nuevos_registros:
                            valores = [nuevo.get(col, None) for col in COLUMNAS_DB]
                            cursor.execute(f"""
                                INSERT INTO calificaciones ({', '.join(COLUMNAS_DB)})
                                VALUES ({', '.join(['?'] * len(COLUMNAS_DB))})
                            """, valores)
                        
                        conn.commit()
                        conn.close()

                        st.success(f"✅ {len(nuevos_registros)} alumnos registrados exitosamente en recursamiento.")
                        
                        # Limpiar session state
                        keys_to_delete = ['df_reprobados_filtrado', 'alumnos_seleccionados', 'fecha_recursamiento', 'df_filtrado_para_recursamiento']
                        for key in keys_to_delete:
                            if key in st.session_state:
                                del st.session_state[key]
                        
                        # Generar Excel para descarga
                        df_excel = pd.DataFrame(nuevos_registros)
                        import io
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            df_excel.to_excel(writer, index=False)
                        output.seek(0)

                        st.download_button(
                            label="📥 Descargar Excel de recursantes",
                            data=output,
                            file_name="alumnos_recursamiento.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                        # Mostrar mensaje de éxito con animación
                        st.balloons()
                        
                    except Exception as e:
                        st.error(f"❌ Error al procesar recursamiento: {str(e)}")
                        st.write("Detalles del error:", e)
            else:
                st.error("❌ No hay registros válidos seleccionados. Por favor, vuelve al paso 2.")

elif menu == "🎓 Promoción de cuatrimestre":
	ejecutar_promocion_cuatrimestre_masiva()
	

# ========= MÓDULO COMPLETO: HISTORIAL POR ALUMNO CON ORGANIZACIÓN POR CUATRIMESTRES =========
# ========= MÓDULO COMPLETO: HISTORIAL POR ALUMNO CON ORGANIZACIÓN POR CUATRIMESTRES =========
elif menu == "🧾 Historial por alumno":
    st.subheader("🧾 Historial académico por alumno")

    df = cargar_datos_db()
    if df.empty:
        st.warning("No hay datos disponibles.")
    else:
        # Barra de búsqueda mejorada
        col1, col2 = st.columns([3, 1])
        with col1:
            busqueda = st.text_input("🔍 Buscar por nombre o matrícula", placeholder="Escribe el nombre o matrícula del estudiante...").strip().lower()
        with col2:
            buscar_exacto = st.checkbox("Búsqueda exacta", help="Buscar coincidencia exacta en lugar de contiene")

        if busqueda:
            # Aplicar búsqueda
            if buscar_exacto:
                df_filtrado = df[
                    (df["nombre"].str.lower() == busqueda) | 
                    (df["matricula"].str.lower() == busqueda)
                ]
            else:
                df_filtrado = df[
                    df["nombre"].str.lower().str.contains(busqueda, na=False) | 
                    df["matricula"].str.lower().str.contains(busqueda, na=False)
                ]

            if not df_filtrado.empty:
                # Mostrar estudiantes encontrados
                estudiantes_unicos = df_filtrado[['matricula', 'nombre', 'carrera']].drop_duplicates()
                
                if len(estudiantes_unicos) > 1:
                    st.markdown("### 👥 Estudiantes encontrados:")
                    estudiante_seleccionado = st.selectbox(
                        "Selecciona un estudiante:",
                        options=estudiantes_unicos.index,
                        format_func=lambda x: f"{estudiantes_unicos.loc[x, 'matricula']} - {estudiantes_unicos.loc[x, 'nombre']} ({estudiantes_unicos.loc[x, 'carrera']})"
                    )
                    matricula_seleccionada = estudiantes_unicos.loc[estudiante_seleccionado, 'matricula']
                    df_estudiante = df_filtrado[df_filtrado['matricula'] == matricula_seleccionada]
                else:
                    df_estudiante = df_filtrado[
                      (df_filtrado['materia'].notna()) & 
                      (df_filtrado['materia'].str.strip() != '') &
                      (df_filtrado['materia'] != 'None')
                       ]

                # ========= VALIDACIÓN CRÍTICA: VERIFICAR SI TIENE MATERIAS CON CALIFICACIONES =========
                # Verificar si el estudiante tiene registros con materias válidas
                materias_validas = df_estudiante[
                    (df_estudiante['materia'].notna()) & 
                    (df_estudiante['materia'].str.strip() != '') &
                    (df_estudiante['materia'] != 'None')
                ]
                
                # Si no hay materias válidas, mostrar mensaje y terminar
                if materias_validas.empty:
                    # Obtener información básica del estudiante desde df_filtrado
                    estudiante_info = df_filtrado.iloc[0] if not df_filtrado.empty else None
                    
                    if estudiante_info is not None:
                        nombre = estudiante_info.get('nombre', 'Sin nombre') if pd.notna(estudiante_info.get('nombre')) else 'Sin nombre'
                        matricula = estudiante_info.get('matricula', 'Sin matrícula') if pd.notna(estudiante_info.get('matricula')) else 'Sin matrícula'
                        carrera = estudiante_info.get('carrera', 'Sin carrera') if pd.notna(estudiante_info.get('carrera')) else 'Sin carrera'
                        
                        # Header del estudiante sin calificaciones
                        st.markdown(f"""
                        <div style="background: linear-gradient(135deg, #ff7675 0%, #d63031 100%); padding: 20px; border-radius: 10px; margin: 20px 0;">
                            <h2 style="color: white; margin: 0;">👤 {nombre}</h2>
                            <p style="color: #fab1a0; margin: 5px 0 0 0;">
                                📊 Matrícula: {matricula} | 🎓 {carrera}
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        # Mensaje informativo
                        st.info(f"""
                        ℹ️ **Estudiante encontrado pero sin historial académico**
                        
                        El estudiante **{nombre}** (Matrícula: {matricula}) está registrado en el sistema, 
                        pero no tiene materias con calificaciones registradas.
                        
                        **Posibles causas:**
                        • Es un estudiante de nuevo ingreso sin materias cursadas
                        • Los registros académicos aún no han sido capturados
                        • Solo tiene registros administrativos sin calificaciones
                        
                        **¿Qué puedes hacer?**
                        • Verificar si las calificaciones han sido capturadas en el sistema
                        • Consultar con el departamento académico
                        • Intentar buscar con una variante del nombre o matrícula
                        """)
                        
                        # Mostrar los datos básicos que sí existen
                        if not df_filtrado.empty:
                            st.markdown("### 📋 Información disponible:")
                            columnas_basicas = ['matricula', 'nombre', 'carrera', 'cuatrimestre', 'fecha_ingreso_materia']
                            columnas_existentes_basicas = [col for col in columnas_basicas if col in df_filtrado.columns]
                            
                            if columnas_existentes_basicas:
                                df_basico = df_filtrado[columnas_existentes_basicas].drop_duplicates()
                                st.dataframe(df_basico, use_container_width=True)
                        
                        # Sugerencias de búsqueda
                        st.markdown("### 💡 Sugerencias:")
                        st.write("• Intenta buscar con variaciones del nombre (apellidos, nombres)")
                        st.write("• Verifica que la matrícula esté escrita correctamente")
                        st.write("• Contacta al administrador del sistema si el problema persiste")
                        
                    else:
                        st.error("❌ No se pudo obtener información del estudiante.")
                    
                    # Terminar la ejecución aquí para estudiantes sin calificaciones
                    st.stop()

                # ========= CONTINÚA CON EL CÓDIGO NORMAL SOLO SI HAY MATERIAS =========
                # Si llegamos aquí, significa que sí hay materias válidas
                df_estudiante = materias_validas  # Usar solo las materias válidas

                # Información del estudiante (ahora sabemos que df_estudiante no está vacío)
                info_estudiante = df_estudiante.iloc[0]
                
                # Validar datos del estudiante
                nombre_estudiante = info_estudiante.get('nombre', 'Sin nombre') if pd.notna(info_estudiante.get('nombre')) else 'Sin nombre'
                matricula_estudiante = info_estudiante.get('matricula', 'Sin matrícula') if pd.notna(info_estudiante.get('matricula')) else 'Sin matrícula'
                carrera_estudiante = info_estudiante.get('carrera', 'Sin carrera') if pd.notna(info_estudiante.get('carrera')) else 'Sin carrera'
                
                # Header del estudiante
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #74b9ff 0%, #0984e3 100%); padding: 20px; border-radius: 10px; margin: 20px 0;">
                    <h2 style="color: white; margin: 0;">👤 {nombre_estudiante}</h2>
                    <p style="color: #dff9fb; margin: 5px 0 0 0;">
                        📊 Matrícula: {matricula_estudiante} | 🎓 {carrera_estudiante}
                    </p>
                </div>
                """, unsafe_allow_html=True)

                # FUNCIÓN: Determinar cuatrimestre histórico basado en materias aprobadas
                def calcular_cuatrimestre_por_materias_aprobadas(df_estudiante_completo):
                    """
                    Calcula el cuatrimestre histórico basado en materias aprobadas acumuladas
                    Regla: Cada 5 materias aprobadas = 1 cuatrimestre
                    """
                    
                    # Ordenar por fecha de ingreso para calcular progresión
                    df_ordenado = df_estudiante_completo.copy()
                    df_ordenado['fecha_ingreso_materia'] = pd.to_datetime(df_ordenado['fecha_ingreso_materia'], errors='coerce')
                    df_ordenado = df_ordenado.sort_values('fecha_ingreso_materia', ascending=True)
                    
                    # Filtrar solo materias válidas
                    df_ordenado = df_ordenado[
                        (df_ordenado['materia'].notna()) & 
                        (df_ordenado['materia'] != '') & 
                        (df_ordenado['materia'] != 'None')
                    ]
                    
                    cuatrimestres_calculados = []
                    materias_aprobadas_acumuladas = 0
                    
                    for idx, row in df_ordenado.iterrows():
                        # Si ya tiene cuatrimestre_historico definido y no está vacío, usarlo
                        if pd.notna(row.get("cuatrimestre_historico")) and str(row.get("cuatrimestre_historico")).strip():
                            cuatrimestre_asignado = str(row["cuatrimestre_historico"]).strip()
                        else:
                            # Verificar si la materia está aprobada
                            calificacion = pd.to_numeric(row.get('calificacion', 0), errors='coerce')
                            estatus = row.get('estatus', '').lower()
                            
                            # Considerar aprobada si calificación >= 6 O estatus es 'aprobado'
                            materia_aprobada = (calificacion >= 6) or (estatus == 'aprobado')
                            
                            if materia_aprobada:
                                materias_aprobadas_acumuladas += 1
                            
                            # Calcular cuatrimestre basado en materias aprobadas acumuladas
                            # Fórmula: cuatrimestre = (materias_aprobadas // 5) + 1
                            cuatrimestre_calculado = (materias_aprobadas_acumuladas // 5) + 1
                            cuatrimestre_asignado = str(cuatrimestre_calculado)
                        
                        cuatrimestres_calculados.append(cuatrimestre_asignado)
                    
                    # Asignar los cuatrimestres calculados de vuelta al DataFrame original
                    df_ordenado['cuatrimestre_display'] = cuatrimestres_calculados
                    
                    # Crear un mapeo para asignar a todo el DataFrame original
                    mapeo_cuatrimestres = dict(zip(df_ordenado.index, df_ordenado['cuatrimestre_display']))
                    
                    return mapeo_cuatrimestres

                # Calcular cuatrimestres basado en materias aprobadas
                try:
                    mapeo_cuatrimestres = calcular_cuatrimestre_por_materias_aprobadas(df_estudiante)
                    df_estudiante['cuatrimestre_display'] = df_estudiante.index.map(mapeo_cuatrimestres)
                    
                    # Rellenar valores nulos con cuatrimestre 1
                    df_estudiante['cuatrimestre_display'] = df_estudiante['cuatrimestre_display'].fillna('1')
                    
                    # Mostrar información del cálculo
                    total_aprobadas = len(df_estudiante[
                        (pd.to_numeric(df_estudiante['calificacion'], errors='coerce') >= 6) |
                        (df_estudiante['estatus'].str.lower() == 'aprobado')
                    ])
                    cuatrimestre_calculado_actual = (total_aprobadas // 5) + 1
                    
                    st.info(f"""
                    📊 **Cálculo de Cuatrimestre:**
                    • Materias aprobadas totales: {total_aprobadas}
                    • Cuatrimestre actual calculado: {cuatrimestre_calculado_actual}
                    • Regla: Cada 5 materias aprobadas = 1 cuatrimestre
                    """)
                    
                except Exception as e:
                    st.warning(f"Error al calcular cuatrimestres por materias aprobadas: {e}")
                    # Fallback: usar cuatrimestre actual del registro
                    df_estudiante['cuatrimestre_display'] = df_estudiante.get('cuatrimestre', 1).astype(str)

                # ==================== RESTO DEL CÓDIGO CONTINÚA IGUAL ====================
                # [El resto del código permanece sin cambios...]

                # ==================== SECCIÓN DE EXPORTACIÓN MEJORADA ====================
                st.markdown("---")

                # Título centrado
                st.markdown("""
                <div style="text-align: center; margin-bottom: 20px;">
                    <h3 style="color: #FDFD96; margin: 0;">📊 Información del Estudiante y Exportación</h3>
                </div>
                """, unsafe_allow_html=True)

                col_info, col_export = st.columns([2, 1])

                with col_info:
                    # Contenedor centrado para las subcolunas
                    subcol1, subcol2, subcol3 = st.columns(3)

                    with subcol1:
                        st.markdown(f"""
                        <div style="text-align: center; padding: 8px;">
                            <div style="font-size: 15px; color: #FDFD96; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 3px;">👤 Matrícula</div>
                            <div style="font-size: 20px; font-weight: 500; color: #888;">{matricula_estudiante}</div>
                        </div>
                        """, unsafe_allow_html=True)

                    with subcol2:
                        st.markdown(f"""
                        <div style="text-align: center; padding: 8px;">
                            <div style="font-size: 15px; color: #FDFD96; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 3px;">🎓 Carrera</div>
                            <div style="font-size: 14px; font-weight: 500; color: #888;">{carrera_estudiante}</div>
                        </div>
                        """, unsafe_allow_html=True)

                    with subcol3:
                        cuatrimestre_actual = info_estudiante.get('cuatrimestre', 'N/A')
                        st.markdown(f"""
                        <div style="text-align: center; padding: 8px;">
                            <div style="font-size: 15px; color: #FDFD96; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 3px;">📚 Cuatrimestre</div>
                            <div style="font-size: 14px; font-weight: 500; color: #888;">{cuatrimestre_actual}</div>
                        </div>
                        """, unsafe_allow_html=True)
            
                with col_export:
                    st.markdown("### 📊 Exportar")
                    
                    # Botón principal para Kardex Oficial
                    if st.button("🎓 DESCARGAR KARDEX OFICIAL", type="primary", use_container_width=True, help="Kardex con formato institucional oficial"):
                        conn = sqlite3.connect(DB_FILE)
                        
                        with st.spinner("📋 Generando Kardex Oficial..."):
                            resultado_kardex = exportar_kardex_formato_oficial(
                                matricula_estudiante, nombre_estudiante, conn
                            )

                            if resultado_kardex:
                                st.success("✅ ¡Kardex Oficial generado exitosamente!")
                        
                                # Información del kardex generado
                                stats = resultado_kardex['estadisticas']
                                st.info(f"""
                                🎯 **Kardex Oficial Generado:**
                                • Materias cursadas: {stats['materias_cursadas']}
                                • Materias aprobadas: {stats['aprobadas']}
                                • Materias reprobadas: {stats['reprobadas']}
                                • Promedio general: {stats['promedio_general']}
                                • Formato institucional oficial con logotipos
                                """)
                        
                                # Botón de descarga principal
                                st.download_button(
                                    label="📋 DESCARGAR KARDEX OFICIAL",
                                    data=resultado_kardex['buffer'],
                                    file_name=resultado_kardex['filename'],
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True,
                                    help="Documento oficial con formato institucional"
                                )
                            else:
                                st.error("❌ No se pudo generar el kardex. Revisa los datos.")

                        conn.close()
                    
                    # Separador visual
                    st.markdown("---")
                    
                    # Botón secundario para Excel por Cuatrimestre (existente)
                    if st.button("📊 Descargar Excel por Cuatrimestre", use_container_width=True, help="Formato detallado por cuatrimestres"):
                        conn = sqlite3.connect(DB_FILE)
                        
                        with st.spinner("📄 Generando Excel con detalle por cuatrimestre..."):
                            resultado_export = exportar_detalle_por_cuatrimestre_excel(
                                matricula_estudiante, nombre_estudiante, conn
                            )

                            if resultado_export:
                                st.success("✅ ¡Excel generado exitosamente!")
                        
                                st.info(f"""
                                📈 **Archivo generado:**
                                • Total de cuatrimestres: {resultado_export['total_cuatrimestres']}
                                • Total de materias: {resultado_export['total_materias']}
                                • Incluye estadísticas por cuatrimestre
                                • Una hoja por cada cuatrimestre con el detalle completo
                                """)
                        
                                # Botón de descarga secundario
                                st.download_button(
                                    label="💾 Descargar Detalle por Cuatrimestre",
                                    data=resultado_export['buffer'],
                                    file_name=resultado_export['filename'],
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True,
                                    help="Descarga el historial organizado por cuatrimestre"
                                )
                            else:
                                st.error("❌ No se pudo generar el archivo. Revisa los datos.")

                        conn.close()
                # ==================== FIN SECCIÓN DE EXPORTACIÓN ====================

                # Pestañas principales - AGREGADA PESTAÑA "Por Cuatrimestres"
                tab1, tab2, tab3, tab4 = st.tabs(["📊 Resumen General", "🎓 Por Cuatrimestres", "📈 Línea de Tiempo", "🎯 Análisis Detallado"])
                
                with tab1:
                    # Métricas del estudiante
                    col1, col2, col3, col4 = st.columns(4)
                    
                    total_materias = len(df_estudiante)
                    materias_aprobadas = len(df_estudiante[pd.to_numeric(df_estudiante['calificacion'], errors='coerce') >= 6])
                    materias_reprobadas = len(df_estudiante[pd.to_numeric(df_estudiante['calificacion'], errors='coerce') < 6])
                    en_recursamiento = len(df_estudiante[df_estudiante['estatus'] == 'recursando'])
                    
                    with col1:
                        st.metric("📚 Total Materias", total_materias)
                    with col2:
                        st.metric("✅ Aprobadas", materias_aprobadas)
                    with col3:
                        st.metric("❌ Reprobadas", materias_reprobadas)
                    with col4:
                        st.metric("🔄 En Recursamiento", en_recursamiento)
                    
                    # Promedio general
                    calificaciones_numericas = pd.to_numeric(df_estudiante['calificacion'], errors='coerce').dropna()
                    if not calificaciones_numericas.empty:
                        promedio = calificaciones_numericas.mean()
                        st.metric("🎯 Promedio General", f"{promedio:.2f}")
                    
                    # Tabla resumen
                    columnas_resumen = [
                        "materia", "grupo", "fecha_ingreso_materia", "calificacion", 
                        "estatus", "tipo_asignacion", "n_recursamientos", "profesor"
                    ]
                    columnas_existentes = [col for col in columnas_resumen if col in df_estudiante.columns]
                    
                    st.markdown("### 📋 Registro Académico Completo")
                    df_mostrar = df_estudiante[columnas_existentes].sort_values('fecha_ingreso_materia', ascending=False)
                    st.dataframe(df_mostrar, use_container_width=True)

                # ==================== NUEVA PESTAÑA: Por Cuatrimestres ====================
                with tab2:
                    st.markdown("### 🎓 Historial Organizado por Cuatrimestres")
                    
                    # Filtrar materias con calificación válida
                    materias_con_calificacion = df_estudiante[
                        (df_estudiante["materia"].notna()) & 
                        (df_estudiante["materia"] != "") &
                        (df_estudiante["calificacion"].notna())
                    ].copy()
                    
                    if not materias_con_calificacion.empty:
                        # Mostrar distribución general por cuatrimestre
                        st.markdown("#### 📊 Resumen por Cuatrimestre")
                        
                        try:
                            # Crear resumen estadístico por cuatrimestre
                            distribucion_cuatrimestre = materias_con_calificacion.groupby("cuatrimestre_display").agg({
                                "materia": "count",
                                "calificacion": [
                                    lambda x: pd.to_numeric(x, errors='coerce').mean(),
                                    lambda x: sum(pd.to_numeric(x, errors='coerce') >= 6)
                                ]
                            }).round(2)
                            
                            distribucion_cuatrimestre.columns = ["Total_Materias", "Promedio", "Aprobadas"]
                            distribucion_cuatrimestre["Reprobadas"] = distribucion_cuatrimestre["Total_Materias"] - distribucion_cuatrimestre["Aprobadas"]
                            distribucion_cuatrimestre["Tasa_Aprobacion"] = (
                                distribucion_cuatrimestre["Aprobadas"] / distribucion_cuatrimestre["Total_Materias"] * 100
                            ).round(1)
                            
                            # Mostrar tabla de resumen
                            st.dataframe(distribucion_cuatrimestre, use_container_width=True)
                            
                            # Detalle expandible por cuatrimestre
                            st.markdown("#### 📚 Detalle por Cuatrimestre")
                            
                            cuatrimestres_disponibles = sorted(
                                materias_con_calificacion["cuatrimestre_display"].unique(),
                                key=lambda x: int(x) if x.isdigit() else 999
                            )
                            
                            for cuatr in cuatrimestres_disponibles:
                                materias_cuatr = materias_con_calificacion[
                                    materias_con_calificacion["cuatrimestre_display"] == cuatr
                                ]
                                
                                # Calcular métricas del cuatrimestre
                                califs_numericas = pd.to_numeric(materias_cuatr["calificacion"], errors='coerce')
                                aprobadas = len(califs_numericas[califs_numericas >= 6])
                                promedio = califs_numericas.mean() if not califs_numericas.empty else 0
                                
                                # Determinar color según el promedio
                                if promedio >= 8.5:
                                    color_promedio = "🟢"  # Verde
                                elif promedio >= 7:
                                    color_promedio = "🟡"  # Amarillo
                                else:
                                    color_promedio = "🔴"  # Rojo
                                
                                with st.expander(f"📚 Cuatrimestre {cuatr} ({len(materias_cuatr)} materias) - Promedio: {promedio:.2f} {color_promedio}"):
                                    
                                    # Métricas del cuatrimestre
                                    col1, col2, col3, col4 = st.columns(4)
                                    with col1:
                                        st.metric("📘 Total", len(materias_cuatr))
                                    with col2:
                                        st.metric("✅ Aprobadas", aprobadas)
                                    with col3:
                                        st.metric("❌ Reprobadas", len(materias_cuatr) - aprobadas)
                                    with col4:
                                        st.metric("📊 Promedio", f"{promedio:.2f}")
                                    
                                    # Tabla de materias del cuatrimestre
                                    columnas_cuatr = ["materia", "calificacion", "estatus", "fecha_ingreso_materia", "tipo_asignacion"]
                                    if "profesor" in materias_cuatr.columns:
                                        columnas_cuatr.append("profesor")
                                    if "grupo" in materias_cuatr.columns:
                                        columnas_cuatr.append("grupo")
                                    
                                    columnas_existentes_cuatr = [col for col in columnas_cuatr if col in materias_cuatr.columns]
                                    
                                    # Ordenar por fecha dentro del cuatrimestre
                                    df_cuatr_ordenado = materias_cuatr[columnas_existentes_cuatr].sort_values(
                                        'fecha_ingreso_materia', ascending=False
                                    )
                                    
                                    st.dataframe(df_cuatr_ordenado, use_container_width=True, key=f"cuatr_{cuatr}")
                                    
                                    # Gráfico de calificaciones del cuatrimestre (si hay más de una materia)
                                    # Sección del gráfico corregida - Reemplazar la parte del gráfico en el código original
                                    # Gráfico de calificaciones del cuatrimestre (si hay más de una materia)
                                    if len(materias_cuatr) > 1:
                                        try:
                                            import plotly.express as px
                                            
                                            # Crear gráfico simple de barras
                                            califs_validas = califs_numericas.dropna()
                                            if not califs_validas.empty:
                                                materias_names = materias_cuatr[califs_numericas.notna()]["materia"].tolist()
                                                
                                                # Truncar nombres de materias si son muy largos
                                                materias_names_short = [
                                                    nombre[:20] + "..." if len(nombre) > 20 else nombre 
                                                    for nombre in materias_names
                                                ]
                                                
                                                fig = px.bar(
                                                    x=materias_names_short,
                                                    y=califs_validas.tolist(),
                                                    title=f"Calificaciones - Cuatrimestre {cuatr}",
                                                    color=califs_validas.tolist(),
                                                    color_continuous_scale="RdYlGn",
                                                    range_color=[0, 10],
                                                    labels={'x': 'Materias', 'y': 'Calificación'}
                                                )
                                                
                                                fig.update_layout(
                                                    xaxis_title="Materias",
                                                    yaxis_title="Calificación",
                                                    showlegend=False,
                                                    height=400,
                                                    yaxis=dict(range=[0, 10]),
                                                    xaxis=dict(tickangle=45)  # Rotación de etiquetas
                                                )
                                                
                                                # CORRECCIÓN: Usar update_xaxes() en lugar de update_xaxis()
                                                fig.update_xaxes(tickangle=45)
                                                
                                                # Agregar línea horizontal en 6 (línea de aprobado)
                                                fig.add_hline(
                                                    y=6, 
                                                    line_dash="dash", 
                                                    line_color="orange", 
                                                    annotation_text="Línea de aprobado (6.0)"
                                                )
                                                
                                                st.plotly_chart(fig, use_container_width=True, key=f"chart_{cuatr}")
                                                
                                        except ImportError:
                                            st.info("📊 Instala plotly para ver gráficos: `pip install plotly`")
                                        except Exception as e:
                                            st.warning(f"No se pudo generar el gráfico: {e}")
                                            # Mostrar un gráfico alternativo simple con streamlit
                                            try:
                                                st.markdown(f"#### 📊 Calificaciones - Cuatrimestre {cuatr}")
                                                
                                                # Crear datos para gráfico de barras simple de streamlit
                                                materias_names = materias_cuatr[califs_numericas.notna()]["materia"].tolist()
                                                califs_values = califs_numericas.dropna().tolist()
                                                
                                                if len(materias_names) == len(califs_values):
                                                    chart_data = pd.DataFrame({
                                                        'Materia': materias_names,
                                                        'Calificación': califs_values
                                                    })
                                                    chart_data = chart_data.set_index('Materia')
                                                    st.bar_chart(chart_data, height=300)
                                                else:
                                                    st.info("No se puede mostrar el gráfico debido a inconsistencias en los datos.")
                                                    
                                            except Exception as fallback_error:
                                                st.warning(f"Tampoco se pudo generar el gráfico alternativo: {fallback_error}")
                                                # Mostrar tabla simple como último recurso
                                                st.markdown("**📋 Resumen de calificaciones:**")
                                                for i, (materia, calif) in enumerate(zip(materias_names, califs_values)):
                                                    st.write(f"• {materia}: {calif}")
                        
                        except Exception as e:
                            st.error(f"Error al procesar datos por cuatrimestre: {e}")
                            st.info("Los datos podrían tener inconsistencias. Intenta el análisis detallado.")
                            
                            # Mostrar datos básicos como respaldo
                            st.markdown("#### 📋 Vista básica por cuatrimestre:")
                            for cuatr in sorted(materias_con_calificacion["cuatrimestre_display"].unique()):
                                materias_cuatr = materias_con_calificacion[
                                    materias_con_calificacion["cuatrimestre_display"] == cuatr
                                ]
                                st.write(f"**Cuatrimestre {cuatr}:** {len(materias_cuatr)} materias")
                                st.dataframe(materias_cuatr[["materia", "calificacion", "estatus"]], use_container_width=True)
                    else:
                        st.warning("No se encontraron materias con calificaciones válidas.")
                        st.info("💡 Verifica que existan registros con materias y calificaciones en la base de datos.")

                with tab3:
                    st.markdown("### 📈 Línea de Tiempo Académica")
                    
                    # Preparar datos para la línea de tiempo
                    df_timeline = df_estudiante.copy()
                    df_timeline['fecha_ingreso_materia'] = pd.to_datetime(df_timeline['fecha_ingreso_materia'])
                    df_timeline = df_timeline.sort_values('fecha_ingreso_materia')
                    
                    # Crear timeline visual
                    for idx, row in df_timeline.iterrows():
                        fecha = row['fecha_ingreso_materia'].strftime('%d/%m/%Y') if pd.notna(row['fecha_ingreso_materia']) else 'Sin fecha'
                        materia = row['materia'] if pd.notna(row['materia']) and row['materia'] else 'Sin materia'
                        calificacion = row['calificacion'] if pd.notna(row['calificacion']) else 'Pendiente'
                        estatus = row['estatus'] if pd.notna(row['estatus']) and row['estatus'] else 'sin_estatus'
                        tipo = row.get('tipo_asignacion', 'regular')
                        tipo = tipo if pd.notna(tipo) and tipo else 'regular'
                        cuatrimestre_display = row.get('cuatrimestre_display', '1')
                        
                        # Determinar color y emoji según el estado
                        if estatus == 'aprobado':
                            color = "#27ae60"
                            emoji = "✅"
                        elif estatus == 'reprobado':
                            color = "#e74c3c"
                            emoji = "❌"
                        elif estatus == 'recursando':
                            color = "#f39c12"
                            emoji = "🔄"
                        else:
                            color = "#95a5a6"
                            emoji = "⏳"
                        
                        # Tipo de asignación
                        tipo_emoji = "🔁" if tipo == "recursamiento" else "📚"
                        tipo_texto = str(tipo).title() if tipo else "Regular"
                        
                        # Valores seguros para mostrar
                        grupo_texto = row.get('grupo', 'N/A') if pd.notna(row.get('grupo')) else 'N/A'
                        profesor_texto = row.get('profesor', 'N/A') if pd.notna(row.get('profesor')) and row.get('profesor') else 'N/A'
                        
                        # Crear card de timeline con información de cuatrimestre
                        st.markdown(f"""
                        <div style="border-left: 4px solid {color}; padding: 15px; margin: 10px 0; background: #f8f9fa; border-radius: 5px;">
                            <div style="display: flex; justify-content: space-between; align-items: center;">
                                <h4 style="margin: 0; color: {color};">{emoji} {materia}</h4>
                                <span style="background: {color}; color: white; padding: 5px 10px; border-radius: 15px; font-size: 12px;">
                                    {fecha}
                                </span>
                            </div>
                            <p style="margin: 5px 0; color: #666;">
                                🎓 Cuatrimestre: {cuatrimestre_display} | 
                                {tipo_emoji} {tipo_texto} | 
                                Grupo: {grupo_texto} | 
                                Calificación: {calificacion}
                            </p>
                            {f"<p style='margin: 0; color: #666; font-size: 12px;'>👨‍🏫 Profesor: {profesor_texto}</p>" if profesor_texto != 'N/A' else ""}
                        </div>
                        """, unsafe_allow_html=True)

                with tab4:
                    st.markdown("### 🎯 Análisis Detallado")
                    
                    # Filtros avanzados
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        tipos_disponibles = ["Todos"] + sorted(df_estudiante["tipo_asignacion"].dropna().unique().tolist())
                        tipo_filtro = st.selectbox("📚 Tipo de asignación", tipos_disponibles)
                    
                    with col2:
                        estatus_disponibles = ["Todos"] + sorted(df_estudiante["estatus"].dropna().unique().tolist())
                        estatus_filtro = st.selectbox("📊 Estatus", estatus_disponibles)
                    
                    with col3:
                        cuatrimestres_disponibles = ["Todos"] + sorted(df_estudiante["cuatrimestre_display"].dropna().unique().tolist())
                        cuatrimestre_filtro = st.selectbox("🎓 Cuatrimestre", cuatrimestres_disponibles)
                    
                    with col4:
                        # Nuevo filtro por rango de calificaciones
                        rango_califs = st.select_slider(
                            "📊 Rango de Calificaciones",
                            options=["Todas", "0-5", "6-7", "8-9", "10"],
                            value="Todas"
                        )
                    
                    # Aplicar filtros
                    df_analisis = df_estudiante.copy()
                    if tipo_filtro != "Todos":
                        df_analisis = df_analisis[df_analisis["tipo_asignacion"] == tipo_filtro]
                    if estatus_filtro != "Todos":
                        df_analisis = df_analisis[df_analisis["estatus"] == estatus_filtro]
                    if cuatrimestre_filtro != "Todos":
                        df_analisis = df_analisis[df_analisis["cuatrimestre_display"] == cuatrimestre_filtro]
                    
                    # Filtro por rango de calificaciones
                    if rango_califs != "Todas":
                        califs_numericas = pd.to_numeric(df_analisis["calificacion"], errors='coerce')
                        if rango_califs == "0-5":
                            df_analisis = df_analisis[califs_numericas < 6]
                        elif rango_califs == "6-7":
                            df_analisis = df_analisis[(califs_numericas >= 6) & (califs_numericas < 8)]
                        elif rango_califs == "8-9":
                            df_analisis = df_analisis[(califs_numericas >= 8) & (califs_numericas < 10)]
                        elif rango_califs == "10":
                            df_analisis = df_analisis[califs_numericas == 10]
                    
                    # Rango de fechas
                    df_analisis["fecha_ingreso_materia"] = pd.to_datetime(df_analisis["fecha_ingreso_materia"], errors="coerce")
                    if not df_analisis["fecha_ingreso_materia"].dropna().empty:
                        fecha_min = df_analisis["fecha_ingreso_materia"].min()
                        fecha_max = df_analisis["fecha_ingreso_materia"].max()
                        fecha_rango = st.date_input(
                            "📅 Rango de fechas",
                            value=[fecha_min, fecha_max],
                            min_value=fecha_min,
                            max_value=fecha_max
                        )
                        
                        if isinstance(fecha_rango, (list, tuple)) and len(fecha_rango) == 2:
                            fecha_inicio, fecha_fin = fecha_rango
                            df_analisis = df_analisis[
                                (df_analisis["fecha_ingreso_materia"] >= pd.to_datetime(fecha_inicio)) &
                                (df_analisis["fecha_ingreso_materia"] <= pd.to_datetime(fecha_fin))
                            ]
                    
                    # Mostrar análisis
                    if not df_analisis.empty:
                        # Convertir fechas para mostrar
                        df_analisis_mostrar = df_analisis.copy()
                        df_analisis_mostrar["fecha_ingreso_materia"] = df_analisis_mostrar["fecha_ingreso_materia"].dt.strftime("%Y-%m-%d")
                        df_analisis_mostrar["fecha_calificacion"] = pd.to_datetime(df_analisis_mostrar["fecha_calificacion"], errors="coerce").dt.strftime("%Y-%m-%d")
                        
                        columnas_analisis = [
                            "cuatrimestre_display", "materia", "grupo", "fecha_ingreso_materia", "fecha_calificacion", 
                            "calificacion", "tipo_asignacion", "estatus", "n_recursamientos", "profesor"
                        ]
                        columnas_existentes = [col for col in columnas_analisis if col in df_analisis_mostrar.columns]
                        
                        # Ordenar por cuatrimestre y luego por fecha
                        df_ordenado = df_analisis_mostrar[columnas_existentes].sort_values(
                            by=["cuatrimestre_display", "fecha_ingreso_materia"], ascending=[True, False]
                        )
                        
                        st.markdown(f"### 📊 Registros filtrados: {len(df_ordenado)}")
                        st.dataframe(df_ordenado, use_container_width=True)
                        
                        # Estadísticas del período filtrado
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("📘 Materias filtradas", len(df_ordenado))
                        with col2:
                            recursadas = len(df_ordenado[df_ordenado["tipo_asignacion"] == "recursamiento"])
                            st.metric("🔁 Recursadas", recursadas)
                        with col3:
                            if "calificacion" in df_ordenado.columns:
                                califs_numericas = pd.to_numeric(df_ordenado["calificacion"], errors="coerce").dropna()
                                if not califs_numericas.empty:
                                    promedio_periodo = califs_numericas.mean()
                                    st.metric("📊 Promedio filtrado", f"{promedio_periodo:.2f}")
                        with col4:
                            cuatrimestres_unicos = df_ordenado["cuatrimestre_display"].nunique() if "cuatrimestre_display" in df_ordenado.columns else 0
                            st.metric("🎓 Cuatrimestres", cuatrimestres_unicos)
                    else:
                        st.info("No hay datos que coincidan con los filtros seleccionados.")
                
                # ==================== DESCARGA DEL HISTORIAL MEJORADA ====================
                # ==================== DESCARGA DEL HISTORIAL CORREGIDA ====================
                st.markdown("### 📥 Descargar Historial")

                col1, col2 = st.columns(2)
                with col1:
                    # Descarga completa CON CUATRIMESTRE HISTÓRICO CORRECTO
                    import io
                    output = io.BytesIO()
                    
                    # PREPARAR DATOS PARA DESCARGA CON CUATRIMESTRE HISTÓRICO
                    df_para_descarga = df_estudiante.copy()
                    
                    # Asegurar que tenemos la columna cuatrimestre_display calculada
                    if 'cuatrimestre_display' not in df_para_descarga.columns:
                        # Si no existe, recalcular usando la misma función
                        try:
                            mapeo_cuatrimestres = calcular_cuatrimestre_por_materias_aprobadas(df_para_descarga)
                            df_para_descarga['cuatrimestre_display'] = df_para_descarga.index.map(mapeo_cuatrimestres)
                            df_para_descarga['cuatrimestre_display'] = df_para_descarga['cuatrimestre_display'].fillna('1')
                        except Exception as e:
                            st.warning(f"No se pudo calcular cuatrimestre histórico: {e}")
                            df_para_descarga['cuatrimestre_display'] = '1'
                    
                    # RENOMBRAR COLUMNAS PARA CLARIDAD EN EXCEL
                    df_para_descarga_renamed = df_para_descarga.copy()
                    
                    # Renombrar columnas clave
                    columnas_rename = {
                        'cuatrimestre': 'cuatrimestre_actual_estudiante',  # El cuatrimestre actual del estudiante
                        'cuatrimestre_display': 'cuatrimestre_historico_materia'  # El cuatrimestre cuando cursó cada materia
                    }
                    
                    df_para_descarga_renamed = df_para_descarga_renamed.rename(columns=columnas_rename)
                    
                    # REORDENAR COLUMNAS para que sea más claro
                    columnas_importantes = [
                        'matricula', 'nombre', 'carrera', 'cuatrimestre_actual_estudiante',
                        'materia', 'cuatrimestre_historico_materia', 'grupo', 
                        'fecha_ingreso_materia', 'fecha_calificacion', 'calificacion', 
                        'estatus', 'tipo_asignacion', 'n_recursamientos', 'profesor'
                    ]
                    
                    # Filtrar solo las columnas que existen
                    columnas_existentes = [col for col in columnas_importantes if col in df_para_descarga_renamed.columns]
                    columnas_restantes = [col for col in df_para_descarga_renamed.columns if col not in columnas_existentes]
                    columnas_finales = columnas_existentes + columnas_restantes
                    
                    df_final = df_para_descarga_renamed[columnas_finales]
                    
                    # ORDENAR POR CUATRIMESTRE HISTÓRICO Y FECHA
                    df_final['cuatrimestre_historico_num'] = pd.to_numeric(df_final['cuatrimestre_historico_materia'], errors='coerce').fillna(1)
                    df_final['fecha_ingreso_sort'] = pd.to_datetime(df_final['fecha_ingreso_materia'], errors='coerce')
                    
                    df_final_ordenado = df_final.sort_values([
                        'cuatrimestre_historico_num', 'fecha_ingreso_sort'
                    ], ascending=[True, True])
                    
                    # Eliminar columnas auxiliares de ordenamiento
                    df_final_ordenado = df_final_ordenado.drop(columns=['cuatrimestre_historico_num', 'fecha_ingreso_sort'])
                    
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        # HOJA PRINCIPAL: Historial completo con cuatrimestre histórico
                        df_final_ordenado.to_excel(writer, sheet_name='Historial_Completo_Historico', index=False)
                        
                        # HOJA DE RESUMEN POR CUATRIMESTRE
                        try:
                            materias_con_calificacion = df_final_ordenado[
                                (df_final_ordenado["materia"].notna()) & 
                                (df_final_ordenado["materia"] != "") &
                                (df_final_ordenado["calificacion"].notna())
                            ].copy()
                            
                            if not materias_con_calificacion.empty:
                                # Crear resumen estadístico por cuatrimestre histórico
                                distribucion_cuatrimestre = materias_con_calificacion.groupby("cuatrimestre_historico_materia").agg({
                                    "materia": "count",
                                    "calificacion": [
                                        lambda x: pd.to_numeric(x, errors='coerce').mean(),
                                        lambda x: sum(pd.to_numeric(x, errors='coerce') >= 6),
                                        lambda x: sum(pd.to_numeric(x, errors='coerce') < 6)
                                    ]
                                }).round(2)
                                
                                distribucion_cuatrimestre.columns = ["Total_Materias", "Promedio", "Aprobadas", "Reprobadas"]
                                distribucion_cuatrimestre["Tasa_Aprobacion_%"] = (
                                    distribucion_cuatrimestre["Aprobadas"] / distribucion_cuatrimestre["Total_Materias"] * 100
                                ).round(1)
                                
                                # Resetear índice para que cuatrimestre sea una columna
                                distribucion_cuatrimestre.reset_index(inplace=True)
                                distribucion_cuatrimestre.to_excel(writer, sheet_name='Resumen_Por_Cuatrimestre', index=False)
                        except Exception as e:
                            st.warning(f"No se pudo crear resumen por cuatrimestre: {e}")
                        
                        # HOJAS INDIVIDUALES POR CUATRIMESTRE HISTÓRICO
                        if 'cuatrimestre_historico_materia' in df_final_ordenado.columns:
                            cuatrimestres_unicos = sorted(
                                df_final_ordenado['cuatrimestre_historico_materia'].unique(),
                                key=lambda x: int(x) if str(x).isdigit() else 999
                            )
                            
                            for cuatr in cuatrimestres_unicos:
                                df_cuatr = df_final_ordenado[df_final_ordenado['cuatrimestre_historico_materia'] == cuatr]
                                
                                # Ordenar por fecha dentro del cuatrimestre
                                df_cuatr_ordenado = df_cuatr.sort_values('fecha_ingreso_materia', ascending=True)
                                
                                sheet_name = f'Cuatrimestre_{cuatr}'[:31]  # Límite de Excel
                                df_cuatr_ordenado.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    output.seek(0)

                    st.download_button(
                        label="📥 Historial completo con cuatrimestre histórico (Excel)",
                        data=output,
                        file_name=f"historial_completo_historico_{matricula_estudiante}_{nombre_estudiante.replace(' ', '_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Descarga el historial con el cuatrimestre histórico calculado para cada materia"
                    )
                    
                    
                with col2:
                    # Descarga solo por cuatrimestres (sin cambios mayores, pero usando datos corregidos)
                    if 'cuatrimestre_historico_materia' in df_final_ordenado.columns:
                        output_cuatr = io.BytesIO()
                        with pd.ExcelWriter(output_cuatr, engine='openpyxl') as writer:
                            # Resumen por cuatrimestre histórico
                            try:
                                materias_con_calificacion = df_final_ordenado[
                                    (df_final_ordenado["materia"].notna()) & 
                                    (df_final_ordenado["calificacion"].notna())
                                ].copy()
                                
                                if not materias_con_calificacion.empty:
                                    distribucion_cuatrimestre = materias_con_calificacion.groupby("cuatrimestre_historico_materia").agg({
                                        "materia": "count",
                                        "calificacion": [
                                            lambda x: pd.to_numeric(x, errors='coerce').mean(),
                                            lambda x: sum(pd.to_numeric(x, errors='coerce') >= 6)
                                        ]
                                    }).round(2)
                                    
                                    distribucion_cuatrimestre.columns = ["Total_Materias", "Promedio", "Aprobadas"]
                                    distribucion_cuatrimestre["Reprobadas"] = distribucion_cuatrimestre["Total_Materias"] - distribucion_cuatrimestre["Aprobadas"]
                                    distribucion_cuatrimestre["Tasa_Aprobacion_%"] = (
                                        distribucion_cuatrimestre["Aprobadas"] / distribucion_cuatrimestre["Total_Materias"] * 100
                                    ).round(1)
                                    
                                    distribucion_cuatrimestre.reset_index(inplace=True)
                                    distribucion_cuatrimestre.to_excel(writer, sheet_name='Resumen_Cuatrimestres', index=False)
                                    
                                    # Detalle por cuatrimestre histórico
                                    cuatrimestres_unicos = sorted(
                                        df_final_ordenado['cuatrimestre_historico_materia'].unique(),
                                        key=lambda x: int(x) if str(x).isdigit() else 999
                                    )
                                    
                                    for cuatr in cuatrimestres_unicos:
                                        df_cuatr = df_final_ordenado[df_final_ordenado['cuatrimestre_historico_materia'] == cuatr]
                                        sheet_name = f'Detalle_Cuatr_{cuatr}'[:31]
                                        df_cuatr.to_excel(writer, sheet_name=sheet_name, index=False)
                                else:
                                    # Si no hay materias con calificación, usar datos básicos
                                    df_final_ordenado.to_excel(writer, sheet_name='Historial', index=False)
                            except Exception as e:
                                # En caso de error, exportar datos básicos
                                df_final_ordenado.to_excel(writer, sheet_name='Historial', index=False)
                                st.warning(f"Se exportaron datos básicos debido a: {e}")
                        
                        output_cuatr.seek(0)
                        
                        st.download_button(
                            label="🎓 Organizado por cuatrimestres históricos (Excel)",
                            data=output_cuatr,
                            file_name=f"historial_cuatrimestres_historicos_{matricula_estudiante}_{nombre_estudiante.replace(' ', '_')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="Descarga organizada por cuatrimestres históricos con resumen estadístico"
                        )
                    else:
                        st.info("💡 El cuatrimestre histórico no está disponible para esta descarga.")

                # MOSTRAR PREVIEW DE CÓMO SE VERÁN LOS DATOS
                if st.checkbox("🔍 Vista previa de datos con cuatrimestre histórico", help="Ver cómo se organizarán los datos en el Excel"):
                    st.markdown("#### 📊 Preview: Primeras 10 filas con cuatrimestre histórico")
                    
                    # Mostrar solo columnas relevantes para el preview
                    columnas_preview = [
                        'materia', 'cuatrimestre_historico_materia', 'calificacion', 
                        'estatus', 'fecha_ingreso_materia', 'tipo_asignacion'
                    ]
                    columnas_preview_existentes = [col for col in columnas_preview if col in df_final_ordenado.columns]
                    
                    df_preview = df_final_ordenado[columnas_preview_existentes].head(10)
                    st.dataframe(df_preview, use_container_width=True)
                    
                    # Mostrar conteo por cuatrimestre histórico
                    if 'cuatrimestre_historico_materia' in df_final_ordenado.columns:
                        conteo_cuatr = df_final_ordenado['cuatrimestre_historico_materia'].value_counts().sort_index()
                        st.markdown("#### 📈 Distribución por cuatrimestre histórico:")
                        for cuatr, count in conteo_cuatr.items():
                            st.write(f"**Cuatrimestre {cuatr}:** {count} materias")
            
# ========= NUEVA FUNCIONALIDAD: MOVER MATERIAS ENTRE CUATRIMESTRES =========
# ========= NUEVA FUNCIONALIDAD: MOVER MATERIAS ENTRE CUATRIMESTRES =========
# ========= NUEVA FUNCIONALIDAD: MOVER MATERIAS ENTRE CUATRIMESTRES =========
elif menu == "🔄 Mover materias de cuatrimestre":
    st.subheader("🔄 Mover materias entre cuatrimestres")
    st.markdown("### Corrige asignaciones incorrectas de cuatrimestre")
    
    df = cargar_datos_db()
    if df.empty:
        st.warning("No hay datos disponibles.")
    else:
        # Pestañas para diferentes opciones
        tab1, tab2 = st.tabs(["👤 Por Estudiante Individual", "👥 Por Grupo Completo"])
        
        # ====================== TAB 1: POR ESTUDIANTE INDIVIDUAL ======================
        with tab1:
            st.markdown("#### 🎯 Seleccionar estudiante específico")
            
            # Búsqueda de estudiante
            col1, col2 = st.columns([3, 1])
            with col1:
                busqueda_estudiante = st.text_input(
                    "🔍 Buscar estudiante por matrícula o nombre", 
                    placeholder="Ejemplo: 0125070026 o González",
                    key="busqueda_estudiante_individual"
                ).strip().lower()
            
            with col2:
                buscar_exacto_est = st.checkbox("Búsqueda exacta", key="estudiante_exacto")
            
            if busqueda_estudiante:
                # Filtrar estudiantes
                estudiantes_encontrados = []
                for idx, row in df.iterrows():
                    try:
                        matricula_str = str(row.get('matricula', '')).lower()
                        nombre_str = str(row.get('nombre', '')).lower()
                        
                        if matricula_str in ['nan', 'none', 'null']:
                            matricula_str = ''
                        if nombre_str in ['nan', 'none', 'null']:
                            nombre_str = ''
                        
                        coincide = False
                        if buscar_exacto_est:
                            if matricula_str == busqueda_estudiante or nombre_str == busqueda_estudiante:
                                coincide = True
                        else:
                            if busqueda_estudiante in matricula_str or busqueda_estudiante in nombre_str:
                                coincide = True
                        
                        if coincide:
                            estudiantes_encontrados.append({
                                'matricula': str(row.get('matricula', 'N/A')),
                                'nombre': str(row.get('nombre', 'N/A')),
                                'carrera': str(row.get('carrera', 'N/A'))
                            })
                    except:
                        continue
                
                # Eliminar duplicados por matrícula
                estudiantes_unicos = {}
                for est in estudiantes_encontrados:
                    if est['matricula'] not in estudiantes_unicos:
                        estudiantes_unicos[est['matricula']] = est
                
                estudiantes_lista = list(estudiantes_unicos.values())
                
                if estudiantes_lista:
                    st.success(f"✅ Se encontraron {len(estudiantes_lista)} estudiantes")
                    
                    # Selectbox para elegir estudiante
                    estudiante_seleccionado = st.selectbox(
                        "Selecciona el estudiante:",
                        estudiantes_lista,
                        format_func=lambda x: f"{x['matricula']} - {x['nombre']} ({x['carrera']})",
                        key="estudiante_select_individual"
                    )
                    
                    if estudiante_seleccionado:
                        matricula_seleccionada = estudiante_seleccionado['matricula']
                        
                        # Obtener materias del estudiante
                        materias_estudiante = df[df['matricula'].astype(str) == matricula_seleccionada].copy()
                        
                        if not materias_estudiante.empty:
                            st.markdown(f"#### 📚 Materias de **{estudiante_seleccionado['nombre']}**")
                            
                            # VERIFICAR COLUMNAS DISPONIBLES
                            st.info(f"📋 Columnas disponibles: {list(materias_estudiante.columns)}")
                            
                            # Detectar columna de cuatrimestre
                            columna_cuatrimestre = None
                            posibles_columnas = ['cuatrimestre', 'cuatrimestre_display', 'cuatrimestre_historico', 'periodo', 'semestre']
                            for col in posibles_columnas:
                                if col in materias_estudiante.columns:
                                    columna_cuatrimestre = col
                                    break
                            
                            # Preguntar al usuario qué columna usar si hay múltiples opciones
                            columnas_cuatrimestre_disponibles = [col for col in ['cuatrimestre', 'cuatrimestre_historico'] if col in materias_estudiante.columns]
                            
                            if len(columnas_cuatrimestre_disponibles) > 1:
                                st.info("Se encontraron múltiples columnas de cuatrimestre:")
                                columna_cuatrimestre = st.selectbox(
                                    "¿Cuál columna quieres usar para mostrar/modificar?",
                                    columnas_cuatrimestre_disponibles,
                                    help="'cuatrimestre' = actual, 'cuatrimestre_historico' = registro histórico",
                                    key="select_columna_cuatrimestre_individual"
                                )
                            
                            if columna_cuatrimestre:
                                st.success(f"✅ Usando columna: {columna_cuatrimestre}")
                            else:
                                st.error("❌ No se encontró columna de cuatrimestre. Columnas disponibles:")
                                st.write(list(materias_estudiante.columns))
                                st.stop()
                            
                            # Mostrar materias actuales
                            with st.expander("📋 Ver todas las materias actuales", expanded=True):
                                columnas_mostrar = ['materia', columna_cuatrimestre, 'calificacion', 'estatus', 'grupo']
                                columnas_existentes = [col for col in columnas_mostrar if col in materias_estudiante.columns]
                                st.dataframe(materias_estudiante[columnas_existentes], use_container_width=True)
                            
                            # Seleccionar materias para mover
                            st.markdown("#### ✅ Seleccionar materias para mover")
                            
                            materias_opciones = []
                            for idx, row in materias_estudiante.iterrows():
                                cuatri_actual = str(row.get(columna_cuatrimestre, 'Sin cuatrimestre'))
                                materia_nombre = str(row.get('materia', 'Sin nombre'))
                                calificacion = str(row.get('calificacion', 'S/C'))
                                
                                materias_opciones.append({
                                    'idx': idx,
                                    'display': f"{materia_nombre} (Cuatrimestre actual: {cuatri_actual}) - Calif: {calificacion}",
                                    'materia': materia_nombre,
                                    'cuatrimestre_actual': cuatri_actual,
                                    'columna_cuatrimestre': columna_cuatrimestre
                                })
                            
                            materias_seleccionadas = st.multiselect(
                                "Elige las materias que quieres mover:",
                                materias_opciones,
                                format_func=lambda x: x['display'],
                                key="materias_mover_individual"
                            )
                            
                            if materias_seleccionadas:
                                # Seleccionar cuatrimestre destino
                                cuatrimestre_destino = st.selectbox(
                                    "🎯 Selecciona el cuatrimestre de destino:",
                                    ["1", "2", "3", "4", "5", "6", "7", "8", "9"],
                                    key="cuatrimestre_destino_individual"
                                )
                                
                                # Mostrar resumen de cambios
                                st.markdown("#### 📝 Resumen de cambios:")
                                for materia in materias_seleccionadas:
                                    st.write(f"• **{materia['materia']}**: Cuatrimestre {materia['cuatrimestre_actual']} → Cuatrimestre {cuatrimestre_destino}")
                                
                                # Confirmación
                                confirmar_individual = st.checkbox(
                                    f"Confirmo mover {len(materias_seleccionadas)} materia(s) al cuatrimestre {cuatrimestre_destino}",
                                    key="confirmar_mover_individual"
                                )
                                
                                if confirmar_individual and st.button("🔄 EJECUTAR MOVIMIENTO", type="primary", key="ejecutar_individual"):
                                    # Agregar la columna seleccionada a cada materia
                                    for materia in materias_seleccionadas:
                                        materia['columna_cuatrimestre'] = columna_cuatrimestre
                                    ejecutar_movimiento_materias(materias_seleccionadas, cuatrimestre_destino, "individual")
                        else:
                            st.warning("No se encontraron materias para este estudiante")
                else:
                    st.warning("No se encontraron estudiantes con esa búsqueda")
        
        # ====================== TAB 2: POR GRUPO COMPLETO ======================
        with tab2:
            st.markdown("#### 👥 Seleccionar grupo completo")
            
            # Obtener grupos únicos
            grupos_disponibles = []
            try:
                grupos_unicos = df['grupo'].dropna().unique()
                for grupo in grupos_unicos:
                    if str(grupo) not in ['nan', 'None', 'null', '']:
                        grupos_disponibles.append(str(grupo))
                grupos_disponibles = sorted(list(set(grupos_disponibles)))
            except:
                grupos_disponibles = []
            
            if grupos_disponibles:
                grupo_seleccionado = st.selectbox(
                    "🎯 Selecciona el grupo:",
                    grupos_disponibles,
                    key="grupo_select_completo"
                )
                
                if grupo_seleccionado:
                                            # Obtener estudiantes del grupo
                    estudiantes_grupo = df[df['grupo'].astype(str) == grupo_seleccionado].copy()
                    
                    if not estudiantes_grupo.empty:
                        # DETECTAR COLUMNA DE CUATRIMESTRE PARA EL GRUPO
                        columnas_cuatrimestre_disponibles = [col for col in ['cuatrimestre', 'cuatrimestre_historico'] if col in estudiantes_grupo.columns]
                        
                        if not columnas_cuatrimestre_disponibles:
                            st.error("❌ No se encontraron columnas de cuatrimestre. Columnas disponibles:")
                            st.write(list(estudiantes_grupo.columns))
                            st.stop()
                        
                        # Si hay múltiples columnas, preguntar cuál usar
                        if len(columnas_cuatrimestre_disponibles) > 1:
                            st.info("Se encontraron múltiples columnas de cuatrimestre:")
                            columna_cuatrimestre_grupo = st.selectbox(
                                "¿Cuál columna quieres usar para mostrar/modificar?",
                                columnas_cuatrimestre_disponibles,
                                help="'cuatrimestre' = actual, 'cuatrimestre_historico' = registro histórico",
                                key="select_columna_cuatrimestre_grupo"
                            )
                        else:
                            columna_cuatrimestre_grupo = columnas_cuatrimestre_disponibles[0]
                        
                        st.info(f"📋 Usando columna de cuatrimestre: **{columna_cuatrimestre_grupo}**")
                        
                        # Mostrar información del grupo
                        num_estudiantes = estudiantes_grupo['matricula'].nunique()
                        num_materias = len(estudiantes_grupo)
                        
                        st.success(f"📊 Grupo **{grupo_seleccionado}**: {num_estudiantes} estudiantes, {num_materias} registros de materias")
                        
                        # Mostrar vista previa del grupo
                        with st.expander("👀 Vista previa del grupo", expanded=False):
                            columnas_mostrar = ['matricula', 'nombre', 'materia', columna_cuatrimestre_grupo, 'calificacion']
                            columnas_existentes = [col for col in columnas_mostrar if col in estudiantes_grupo.columns]
                            st.dataframe(estudiantes_grupo[columnas_existentes], use_container_width=True)
                        
                        # Opciones de filtrado
                        st.markdown("#### 🎛️ Opciones de filtrado")
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            # Filtrar por cuatrimestre actual
                            try:
                                cuatrimestres_actuales = estudiantes_grupo[columna_cuatrimestre_grupo].dropna().unique()
                                cuatrimestres_actuales = [str(c) for c in cuatrimestres_actuales if str(c) not in ['nan', 'None']]
                            except:
                                cuatrimestres_actuales = []
                            
                            if cuatrimestres_actuales:
                                cuatrimestre_filtro = st.selectbox(
                                    "Mover solo materias del cuatrimestre:",
                                    ["Todas las materias"] + sorted(cuatrimestres_actuales),
                                    key="cuatrimestre_filtro_grupo"
                                )
                            else:
                                cuatrimestre_filtro = "Todas las materias"
                        
                        with col2:
                            # Filtrar por materia específica
                            materias_disponibles = estudiantes_grupo['materia'].dropna().unique()
                            materias_disponibles = [str(m) for m in materias_disponibles if str(m) not in ['nan', 'None']]
                            
                            if materias_disponibles:
                                materia_especifica = st.selectbox(
                                    "Mover solo una materia específica:",
                                    ["Todas las materias"] + sorted(materias_disponibles),
                                    key="materia_filtro_grupo"
                                )
                            else:
                                materia_especifica = "Todas las materias"
                        
                        # Aplicar filtros
                        materias_filtradas = estudiantes_grupo.copy()
                        
                        if cuatrimestre_filtro != "Todas las materias":
                            materias_filtradas = materias_filtradas[
                                materias_filtradas[columna_cuatrimestre_grupo].astype(str) == cuatrimestre_filtro
                            ]
                        
                        if materia_especifica != "Todas las materias":
                            materias_filtradas = materias_filtradas[
                                materias_filtradas['materia'].astype(str) == materia_especifica
                            ]
                        
                        if not materias_filtradas.empty:
                            st.info(f"📊 Con los filtros aplicados: {len(materias_filtradas)} registros serán movidos")
                            
                            # Seleccionar cuatrimestre destino
                            cuatrimestre_destino_grupo = st.selectbox(
                                "🎯 Cuatrimestre de destino:",
                                ["1", "2", "3", "4", "5", "6", "7", "8", "9"],
                                key="cuatrimestre_destino_grupo"
                            )
                            
                            # Resumen de cambios
                            st.markdown("#### 📝 Resumen del movimiento:")
                            st.write(f"• **Grupo:** {grupo_seleccionado}")
                            st.write(f"• **Registros a mover:** {len(materias_filtradas)}")
                            st.write(f"• **Cuatrimestre destino:** {cuatrimestre_destino_grupo}")
                            
                            if cuatrimestre_filtro != "Todas las materias":
                                st.write(f"• **Filtro cuatrimestre:** Solo del cuatrimestre {cuatrimestre_filtro}")
                            if materia_especifica != "Todas las materias":
                                st.write(f"• **Filtro materia:** Solo {materia_especifica}")
                            
                            # Confirmación
                            confirmar_grupo = st.checkbox(
                                f"Confirmo mover {len(materias_filtradas)} registros del grupo {grupo_seleccionado} al cuatrimestre {cuatrimestre_destino_grupo}",
                                key="confirmar_mover_grupo"
                            )
                            
                            if confirmar_grupo and st.button("🔄 EJECUTAR MOVIMIENTO GRUPAL", type="primary", key="ejecutar_grupo"):
                                ejecutar_movimiento_materias(materias_filtradas, cuatrimestre_destino_grupo, "grupo", columna_cuatrimestre_grupo)
                        else:
                            st.warning("No hay registros que coincidan con los filtros aplicados")
            else:
                st.warning("No se encontraron grupos en la base de datos")         

# Agregar después del case de "🧾 Historial por alumno":
elif menu == "🎓 Historial de cuatrimestre":
    mostrar_historial_cuatrimestre_alumno()
    
elif menu == "🎓 Limpiar Base de Datos":
    limpiar_tipos_cuatrimestre()   

# ========= MÓDULO COMPLETO: GENERACIÓN DE MATERIA CON VALIDACIONES Y VISTA PREVIA =========
elif menu == "📚 Generación de materia":
    st.subheader("📚 Generación de Materia - Integración de Alumnos")
    
    df = cargar_datos_db()
    if df.empty:
        st.warning("No hay datos en la base de datos.")
    else:
        # FUNCIÓN: Validar alumnos ya aprobados
        def validar_alumnos_aprobados(df_candidatos, materia_objetivo):
            """
            Valida si algún alumno candidato ya aprobó la materia objetivo
            Retorna: df_validos, df_conflictos
            """
            if df_candidatos.empty:
                return df_candidatos, pd.DataFrame()
            
            conflictos = []
            indices_conflictos = []
            
            for idx, alumno in df_candidatos.iterrows():
                matricula = alumno.get('matricula', '')
                
                # Buscar si este alumno ya aprobó la materia objetivo
                aprobaciones_previas = df[
                    (df['matricula'] == matricula) &
                    (df['materia'] == materia_objetivo) &
                    (pd.to_numeric(df['calificacion'], errors='coerce') >= 6) &
                    (df['estatus'] == 'aprobado')
                ]
                
                if not aprobaciones_previas.empty:
                    # Obtener información de la aprobación previa
                    aprobacion_info = aprobaciones_previas.iloc[0]
                    conflicto_info = {
                        'matricula': matricula,
                        'nombre': alumno.get('nombre', 'N/A'),
                        'grupo_actual': alumno.get('grupo', 'N/A'),
                        'materia': materia_objetivo,
                        'calificacion_previa': aprobacion_info.get('calificacion', 'N/A'),
                        'fecha_aprobacion': aprobacion_info.get('fecha_calificacion', 'N/A'),
                        'profesor_previo': aprobacion_info.get('profesor', 'N/A'),
                        'estatus_actual_candidato': alumno.get('estatus', 'N/A')
                    }
                    conflictos.append(conflicto_info)
                    indices_conflictos.append(idx)
            
            # Crear DataFrames de resultado
            df_conflictos = pd.DataFrame(conflictos) if conflictos else pd.DataFrame()
            df_validos = df_candidatos.drop(indices_conflictos) if indices_conflictos else df_candidatos.copy()
            
            return df_validos, df_conflictos
        
        # Filtros principales
        st.markdown("""
        <div style="background: linear-gradient(135deg, #6c5ce7 0%, #a29bfe 100%); padding: 20px; border-radius: 10px; margin: 20px 0;">
            <h3 style="color: white; margin: 0; text-align: center;">🎯 Configuración de Generación</h3>
            <p style="color: #f8f9ff; margin: 5px 0 0 0; text-align: center;">Configura los filtros para generar una nueva materia</p>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            carreras = sorted(df["carrera"].dropna().unique().tolist())
            carrera_seleccionada = st.selectbox("🎓 Selecciona la carrera", carreras, key="generacion_carrera")
        
        with col2:
            df["fecha_ingreso_original"] = pd.to_datetime(df["fecha_ingreso_original"], errors="coerce")
            fechas_ingreso = sorted(
                df[df["carrera"] == carrera_seleccionada]["fecha_ingreso_original"]
                .dropna().dt.strftime("%Y-%m-%d").unique().tolist()
            )
            fecha_ingreso_alumnos = st.selectbox("📅 Fecha de ingreso de alumnos", fechas_ingreso, key="fecha_ingreso_alumnos")
        
        with col3:
            df["fecha_ingreso_materia"] = pd.to_datetime(df["fecha_ingreso_materia"], errors="coerce")
            fechas_materia = sorted(
                df[df["carrera"] == carrera_seleccionada]["fecha_ingreso_materia"]
                .dropna().dt.strftime("%Y-%m-%d").unique().tolist()
            )
            fecha_ingreso_materia = st.selectbox("📚 Fecha de ingreso a materia", fechas_materia, key="fecha_ingreso_materia")
        
        # Configuración de materia con validación
        materia_nombre = st.text_input("📝 Nombre de la materia a generar", key="nombre_materia_generar")
        
        if materia_nombre:
            # Verificar si la materia ya existe
            materias_existentes = df['materia'].dropna().unique()
            if materia_nombre in materias_existentes:
                st.warning(f"⚠️ **Atención:** La materia '{materia_nombre}' ya existe en el sistema. Se aplicarán validaciones adicionales.")
                
                # Estadísticas de la materia existente
                df_materia_existente = df[df['materia'] == materia_nombre]
                total_registros = len(df_materia_existente)
                aprobados = len(df_materia_existente[
                    (pd.to_numeric(df_materia_existente['calificacion'], errors='coerce') >= 6) &
                    (df_materia_existente['estatus'] == 'aprobado')
                ])
                
                st.info(f"📊 **Materia existente:** {total_registros} registros totales, {aprobados} alumnos ya aprobados")
            
            # Convertir fechas para filtrado
            fecha_ingreso_dt = pd.to_datetime(fecha_ingreso_alumnos)
            fecha_materia_dt = pd.to_datetime(fecha_ingreso_materia)
            
            # Buscar datos base
            nuevos_ingresos = df[
                (df["carrera"] == carrera_seleccionada) &
                (df["fecha_ingreso_original"] == fecha_ingreso_dt) &
                (df["calificacion"].isna()) &
                (df["materia"].isna() | (df["materia"] == "")) &
                (df["estatus"].isna() | (df["estatus"] == "")) &
                (df["tipo_asignacion"] != "recursamiento")
            ].copy()
            
            recursamiento_general = df[
                (df["carrera"] == carrera_seleccionada) &
                (df["fecha_ingreso_materia"] == fecha_materia_dt) &
                (
                    (df["estatus"] == "recursando") | 
                    (df["tipo_asignacion"] == "recursamiento") |
                    (df["estatus"] == "reprobado")
                )
            ].copy()
            
            aprobados = df[
                (df["carrera"] == carrera_seleccionada) &
                (df["fecha_ingreso_materia"] == fecha_materia_dt) &
                (pd.to_numeric(df["calificacion"], errors="coerce") >= 6) &
                (df["estatus"] == "aprobado")
            ].copy()
            
            # Mostrar resultados en tabs
            tab1, tab2, tab3, tab4 = st.tabs(["🆕 Nuevos Ingresos", "🔁 Recursamiento", "✅ Aprobados", "📦 Integraciones"])
            
            with tab1:
                st.markdown(f"### 🆕 Alumnos de Nuevo Ingreso: {len(nuevos_ingresos)}")
                st.info(f"📅 Fecha de ingreso: {fecha_ingreso_alumnos}")
                
                # VALIDACIÓN PARA NUEVOS INGRESOS
                if not nuevos_ingresos.empty:
                    nuevos_validos, nuevos_conflictos = validar_alumnos_aprobados(nuevos_ingresos, materia_nombre)
                    
                    if not nuevos_conflictos.empty:
                        st.error(f"🚨 **CONFLICTO DETECTADO:** {len(nuevos_conflictos)} alumno(s) ya aprobaron '{materia_nombre}'")
                        
                        with st.expander("⚠️ Ver alumnos con conflictos", expanded=True):
                            st.markdown("**Alumnos que YA aprobaron esta materia:**")
                            st.dataframe(nuevos_conflictos, use_container_width=True)
                            st.warning("Estos alumnos NO deberían ser asignados nuevamente a esta materia.")
                    
                    if not nuevos_validos.empty:
                        st.success(f"✅ Alumnos válidos para asignar: {len(nuevos_validos)}")
                        
                        # Métricas
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("👥 Total Válidos", len(nuevos_validos))
                        with col2:
                            grupos_nuevos = nuevos_validos['grupo'].nunique()
                            st.metric("📚 Grupos", grupos_nuevos)
                            
                            if grupos_nuevos == 1:
                                grupo_principal = nuevos_validos['grupo'].iloc[0]
                                st.success(f"🎯 **Grupo principal:** {grupo_principal}")
                            elif grupos_nuevos > 1:
                                grupo_mas_comun = nuevos_validos['grupo'].mode().iloc[0]
                                st.warning(f"⚠️ **Múltiples grupos detectados**")
                                st.info(f"🎯 **Grupo más común:** {grupo_mas_comun}")
                        with col3:
                            con_email = len(nuevos_validos[nuevos_validos['usuario_email_nova'].notna()])
                            st.metric("📧 Con Email", con_email)
                        
                        # Tabla de alumnos válidos
                        columnas_nuevos = ["matricula", "nombre", "grupo", "carrera", "cuatrimestre", "email_personal", "usuario_email_nova", "fecha_ingreso_original", "estatus"]
                        st.dataframe(nuevos_validos[columnas_nuevos], use_container_width=True)
                        
                        # Exportación solo de válidos
                        nuevos_export = nuevos_validos.copy()
                        nuevos_export["materia"] = materia_nombre
                        nuevos_export["fecha_ingreso_materia"] = fecha_ingreso_materia
                        nuevos_export["tipo_alumno"] = "nuevo_ingreso"
                        
                        import io
                        output_nuevos = io.BytesIO()
                        with pd.ExcelWriter(output_nuevos, engine='openpyxl') as writer:
                            nuevos_export[["matricula", "nombre", "grupo", "materia", "carrera", "cuatrimestre", "email_personal", "usuario_email_nova", "fecha_ingreso_materia", "tipo_alumno"]].to_excel(writer, index=False, sheet_name='Nuevos_Validos')
                        output_nuevos.seek(0)
                        
                        st.download_button(
                            label="📥 Descargar Excel - Nuevos Ingresos Válidos",
                            data=output_nuevos,
                            file_name=f"nuevos_validos_{materia_nombre}_{carrera_seleccionada}_{fecha_ingreso_alumnos}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    else:
                        st.warning("⚠️ Todos los alumnos de nuevo ingreso ya aprobaron esta materia.")
                else:
                    st.info("No se encontraron alumnos de nuevo ingreso.")
            
            with tab2:
                st.markdown(f"### 🔁 Alumnos en Recursamiento")
                st.info(f"📚 Fecha de ingreso a materia: {fecha_ingreso_materia}")
                
                # Filtro específico para recursamiento
                if not recursamiento_general.empty:
                    materias_recursamiento = sorted(recursamiento_general["materia"].dropna().unique().tolist())
                    
                    st.markdown("#### 🎯 Filtrar por materia específica:")
                    col_filtro1, col_filtro2 = st.columns([2, 1])
                    
                    with col_filtro1:
                        materia_recursamiento = st.selectbox(
                            "📘 Selecciona la materia para recursamiento:",
                            ["Todas las materias"] + materias_recursamiento,
                            key="filtro_materia_recursamiento"
                        )
                    
                    with col_filtro2:
                        if materia_recursamiento != "Todas las materias":
                            count_materia = len(recursamiento_general[recursamiento_general["materia"] == materia_recursamiento])
                            st.metric("👥 Alumnos", count_materia)
                    
                    # Aplicar filtro
                    if materia_recursamiento != "Todas las materias":
                        recursamiento_filtrado = recursamiento_general[recursamiento_general["materia"] == materia_recursamiento].copy()
                        
                        # VALIDACIÓN CRÍTICA
                        if materia_recursamiento == materia_nombre:
                            st.error("🚨 **ALERTA CRÍTICA:** Estás intentando asignar recursamiento de la misma materia que están generando!")
                            
                            recursamiento_validos, recursamiento_conflictos = validar_alumnos_aprobados(recursamiento_filtrado, materia_nombre)
                            
                            if not recursamiento_conflictos.empty:
                                st.error(f"❌ **{len(recursamiento_conflictos)} alumno(s) YA APROBARON '{materia_nombre}'**")
                                
                                with st.expander("🚨 ALUMNOS QUE YA APROBARON - NO PUEDEN RECURSAR", expanded=True):
                                    st.markdown("**❌ Estos alumnos NO deben recursar porque ya aprobaron:**")
                                    st.dataframe(recursamiento_conflictos[['matricula', 'nombre', 'calificacion_previa', 'fecha_aprobacion', 'profesor_previo']], use_container_width=True)
                                    
                                    st.error("🚫 **ACCIÓN REQUERIDA:** Estos alumnos deben ser excluidos del recursamiento.")
                            
                            recursamiento = recursamiento_validos
                            
                            if recursamiento.empty:
                                st.warning("⚠️ No hay alumnos válidos para recursamiento después de aplicar validaciones.")
                            else:
                                st.success(f"✅ Alumnos válidos para recursamiento: {len(recursamiento)}")
                        else:
                            recursamiento = recursamiento_filtrado
                            st.success(f"🔍 Filtro aplicado: {materia_recursamiento}")
                    else:
                        recursamiento = recursamiento_general.copy()
                else:
                    recursamiento = pd.DataFrame()
                
                # Mostrar información del recursamiento
                if not recursamiento.empty:
                    # Métricas
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("🔄 Total Recursamiento", len(recursamiento))
                    with col2:
                        if 'n_recursamientos' in recursamiento.columns:
                            promedio_recursamientos = recursamiento['n_recursamientos'].mean()
                            st.metric("📈 Promedio Recursamientos", f"{promedio_recursamientos:.1f}")
                        else:
                            st.metric("📈 Promedio Recursamientos", "N/A")
                    with col3:
                        materias_recursamiento_count = recursamiento['materia'].nunique()
                        st.metric("📚 Materias", materias_recursamiento_count)
                    
                    # Información sobre grupos actuales de recursantes
                    grupos_recursantes = recursamiento['grupo'].unique()
                    st.markdown("#### 📋 Grupos actuales de recursantes:")
                    for grupo in sorted(grupos_recursantes):
                        count = len(recursamiento[recursamiento['grupo'] == grupo])
                        materia_grupo = recursamiento[recursamiento['grupo'] == grupo]['materia'].iloc[0] if count > 0 else "N/A"
                        st.write(f"• **{grupo}:** {count} alumnos - Materia: {materia_grupo}")
                    
                    # Análisis de homogeneidad del grupo
                    if 'materia_recursamiento' in locals() and materia_recursamiento != "Todas las materias":
                        st.success("✅ **Grupo homogéneo:** Todos los alumnos recursarán la misma materia")
                        if len(grupos_recursantes) == 1:
                            st.success(f"✅ **Grupo único:** Todos pertenecen al grupo {grupos_recursantes[0]}")
                        else:
                            st.warning(f"⚠️ **Múltiples grupos:** Los alumnos vienen de {len(grupos_recursantes)} grupos diferentes")
                    else:
                        materias_diferentes = recursamiento['materia'].nunique()
                        st.warning(f"⚠️ **Grupo heterogéneo:** Hay alumnos de {materias_diferentes} materias diferentes")
                    
                    # Tabla
                    columnas_recursamiento = ["matricula", "nombre", "grupo", "materia", "carrera", "cuatrimestre", "email_personal", "usuario_email_nova", "fecha_ingreso_materia", "n_recursamientos", "estatus"]
                    columnas_existentes_rec = [col for col in columnas_recursamiento if col in recursamiento.columns]
                    st.dataframe(recursamiento[columnas_existentes_rec], use_container_width=True)
                    
                    # Exportación
                    recursamiento_export = recursamiento.copy()
                    recursamiento_export["tipo_alumno"] = "recursamiento"
                    
                    import io
                    output_recursamiento = io.BytesIO()
                    with pd.ExcelWriter(output_recursamiento, engine='openpyxl') as writer:
                        recursamiento_export[columnas_existentes_rec + ["tipo_alumno"]].to_excel(writer, index=False, sheet_name='Recursamiento_Valido')
                    output_recursamiento.seek(0)
                    
                    materia_archivo = materia_recursamiento.replace(" ", "_") if 'materia_recursamiento' in locals() and materia_recursamiento != "Todas las materias" else "todas_materias"
                    st.download_button(
                        label="📥 Descargar Excel - Recursamiento Válido",
                        data=output_recursamiento,
                        file_name=f"recursamiento_valido_{materia_archivo}_{carrera_seleccionada}_{fecha_ingreso_materia}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    # Información adicional sobre el filtro
                    if 'materia_recursamiento' in locals() and materia_recursamiento != "Todas las materias":
                        st.markdown("---")
                        st.info(f"""
                        💡 **Información del filtro aplicado:**
                        - Materia seleccionada: **{materia_recursamiento}**
                        - Alumnos que recursarán esta materia: **{len(recursamiento)}**
                        - Ideal para formar un grupo homogéneo de recursamiento
                        """)
                else:
                    st.info("No hay alumnos válidos para recursamiento.")
            
            with tab3:
                st.markdown(f"### ✅ Alumnos Aprobados: {len(aprobados)}")
                st.info(f"📚 Fecha de ingreso a materia: {fecha_ingreso_materia}")
                
                if not aprobados.empty:
                    # VALIDACIÓN: Verificar si hay aprobados de la misma materia
                    aprobados_misma_materia = aprobados[aprobados['materia'] == materia_nombre]
                    
                    if not aprobados_misma_materia.empty:
                        st.warning(f"⚠️ **Advertencia:** {len(aprobados_misma_materia)} alumno(s) ya aprobaron '{materia_nombre}' anteriormente")
                        
                        with st.expander("👀 Ver alumnos que ya aprobaron esta materia", expanded=False):
                            st.dataframe(aprobados_misma_materia[['matricula', 'nombre', 'calificacion', 'fecha_calificacion', 'profesor']], use_container_width=True)
                    
                    # Métricas generales
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("✅ Total Aprobados", len(aprobados))
                    with col2:
                        if 'calificacion' in aprobados.columns:
                            promedio_califs = pd.to_numeric(aprobados['calificacion'], errors='coerce').mean()
                            st.metric("📊 Promedio", f"{promedio_califs:.2f}")
                        else:
                            st.metric("📊 Promedio", "N/A")
                    with col3:
                        materias_aprobados = aprobados['materia'].nunique()
                        st.metric("📚 Materias", materias_aprobados)
                    
                    # Tabla
                    columnas_aprobados = ["matricula", "nombre", "grupo", "materia", "carrera", "cuatrimestre", "calificacion", "email_personal", "usuario_email_nova", "fecha_ingreso_materia", "profesor"]
                    columnas_existentes_apr = [col for col in columnas_aprobados if col in aprobados.columns]
                    st.dataframe(aprobados[columnas_existentes_apr], use_container_width=True)
                    
                    # Exportación
                    aprobados_export = aprobados.copy()
                    aprobados_export["tipo_alumno"] = "aprobado"
                    
                    import io
                    output_aprobados = io.BytesIO()
                    with pd.ExcelWriter(output_aprobados, engine='openpyxl') as writer:
                        aprobados_export[columnas_existentes_apr + ["tipo_alumno"]].to_excel(writer, index=False, sheet_name='Aprobados')
                    output_aprobados.seek(0)
                    
                    st.download_button(
                        label="📥 Descargar Excel - Aprobados",
                        data=output_aprobados,
                        file_name=f"aprobados_{materia_nombre}_{carrera_seleccionada}_{fecha_ingreso_materia}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.info("No se encontraron alumnos aprobados con los filtros seleccionados.")
            
            with tab4:
                st.markdown("### 📦 Integraciones Disponibles")
                
                # Aplicar validaciones a los datos para integración
                nuevos_para_integracion = nuevos_ingresos.copy()
                recursamiento_para_integracion = recursamiento.copy() if 'recursamiento' in locals() else pd.DataFrame()
                
                if not nuevos_para_integracion.empty:
                    nuevos_para_integracion, _ = validar_alumnos_aprobados(nuevos_para_integracion, materia_nombre)
                
                if not recursamiento_para_integracion.empty:
                    recursamiento_para_integracion, conflictos_rec = validar_alumnos_aprobados(recursamiento_para_integracion, materia_nombre)
                    
                    if not conflictos_rec.empty:
                        st.error(f"⚠️ Se excluyeron {len(conflictos_rec)} alumnos de recursamiento que ya aprobaron la materia")
                
                # Mostrar información sobre el recursamiento filtrado
                if 'materia_recursamiento' in locals() and materia_recursamiento != "Todas las materias":
                    st.info(f"🎯 **Recursamiento filtrado:** Las integraciones usarán solo alumnos que recursarán '{materia_recursamiento}'")
                
                # Verificar integraciones posibles
                puede_integrar_nuevos_recursamiento = (fecha_ingreso_alumnos == fecha_ingreso_materia) and not nuevos_para_integracion.empty and not recursamiento_para_integracion.empty
                puede_integrar_aprobados_recursamiento = not aprobados.empty and not recursamiento_para_integracion.empty
                
                if puede_integrar_nuevos_recursamiento:
                    # Determinar el grupo de integración
                    grupos_nuevos = nuevos_para_integracion['grupo'].unique()
                    
                    if len(grupos_nuevos) == 1:
                        grupo_integracion = grupos_nuevos[0]
                        st.success(f"✅ Se puede integrar: Nuevos Ingresos + Recursamiento")
                        st.info(f"🎯 **Grupo de integración:** {grupo_integracion} (los recursantes adoptarán este grupo)")
                    else:
                        st.warning("⚠️ Los nuevos ingresos tienen múltiples grupos. Selecciona el grupo principal:")
                        grupo_integracion = st.selectbox(
                            "🎯 Grupo para la integración:",
                            sorted(grupos_nuevos),
                            key="grupo_integracion_nuevos_rec"
                        )
                    
                    # ========= VISTA PREVIA RESTAURADA =========
                    with st.expander("👀 Vista previa de integración con cambios de grupo", expanded=False):
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.markdown("**🆕 Nuevos Ingresos (mantienen su grupo):**")
                            nuevos_preview = nuevos_para_integracion[["matricula", "nombre", "grupo", "carrera"]].copy()
                            nuevos_preview["tipo"] = "Nuevo Ingreso"
                            nuevos_preview["grupo_final"] = nuevos_preview["grupo"]
                            st.dataframe(nuevos_preview[["matricula", "nombre", "grupo_final", "tipo"]], use_container_width=True)
                        
                        with col2:
                            st.markdown(f"**🔁 Recursamiento (cambiarán a grupo {grupo_integracion}):**")
                            recursamiento_preview = recursamiento_para_integracion[["matricula", "nombre", "grupo", "carrera"]].copy()
                            # Agregar columna de materia si existe
                            if 'materia' in recursamiento_para_integracion.columns:
                                recursamiento_preview["materia"] = recursamiento_para_integracion["materia"]
                            
                            recursamiento_preview["tipo"] = "Recursamiento"
                            recursamiento_preview["grupo_anterior"] = recursamiento_preview["grupo"]
                            recursamiento_preview["grupo_final"] = grupo_integracion
                            
                            # Mostrar columnas relevantes
                            if 'materia' in recursamiento_preview.columns:
                                st.dataframe(recursamiento_preview[["matricula", "nombre", "materia", "grupo_anterior", "grupo_final", "tipo"]], use_container_width=True)
                            else:
                                st.dataframe(recursamiento_preview[["matricula", "nombre", "grupo_anterior", "grupo_final", "tipo"]], use_container_width=True)
                            
                            # Mostrar estadísticas de cambios
                            cambios_grupo = len(recursamiento_para_integracion[recursamiento_para_integracion['grupo'] != grupo_integracion])
                            st.metric("🔄 Cambios de grupo", cambios_grupo)
                    
                    # SECCIÓN CORREGIDA - Busca el botón "Generar Integración: Nuevos + Recursamiento" 
# y reemplaza desde ahí hasta el "st.balloons()"

                    if st.button("📦 Generar Integración: Nuevos + Recursamiento", key="integrar_nuevos_recursamiento"):
                        try:
                            # Preparar datos de nuevos ingresos (mantienen su grupo)
                            nuevos_para_integrar = nuevos_para_integracion.copy()
                            nuevos_para_integrar["materia"] = materia_nombre
                            nuevos_para_integrar["fecha_ingreso_materia"] = fecha_ingreso_materia
                            nuevos_para_integrar["tipo_alumno"] = "nuevo_ingreso"
                            nuevos_para_integrar["calificacion"] = None
                            nuevos_para_integrar["estatus"] = "activo"
                            nuevos_para_integrar["tipo_asignacion"] = "regular"
                            
                            # Preparar datos de recursamiento (CAMBIAN SU GRUPO)
                            recursamiento_para_integrar = recursamiento_para_integracion.copy()
                            recursamiento_para_integrar["tipo_alumno"] = "recursamiento"
                            # CAMBIO IMPORTANTE: Los recursantes adoptan el grupo de los nuevos ingresos
                            recursamiento_para_integrar["grupo"] = grupo_integracion
                            
                            # Combinar datos
                            columnas_comunes = ["matricula", "nombre", "grupo", "materia", "carrera", "cuatrimestre", "email_personal", "usuario_email_nova", "fecha_ingreso_materia", "tipo_alumno"]
                            
                            # Asegurar que ambos DataFrames tengan las mismas columnas
                            for col in columnas_comunes:
                                if col not in nuevos_para_integrar.columns:
                                    nuevos_para_integrar[col] = ""
                                if col not in recursamiento_para_integrar.columns:
                                    recursamiento_para_integrar[col] = ""
                            
                            df_integrado = pd.concat([
                                nuevos_para_integrar[columnas_comunes],
                                recursamiento_para_integrar[columnas_comunes]
                            ], ignore_index=True)
                            
                            # Mostrar vista previa
                            st.markdown("#### 👀 Vista previa de la integración generada:")
                            st.success(f"🎯 **Todos los alumnos ahora pertenecen al grupo: {grupo_integracion}**")
                            
                            # Separar por tipo para mostrar claramente
                            df_nuevos_final = df_integrado[df_integrado['tipo_alumno'] == 'nuevo_ingreso']
                            df_recursantes_final = df_integrado[df_integrado['tipo_alumno'] == 'recursamiento']
                            
                            col_prev1, col_prev2 = st.columns(2)
                            with col_prev1:
                                st.markdown(f"**🆕 Nuevos ingresos en grupo {grupo_integracion}: {len(df_nuevos_final)}**")
                                st.dataframe(df_nuevos_final[['matricula', 'nombre', 'grupo', 'tipo_alumno']], use_container_width=True)
                            
                            with col_prev2:
                                st.markdown(f"**🔁 Recursantes integrados en grupo {grupo_integracion}: {len(df_recursantes_final)}**")
                                st.dataframe(df_recursantes_final[['matricula', 'nombre', 'grupo', 'tipo_alumno']], use_container_width=True)
                            
                            # Información sobre validaciones aplicadas
                            if 'materia_recursamiento' in locals() and materia_recursamiento != "Todas las materias":
                                st.info(f"🔍 **Recursamiento filtrado:** Solo incluye alumnos que recursarán '{materia_recursamiento}'")
                            
                            st.dataframe(df_integrado, use_container_width=True)
                            
                            # Botón de descarga
                            import io
                            output_integrado = io.BytesIO()
                            with pd.ExcelWriter(output_integrado, engine='openpyxl') as writer:
                                df_integrado.to_excel(writer, index=False, sheet_name='Nuevos_y_Recursamiento')
                            output_integrado.seek(0)
                            
                            # Nombre de archivo que incluye info del filtro
                            materia_info = materia_recursamiento.replace(" ", "_") if 'materia_recursamiento' in locals() and materia_recursamiento != "Todas las materias" else "mixto"
                            
                            st.download_button(
                                label="📥 Descargar Integración: Nuevos + Recursamiento",
                                data=output_integrado,
                                file_name=f"integracion_nuevos_recursamiento_{materia_nombre}_{materia_info}_{carrera_seleccionada}_grupo_{grupo_integracion}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            
                            st.balloons()
                            
                        except Exception as e:
                            st.error(f"❌ Error al generar integración: {str(e)}")		
            
# ========= NUEVA SECCIÓN: ESTATUS DE ALUMNOS =========
elif menu == "👥 Estatus de alumnos":
    st.subheader("👥 Estatus general de alumnos")
    
    df = cargar_datos_db()
    if df.empty:
        st.warning("No hay datos disponibles.")
    else:
        # Filtros principales
        col1, col2 = st.columns(2)
        with col1:
            carreras = ["Todas"] + sorted(df["carrera"].dropna().unique().tolist())
            carrera_filtro = st.selectbox("🎓 Filtrar por carrera", carreras, key="estatus_carrera")
        
        with col2:
            grupos = ["Todos"] + sorted(df["grupo"].dropna().unique().tolist())
            grupo_filtro = st.selectbox("👥 Filtrar por grupo", grupos, key="estatus_grupo")
        
        # Aplicar filtros
        df_filtrado = df.copy()
        if carrera_filtro != "Todas":
            df_filtrado = df_filtrado[df_filtrado["carrera"] == carrera_filtro]
        if grupo_filtro != "Todos":
            df_filtrado = df_filtrado[df_filtrado["grupo"] == grupo_filtro]
        
        if df_filtrado.empty:
            st.warning("No hay datos con los filtros seleccionados.")
        else:
            # Analizar estatus por alumno (último registro por alumno)
            df_filtrado = df_filtrado.sort_values(['matricula', 'fecha_ingreso_materia'], ascending=[True, False])
            df_ultimo_estatus = df_filtrado.drop_duplicates(subset=['matricula'], keep='first')
            
            # Clasificar alumnos
            alumnos_activos = []
            alumnos_inactivos = []
            alumnos_recursado = []
            
            for _, alumno in df_ultimo_estatus.iterrows():
                matricula = alumno['matricula']
                
                # Verificar si tiene algún registro de éxito académico
                registros_alumno = df_filtrado[df_filtrado['matricula'] == matricula]
                
                # Condiciones para ACTIVO
                tiene_aprobado = len(registros_alumno[pd.to_numeric(registros_alumno['calificacion'], errors='coerce') >= 6]) > 0
                esta_recursando = len(registros_alumno[registros_alumno['estatus'] == 'recursando']) > 0
                tiene_recursamiento = len(registros_alumno[registros_alumno['tipo_asignacion'] == 'recursamiento']) > 0
                
                if tiene_aprobado or esta_recursando or tiene_recursamiento:
                    alumnos_activos.append(alumno)
                    # Verificar si este alumno activo ha recursado
                    if tiene_recursamiento or esta_recursando:
                        alumnos_recursado.append(alumno)
                else:
                    alumnos_inactivos.append(alumno)
            
            # Convertir a DataFrames
            df_activos = pd.DataFrame(alumnos_activos) if alumnos_activos else pd.DataFrame()
            df_inactivos = pd.DataFrame(alumnos_inactivos) if alumnos_inactivos else pd.DataFrame()
            df_recursado = pd.DataFrame(alumnos_recursado) if alumnos_recursado else pd.DataFrame()
            
            # Métricas principales
            total_activos = len(df_activos)
            total_inactivos = len(df_inactivos)
            total_recursado = len(df_recursado)
            total_general = total_activos + total_inactivos
            
            # Mostrar métricas
            st.markdown("""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 20px; border-radius: 10px; margin: 20px 0;">
                <h3 style="color: white; margin: 0; text-align: center;">📊 RESUMEN GENERAL DE ALUMNOS</h3>
            </div>
            """, unsafe_allow_html=True)
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("🟢 Alumnos Activos", total_activos)
            with col2:
                st.metric("🔴 Alumnos Inactivos", total_inactivos)
            with col3:
                st.metric("🔄 Han Recursado", total_recursado)
            with col4:
                st.metric("📊 Total General", total_general)
            
            # Calcular porcentajes
            if total_general > 0:
                porcentaje_activos = (total_activos / total_general) * 100
                porcentaje_inactivos = (total_inactivos / total_general) * 100
                porcentaje_recursado = (total_recursado / total_general) * 100
                
                st.markdown(f"""
                <div style="background: #f8f9fa; padding: 15px; border-radius: 8px; margin: 10px 0;">
                    <p style="margin: 0; text-align: center;">
                        <strong>📈 Distribución:</strong> 
                        🟢 {porcentaje_activos:.1f}% Activos | 
                        🔴 {porcentaje_inactivos:.1f}% Inactivos | 
                        🔄 {porcentaje_recursado:.1f}% Han Recursado
                    </p>
                </div>
                """, unsafe_allow_html=True)
            
            # Tabs para mostrar detalles
            tab1, tab2, tab3 = st.tabs(["🟢 Alumnos Activos", "🔴 Alumnos Inactivos", "🔄 Han Recursado"])
            
            with tab1:
                st.markdown(f"### 🟢 Alumnos Activos: {total_activos}")
                st.info("💡 **Activos:** Alumnos con al menos una calificación ≥6, o en recursamiento, o que han recursado.")
                
                if not df_activos.empty:
                    # Preparar columnas para mostrar
                    columnas_mostrar = [
                        "matricula", "nombre", "grupo", "carrera", "cuatrimestre", 
                        "materia", "calificacion", "estatus", "tipo_asignacion", 
                        "fecha_ingreso_materia", "email_personal"
                    ]
                    columnas_existentes = [col for col in columnas_mostrar if col in df_activos.columns]
                    
                    st.dataframe(df_activos[columnas_existentes], use_container_width=True)
                    
                    # Botón de descarga para activos
                    import io
                    output_activos = io.BytesIO()
                    with pd.ExcelWriter(output_activos, engine='openpyxl') as writer:
                        df_activos[columnas_existentes].to_excel(writer, index=False, sheet_name='Alumnos_Activos')
                    output_activos.seek(0)
                    
                    st.download_button(
                        label="📥 Descargar Excel - Alumnos Activos",
                        data=output_activos,
                        file_name=f"alumnos_activos_{carrera_filtro}_{grupo_filtro}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.info("No hay alumnos activos con los filtros seleccionados.")
            
            with tab2:
                st.markdown(f"### 🔴 Alumnos Inactivos: {total_inactivos}")
                st.warning("⚠️ **Inactivos:** Alumnos que solo tienen reprobaciones y NO han recursado.")
                
                if not df_inactivos.empty:
                    # Preparar columnas para mostrar
                    columnas_mostrar = [
                        "matricula", "nombre", "grupo", "carrera", "cuatrimestre", 
                        "materia", "calificacion", "estatus", "tipo_asignacion", 
                        "fecha_ingreso_materia", "email_personal"
                    ]
                    columnas_existentes = [col for col in columnas_mostrar if col in df_inactivos.columns]
                    
                    st.dataframe(df_inactivos[columnas_existentes], use_container_width=True)
                    
                    # Botón de descarga para inactivos
                    import io
                    output_inactivos = io.BytesIO()
                    with pd.ExcelWriter(output_inactivos, engine='openpyxl') as writer:
                        df_inactivos[columnas_existentes].to_excel(writer, index=False, sheet_name='Alumnos_Inactivos')
                    output_inactivos.seek(0)
                    
                    st.download_button(
                        label="📥 Descargar Excel - Alumnos Inactivos",
                        data=output_inactivos,
                        file_name=f"alumnos_inactivos_{carrera_filtro}_{grupo_filtro}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.success("✅ No hay alumnos inactivos con los filtros seleccionados.")
            
            with tab3:
                st.markdown(f"### 🔄 Alumnos que han Recursado: {total_recursado}")
                st.info("📚 **Han Recursado:** Alumnos activos que tienen al menos un registro de recursamiento.")
                
                if not df_recursado.empty:
                    # Preparar columnas para mostrar
                    columnas_mostrar = [
                        "matricula", "nombre", "grupo", "carrera", "cuatrimestre", 
                        "materia", "calificacion", "estatus", "tipo_asignacion", 
                        "n_recursamientos", "fecha_ingreso_materia", "email_personal"
                    ]
                    columnas_existentes = [col for col in columnas_mostrar if col in df_recursado.columns]
                    
                    st.dataframe(df_recursado[columnas_existentes], use_container_width=True)
                    
                    # Botón de descarga para recursamiento
                    import io
                    output_recursado = io.BytesIO()
                    with pd.ExcelWriter(output_recursado, engine='openpyxl') as writer:
                        df_recursado[columnas_existentes].to_excel(writer, index=False, sheet_name='Alumnos_Recursado')
                    output_recursado.seek(0)
                    
                    st.download_button(
                        label="📥 Descargar Excel - Alumnos que han Recursado",
                        data=output_recursado,
                        file_name=f"alumnos_recursado_{carrera_filtro}_{grupo_filtro}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.info("No hay alumnos que hayan recursado con los filtros seleccionados.")
            
            # Resumen adicional
            st.markdown("---")
            st.markdown("### 📋 Definiciones:")
            st.markdown("""
            - **🟢 ACTIVO:** Alumno que tiene al menos:
              - Una calificación ≥ 6 (aprobado), O
              - Estatus = "recursando", O
              - Tipo asignación = "recursamiento"
            
            - **🔴 INACTIVO:** Alumno que solo tiene:
              - Calificaciones < 6 (reprobado) Y
              - Sin registros de recursamiento
            
            - **🔄 HAN RECURSADO:** Subconjunto de activos que tienen:
              - Al menos un registro con tipo_asignacion = "recursamiento" O
              - Estatus = "recursando"
            """)
            
# ========= SOLUCIÓN DEFINITIVA: EDITAR REGISTROS SIN VALUEERROR =========
elif menu == "✏️ Editar registros":
    st.subheader("✏️ Edición de registros existentes")
    
    df = cargar_datos_db()
    if df.empty:
        st.warning("No hay datos disponibles.")
    else:
        # Barra de búsqueda
        col1, col2 = st.columns([3, 1])
        with col1:
            busqueda = st.text_input("🔍 Buscar por matrícula, nombre o materia", 
                                   placeholder="Ejemplo: 0125070026 o González o Matemáticas",
                                   key="busqueda_edicion").strip().lower()
        with col2:
            buscar_exacto = st.checkbox("Búsqueda exacta", help="Buscar coincidencia exacta", key="edicion_exacto")

        if busqueda:
            # SOLUCIÓN AL VALUEERROR: Evitar completamente las comparaciones de Series
            try:
                # Crear lista de índices que coinciden
                indices_coincidentes = []
                
                for idx, row in df.iterrows():
                    try:
                        # Convertir valores a string de forma segura
                        matricula_str = str(row.get('matricula', '')).lower()
                        nombre_str = str(row.get('nombre', '')).lower()
                        materia_str = str(row.get('materia', '')).lower()
                        
                        # Limpiar valores NaN
                        if matricula_str in ['nan', 'none', 'null']:
                            matricula_str = ''
                        if nombre_str in ['nan', 'none', 'null']:
                            nombre_str = ''
                        if materia_str in ['nan', 'none', 'null']:
                            materia_str = ''
                        
                        # Aplicar lógica de búsqueda
                        coincide = False
                        if buscar_exacto:
                            if (matricula_str == busqueda or 
                                nombre_str == busqueda or 
                                materia_str == busqueda):
                                coincide = True
                        else:
                            if (busqueda in matricula_str or 
                                busqueda in nombre_str or 
                                busqueda in materia_str):
                                coincide = True
                        
                        if coincide:
                            indices_coincidentes.append(idx)
                    
                    except Exception as e:
                        # Si hay error en un registro, continúa con el siguiente
                        continue
                
                # Crear DataFrame filtrado usando los índices encontrados
                if indices_coincidentes:
                    df_filtrado = df.loc[indices_coincidentes].copy()
                else:
                    df_filtrado = pd.DataFrame()
                
            except Exception as e:
                st.error(f"Error en la búsqueda: {str(e)}")
                df_filtrado = pd.DataFrame()

            if not df_filtrado.empty:
                st.success(f"✅ Se encontraron {len(df_filtrado)} registros")
                
                # Mostrar registros encontrados
                columnas_mostrar = [
                    "matricula", "nombre", "grupo", "materia", "calificacion", 
                    "estatus", "profesor", "fecha_calificacion", "tipo_asignacion", "fecha_ingreso_materia"
                ]
                columnas_existentes = [col for col in columnas_mostrar if col in df_filtrado.columns]
                
                with st.expander("👀 Registros encontrados", expanded=True):
                    st.dataframe(df_filtrado[columnas_existentes], use_container_width=True)
                
                # Seleccionar registro para editar
                st.markdown("### 📝 Seleccionar registro para editar:")
                
                # Crear opciones para el selectbox de forma segura
                opciones_registro = []
                for idx, row in df_filtrado.iterrows():
                    try:
                        # Obtener valores de forma completamente segura
                        matricula = str(row.get('matricula', 'Sin matrícula'))
                        nombre = str(row.get('nombre', 'Sin nombre'))
                        materia = str(row.get('materia', 'Sin materia'))
                        calificacion = str(row.get('calificacion', 'Sin calificación'))
                        fecha = str(row.get('fecha_ingreso_materia', 'Sin fecha'))
                        
                        # Limpiar valores problemáticos
                        for var in [matricula, nombre, materia, calificacion, fecha]:
                            if var in ['None', 'nan', 'NaN', 'null']:
                                var = 'N/A'
                        
                        display_text = f"{matricula} - {nombre} - {materia} - Calif: {calificacion} - Fecha: {fecha}"
                        
                        opciones_registro.append({
                            'id': idx,
                            'display': display_text,
                            'data': row.to_dict()  # Convertir a dict para evitar problemas de Series
                        })
                    except Exception as e:
                        # Si hay error, crear una opción básica
                        opciones_registro.append({
                            'id': idx,
                            'display': f"Registro {idx} (Error al procesar)",
                            'data': {}
                        })
                
                if opciones_registro:
                    opcion_seleccionada = st.selectbox(
                        "Selecciona el registro a editar:",
                        opciones_registro,
                        format_func=lambda x: x['display'],
                        key="registro_editar_select"
                    )
                    
                    if opcion_seleccionada and opcion_seleccionada['data']:
                        registro_original = opcion_seleccionada['data']
                        registro_id = opcion_seleccionada['id']
                        
                        # Formulario de edición
                        st.markdown("### ✏️ Editar registro seleccionado:")
                        
                        # Funciones auxiliares más robustas
                        def safe_get_str(data, key, default=''):
                            try:
                                value = data.get(key, default)
                                if pd.isna(value) or value is None or str(value).lower() in ['nan', 'none', 'null']:
                                    return default
                                return str(value)
                            except:
                                return default
                        
                        def safe_get_float(data, key, default=0.0):
                            try:
                                value = data.get(key, default)
                                if pd.isna(value) or value is None or str(value).lower() in ['nan', 'none', 'null', '']:
                                    return default
                                return float(value)
                            except:
                                return default
                        
                        def safe_get_int(data, key, default=0):
                            try:
                                value = data.get(key, default)
                                if pd.isna(value) or value is None or str(value).lower() in ['nan', 'none', 'null', '']:
                                    return default
                                return int(float(value))  # Convertir via float para manejar decimales
                            except:
                                return default
                        
                        # Crear columnas para el formulario
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.markdown("**📚 Información Académica:**")
                            
                            # Obtener valores originales de forma segura
                            materia_original = safe_get_str(registro_original, 'materia')
                            calificacion_original = safe_get_float(registro_original, 'calificacion', 0.0)
                            estatus_original = safe_get_str(registro_original, 'estatus')
                            profesor_original = safe_get_str(registro_original, 'profesor')
                            
                            # Campos de entrada
                            nueva_materia = st.text_input("Materia", value=materia_original, key="edit_materia_input")
                            nueva_calificacion = st.number_input("Calificación", 
                                                               min_value=0.0, max_value=10.0, step=0.1,
                                                               value=calificacion_original,
                                                               key="edit_calificacion_input")
                            
                            # Estatus con índice seguro
                            estatus_options = ["aprobado", "reprobado", "recursando", ""]
                            estatus_index = 3  # Default
                            if estatus_original in estatus_options:
                                estatus_index = estatus_options.index(estatus_original)
                            
                            nuevo_estatus = st.selectbox("Estatus", estatus_options, 
                                                       index=estatus_index, key="edit_estatus_input")
                            nuevo_profesor = st.text_input("Profesor", value=profesor_original, key="edit_profesor_input")
                        
                        with col2:
                            st.markdown("**📋 Información Adicional:**")
                            
                            # Valores adicionales
                            grupo_original = safe_get_str(registro_original, 'grupo')
                            tipo_original = safe_get_str(registro_original, 'tipo_asignacion')
                            recursamientos_original = safe_get_int(registro_original, 'n_recursamientos', 0)
                            
                            nuevo_grupo = st.text_input("Grupo", value=grupo_original, key="edit_grupo_input")
                            
                            # Tipo con índice seguro
                            tipo_options = ["regular", "recursamiento", ""]
                            tipo_index = 2  # Default
                            if tipo_original in tipo_options:
                                tipo_index = tipo_options.index(tipo_original)
                            
                            nuevo_tipo = st.selectbox("Tipo de asignación", tipo_options, 
                                                     index=tipo_index, key="edit_tipo_input")
                            
                            # Fecha segura
                            fecha_value = None
                            try:
                                fecha_str = safe_get_str(registro_original, 'fecha_calificacion')
                                if fecha_str and fecha_str != '':
                                    fecha_value = pd.to_datetime(fecha_str).date()
                            except:
                                fecha_value = None
                            
                            nueva_fecha_calif = st.date_input("Fecha de calificación", 
                                                            value=fecha_value, key="edit_fecha_input")
                            nuevos_recursamientos = st.number_input("Número de recursamientos",
                                                                  min_value=0, max_value=10, step=1,
                                                                  value=recursamientos_original,
                                                                  key="edit_recursamientos_input")
                        
                        # Información NO editable
                        st.markdown("### 🔒 Información NO editable (por seguridad):")
                        col_info1, col_info2, col_info3 = st.columns(3)
                        with col_info1:
                            st.info(f"**Matrícula:** {safe_get_str(registro_original, 'matricula', 'N/A')}")
                            st.info(f"**Nombre:** {safe_get_str(registro_original, 'nombre', 'N/A')}")
                        with col_info2:
                            st.info(f"**Carrera:** {safe_get_str(registro_original, 'carrera', 'N/A')}")
                            st.info(f"**Email:** {safe_get_str(registro_original, 'usuario_email_nova', 'N/A')}")
                        with col_info3:
                            st.info(f"**Fecha ingreso materia:** {safe_get_str(registro_original, 'fecha_ingreso_materia', 'N/A')}")
                            st.info(f"**ID registro:** {registro_id}")
                        
                        # Detectar cambios de forma completamente segura
                        cambios_detectados = False
                        cambios_lista = []
                        
                        # Comparaciones directas sin usar pandas
                        if nueva_materia != materia_original:
                            cambios_detectados = True
                            cambios_lista.append(f"• **Materia:** '{materia_original}' → '{nueva_materia}'")
                        
                        if abs(nueva_calificacion - calificacion_original) > 0.001:
                            cambios_detectados = True
                            cambios_lista.append(f"• **Calificación:** {calificacion_original} → {nueva_calificacion}")
                        
                        if nuevo_estatus != estatus_original:
                            cambios_detectados = True
                            cambios_lista.append(f"• **Estatus:** '{estatus_original}' → '{nuevo_estatus}'")
                        
                        if nuevo_profesor != profesor_original:
                            cambios_detectados = True
                            cambios_lista.append(f"• **Profesor:** '{profesor_original}' → '{nuevo_profesor}'")
                        
                        if nuevo_grupo != grupo_original:
                            cambios_detectados = True
                            cambios_lista.append(f"• **Grupo:** '{grupo_original}' → '{nuevo_grupo}'")
                        
                        if nuevo_tipo != tipo_original:
                            cambios_detectados = True
                            cambios_lista.append(f"• **Tipo:** '{tipo_original}' → '{nuevo_tipo}'")
                        
                        # Comparar fechas
                        fecha_original_str = str(fecha_value) if fecha_value else 'Sin fecha'
                        fecha_nueva_str = str(nueva_fecha_calif) if nueva_fecha_calif else 'Sin fecha'
                        if fecha_original_str != fecha_nueva_str:
                            cambios_detectados = True
                            cambios_lista.append(f"• **Fecha:** {fecha_original_str} → {fecha_nueva_str}")
                        
                        if nuevos_recursamientos != recursamientos_original:
                            cambios_detectados = True
                            cambios_lista.append(f"• **Recursamientos:** {recursamientos_original} → {nuevos_recursamientos}")
                        
                        # Mostrar cambios
                        if cambios_detectados:
                            st.warning("⚠️ Se detectaron cambios en el registro")
                            
                            with st.expander("📋 Resumen de cambios", expanded=True):
                                for cambio in cambios_lista:
                                    st.write(cambio)
                            
                            # Confirmación
                            confirmar = st.checkbox("Confirmo que quiero guardar estos cambios", 
                                                   key="confirmar_cambios_checkbox")
                            
                            if confirmar and st.button("💾 GUARDAR CAMBIOS", type="primary", key="guardar_cambios_btn"):
                                try:
                                    # Respaldo
                                    if 'respaldo_edicion' not in st.session_state:
                                        st.session_state["respaldo_edicion"] = cargar_datos_db()
                                    
                                    # Actualizar BD
                                    conn = sqlite3.connect(DB_FILE)
                                    cursor = conn.cursor()
                                    
                                    # Preparar valores
                                    valores_update = {
                                        'materia': nueva_materia if nueva_materia.strip() else None,
                                        'calificacion': nueva_calificacion if nueva_calificacion > 0 else None,
                                        'estatus': nuevo_estatus if nuevo_estatus.strip() else None,
                                        'profesor': nuevo_profesor if nuevo_profesor.strip() else None,
                                        'grupo': nuevo_grupo if nuevo_grupo.strip() else None,
                                        'tipo_asignacion': nuevo_tipo if nuevo_tipo.strip() else None,
                                        'fecha_calificacion': nueva_fecha_calif.strftime('%Y-%m-%d') if nueva_fecha_calif else None,
                                        'n_recursamientos': nuevos_recursamientos
                                    }
                                    
                                    # Identificadores únicos
                                    matricula_id = safe_get_str(registro_original, 'matricula')
                                    fecha_ingreso_id = safe_get_str(registro_original, 'fecha_ingreso_materia')
                                    
                                    # UPDATE usando identificadores únicos
                                    cursor.execute("""
                                        UPDATE calificaciones 
                                        SET materia = ?, calificacion = ?, estatus = ?, profesor = ?, 
                                            grupo = ?, tipo_asignacion = ?, fecha_calificacion = ?, n_recursamientos = ?
                                        WHERE matricula = ? AND fecha_ingreso_materia = ?
                                    """, (
                                        valores_update['materia'],
                                        valores_update['calificacion'],
                                        valores_update['estatus'],
                                        valores_update['profesor'],
                                        valores_update['grupo'],
                                        valores_update['tipo_asignacion'],
                                        valores_update['fecha_calificacion'],
                                        valores_update['n_recursamientos'],
                                        matricula_id,
                                        fecha_ingreso_id
                                    ))
                                    
                                    if cursor.rowcount > 0:
                                        conn.commit()
                                        st.success("✅ Registro actualizado exitosamente")
                                        st.balloons()
                                    else:
                                        st.error("❌ No se encontró el registro para actualizar")
                                    
                                    conn.close()
                                    
                                except Exception as e:
                                    st.error(f"❌ Error al actualizar: {str(e)}")
                        else:
                            st.info("ℹ️ No se han detectado cambios en el registro")
                
            else:
                st.warning("❌ No se encontraron registros con esa búsqueda")
                st.info("💡 Intenta con términos más generales")
        
        else:
            st.info("🔍 Ingresa un término de búsqueda para encontrar registros")
            st.markdown("""
            ### 💡 **Consejos de búsqueda:**
            - **Por matrícula:** 0125070026
            - **Por nombre:** González o Patricia  
            - **Por materia:** Matemáticas o Programación
            """)

# Botón simple para deshacer (sin errores)
if menu == "✏️ Editar registros" and "respaldo_edicion" in st.session_state:
    st.markdown("---")
    if st.button("🔙 Deshacer último cambio", type="secondary", key="deshacer_simple"):
        try:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM calificaciones")
            for _, row in st.session_state["respaldo_edicion"].iterrows():
                valores = [row.get(col, None) for col in COLUMNAS_DB]
                cursor.execute(f"""
                    INSERT INTO calificaciones ({', '.join(COLUMNAS_DB)})
                    VALUES ({', '.join(['?'] * len(COLUMNAS_DB))})
                """, valores)
            conn.commit()
            conn.close()
            st.success("✅ Cambios revertidos")
            del st.session_state["respaldo_edicion"]
            st.rerun()
        except Exception as e:
            st.error(f"Error: {str(e)}")
            
elif menu == "🗑️ Eliminar registro":
    st.subheader("🗑️ Eliminar registro individual")
    
    df = cargar_datos_db()
    if df.empty:
        st.warning("No hay datos disponibles.")
    else:
        # Barra de búsqueda
        col1, col2 = st.columns([3, 1])
        with col1:
            busqueda_eliminar = st.text_input("🔍 Buscar por nombre, materia o grupo", 
                                            placeholder="Ejemplo: González, Matemáticas o G1A",
                                            key="busqueda_eliminar").strip().lower()
        with col2:
            buscar_exacto_eliminar = st.checkbox("Búsqueda exacta", help="Buscar coincidencia exacta", key="eliminar_exacto")

        if busqueda_eliminar:
            # Búsqueda segura similar a la sección de editar
            try:
                indices_coincidentes = []
                
                for idx, row in df.iterrows():
                    try:
                        # Convertir valores a string de forma segura
                        nombre_str = str(row.get('nombre', '')).lower()
                        materia_str = str(row.get('materia', '')).lower()
                        grupo_str = str(row.get('grupo', '')).lower()
                        
                        # Limpiar valores NaN
                        if nombre_str in ['nan', 'none', 'null']:
                            nombre_str = ''
                        if materia_str in ['nan', 'none', 'null']:
                            materia_str = ''
                        if grupo_str in ['nan', 'none', 'null']:
                            grupo_str = ''
                        
                        # Aplicar lógica de búsqueda
                        coincide = False
                        if buscar_exacto_eliminar:
                            if (nombre_str == busqueda_eliminar or 
                                materia_str == busqueda_eliminar or 
                                grupo_str == busqueda_eliminar):
                                coincide = True
                        else:
                            if (busqueda_eliminar in nombre_str or 
                                busqueda_eliminar in materia_str or 
                                busqueda_eliminar in grupo_str):
                                coincide = True
                        
                        if coincide:
                            indices_coincidentes.append(idx)
                    
                    except Exception as e:
                        continue
                
                # Crear DataFrame filtrado
                if indices_coincidentes:
                    df_filtrado = df.loc[indices_coincidentes].copy()
                else:
                    df_filtrado = pd.DataFrame()
                
            except Exception as e:
                st.error(f"Error en la búsqueda: {str(e)}")
                df_filtrado = pd.DataFrame()

            if not df_filtrado.empty:
                st.success(f"✅ Se encontraron {len(df_filtrado)} registros")
                
                # Mostrar registros encontrados
                columnas_mostrar = [
                    "matricula", "nombre", "grupo", "materia", "calificacion", 
                    "estatus", "profesor", "fecha_calificacion", "tipo_asignacion", "fecha_ingreso_materia"
                ]
                columnas_existentes = [col for col in columnas_mostrar if col in df_filtrado.columns]
                
                with st.expander("👀 Registros encontrados", expanded=True):
                    st.dataframe(df_filtrado[columnas_existentes], use_container_width=True)
                
                # Reemplaza desde "# Seleccionar registro para eliminar" hasta antes de "# Información sobre la funcionalidad"

                # Seleccionar registro para eliminar
                st.markdown("### 🎯 Seleccionar registro para eliminar:")
                st.warning("⚠️ Esta acción eliminará permanentemente el registro seleccionado")
                
                # Crear opciones más legibles para el selectbox
                opciones_eliminar = []
                for idx, row in df_filtrado.iterrows():
                    try:
                        # Función auxiliar para valores seguros
                        def safe_str(value, default='N/A'):
                            if pd.isna(value) or value is None or str(value).lower() in ['nan', 'none', 'null']:
                                return default
                            return str(value)
                        
                        matricula = safe_str(row.get('matricula'))
                        nombre = safe_str(row.get('nombre'))
                        materia = safe_str(row.get('materia'))
                        grupo = safe_str(row.get('grupo'))
                        calificacion = safe_str(row.get('calificacion'))
                        
                        # Formato más limpio y legible
                        display_text = f"{matricula} | {nombre} | {materia} | Grupo: {grupo} | Calif: {calificacion}"
                        
                        opciones_eliminar.append({
                            'id': idx,
                            'display': display_text,
                            'data': row.to_dict()
                        })
                    except Exception as e:
                        opciones_eliminar.append({
                            'id': idx,
                            'display': f"Registro {idx} (Error al procesar)",
                            'data': {}
                        })
                
                if opciones_eliminar:
                    # Selectbox con mejor formato
                    st.markdown("#### 📝 Registro a eliminar:")
                    registro_a_eliminar = st.selectbox(
                        "Elige el registro:",
                        opciones_eliminar,
                        format_func=lambda x: x['display'],
                        key="registro_eliminar_select",
                        help="Selecciona el registro específico que deseas eliminar"
                    )
                    
                    if registro_a_eliminar and registro_a_eliminar['data']:
                        registro_datos = registro_a_eliminar['data']
                        registro_idx = registro_a_eliminar['id']
                        
                        # Separador visual
                        st.markdown("---")
                        
                        # Mostrar información detallada del registro con mejor diseño
                        st.markdown("### 📋 Vista previa del registro a eliminar")
                        
                        # Usar un contenedor con borde
                        with st.container():
                            # Información básica destacada
                            col_main1, col_main2 = st.columns(2)
                            
                            with col_main1:
                                matricula = str(registro_datos.get('matricula', 'N/A'))
                                nombre = str(registro_datos.get('nombre', 'N/A'))
                                st.markdown(f"""
                                <div style="background: #f0f2f6; padding: 15px; border-radius: 10px; margin: 10px 0;">
                                    <h4 style="margin: 0; color: #1f77b4;">👤 ESTUDIANTE</h4>
                                    <p style="margin: 5px 0; font-size: 16px;"><strong>Matrícula:</strong> {matricula}</p>
                                    <p style="margin: 5px 0; font-size: 16px;"><strong>Nombre:</strong> {nombre}</p>
                                </div>
                                """, unsafe_allow_html=True)
                            
                            with col_main2:
                                materia = str(registro_datos.get('materia', 'N/A'))
                                grupo = str(registro_datos.get('grupo', 'N/A'))
                                calificacion = str(registro_datos.get('calificacion', 'N/A'))
                                st.markdown(f"""
                                <div style="background: #e8f5e8; padding: 15px; border-radius: 10px; margin: 10px 0;">
                                    <h4 style="margin: 0; color: #2ca02c;">📚 ACADÉMICO</h4>
                                    <p style="margin: 5px 0; font-size: 16px;"><strong>Materia:</strong> {materia}</p>
                                    <p style="margin: 5px 0; font-size: 16px;"><strong>Grupo:</strong> {grupo}</p>
                                    <p style="margin: 5px 0; font-size: 16px;"><strong>Calificación:</strong> {calificacion}</p>
                                </div>
                                """, unsafe_allow_html=True)
                        
                        # Detalles adicionales en tabs organizados
                        st.markdown("#### 📊 Información Detallada")
                        tab1, tab2, tab3 = st.tabs(["🏫 Información Académica", "📅 Fechas", "👨‍🏫 Otros Datos"])
                        
                        with tab1:
                            col1, col2 = st.columns(2)
                            with col1:
                                carrera = str(registro_datos.get('carrera', 'N/A'))
                                estatus = str(registro_datos.get('estatus', 'N/A'))
                                tipo = str(registro_datos.get('tipo_asignacion', 'N/A'))
                                st.write(f"**🎓 Carrera:** {carrera}")
                                st.write(f"**📊 Estatus:** {estatus}")
                                st.write(f"**📋 Tipo asignación:** {tipo}")
                            
                            with col2:
                                cuatrimestre = str(registro_datos.get('cuatrimestre', 'N/A'))
                                recursamientos = str(registro_datos.get('n_recursamientos', '0'))
                                origen = str(registro_datos.get('origen_asignacion', 'N/A'))
                                st.write(f"**📖 Cuatrimestre:** {cuatrimestre}")
                                st.write(f"**🔄 Recursamientos:** {recursamientos}")
                                st.write(f"**🏷️ Origen:** {origen}")
                        
                        with tab2:
                            col1, col2 = st.columns(2)
                            with col1:
                                fecha_ingreso_materia = str(registro_datos.get('fecha_ingreso_materia', 'N/A'))
                                fecha_ingreso_original = str(registro_datos.get('fecha_ingreso_original', 'N/A'))
                                st.write(f"**📅 Ingreso a materia:** {fecha_ingreso_materia}")
                                st.write(f"**📅 Ingreso original:** {fecha_ingreso_original}")
                            
                            with col2:
                                fecha_calificacion = str(registro_datos.get('fecha_calificacion', 'N/A'))
                                fecha_recursamiento = str(registro_datos.get('fecha_recursamiento', 'N/A'))
                                st.write(f"**📅 Calificación:** {fecha_calificacion}")
                                st.write(f"**📅 Recursamiento:** {fecha_recursamiento}")
                        
                        with tab3:
                            col1, col2 = st.columns(2)
                            with col1:
                                profesor = str(registro_datos.get('profesor', 'N/A'))
                                email_personal = str(registro_datos.get('email_personal', 'N/A'))
                                st.write(f"**👨‍🏫 Profesor:** {profesor}")
                                st.write(f"**📧 Email personal:** {email_personal}")
                            
                            with col2:
                                email_nova = str(registro_datos.get('usuario_email_nova', 'N/A'))
                                contraseña = str(registro_datos.get('contraseña', 'N/A'))
                                st.write(f"**📧 Email Nova:** {email_nova}")
                                st.write(f"**🔐 Contraseña:** {contraseña}")
                        
                        # Separador antes de la confirmación
                        st.markdown("---")
                        
                        # Área de confirmación con mejor diseño
                        st.markdown("### ⚠️ ZONA DE ELIMINACIÓN")
                        
                        # Contenedor de advertencia destacado
                        st.markdown("""
                        <div style="background: #ffebee; border: 2px solid #f44336; border-radius: 10px; padding: 20px; margin: 20px 0;">
                            <h3 style="color: #d32f2f; margin: 0 0 10px 0;">🚨 ADVERTENCIA CRÍTICA</h3>
                            <ul style="color: #d32f2f; margin: 0;">
                                <li><strong>Esta acción es IRREVERSIBLE</strong></li>
                                <li><strong>El registro se eliminará PERMANENTEMENTE</strong></li>
                                <li><strong>NO se puede recuperar después</strong></li>
                                <li><strong>La base de datos se actualizará inmediatamente</strong></li>
                            </ul>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        # Pasos de confirmación numerados
                        st.markdown("#### 📝 Pasos para confirmar la eliminación:")
                        
                        # Paso 1: Checkbox principal
                        st.markdown("**Paso 1:** Confirma que entiendes las consecuencias")
                        confirmar_eliminacion = st.checkbox(
                            f"☑️ Entiendo que eliminaré PERMANENTEMENTE el registro de '{nombre}' (Matrícula: {matricula})",
                            key="confirmar_eliminacion_individual",
                            help="Este checkbox confirma que entiendes que la acción es irreversible"
                        )
                        
                        # Paso 2: Campo de verificación (solo si paso 1 está confirmado)
                        if confirmar_eliminacion:
                            st.markdown("**Paso 2:** Verificación de seguridad")
                            col_verif1, col_verif2 = st.columns([2, 1])
                            
                            with col_verif1:
                                verificacion = st.text_input(
                                    "🔒 Escribe exactamente 'ELIMINAR' para continuar:",
                                    key="verificacion_eliminar",
                                    help="Medida de seguridad adicional para evitar eliminaciones accidentales",
                                    placeholder="Escribe: ELIMINAR"
                                )
                            
                            with col_verif2:
                                if verificacion == "ELIMINAR":
                                    st.success("✅ Verificación correcta")
                                elif verificacion and verificacion != "ELIMINAR":
                                    st.error("❌ Debe ser exactamente 'ELIMINAR'")
                            
                            # Paso 3: Botón final (solo si verificación es correcta)
                            if verificacion == "ELIMINAR":
                                st.markdown("**Paso 3:** Eliminación final")
                                
                                # Botón con estilo llamativo
                                col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
                                with col_btn2:
                                    if st.button(
                                        "🗑️ ELIMINAR REGISTRO DEFINITIVAMENTE", 
                                        type="primary", 
                                        key="eliminar_definitivo",
                                        help=f"Eliminar definitivamente el registro de {nombre}",
                                        use_container_width=True
                                    ):
                                        try:
                                            # Conectar a la base de datos
                                            conn = sqlite3.connect(DB_FILE)
                                            cursor = conn.cursor()
                                            
                                            # Preparar identificadores únicos para la eliminación
                                            matricula_id = str(registro_datos.get('matricula', ''))
                                            fecha_ingreso_id = str(registro_datos.get('fecha_ingreso_materia', ''))
                                            materia_id = str(registro_datos.get('materia', ''))
                                            
                                            # Eliminar registro específico
                                            cursor.execute("""
                                                DELETE FROM calificaciones 
                                                WHERE matricula = ? AND fecha_ingreso_materia = ? AND materia = ?
                                            """, (matricula_id, fecha_ingreso_id, materia_id))
                                            
                                            registros_eliminados = cursor.rowcount
                                            conn.commit()
                                            conn.close()
                                            
                                            if registros_eliminados > 0:
                                                # Mensaje de éxito con información detallada
                                                st.success("🎉 ¡Registro eliminado exitosamente!")
                                                
                                                # Información de lo eliminado en un contenedor verde
                                                st.markdown(f"""
                                                <div style="background: #e8f5e8; border: 2px solid #4caf50; border-radius: 10px; padding: 15px; margin: 10px 0;">
                                                    <h4 style="color: #2e7d32; margin: 0 0 10px 0;">✅ Registro Eliminado:</h4>
                                                    <p style="margin: 5px 0; color: #2e7d32;"><strong>Estudiante:</strong> {nombre}</p>
                                                    <p style="margin: 5px 0; color: #2e7d32;"><strong>Matrícula:</strong> {matricula}</p>
                                                    <p style="margin: 5px 0; color: #2e7d32;"><strong>Materia:</strong> {materia}</p>
                                                    <p style="margin: 5px 0; color: #2e7d32;"><strong>Registros eliminados:</strong> {registros_eliminados}</p>
                                                </div>
                                                """, unsafe_allow_html=True)
                                                
                                                st.balloons()
                                                
                                                # Botón para continuar
                                                if st.button("🔄 Continuar eliminando registros", key="continuar_eliminando"):
                                                    st.rerun()
                                            else:
                                                st.error("❌ No se pudo eliminar el registro. Puede que ya haya sido eliminado.")
                                            
                                        except Exception as e:
                                            st.error(f"❌ Error al eliminar el registro: {str(e)}")
        
        # Información sobre la funcionalidad
        st.markdown("---")
        st.markdown("### ℹ️ Información sobre esta función")
        st.info("""
        **🎯 ¿Para qué sirve esta función?**
        - Eliminar registros individuales específicos de la base de datos
        - Corregir errores de captura eliminando registros incorrectos
        - Limpiar datos duplicados o no deseados
        
        **🔒 Medidas de seguridad:**
        - Búsqueda y selección específica del registro
        - Vista previa completa antes de eliminar
        - Confirmación obligatoria con checkbox
        - Verificación adicional escribiendo 'ELIMINAR'
        - Eliminación permanente e irreversible
        
        **⚠️ Importante:**
        - Esta acción NO se puede deshacer
        - Solo elimina el registro específico seleccionado
        - Actualiza inmediatamente la base de datos
        """)            
            
elif menu == "🗑️ Eliminar recursamientos con calificación":
    st.subheader("🗑️ Eliminar recursamientos con calificación")

    df = pd.read_sql_query("SELECT * FROM calificaciones", sqlite3.connect(DB_FILE))
    df = df[df["tipo_asignacion"] == "recursamiento"]
    df = df[pd.to_numeric(df["calificacion"], errors="coerce").notna()]

    if df.empty:
        st.info("No hay recursamientos con calificación para mostrar.")
    else:
        carreras = ["Todas"] + sorted(df["carrera"].dropna().unique().tolist())
        materias = ["Todas"] + sorted(df["materia"].dropna().unique().tolist())
        grupos = ["Todos"] + sorted(df["grupo"].dropna().unique().tolist())
        fechas = ["Todas"] + sorted(df["fecha_ingreso_materia"].dropna().unique().tolist())

        col1, col2 = st.columns(2)
        carrera_sel = col1.selectbox("🎓 Filtrar por carrera", carreras)
        materia_sel = col2.selectbox("📘 Filtrar por materia", materias)
        col3, col4 = st.columns(2)
        grupo_sel = col3.selectbox("👥 Filtrar por grupo", grupos)
        fecha_sel = col4.selectbox("📅 Filtrar por fecha", fechas)

        if carrera_sel != "Todas":
            df = df[df["carrera"] == carrera_sel]
        if materia_sel != "Todas":
            df = df[df["materia"] == materia_sel]
        if grupo_sel != "Todos":
            df = df[df["grupo"] == grupo_sel]
        if fecha_sel != "Todas":
            df = df[df["fecha_ingreso_materia"] == fecha_sel]

        if df.empty:
            st.warning("No hay registros que coincidan con los filtros seleccionados.")
        else:
            st.write(f"Registros encontrados: {len(df)}")
            df = df.reset_index(drop=True)
            
            # Mostrar tabla de registros
            columnas_mostrar = [
                "matricula", "nombre", "grupo", "materia", "calificacion", 
                "fecha_ingreso_materia", "estatus", "profesor", "n_recursamientos"
            ]
            columnas_existentes = [col for col in columnas_mostrar if col in df.columns]
            
            st.markdown("### 📋 Recursamientos con calificación:")
            st.dataframe(df[columnas_existentes], use_container_width=True)
            
            # Selección múltiple para eliminar
            seleccionados = st.multiselect(
                "Selecciona los registros a eliminar:",
                options=df.index,
                format_func=lambda i: f"{df.at[i, 'matricula']} - {df.at[i, 'materia']} - Calif: {df.at[i, 'calificacion']} - {df.at[i, 'fecha_ingreso_materia']}"
            )

            if seleccionados:
                st.warning(f"⚠️ Se eliminarán {len(seleccionados)} registros de recursamiento con calificación.")
                
                # Mostrar registros que se van a eliminar
                with st.expander("👀 Registros que se eliminarán", expanded=False):
                    df_eliminar = df.loc[seleccionados]
                    st.dataframe(df_eliminar[columnas_existentes], use_container_width=True)
                
                # Confirmación
                confirmar_eliminacion = st.checkbox(
                    f"Confirmo que quiero eliminar definitivamente {len(seleccionados)} registros de recursamiento",
                    key="confirmar_eliminar_recursamientos"
                )
                
                if confirmar_eliminacion and st.button("❌ ELIMINAR REGISTROS SELECCIONADOS", type="primary"):
                    try:
                        conn = sqlite3.connect(DB_FILE)
                        cursor = conn.cursor()
                        eliminados = 0
                        
                        for i in seleccionados:
                            row = df.loc[i]
                            cursor.execute(
                                """
                                DELETE FROM calificaciones
                                WHERE matricula = ? AND materia = ? AND grupo = ? 
                                AND fecha_ingreso_materia = ? AND tipo_asignacion = 'recursamiento'
                                AND calificacion IS NOT NULL
                                """,
                                (row["matricula"], row["materia"], row["grupo"], row["fecha_ingreso_materia"])
                            )
                            eliminados += cursor.rowcount
                        
                        conn.commit()
                        conn.close()
                        
                        if eliminados > 0:
                            st.success(f"✅ {eliminados} registros eliminados exitosamente")
                            st.balloons()
                            st.rerun()
                        else:
                            st.error("❌ No se pudo eliminar ningún registro")
                            
                    except Exception as e:
                        st.error(f"❌ Error al eliminar registros: {str(e)}")
            
            # Información adicional
            st.markdown("---")
            st.markdown("### ℹ️ Información sobre esta función")
            st.info("""
            **📌 ¿Qué hace esta función?**
            - Muestra solo registros de recursamiento que YA tienen una calificación asignada
            - Permite eliminar estos registros de la base de datos
            - Útil para limpiar datos duplicados o corregir errores de captura
            
            **⚠️ Importante:**
            - Esta acción es irreversible
            - Solo se muestran recursamientos con calificación
            - Se requiere confirmación explícita antes de eliminar
            """)
    
    # Estadísticas generales
    st.markdown("---")
    st.markdown("### 📊 Estadísticas Generales")
    
    try:
        df_all = pd.read_sql_query("SELECT * FROM calificaciones", sqlite3.connect(DB_FILE))
        
        # Contar recursamientos
        total_recursamientos = len(df_all[df_all["tipo_asignacion"] == "recursamiento"])
        recursamientos_con_calif = len(df_all[
            (df_all["tipo_asignacion"] == "recursamiento") & 
            (pd.to_numeric(df_all["calificacion"], errors="coerce").notna())
        ])
        recursamientos_sin_calif = total_recursamientos - recursamientos_con_calif
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("📊 Total Recursamientos", total_recursamientos)
        with col2:
            st.metric("✅ Con Calificación", recursamientos_con_calif)
        with col3:
            st.metric("⏳ Sin Calificación", recursamientos_sin_calif)
            
    except Exception as e:
        st.error(f"Error al calcular estadísticas: {str(e)}")   
        
                 
# ========= VER INTEGRACIONES REALES - DATOS HISTÓRICOS =========
elif menu == "📋 Ver integraciones":
    st.subheader("📋 Ver Integraciones Realizadas")
    
    df = cargar_datos_db()
    if df.empty:
        st.warning("No hay datos disponibles.")
    else:
        # Header informativo
        st.markdown("""
        <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 20px; border-radius: 10px; margin: 20px 0;">
            <h3 style="color: white; margin: 0; text-align: center;">🔍 Consultar Integraciones Históricas</h3>
            <p style="color: #f8f9ff; margin: 5px 0 0 0; text-align: center;">Visualiza las integraciones reales que ya se realizaron</p>
        </div>
        """, unsafe_allow_html=True)
        
        # Filtros principales en 3 columnas
        col1, col2, col3 = st.columns(3)
        
        with col1:
            # Filtro por carrera
            carreras = sorted(df["carrera"].dropna().unique().tolist())
            carrera_consulta = st.selectbox("🎓 Filtrar por carrera", ["Todas"] + carreras, key="consulta_carrera")
        
        with col2:
            # Filtro por fechas de materia disponibles
            df["fecha_ingreso_materia"] = pd.to_datetime(df["fecha_ingreso_materia"], errors="coerce")
            
            # Filtrar por carrera si está seleccionada
            if carrera_consulta != "Todas":
                df_fechas = df[df["carrera"] == carrera_consulta]
            else:
                df_fechas = df
            
            fechas_materia = sorted(
                df_fechas["fecha_ingreso_materia"].dropna().dt.strftime("%Y-%m-%d").unique().tolist()
            )
            
            if fechas_materia:
                fecha_consulta = st.selectbox("📅 Selecciona fecha de integración", fechas_materia, key="fecha_consulta")
            else:
                st.warning("No hay fechas de materia disponibles.")
                fecha_consulta = None
        
        with col3:
            # Filtro por materia
            if fecha_consulta:
                fecha_consulta_dt = pd.to_datetime(fecha_consulta)
                
                # Filtrar datos por fecha y carrera para obtener materias
                if carrera_consulta != "Todas":
                    df_temp = df[
                        (df["fecha_ingreso_materia"] == fecha_consulta_dt) &
                        (df["carrera"] == carrera_consulta)
                    ]
                else:
                    df_temp = df[df["fecha_ingreso_materia"] == fecha_consulta_dt]
                
                materias_disponibles = sorted(df_temp["materia"].dropna().unique().tolist())
                
                if materias_disponibles:
                    materia_consulta = st.selectbox("📘 Filtrar por materia", ["Todas"] + materias_disponibles, key="materia_consulta")
                else:
                    st.warning("No hay materias disponibles")
                    materia_consulta = "Todas"
            else:
                materia_consulta = "Todas"
        
        if fecha_consulta:
            # Convertir fecha para filtrado
            fecha_consulta_dt = pd.to_datetime(fecha_consulta)
            
            # Filtrar datos por fecha, carrera y materia
            df_integracion = df[df["fecha_ingreso_materia"] == fecha_consulta_dt].copy()
            
            if carrera_consulta != "Todas":
                df_integracion = df_integracion[df_integracion["carrera"] == carrera_consulta]
            
            if materia_consulta != "Todas":
                df_integracion = df_integracion[df_integracion["materia"] == materia_consulta]
            
            if df_integracion.empty:
                st.warning(f"No se encontraron registros para los filtros seleccionados")
            else:
                # LÓGICA CORREGIDA: Identificar ALUMNOS REALES de la integración histórica
                
                # 1. NUEVOS INGRESOS REALES: Alumnos que se registraron por primera vez en esa fecha
                # (Su fecha_ingreso_original es igual a fecha_ingreso_materia)
                nuevos_ingresos_reales = df_integracion[
                    pd.to_datetime(df_integracion["fecha_ingreso_original"]) == pd.to_datetime(df_integracion["fecha_ingreso_materia"])
                ].copy()
                
                # 2. RECURSAMIENTO REAL: Alumnos que se integraron desde recursamiento
                # (Su fecha_ingreso_original es ANTERIOR a fecha_ingreso_materia)
                recursamiento_real = df_integracion[
                    pd.to_datetime(df_integracion["fecha_ingreso_original"]) < pd.to_datetime(df_integracion["fecha_ingreso_materia"])
                ].copy()
                
                # También incluir explícitamente los marcados como recursamiento
                recursamiento_explicito = df_integracion[
                    (df_integracion["tipo_asignacion"] == "recursamiento") |
                    (df_integracion["estatus"] == "recursando")
                ].copy()
                
                # Combinar recursamiento real con explícito (sin duplicados)
                if not recursamiento_explicito.empty:
                    recursamiento_real = pd.concat([recursamiento_real, recursamiento_explicito]).drop_duplicates(subset=['matricula'])
                
                # Información de filtros aplicados
                filtros_info = []
                if carrera_consulta != "Todas":
                    filtros_info.append(f"Carrera: {carrera_consulta}")
                if materia_consulta != "Todas":
                    filtros_info.append(f"Materia: {materia_consulta}")
                
                if filtros_info:
                    st.info(f"🔍 **Filtros aplicados:** {' | '.join(filtros_info)}")
                
                # Mostrar estadísticas de la integración real
                st.markdown(f"### 📊 Integración Histórica del {fecha_consulta}")
                
                col_stats1, col_stats2, col_stats3, col_stats4 = st.columns(4)
                with col_stats1:
                    st.metric("👥 Total Registros", len(df_integracion))
                with col_stats2:
                    st.metric("🆕 Nuevos Ingresos", len(nuevos_ingresos_reales))
                with col_stats3:
                    st.metric("🔁 Recursamiento", len(recursamiento_real))
                with col_stats4:
                    grupos_diferentes = df_integracion['grupo'].nunique()
                    st.metric("📚 Grupos", grupos_diferentes)
                
                # Verificar si hubo integración real
                hubo_integracion = len(nuevos_ingresos_reales) > 0 and len(recursamiento_real) > 0
                
                if hubo_integracion:
                    # MOSTRAR LA INTEGRACIÓN REAL COMO EN GENERACIÓN DE MATERIA
                    
                    # Determinar el grupo de integración (el grupo final donde quedaron)
                    grupos_finales = df_integracion['grupo'].value_counts()
                    grupo_principal = grupos_finales.index[0]  # El grupo con más alumnos
                    
                    st.success(f"✅ Se detectó integración: Nuevos Ingresos + Recursamiento")
                    st.info(f"🎯 **Grupo de integración utilizado:** {grupo_principal}")
                    
                    # Detectar cambios de grupo en recursamiento
                    cambios_grupo = 0
                    if not recursamiento_real.empty:
                        # Para cada alumno de recursamiento, verificar si cambió de grupo
                        for _, alumno in recursamiento_real.iterrows():
                            matricula = alumno.get('matricula', '')
                            grupo_actual = alumno.get('grupo', '')
                            
                            # Buscar registros previos del mismo alumno
                            registros_previos = df[
                                (df['matricula'] == matricula) &
                                (df['fecha_ingreso_materia'] < fecha_consulta_dt)
                            ]
                            
                            if not registros_previos.empty:
                                grupo_anterior = registros_previos.iloc[-1].get('grupo', '')
                                if grupo_anterior != grupo_actual and grupo_anterior:
                                    cambios_grupo += 1
                    
                    # VISTA PREVIA EXACTA COMO EN GENERACIÓN DE MATERIA
                    with st.expander("👀 Vista previa de integración con cambios de grupo", expanded=True):
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.markdown("**🆕 Nuevos Ingresos (mantienen su grupo):**")
                            # Mostrar solo matrícula y nombre
                            if not nuevos_ingresos_reales.empty:
                                nuevos_display = nuevos_ingresos_reales[["matricula", "nombre"]].copy()
                                st.dataframe(nuevos_display, use_container_width=True)
                            else:
                                st.info("No hay nuevos ingresos en esta integración")
                        
                        with col2:
                            st.markdown(f"**🔁 Recursamiento (cambiaron a grupo {grupo_principal}):**")
                            # Mostrar solo matrícula y nombre
                            if not recursamiento_real.empty:
                                recursamiento_display = recursamiento_real[["matricula", "nombre"]].copy()
                                st.dataframe(recursamiento_display, use_container_width=True)
                                
                                # Mostrar métrica de cambios de grupo
                                st.metric("🔄 Cambios de grupo", cambios_grupo)
                            else:
                                st.info("No hay recursamiento en esta integración")
                    
                    # Información adicional detallada
                    st.markdown("### 📋 Detalles de la Integración Realizada")
                    
                    col_det1, col_det2 = st.columns(2)
                    
                    with col_det1:
                        st.markdown("**🆕 Información de Nuevos Ingresos:**")
                        if not nuevos_ingresos_reales.empty:
                            fecha_ingreso_original = nuevos_ingresos_reales["fecha_ingreso_original"].iloc[0]
                            grupos_nuevos = nuevos_ingresos_reales['grupo'].value_counts()
                            
                            st.write(f"• **Fecha de ingreso:** {fecha_ingreso_original}")
                            st.write(f"• **Total:** {len(nuevos_ingresos_reales)} alumnos")
                            st.write("• **Distribución por grupo:**")
                            for grupo, count in grupos_nuevos.items():
                                st.write(f"  - {grupo}: {count} alumnos")
                        else:
                            st.write("No hay nuevos ingresos en esta fecha")
                    
                    with col_det2:
                        st.markdown("**🔁 Información de Recursamiento:**")
                        if not recursamiento_real.empty:
                            materias_recursamiento = recursamiento_real["materia"].value_counts()
                            grupos_finales_rec = recursamiento_real['grupo'].value_counts()
                            
                            st.write(f"• **Total:** {len(recursamiento_real)} alumnos")
                            st.write("• **Materias recursadas:**")
                            for materia, count in materias_recursamiento.items():
                                st.write(f"  - {materia}: {count} alumnos")
                            st.write("• **Grupos finales:**")
                            for grupo, count in grupos_finales_rec.items():
                                st.write(f"  - {grupo}: {count} alumnos")
                            if cambios_grupo > 0:
                                st.write(f"• **Cambios de grupo realizados:** {cambios_grupo}")
                        else:
                            st.write("No hay recursamiento en esta fecha")
                
                else:
                    # Mostrar información cuando no hubo integración mixta
                    if len(nuevos_ingresos_reales) > 0 and len(recursamiento_real) == 0:
                        st.info("📋 **Tipo:** Solo Nuevos Ingresos (sin integración)")
                        
                        with st.expander("👀 Vista de Nuevos Ingresos", expanded=True):
                            st.markdown("**🆕 Nuevos Ingresos:**")
                            if not nuevos_ingresos_reales.empty:
                                nuevos_display = nuevos_ingresos_reales[["matricula", "nombre", "grupo"]].copy()
                                st.dataframe(nuevos_display, use_container_width=True)
                                
                                # Estadísticas
                                grupos_stats = nuevos_ingresos_reales['grupo'].value_counts()
                                st.markdown("**📊 Estadísticas:**")
                                for grupo, count in grupos_stats.items():
                                    st.write(f"• Grupo {grupo}: {count} alumnos")
                    
                    elif len(nuevos_ingresos_reales) == 0 and len(recursamiento_real) > 0:
                        st.info("📋 **Tipo:** Solo Recursamiento (sin integración)")
                        
                        with st.expander("👀 Vista de Recursamiento", expanded=True):
                            st.markdown("**🔁 Recursamiento:**")
                            if not recursamiento_real.empty:
                                recursamiento_display = recursamiento_real[["matricula", "nombre", "grupo", "materia"]].copy()
                                st.dataframe(recursamiento_display, use_container_width=True)
                                
                                # Estadísticas
                                materias_stats = recursamiento_real['materia'].value_counts()
                                st.markdown("**📊 Estadísticas:**")
                                for materia, count in materias_stats.items():
                                    st.write(f"• {materia}: {count} alumnos")
                    
                    else:
                        st.warning("No se encontraron registros válidos para mostrar.")
                
                # Opciones de descarga
                st.markdown("---")
                st.markdown("### 📥 Descargar Integración Histórica")
                
                col_desc1, col_desc2 = st.columns(2)
                
                with col_desc1:
                    # Descarga completa
                    df_descarga = df_integracion.copy()
                    
                    # Agregar columna indicando el tipo real
                    def determinar_tipo_real(row):
                        fecha_orig = pd.to_datetime(row['fecha_ingreso_original'])
                        fecha_mat = pd.to_datetime(row['fecha_ingreso_materia'])
                        
                        if fecha_orig == fecha_mat:
                            return "Nuevo_Ingreso"
                        elif fecha_orig < fecha_mat:
                            return "Recursamiento"
                        else:
                            return "Otros"
                    
                    df_descarga["tipo_real"] = df_descarga.apply(determinar_tipo_real, axis=1)
                    
                    import io
                    output_completo = io.BytesIO()
                    with pd.ExcelWriter(output_completo, engine='openpyxl') as writer:
                        df_descarga.to_excel(writer, index=False, sheet_name='Integracion_Historica')
                    output_completo.seek(0)
                    
                    # Nombre de archivo
                    nombre_archivo = f"integracion_historica_{fecha_consulta}"
                    if carrera_consulta != "Todas":
                        nombre_archivo += f"_{carrera_consulta.replace(' ', '_')}"
                    if materia_consulta != "Todas":
                        nombre_archivo += f"_{materia_consulta.replace(' ', '_')}"
                    nombre_archivo += ".xlsx"
                    
                    st.download_button(
                        label="📥 Descargar Integración Histórica",
                        data=output_completo,
                        file_name=nombre_archivo,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
                with col_desc2:
                    # Información sobre la descarga
                    st.markdown("**📁 El archivo incluye:**")
                    st.write(f"• {len(df_integracion)} registros totales")
                    if len(nuevos_ingresos_reales) > 0:
                        st.write(f"• {len(nuevos_ingresos_reales)} nuevos ingresos reales")
                    if len(recursamiento_real) > 0:
                        st.write(f"• {len(recursamiento_real)} recursamiento real")
                    st.write("• Columna 'tipo_real' identificando el tipo")
                    st.write("• Toda la información académica histórica")
        
        # Información sobre cómo funciona
        st.markdown("---")
        st.markdown("### ℹ️ Cómo Funciona esta Consulta")
        
        col_info1, col_info2 = st.columns(2)
        
        with col_info1:
            st.markdown("**🔍 Detección de Nuevos Ingresos:**")
            st.write("• Alumnos cuya fecha_ingreso_original = fecha_ingreso_materia")
            st.write("• Representa su primera vez en el sistema")
            st.write("• Se integraron directamente en esa fecha")
        
        with col_info2:
            st.markdown("**🔍 Detección de Recursamiento:**")
            st.write("• Alumnos cuya fecha_ingreso_original < fecha_ingreso_materia")
            st.write("• Ya estaban en el sistema anteriormente")
            st.write("• Se integraron desde recursamiento en esa fecha")
        
        st.info("""
        **📋 Esta sección muestra integraciones REALES que ya ocurrieron:**
        - Identifica automáticamente qué alumnos eran nuevos vs recursamiento
        - Muestra la integración exacta como se realizó históricamente
        - Detecta cambios de grupo que se hicieron en el proceso
        - Permite consultar cualquier integración pasada por fecha y filtros
        """)
        
elif menu == "🔍 Análisis Aprobados vs Recursadores":
    analisis_aprobados_vs_recursadores() 

# ================================================================================================
# PASO 2: AGREGAR EN EL MENÚ PRINCIPAL - CÓDIGO COMPLETO
# ================================================================================================

elif menu == "👨‍🏫 Tutorías":
    st.header("👨‍🏫 Sistema de Tutorías")
    
    # Inicializar tablas de tutorías si no existen
    inicializar_tablas_tutoria()
    
    # Submenu de tutorías
    submenu_tutoria = st.selectbox("Selecciona una opción", [
        "📋 Vista General",
        "👨‍🏫 Gestión de Tutores",
        "📚 Asignación Tutor-Grupo", 
        "🎯 Asignaciones Manuales",
        "📊 Reportes de Tutorías"
    ])
    
    # ========================================================================================
    # VISTA GENERAL
    # ========================================================================================
    if submenu_tutoria == "📋 Vista General":
        st.subheader("📋 Vista General del Sistema de Tutorías")
        
        conn = sqlite3.connect(DB_FILE)
        
        # Estadísticas generales
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            total_tutores = pd.read_sql_query("SELECT COUNT(*) as total FROM tutores WHERE activo = 1", conn).iloc[0]['total']
            st.metric("👨‍🏫 Tutores Activos", total_tutores)
        
        with col2:
            total_asignaciones = pd.read_sql_query("SELECT COUNT(*) as total FROM asignaciones_tutoria WHERE activo = 1", conn).iloc[0]['total']
            st.metric("📚 Grupos Asignados", total_asignaciones)
        
        with col3:
            total_manuales = pd.read_sql_query("SELECT COUNT(*) as total FROM asignaciones_manuales WHERE activo = 1", conn).iloc[0]['total']
            st.metric("🎯 Asignaciones Manuales", total_manuales)
        
        with col4:
            total_alumnos_tutoria = pd.read_sql_query("""
                SELECT COUNT(DISTINCT matricula) as total FROM (
                    SELECT DISTINCT c.matricula
                    FROM calificaciones c
                    INNER JOIN asignaciones_tutoria at ON c.grupo = at.grupo AND c.carrera = at.carrera
                    WHERE at.activo = 1 AND c.grupo IS NOT NULL
                    
                    UNION
                    
                    SELECT DISTINCT matricula
                    FROM asignaciones_manuales
                    WHERE activo = 1
                )
            """, conn).iloc[0]['total']
            st.metric("👥 Alumnos con Tutor", total_alumnos_tutoria)
        
        st.markdown("---")
        
        # Lista de tutores con sus asignaciones
        st.subheader("📋 Resumen por Tutor")
        
        tutores_info = pd.read_sql_query("""
            SELECT 
                t.id,
                t.nombre_tutor,
                t.email_tutor,
                COUNT(DISTINCT at.grupo || '-' || at.carrera) as grupos_asignados,
                COUNT(DISTINCT am.matricula) as asignaciones_manuales
            FROM tutores t
            LEFT JOIN asignaciones_tutoria at ON t.id = at.tutor_id AND at.activo = 1
            LEFT JOIN asignaciones_manuales am ON t.id = am.tutor_id AND am.activo = 1
            WHERE t.activo = 1
            GROUP BY t.id, t.nombre_tutor, t.email_tutor
            ORDER BY t.nombre_tutor
        """, conn)
        
        if not tutores_info.empty:
            for _, tutor in tutores_info.iterrows():
                with st.expander(f"👨‍🏫 {tutor['nombre_tutor']} - {tutor['grupos_asignados']} grupos, {tutor['asignaciones_manuales']} manuales"):
                    
                    # Mostrar alumnos del tutor
                    alumnos_tutor = obtener_alumnos_por_tutor(tutor['id'])
                    
                    if alumnos_tutor:
                        st.write("**Alumnos asignados:**")
                        df_alumnos = pd.DataFrame(alumnos_tutor, columns=['Matrícula', 'Nombre', 'Grupo', 'Carrera', 'Tipo'])
                        st.dataframe(df_alumnos, use_container_width=True)
                    else:
                        st.info("No tiene alumnos asignados actualmente")
        else:
            st.info("No hay tutores registrados en el sistema")
        
        conn.close()
    
    # ========================================================================================
    # GESTIÓN DE TUTORES
    # ========================================================================================
    elif submenu_tutoria == "👨‍🏫 Gestión de Tutores":
        st.subheader("👨‍🏫 Gestión de Tutores")
        
        tab1, tab2 = st.tabs(["➕ Agregar Tutor", "📝 Gestionar Tutores"])
        
        with tab1:
            st.subheader("➕ Agregar Nuevo Tutor")
            
            with st.form("form_nuevo_tutor"):
                col1, col2 = st.columns(2)
                with col1:
                    nombre_tutor = st.text_input("Nombre completo del tutor *", placeholder="Ej: Dr. Juan Pérez")
                    email_tutor = st.text_input("Email", placeholder="juan.perez@universidad.edu")
                with col2:
                    telefono_tutor = st.text_input("Teléfono", placeholder="555-123-4567")
                
                observaciones = st.text_area("Observaciones", placeholder="Información adicional del tutor...")
                
                submitted = st.form_submit_button("➕ Agregar Tutor", type="primary")
                
                if submitted:
                    if nombre_tutor.strip():
                        try:
                            conn = sqlite3.connect(DB_FILE)
                            cursor = conn.cursor()
                            cursor.execute("""
                                INSERT INTO tutores (nombre_tutor, email_tutor, telefono, observaciones)
                                VALUES (?, ?, ?, ?)
                            """, (nombre_tutor.strip(), email_tutor.strip(), telefono_tutor.strip(), observaciones.strip()))
                            conn.commit()
                            conn.close()
                            
                            # NOTIFICACIÓN MEJORADA
                            st.success(f"✅ ¡Tutor agregado exitosamente!")
                            st.balloons()  # Animación de celebración
                            st.info(f"📝 **Nombre:** {nombre_tutor}\n📧 **Email:** {email_tutor or 'No especificado'}")
                            time.sleep(1)  # Pausa para mostrar mensaje
                            st.rerun()
                        except sqlite3.IntegrityError:
                            st.error("❌ Ya existe un tutor con ese nombre")
                        except Exception as e:
                            st.error(f"❌ Error al agregar tutor: {e}")
                    else:
                        st.error("❌ El nombre del tutor es obligatorio")
        
        with tab2:
            st.subheader("📝 Gestionar Tutores Existentes")
            
            conn = sqlite3.connect(DB_FILE)
            tutores_df = pd.read_sql_query("""
                SELECT id, nombre_tutor, email_tutor, telefono, fecha_registro, observaciones, activo
                FROM tutores
                ORDER BY nombre_tutor
            """, conn)
            conn.close()
            
            if not tutores_df.empty:
                for _, tutor in tutores_df.iterrows():
                    estado_icon = "✅" if tutor['activo'] else "❌"
                    
                    with st.expander(f"{estado_icon} {tutor['nombre_tutor']} (ID: {tutor['id']})"):
                        col1, col2, col3 = st.columns([2, 2, 1])
                        
                        with col1:
                            st.write(f"**📧 Email:** {tutor['email_tutor'] or 'No especificado'}")
                            st.write(f"**📱 Teléfono:** {tutor['telefono'] or 'No especificado'}")
                        
                        with col2:
                            st.write(f"**📅 Registro:** {tutor['fecha_registro']}")
                            st.write(f"**💼 Estado:** {'Activo' if tutor['activo'] else 'Inactivo'}")
                        
                        with col3:
                            # Botón de activar/desactivar
                            nuevo_estado = not tutor['activo']
                            texto_boton = "🔄 Activar" if nuevo_estado else "⏸️ Desactivar"
                            
                            if st.button(texto_boton, key=f"toggle_{tutor['id']}"):
                                conn = sqlite3.connect(DB_FILE)
                                cursor = conn.cursor()
                                cursor.execute("UPDATE tutores SET activo = ? WHERE id = ?", (nuevo_estado, tutor['id']))
                                conn.commit()
                                conn.close()
                                
                                # NOTIFICACIÓN MEJORADA
                                accion = "activado" if nuevo_estado else "desactivado"
                                st.success(f"✅ Tutor **{accion}** correctamente")
                                st.info(f"👨‍🏫 {tutor['nombre_tutor']} ha sido {accion}")
                                time.sleep(1)
                                st.rerun()
                        
                        if tutor['observaciones']:
                            st.write(f"**📝 Observaciones:** {tutor['observaciones']}")
            else:
                st.info("No hay tutores registrados en el sistema")
    
    # ========================================================================================
    # ASIGNACIÓN TUTOR-GRUPO
    # ========================================================================================
    elif submenu_tutoria == "📚 Asignación Tutor-Grupo":
        st.subheader("📚 Asignación Tutor-Grupo")
        
        tab1, tab2 = st.tabs(["➕ Nueva Asignación", "📋 Gestionar Asignaciones"])
        
        with tab1:
            st.subheader("➕ Asignar Tutor a Grupo")
            
            tutores = obtener_tutores_activos()
            grupos = obtener_grupos_disponibles()
            
            if not tutores:
                st.warning("⚠️ No hay tutores activos. Primero agrega tutores en la sección 'Gestión de Tutores'")
            elif not grupos:
                st.warning("⚠️ No hay grupos disponibles en el sistema")
            else:
                with st.form("form_asignar_grupo"):
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        tutor_options = {f"{t[1]} (ID: {t[0]})": t[0] for t in tutores}
                        tutor_selected = st.selectbox("Seleccionar Tutor", list(tutor_options.keys()))
                        tutor_id = tutor_options[tutor_selected]
                    
                    with col2:
                        grupo_options = {f"Grupo {g[0]} - {g[1]}": (g[0], g[1]) for g in grupos}
                        grupo_selected = st.selectbox("Seleccionar Grupo", list(grupo_options.keys()))
                        grupo_num, carrera = grupo_options[grupo_selected]
                    
                    # Verificar conflictos
                    if verificar_conflicto_asignacion(grupo_num, carrera):
                        st.error("⚠️ Este grupo ya tiene un tutor asignado. Puedes reasignar desde 'Gestionar Asignaciones'")
                        disabled = True
                    else:
                        st.success("✅ Grupo disponible para asignación")
                        disabled = False
                    
                    submitted = st.form_submit_button("📚 Asignar", disabled=disabled, type="primary")
                    
                    if submitted and not disabled:
                        try:
                            conn = sqlite3.connect(DB_FILE)
                            cursor = conn.cursor()
                            cursor.execute("""
                                INSERT INTO asignaciones_tutoria (tutor_id, grupo, carrera)
                                VALUES (?, ?, ?)
                            """, (tutor_id, grupo_num, carrera))
                            conn.commit()
                            conn.close()
                            
                            # NOTIFICACIÓN MEJORADA
                            st.success(f"✅ ¡Asignación realizada exitosamente!")
                            st.balloons()  # Animación de celebración
                            st.info(f"👨‍🏫 **Tutor:** {tutor_selected.split(' (')[0]}\n📚 **Grupo:** {grupo_selected}")
                            
                            # Mostrar cuántos alumnos tiene ahora el tutor
                            alumnos_count = len(obtener_alumnos_por_tutor(tutor_id))
                            st.info(f"👥 El tutor ahora tiene **{alumnos_count} alumnos** asignados")
                            
                            time.sleep(2)  # Pausa para mostrar mensaje
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Error al asignar: {e}")
        
        with tab2:
            st.subheader("📋 Gestionar Asignaciones Existentes")
            
            conn = sqlite3.connect(DB_FILE)
            asignaciones_df = pd.read_sql_query("""
                SELECT 
                    at.id,
                    t.nombre_tutor,
                    at.grupo,
                    at.carrera,
                    at.fecha_asignacion,
                    at.activo,
                    COUNT(DISTINCT c.matricula) as total_alumnos
                FROM asignaciones_tutoria at
                JOIN tutores t ON at.tutor_id = t.id
                LEFT JOIN calificaciones c ON at.grupo = c.grupo AND at.carrera = c.carrera
                GROUP BY at.id, t.nombre_tutor, at.grupo, at.carrera, at.fecha_asignacion, at.activo
                ORDER BY t.nombre_tutor, at.carrera, at.grupo
            """, conn)
            conn.close()
            
            if not asignaciones_df.empty:
                for _, asignacion in asignaciones_df.iterrows():
                    estado_icon = "✅" if asignacion['activo'] else "❌"
                    
                    with st.expander(f"{estado_icon} {asignacion['nombre_tutor']} → Grupo {asignacion['grupo']} ({asignacion['carrera']}) - {asignacion['total_alumnos']} alumnos"):
                        col1, col2 = st.columns([3, 1])
                        
                        with col1:
                            st.write(f"**📅 Fecha asignación:** {asignacion['fecha_asignacion']}")
                            st.write(f"**👥 Total alumnos:** {asignacion['total_alumnos']}")
                            st.write(f"**💼 Estado:** {'Activa' if asignacion['activo'] else 'Inactiva'}")
                        
                        with col2:
                            # Botón de activar/desactivar
                            nuevo_estado = not asignacion['activo']
                            texto_boton = "🔄 Activar" if nuevo_estado else "⏸️ Desactivar"
                            
                            if st.button(texto_boton, key=f"toggle_asig_{asignacion['id']}"):
                                conn = sqlite3.connect(DB_FILE)
                                cursor = conn.cursor()
                                cursor.execute("UPDATE asignaciones_tutoria SET activo = ? WHERE id = ?", (nuevo_estado, asignacion['id']))
                                conn.commit()
                                conn.close()
                                st.rerun()
            else:
                st.info("No hay asignaciones registradas")
    
    # ========================================================================================
    # ASIGNACIONES MANUALES
    # ========================================================================================
    elif submenu_tutoria == "🎯 Asignaciones Manuales":
        st.subheader("🎯 Asignaciones Manuales")
        st.info("💡 Esta sección permite asignar tutores a alumnos específicos que no tienen grupo o casos especiales")
        
        tab1, tab2 = st.tabs(["➕ Nueva Asignación Manual", "📋 Gestionar Asignaciones"])
        
        with tab1:
            st.subheader("➕ Asignación Manual de Alumno")
            
            tutores = obtener_tutores_activos()
            alumnos_sin_grupo = obtener_alumnos_sin_grupo()
            
            # También mostrar todos los alumnos para casos especiales
            conn = sqlite3.connect(DB_FILE)
            todos_alumnos = pd.read_sql_query("""
                SELECT DISTINCT matricula, nombre, carrera, grupo
                FROM calificaciones
                ORDER BY nombre
            """, conn).values.tolist()
            conn.close()
            
            if not tutores:
                st.warning("⚠️ No hay tutores activos disponibles")
            else:
                with st.form("form_asignacion_manual"):
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        tutor_options = {f"{t[1]} (ID: {t[0]})": t[0] for t in tutores}
                        tutor_selected = st.selectbox("Seleccionar Tutor", list(tutor_options.keys()))
                        tutor_id = tutor_options[tutor_selected]
                    
                    with col2:
                        # Opción para elegir tipo de alumno
                        tipo_alumno = st.radio("Tipo de alumno", ["Sin grupo", "Todos los alumnos"])
                    
                    if tipo_alumno == "Sin grupo":
                        if alumnos_sin_grupo:
                            alumno_options = {f"{a[0]} - {a[1]} ({a[2]})": a[0] for a in alumnos_sin_grupo}
                            alumno_selected = st.selectbox("Seleccionar Alumno Sin Grupo", list(alumno_options.keys()))
                            matricula = alumno_options[alumno_selected]
                        else:
                            st.info("✅ No hay alumnos sin grupo")
                            matricula = None
                    else:
                        alumno_options = {f"{a[0]} - {a[1]} ({a[2]}) - Grupo: {a[3] or 'Sin grupo'}": a[0] for a in todos_alumnos}
                        alumno_selected = st.selectbox("Seleccionar Cualquier Alumno", list(alumno_options.keys()))
                        matricula = alumno_options[alumno_selected]
                    
                    observaciones = st.text_area("Observaciones", placeholder="Motivo de la asignación manual...")
                    
                    # Verificar si ya existe asignación
                    if matricula:
                        conn = sqlite3.connect(DB_FILE)
                        cursor = conn.cursor()
                        cursor.execute("SELECT id FROM asignaciones_manuales WHERE matricula = ? AND activo = 1", (matricula,))
                        existe = cursor.fetchone()
                        conn.close()
                        
                        if existe:
                            st.warning("⚠️ Este alumno ya tiene una asignación manual activa")
                            disabled = True
                        else:
                            disabled = False
                    else:
                        disabled = True
                    
                    submitted = st.form_submit_button("🎯 Asignar Manualmente", disabled=disabled, type="primary")
                    
                    if submitted and not disabled:
                        try:
                            conn = sqlite3.connect(DB_FILE)
                            cursor = conn.cursor()
                            cursor.execute("""
                                INSERT INTO asignaciones_manuales (matricula, tutor_id, observaciones)
                                VALUES (?, ?, ?)
                            """, (matricula, tutor_id, observaciones.strip()))
                            conn.commit()
                            conn.close()
                            
                            # NOTIFICACIÓN MEJORADA
                            st.success(f"✅ ¡Asignación manual realizada exitosamente!")
                            st.balloons()  # Animación de celebración
                            st.info(f"👨‍🏫 **Tutor:** {tutor_selected.split(' (')[0]}\n👤 **Alumno:** {alumno_selected}")
                            
                            # Mostrar total de alumnos del tutor
                            alumnos_count = len(obtener_alumnos_por_tutor(tutor_id))
                            st.info(f"👥 El tutor ahora tiene **{alumnos_count} alumnos** asignados")
                            
                            time.sleep(2)
                            st.rerun()
                        except Exception as e:
                            st.error(f"❌ Error al realizar asignación: {e}")
        
        with tab2:
            st.subheader("📋 Gestionar Asignaciones Manuales")
            
            conn = sqlite3.connect(DB_FILE)
            manuales_df = pd.read_sql_query("""
                SELECT 
                    am.id,
                    am.matricula,
                    c.nombre,
                    c.carrera,
                    c.grupo,
                    t.nombre_tutor,
                    am.fecha_asignacion,
                    am.observaciones,
                    am.activo
                FROM asignaciones_manuales am
                JOIN tutores t ON am.tutor_id = t.id
                LEFT JOIN calificaciones c ON am.matricula = c.matricula
                GROUP BY am.id, am.matricula, c.nombre, c.carrera, c.grupo, t.nombre_tutor, am.fecha_asignacion, am.observaciones, am.activo
                ORDER BY t.nombre_tutor, c.nombre
            """, conn)
            conn.close()
            
            if not manuales_df.empty:
                for _, manual in manuales_df.iterrows():
                    estado_icon = "✅" if manual['activo'] else "❌"
                    grupo_info = f"Grupo {manual['grupo']}" if manual['grupo'] else "Sin grupo"
                    
                    with st.expander(f"{estado_icon} {manual['nombre_tutor']} → {manual['nombre']} ({manual['matricula']}) - {grupo_info}"):
                        col1, col2 = st.columns([3, 1])
                        
                        with col1:
                            st.write(f"**🎓 Carrera:** {manual['carrera']}")
                            st.write(f"**📅 Fecha asignación:** {manual['fecha_asignacion']}")
                            if manual['observaciones']:
                                st.write(f"**📝 Observaciones:** {manual['observaciones']}")
                            st.write(f"**💼 Estado:** {'Activa' if manual['activo'] else 'Inactiva'}")
                        
                        with col2:
                            # Botón de activar/desactivar
                            nuevo_estado = not manual['activo']
                            texto_boton = "🔄 Activar" if nuevo_estado else "⏸️ Desactivar"
                            
                            if st.button(texto_boton, key=f"toggle_manual_{manual['id']}"):
                                conn = sqlite3.connect(DB_FILE)
                                cursor = conn.cursor()
                                cursor.execute("UPDATE asignaciones_manuales SET activo = ? WHERE id = ?", (nuevo_estado, manual['id']))
                                conn.commit()
                                conn.close()
                                st.rerun()
            else:
                st.info("No hay asignaciones manuales registradas")
    
    # ========================================================================================
    # REPORTES
    # ========================================================================================
    elif submenu_tutoria == "📊 Reportes de Tutorías":
        st.subheader("📊 Reportes de Tutorías")
        
        conn = sqlite3.connect(DB_FILE)
        
        # Filtros
        col1, col2 = st.columns(2)
        with col1:
            tutores_reporte = pd.read_sql_query("SELECT id, nombre_tutor FROM tutores WHERE activo = 1 ORDER BY nombre_tutor", conn)
            tutor_filtro = st.selectbox("Filtrar por tutor", ["Todos"] + tutores_reporte['nombre_tutor'].tolist())
        
        with col2:
            carreras_reporte = pd.read_sql_query("SELECT DISTINCT carrera FROM calificaciones ORDER BY carrera", conn)
            carrera_filtro = st.selectbox("Filtrar por carrera", ["Todas"] + carreras_reporte['carrera'].tolist())
    
        
        # Reporte completo
        query_reporte_sin_duplicados = """
            WITH asignaciones_completas AS (
                SELECT DISTINCT
                    t.nombre_tutor,
                    'Grupo' as tipo_asignacion,
                    at.carrera,
                    at.grupo,
                    c.matricula,
                    c.nombre as nombre_alumno,
                    at.fecha_asignacion,
                    1 as prioridad
                FROM asignaciones_tutoria at
                JOIN tutores t ON at.tutor_id = t.id
                JOIN calificaciones c ON at.grupo = c.grupo AND at.carrera = c.carrera
                WHERE at.activo = 1
                
                UNION ALL
                
                SELECT DISTINCT
                    t.nombre_tutor,
                    'Manual' as tipo_asignacion,
                    c.carrera,
                    COALESCE(c.grupo, 'Sin grupo') as grupo,
                    am.matricula,
                    c.nombre as nombre_alumno,
                    am.fecha_asignacion,
                    2 as prioridad
                FROM asignaciones_manuales am
                JOIN tutores t ON am.tutor_id = t.id
                JOIN calificaciones c ON am.matricula = c.matricula
                WHERE am.activo = 1
            )
            SELECT DISTINCT
                nombre_tutor,
                tipo_asignacion,
                carrera,
                grupo,
                matricula,
                nombre_alumno,
                fecha_asignacion
            FROM (
                SELECT *,
                    ROW_NUMBER() OVER (PARTITION BY matricula ORDER BY prioridad) as rn
                FROM asignaciones_completas
            ) ranked
            WHERE rn = 1
            ORDER BY nombre_tutor, carrera, grupo, nombre_alumno
        """
        
        # Reporte completo SIN DUPLICADOS
        reporte_df = pd.read_sql_query(query_reporte_sin_duplicados, conn)
        
        # Aplicar filtros
        filtros_aplicados = []
        if tutor_filtro != "Todos":
            reporte_df = reporte_df[reporte_df['nombre_tutor'] == tutor_filtro]
            filtros_aplicados.append(f"Tutor: {tutor_filtro}")
        
        if carrera_filtro != "Todas":
            reporte_df = reporte_df[reporte_df['carrera'] == carrera_filtro]
            filtros_aplicados.append(f"Carrera: {carrera_filtro}")
        
        filtros_texto = ", ".join(filtros_aplicados) if filtros_aplicados else "Sin filtros"
        
        if not reporte_df.empty:
            st.subheader(f"📋 Reporte de Asignaciones ({len(reporte_df)} registros únicos)")
            
            # Mostrar estadísticas del filtro
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("👥 Total Alumnos", len(reporte_df['matricula'].unique()))
            with col2:
                st.metric("👨‍🏫 Tutores Involucrados", len(reporte_df['nombre_tutor'].unique()))
            with col3:
                st.metric("🎓 Carreras", len(reporte_df['carrera'].unique()))
            
            # Alerta sobre duplicados eliminados
            st.info("ℹ️ Este reporte muestra cada alumno una sola vez. Prioridad: Asignación por grupo > Asignación manual")
            
            # Mostrar tabla completa
            st.dataframe(
                reporte_df[['nombre_tutor', 'tipo_asignacion', 'carrera', 'grupo', 'matricula', 'nombre_alumno', 'fecha_asignacion']],
                column_config={
                    'nombre_tutor': 'Tutor',
                    'tipo_asignacion': 'Tipo',
                    'carrera': 'Carrera',
                    'grupo': 'Grupo',
                    'matricula': 'Matrícula',
                    'nombre_alumno': 'Alumno',
                    'fecha_asignacion': 'Fecha Asignación'
                },
                use_container_width=True
            )
            
            # Botones de descarga
            col1, col2 = st.columns(2)
            
            with col1:
                # Botón CSV
                csv = reporte_df.to_csv(index=False)
                st.download_button(
                    label="📥 Descargar CSV",
                    data=csv,
                    file_name=f"reporte_tutorias_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    mime="text/csv"
                )
            
            with col2:
                # Botón Excel
                excel_data = exportar_reporte_excel(reporte_df, filtros_texto)
                if excel_data:
                    st.download_button(
                        label="📊 Descargar Excel",
                        data=excel_data,
                        file_name=f"reporte_tutorias_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            
            # Gráficos de resumen
            st.markdown("---")
            st.subheader("📈 Análisis Visual")
            
            col1, col2 = st.columns(2)
            
            with col1:
                # Gráfico de alumnos por tutor
                alumnos_por_tutor = reporte_df.groupby('nombre_tutor').size().reset_index(name='cantidad_alumnos')
                
                fig_tutores = px.bar(
                    alumnos_por_tutor, 
                    x='nombre_tutor', 
                    y='cantidad_alumnos',
                    title="Alumnos por Tutor (Sin Duplicados)",
                    labels={'nombre_tutor': 'Tutor', 'cantidad_alumnos': 'Cantidad de Alumnos'}
                )
                fig_tutores.update_xaxes(tickangle=45)
                st.plotly_chart(fig_tutores, use_container_width=True)
            
            with col2:
                # Gráfico de tipo de asignaciones
                tipo_asignaciones = reporte_df['tipo_asignacion'].value_counts().reset_index()
                
                fig_tipos = px.pie(
                    tipo_asignaciones, 
                    values='count', 
                    names='tipo_asignacion',
                    title="Distribución por Tipo de Asignación"
                )
                st.plotly_chart(fig_tipos, use_container_width=True)
        
        else:
            st.info("No hay datos que coincidan con los filtros seleccionados")
        
        conn.close()        
            
elif menu == "📊 Resumen por grupo":
    st.subheader("📊 Resumen académico por grupo")

    df = cargar_datos_db()
    if df.empty:
        st.warning("No hay datos disponibles.")
    else:
        # Filtros principales
        col1, col2 = st.columns(2)
        
        with col1:
            carreras = sorted(df["carrera"].dropna().unique().tolist())
            carrera_sel = st.selectbox("🎓 Selecciona la carrera", carreras, key="resumen_carrera")

        with col2:
            grupos = sorted(df[df["carrera"] == carrera_sel]["grupo"].dropna().unique().tolist())
            grupo_sel = st.selectbox("👥 Selecciona el grupo", grupos, key="resumen_grupo")

        # Filtros adicionales
        col3, col4 = st.columns(2)
        
        with col3:
            materias = sorted(df[(df["carrera"] == carrera_sel) & (df["grupo"] == grupo_sel)]["materia"].dropna().unique().tolist())
            materia_sel = st.selectbox("📘 Selecciona la materia", materias, key="resumen_materia")

        with col4:
            fechas = sorted(df[(df["carrera"] == carrera_sel) & (df["grupo"] == grupo_sel) & (df["materia"] == materia_sel)]["fecha_ingreso_materia"].dropna().unique().tolist())
            fecha_sel = st.selectbox("📅 Selecciona la fecha de ingreso", fechas, key="resumen_fecha")

        # Filtrar datos
        df_filtrado = df[
            (df["carrera"] == carrera_sel) &
            (df["grupo"] == grupo_sel) &
            (df["materia"] == materia_sel) &
            (df["fecha_ingreso_materia"] == fecha_sel)
        ]
        
        # Remover duplicados - mantener el registro más reciente por alumno
        df_filtrado = df_filtrado.sort_values(by="fecha_calificacion", ascending=False)
        df_filtrado = df_filtrado.drop_duplicates(subset=["matricula"], keep="first").copy()

        if df_filtrado.empty:
            st.warning("No hay alumnos con esos filtros.")
        else:
            # Convertir calificaciones a numérico
            df_filtrado["calificacion"] = pd.to_numeric(df_filtrado["calificacion"], errors="coerce")

            # Calcular estadísticas
            total = len(df_filtrado)
            aprobados = len(df_filtrado[df_filtrado["calificacion"] >= 6])
            reprobados_total = len(df_filtrado[df_filtrado["calificacion"] < 6])
            reprobados_regular = len(df_filtrado[(df_filtrado["calificacion"] < 6) & (df_filtrado["tipo_asignacion"] == "regular")])
            reprobados_recursamiento = len(df_filtrado[(df_filtrado["calificacion"] < 6) & (df_filtrado["tipo_asignacion"] == "recursamiento")])
            recursando = len(df_filtrado[(df_filtrado["tipo_asignacion"] == "recursamiento") & (df_filtrado["calificacion"].isna())])
            pendientes = len(df_filtrado[df_filtrado["calificacion"].isna()])
            promedio = df_filtrado["calificacion"].mean()

            # Mostrar métricas en un diseño atractivo
            st.markdown("""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 20px; border-radius: 10px; margin: 20px 0;">
                <h3 style="color: white; margin: 0; text-align: center;">📊 ESTADÍSTICAS DEL GRUPO</h3>
                <p style="color: #f8f9ff; margin: 5px 0 0 0; text-align: center;">Resumen académico completo</p>
            </div>
            """, unsafe_allow_html=True)

            # Primera fila de métricas
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("👥 Total alumnos", total)
            with col2:
                st.metric("✅ Aprobados", aprobados)
            with col3:
                st.metric("❌ Reprobados (total)", reprobados_total)

            # Segunda fila de métricas
            col4, col5, col6 = st.columns(3)
            with col4:
                st.metric("📘 Reprobados (regular)", reprobados_regular)
            with col5:
                st.metric("🔁 Reprobados (recursamiento)", reprobados_recursamiento)
            with col6:
                st.metric("🕓 Recursando", recursando)

            # Promedio general
            col7, col8, col9 = st.columns(3)
            with col8:  # Centrar en la columna del medio
                if not pd.isna(promedio):
                    # Determinar color según el promedio
                    if promedio >= 8:
                        color = "#27ae60"  # Verde
                    elif promedio >= 7:
                        color = "#f39c12"  # Amarillo
                    else:
                        color = "#e74c3c"  # Rojo
                    
                    st.markdown(f"""
                    <div style="background: {color}; padding: 15px; border-radius: 10px; text-align: center; margin: 10px 0;">
                        <h3 style="color: white; margin: 0;">📊 Promedio General</h3>
                        <h2 style="color: white; margin: 5px 0 0 0;">{promedio:.2f}</h2>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.metric("📊 Promedio general", "N/A")

            # Calcular porcentajes
            if total > 0:
                porcentaje_aprobados = (aprobados / total) * 100
                porcentaje_reprobados = (reprobados_total / total) * 100
                porcentaje_pendientes = (pendientes / total) * 100
                
                st.markdown(f"""
                <div style="background: #f8f9fa; padding: 15px; border-radius: 8px; margin: 20px 0;">
                    <h4 style="margin: 0; text-align: center; color: #333;">📈 Distribución Porcentual</h4>
                    <p style="margin: 10px 0 0 0; text-align: center;">
                        ✅ {porcentaje_aprobados:.1f}% Aprobados | 
                        ❌ {porcentaje_reprobados:.1f}% Reprobados | 
                        🕓 {porcentaje_pendientes:.1f}% Pendientes
                    </p>
                </div>
                """, unsafe_allow_html=True)

            # Tabla de todos los alumnos
            st.markdown("### 📋 Tabla Completa del Grupo")
            st.info("💡 Incluye todos los alumnos: aprobados, reprobados, recursantes y pendientes")

            columnas_mostrar = [
                "matricula", "nombre", "grupo", "materia", "fecha_ingreso_materia",
                "calificacion", "estatus", "tipo_asignacion", "n_recursamientos", "profesor"
            ]
            
            # Verificar qué columnas existen
            columnas_existentes = [col for col in columnas_mostrar if col in df_filtrado.columns]
            
            # Ordenar por calificación (descendente) y luego por nombre
            df_mostrar = df_filtrado.copy()
            df_mostrar = df_mostrar.sort_values(['calificacion', 'nombre'], ascending=[False, True], na_position='last')
            
            st.dataframe(df_mostrar[columnas_existentes], use_container_width=True)

            # Análisis adicional
            st.markdown("### 📈 Análisis Detallado")
            
            # Tabs para diferentes análisis
            tab1, tab2, tab3 = st.tabs(["🎯 Por Rendimiento", "📊 Por Tipo", "👨‍🏫 Por Profesor"])
            
            with tab1:
                st.markdown("#### 🎯 Clasificación por Rendimiento")
                
                if not df_filtrado["calificacion"].isna().all():
                    # Rangos de calificación
                    excelente = len(df_filtrado[df_filtrado["calificacion"] >= 9])
                    muy_bueno = len(df_filtrado[(df_filtrado["calificacion"] >= 8) & (df_filtrado["calificacion"] < 9)])
                    bueno = len(df_filtrado[(df_filtrado["calificacion"] >= 7) & (df_filtrado["calificacion"] < 8)])
                    suficiente = len(df_filtrado[(df_filtrado["calificacion"] >= 6) & (df_filtrado["calificacion"] < 7)])
                    insuficiente = len(df_filtrado[df_filtrado["calificacion"] < 6])
                    
                    col_r1, col_r2, col_r3, col_r4, col_r5 = st.columns(5)
                    with col_r1:
                        st.metric("🌟 Excelente (9-10)", excelente)
                    with col_r2:
                        st.metric("⭐ Muy Bueno (8-8.9)", muy_bueno)
                    with col_r3:
                        st.metric("👍 Bueno (7-7.9)", bueno)
                    with col_r4:
                        st.metric("✔️ Suficiente (6-6.9)", suficiente)
                    with col_r5:
                        st.metric("❌ Insuficiente (<6)", insuficiente)
                else:
                    st.info("No hay calificaciones registradas para análisis de rendimiento.")
            
            with tab2:
                st.markdown("#### 📊 Clasificación por Tipo de Asignación")
                
                # Contar por tipo de asignación
                if "tipo_asignacion" in df_filtrado.columns:
                    tipos_count = df_filtrado["tipo_asignacion"].value_counts()
                    
                    for tipo, count in tipos_count.items():
                        tipo_emoji = "🔁" if tipo == "recursamiento" else "📚"
                        st.write(f"{tipo_emoji} **{tipo.title()}:** {count} alumnos")
                else:
                    st.info("No hay información de tipo de asignación disponible.")
            
            with tab3:
                st.markdown("#### 👨‍🏫 Rendimiento por Profesor")
                
                if "profesor" in df_filtrado.columns and not df_filtrado["profesor"].isna().all():
                    profesores_stats = df_filtrado.groupby("profesor").agg({
                        "calificacion": ["count", "mean"],
                        "matricula": "count"
                    }).round(2)
                    
                    st.write("📊 **Estadísticas por profesor:**")
                    for profesor in df_filtrado["profesor"].dropna().unique():
                        alumnos_prof = df_filtrado[df_filtrado["profesor"] == profesor]
                        total_prof = len(alumnos_prof)
                        promedio_prof = alumnos_prof["calificacion"].mean()
                        aprobados_prof = len(alumnos_prof[alumnos_prof["calificacion"] >= 6])
                        
                        if not pd.isna(promedio_prof):
                            st.write(f"• **{profesor}:** {total_prof} alumnos | Promedio: {promedio_prof:.2f} | Aprobados: {aprobados_prof}")
                        else:
                            st.write(f"• **{profesor}:** {total_prof} alumnos | Sin calificaciones")
                else:
                    st.info("No hay información de profesores disponible.")

            # Botón de descarga
            st.markdown("### 📥 Descargar Reporte")
            
            # Preparar datos para Excel
            df_export = df_mostrar[columnas_existentes].copy()
            
            import io
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_export.to_excel(writer, index=False, sheet_name='Resumen_Grupo')
            output.seek(0)

            nombre_archivo = f"resumen_{carrera_sel}_{grupo_sel}_{materia_sel}_{fecha_sel}.xlsx"
            nombre_archivo = nombre_archivo.replace(" ", "_").replace("/", "-")

            st.download_button(
                label="📥 Descargar Excel del grupo completo",
                data=output,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
       
            

            # Información adicional del grupo
            st.markdown("### ℹ️ Información del Grupo")
            st.info(f"""
            **📊 Resumen del grupo seleccionado:**
            - **🎓 Carrera:** {carrera_sel}
            - **👥 Grupo:** {grupo_sel}
            - **📘 Materia:** {materia_sel}
            - **📅 Fecha de ingreso:** {fecha_sel}
            - **👨‍👩‍👧‍👦 Total de alumnos:** {total}
            - **📈 Tasa de aprobación:** {(aprobados/total*100):.1f}% ({aprobados}/{total})
            """)


          
